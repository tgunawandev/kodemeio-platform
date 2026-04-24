"""Report management commands.

Includes both Odoo built-in ir.actions.report commands (list, render, templates,
generate) and custom report_management framework commands (types, run, instances,
view, export, download).
"""

from __future__ import annotations

import base64
import contextlib
from datetime import date, timedelta
from pathlib import Path
from typing import Annotated

import typer

from kctl_odoo.core.callbacks import AppContext
from kctl_odoo.core.exceptions import RPCError

app = typer.Typer(help="Manage Odoo reports and generate diagnostic reports.")

# Backward-compatible aliases for old short codes
CODE_ALIASES: dict[str, str] = {
    "pnl": "profit_loss",
    "bs": "balance_sheet",
    "cf": "cash_flow",
    "tb": "trial_balance",
    "cma": "cost_management",
    "ar_aging": "receivable_aging",
    "ap_aging": "payable_aging",
    "cashflow_recon": "cashflow_reconciliation",
    "inventory_fg": "inventory_finished_goods",
    "inventory_rm": "inventory_raw_materials",
    "mrp_oee": "mrp_equipment_effectiveness",
    "mrp_wip": "mrp_work_in_progress",
    "logistics_otd": "logistics_on_time_delivery",
    "logistics_otif": "logistics_on_time_in_full",
}


def _resolve_type_code(type_code: str) -> str:
    """Resolve a report type code, applying backward-compatible aliases."""
    return CODE_ALIASES.get(type_code, type_code)


# Report types → diagnose section mappings
REPORT_TYPES: dict[str, list[str] | None] = {
    "health": ["server_health", "configuration", "mail", "storage", "dr"],
    "finance": ["financial_health", "data_integrity", "kpis"],
    "security": ["security"],
    "inventory": ["inventory_health", "data_integrity", "kpis"],
    "operations": ["kpis", "mail", "data_integrity"],
    "full": None,  # all sections
}


@app.command("list")
def list_(
    ctx: typer.Context,
    model: Annotated[str | None, typer.Option("--model", "-m", help="Filter by model name")] = None,
    limit: Annotated[int, typer.Option("--limit", "-l", help="Max results")] = 50,
) -> None:
    """List available reports."""
    actx: AppContext = ctx.obj
    out = actx.output
    c = actx.client

    domain: list = []
    if model:
        domain.append(("model", "=", model))

    reports = c.search_read(
        "ir.actions.report",
        domain=domain,
        fields=["id", "name", "report_name", "model", "report_type", "print_report_name", "binding_model_id"],
        limit=limit,
        order="model, name",
    )

    if not reports:
        out.info("No reports found.")
        if actx.json_mode:
            out.raw_json([])
        return

    rows = []
    json_data = []
    for r in reports:
        report_type = r.get("report_type", "")
        type_display = {
            "qweb-pdf": "[green]PDF[/green]",
            "qweb-html": "[cyan]HTML[/cyan]",
            "qweb-text": "[dim]Text[/dim]",
        }.get(report_type, report_type)

        binding = r.get("binding_model_id")
        binding_display = binding[1] if isinstance(binding, list) else ""

        rows.append(
            [
                str(r["id"]),
                (r.get("name") or "")[:35],
                r.get("report_name", ""),
                r.get("model", ""),
                type_display,
                (r.get("print_report_name") or "")[:30],
            ]
        )
        json_data.append(
            {
                "id": r["id"],
                "name": r.get("name"),
                "report_name": r.get("report_name"),
                "model": r.get("model"),
                "report_type": report_type,
                "print_report_name": r.get("print_report_name"),
                "binding_model": binding_display or None,
            }
        )

    out.table(
        f"Reports ({len(reports)})",
        [
            ("ID", "cyan"),
            ("Name", ""),
            ("Report Name", ""),
            ("Model", ""),
            ("Type", ""),
            ("Print Name", "dim"),
        ],
        rows,
        data_for_json=json_data,
    )


@app.command()
def render(
    ctx: typer.Context,
    report_name: Annotated[str, typer.Argument(help="Report technical name (e.g. account.report_invoice)")],
    record_ids: Annotated[str, typer.Argument(help="Comma-separated record IDs")],
    fmt: Annotated[str, typer.Option("--format", "-f", help="Output format: pdf or html")] = "pdf",
    output: Annotated[str | None, typer.Option("--output", "-o", help="Output file path")] = None,
) -> None:
    """Render a report for given record IDs and save to file."""
    actx: AppContext = ctx.obj
    out = actx.output
    c = actx.client

    if fmt not in ("pdf", "html"):
        out.error("Format must be 'pdf' or 'html'.")
        raise typer.Exit(1)

    ids = [int(i.strip()) for i in record_ids.split(",")]

    # Find the report action to get the model
    reports = c.search_read(
        "ir.actions.report",
        domain=[("report_name", "=", report_name)],
        fields=["id", "name", "model", "report_type"],
    )
    if not reports:
        out.error(f"Report not found: {report_name}")
        raise typer.Exit(1)

    report_info = reports[0]

    # Use _render first (Odoo 17+), fall back to specific method for older versions
    fallback_method = "render_qweb_pdf" if fmt == "pdf" else "render_qweb_html"
    try:
        result = c.execute_kw(
            "ir.actions.report",
            "_render",
            [[report_info["id"]], report_name, ids],
        )
    except RPCError:
        # Fallback: try the specific render method on the report model
        try:
            result = c.execute_kw(
                "ir.actions.report",
                fallback_method,
                [[report_info["id"]], ids],
            )
        except RPCError as e2:
            out.error(f"Failed to render report: {e2.detail}")
            raise typer.Exit(1) from e2

    # Result is typically [base64_content, report_type] or just base64 content
    if isinstance(result, list) and len(result) >= 1:
        content = result[0]
    elif isinstance(result, str):
        content = result
    else:
        out.error("Unexpected response format from report rendering.")
        raise typer.Exit(1)

    # Decode if base64 encoded
    if isinstance(content, str):
        try:
            file_data = base64.b64decode(content)
        except Exception:
            # Might be raw content
            file_data = content.encode("utf-8")
    elif isinstance(content, bytes):
        file_data = content
    else:
        out.error("Could not decode report content.")
        raise typer.Exit(1)

    ext = "pdf" if fmt == "pdf" else "html"
    if output:
        output_path = Path(output)
    else:
        safe_name = report_name.replace(".", "_")
        output_path = Path(f"{safe_name}_{'-'.join(str(i) for i in ids)}.{ext}")

    output_path.write_bytes(file_data)
    out.success(f"Report saved to {output_path} ({len(file_data)} bytes)")
    if actx.json_mode:
        out.raw_json(
            {
                "report_name": report_name,
                "record_ids": ids,
                "format": fmt,
                "file": str(output_path),
                "size": len(file_data),
            }
        )


@app.command()
def templates(
    ctx: typer.Context,
    model: Annotated[str | None, typer.Option("--model", "-m", help="Filter by related model")] = None,
    limit: Annotated[int, typer.Option("--limit", "-l", help="Max results")] = 50,
) -> None:
    """List QWeb report templates."""
    actx: AppContext = ctx.obj
    out = actx.output
    c = actx.client

    domain: list = [("type", "=", "qweb")]
    if model:
        domain.append(("model", "=", model))

    views = c.search_read(
        "ir.ui.view",
        domain=domain,
        fields=["id", "name", "key", "model", "active", "priority"],
        limit=limit,
        order="model, name",
    )

    if not views:
        out.info("No QWeb templates found.")
        if actx.json_mode:
            out.raw_json([])
        return

    rows = []
    json_data = []
    for v in views:
        status = "[green]active[/green]" if v.get("active") else "[dim]inactive[/dim]"
        rows.append(
            [
                str(v["id"]),
                (v.get("name") or "")[:35],
                v.get("key", "") or "-",
                v.get("model", "") or "-",
                str(v.get("priority", "")),
                status,
            ]
        )
        json_data.append(
            {
                "id": v["id"],
                "name": v.get("name"),
                "key": v.get("key"),
                "model": v.get("model"),
                "priority": v.get("priority"),
                "active": v.get("active"),
            }
        )

    out.table(
        f"QWeb Templates ({len(views)})",
        [("ID", "cyan"), ("Name", ""), ("Key", ""), ("Model", ""), ("Priority", "dim"), ("Status", "")],
        rows,
        data_for_json=json_data,
    )


@app.command("generate")
def generate(
    ctx: typer.Context,
    report_type: Annotated[
        str, typer.Argument(help="Report type: health, finance, security, inventory, operations, full")
    ],
    output: Annotated[str | None, typer.Option("--output", "-o", help="Write to file (.md, .json, .html)")] = None,
    quick: Annotated[bool, typer.Option("--quick", help="Skip slow checks")] = False,
) -> None:
    """Generate a focused diagnostic report by domain.

    Maps business-friendly report names to diagnose sections.

    Examples:
        kctl-odoo report generate health
        kctl-odoo report generate finance --output finance-march.md
        kctl-odoo report generate security --output security.html
        kctl-odoo report generate full --output audit.json
        kctl-odoo report generate inventory --quick
    """
    if report_type not in REPORT_TYPES:
        valid = ", ".join(REPORT_TYPES.keys())
        typer.echo(f"Unknown report type '{report_type}'. Valid: {valid}", err=True)
        raise typer.Exit(1)

    sections = REPORT_TYPES[report_type]
    section_arg = ",".join(sections) if sections else None

    # Import and delegate to diagnose engine
    from kctl_odoo.commands.diagnose import run_diagnose_core

    run_diagnose_core(ctx, output=output, quick=quick, section_filter=section_arg)


# ===================================================================
# report_management framework commands
# ===================================================================


def _check_report_management(c: object, out: object) -> bool:
    """Return True if report_management module is available, else log info."""
    try:
        c.search_count("report.type.registry", [])  # type: ignore[attr-defined]
        return True
    except RPCError:
        out.info("report_management module is not installed.")  # type: ignore[attr-defined]
        return False


def _default_date_range() -> tuple[str, str]:
    """Return (date_from, date_to) for last complete month as ISO strings."""
    today = date.today()
    first_of_month = today.replace(day=1)
    last_month_end = first_of_month - timedelta(days=1)
    last_month_start = last_month_end.replace(day=1)
    return last_month_start.isoformat(), last_month_end.isoformat()


def _fiscal_year_start(ref_date: str) -> str:
    """Return January 1st of the fiscal year for *ref_date* (ISO string)."""
    return date.fromisoformat(ref_date).replace(month=1, day=1).isoformat()


def _m2o(val: object) -> str:
    """Format an M2O field — Odoo returns [id, name] lists."""
    if isinstance(val, list):
        return str(val[1]) if len(val) > 1 else str(val[0])
    return str(val or "")


def _parse_accurate_xlsx(path: Path) -> dict:
    """Parse an Accurate XLSX export into a structured dict."""
    import re

    import openpyxl

    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active

    rows_data = []
    for row in ws.iter_rows(max_col=8, values_only=True):
        rows_data.append(list(row))
    wb.close()

    if len(rows_data) < 5:
        raise ValueError(f"Accurate XLSX too short ({len(rows_data)} rows)")

    company = str(rows_data[0][1] or "").strip()
    title = str(rows_data[1][1] or "").strip()
    date_str = str(rows_data[2][1] or "").strip()

    title_lower = title.lower()
    report_type = None
    if "laba" in title_lower or "rugi" in title_lower:
        report_type = "profit_loss"
    elif "neraca" in title_lower:
        report_type = "balance_sheet"
    elif "trial" in title_lower or "saldo" in title_lower:
        report_type = "trial_balance"
    elif "arus kas" in title_lower or "cash flow" in title_lower:
        report_type = "cash_flow"

    MONTHS_ID = {
        "jan": 1,
        "feb": 2,
        "mar": 3,
        "apr": 4,
        "mei": 5,
        "jun": 6,
        "jul": 7,
        "agu": 8,
        "ags": 8,
        "sep": 9,
        "okt": 10,
        "nov": 11,
        "des": 12,
        "january": 1,
        "february": 2,
        "march": 3,
        "april": 4,
        "may": 5,
        "june": 6,
        "july": 7,
        "august": 8,
        "september": 9,
        "october": 10,
        "november": 11,
        "december": 12,
    }

    date_from = None
    date_to = None

    m = re.search(r"(\d{1,2})\s+(\w+)\s+(\d{4})\s+s/d\s+(\d{1,2})\s+(\w+)\s+(\d{4})", date_str)
    if m:
        mon1 = MONTHS_ID.get(m.group(2).lower()[:3], 1)
        mon2 = MONTHS_ID.get(m.group(5).lower()[:3], 1)
        date_from = f"{m.group(3)}-{mon1:02d}-{int(m.group(1)):02d}"
        date_to = f"{m.group(6)}-{mon2:02d}-{int(m.group(4)):02d}"
    else:
        m2 = re.search(r"(\d{1,2})\s+(\w+)\s+(\d{4})", date_str)
        if m2:
            mon = MONTHS_ID.get(m2.group(2).lower()[:3], 1)
            date_to = f"{m2.group(3)}-{mon:02d}-{int(m2.group(1)):02d}"

    lines = []
    for row in rows_data[5:]:
        desc = row[2] if len(row) > 2 else None
        amt = row[4] if len(row) > 4 else None

        if desc is None and amt is None:
            continue

        desc_str = str(desc or "")
        if "ACCURATE" in desc_str or "Tercetak" in desc_str or "Halaman" in desc_str:
            continue
        if not desc_str.strip():
            continue

        stripped = desc_str.lstrip()
        indent = (len(desc_str) - len(stripped)) // 4

        amount = 0.0
        if amt is not None:
            try:
                amount = float(amt)
            except (ValueError, TypeError):
                amount = 0.0

        lines.append({"label": stripped, "amount": amount, "indent": indent})

    return {
        "company": company,
        "title": title,
        "report_type": report_type,
        "date_from": date_from,
        "date_to": date_to,
        "lines": lines,
    }


@app.command("types")
def report_types(
    ctx: typer.Context,
    category: Annotated[
        str | None,
        typer.Option("--category", "-c", help="Filter by category (financial, aging, sales, inventory, mrp, sfa)"),
    ] = None,
    limit: Annotated[int, typer.Option("--limit", "-l", help="Max results")] = 50,
) -> None:
    """List report types from the report_management registry.

    These are custom report types defined in report.type.registry, NOT Odoo's
    built-in ir.actions.report. Use 'report list' for built-in reports.

    Examples:
        kctl-odoo report types
        kctl-odoo report types --category financial
        kctl-odoo report types --json
    """
    actx: AppContext = ctx.obj
    out = actx.output
    c = actx.client

    if not _check_report_management(c, out):
        if actx.json_mode:
            out.raw_json([])
        return

    domain: list = []
    if category:
        domain.append(("category", "=", category))

    try:
        records = c.search_read(
            "report.type.registry",
            domain=domain,
            fields=["id", "name", "code", "category", "handler_model", "date_mode", "active"],
            limit=limit,
            order="category, sequence, name",
        )
    except RPCError as e:
        out.error(f"Failed to fetch report types: {e.detail}")
        raise typer.Exit(1) from e

    if not records:
        out.info("No report types found.")
        if actx.json_mode:
            out.raw_json([])
        return

    rows = []
    json_data = []
    for r in records:
        cat = r.get("category", "") or ""
        cat_display = {
            "financial": "[green]financial[/green]",
            "aging": "[yellow]aging[/yellow]",
            "management": "[cyan]management[/cyan]",
            "custom": "[dim]custom[/dim]",
        }.get(cat, cat)
        status = "[green]active[/green]" if r.get("active") else "[dim]inactive[/dim]"

        rows.append(
            [
                str(r["id"]),
                r.get("code", ""),
                (r.get("name") or "")[:35],
                cat_display,
                r.get("date_mode", ""),
                (r.get("handler_model") or "")[:30],
                status,
            ]
        )
        json_data.append(
            {
                "id": r["id"],
                "code": r.get("code"),
                "name": r.get("name"),
                "category": cat,
                "date_mode": r.get("date_mode"),
                "handler_model": r.get("handler_model"),
                "active": r.get("active"),
            }
        )

    out.table(
        f"Report Types ({len(records)})",
        [
            ("ID", "cyan"),
            ("Code", ""),
            ("Name", ""),
            ("Category", ""),
            ("Date Mode", "dim"),
            ("Handler", "dim"),
            ("Status", ""),
        ],
        rows,
        data_for_json=json_data,
    )


@app.command("run")
def report_run(
    ctx: typer.Context,
    type_code: Annotated[
        str, typer.Argument(help="Report type code (e.g. profit_loss, balance_sheet, receivable_aging)")
    ],
    date_from: Annotated[str | None, typer.Option("--date-from", help="Start date (YYYY-MM-DD)")] = None,
    date_to: Annotated[str | None, typer.Option("--date-to", help="End date (YYYY-MM-DD)")] = None,
    as_of_date: Annotated[
        str | None,
        typer.Option("--date", help="As-of / cumulative date (for balance_sheet, receivable_aging, payable_aging)"),
    ] = None,
    company: Annotated[str | None, typer.Option("--company", help="Company name or ID")] = None,
    dry_run: Annotated[bool, typer.Option("--dry-run", help="Show what would be created without executing")] = False,
    force: Annotated[bool, typer.Option("--force", "-y", help="Skip confirmation")] = False,
) -> None:
    """Generate a report using the report_management framework.

    Creates a report.wizard, sets parameters, calls button_export_html to
    create a report.instance, then displays the instance metadata.

    Old short codes (pnl, bs, cf, tb, cma, ar_aging, ap_aging) are still
    accepted via backward-compatible aliases.

    Examples:
        kctl-odoo report run profit_loss --date-from 2026-01-01 --date-to 2026-01-31
        kctl-odoo report run balance_sheet --date 2026-03-31
        kctl-odoo report run receivable_aging --date 2026-03-27
        kctl-odoo report run trial_balance --dry-run
        kctl-odoo report run pnl --date-from 2026-01-01 --date-to 2026-01-31  # alias
    """
    actx: AppContext = ctx.obj
    out = actx.output
    c = actx.client

    if not _check_report_management(c, out):
        return

    # Resolve aliases (e.g. pnl → profit_loss)
    type_code = _resolve_type_code(type_code)

    # 1. Resolve report type from registry
    try:
        reg_records = c.search_read(
            "report.type.registry",
            domain=[("code", "=", type_code)],
            fields=["id", "name", "code", "date_mode", "category"],
            limit=1,
        )
    except RPCError as e:
        out.error(f"Failed to query report type registry: {e.detail}")
        raise typer.Exit(1) from e

    if not reg_records:
        out.error(f"Report type not found: {type_code}")
        raise typer.Exit(1)

    reg = reg_records[0]
    reg.get("date_mode", "period")

    # 2. Find a template for this report type
    try:
        tpl_records = c.search_read(
            "report.template",
            domain=[("report_type_id", "=", reg["id"])],
            fields=["id", "name", "code"],
            limit=1,
            order="sequence, id",
        )
    except RPCError as e:
        out.error(f"Failed to query report templates: {e.detail}")
        raise typer.Exit(1) from e

    if not tpl_records:
        # Fallback: try matching by the selection field
        with contextlib.suppress(RPCError):
            tpl_records = c.search_read(
                "report.template",
                domain=[("report_type", "=", type_code)],
                fields=["id", "name", "code"],
                limit=1,
                order="sequence, id",
            )

    if not tpl_records:
        out.error(f"No report template found for type '{type_code}'.")
        raise typer.Exit(1)

    tpl = tpl_records[0]

    # 3. Resolve dates
    if as_of_date:
        # Cumulative / as_of: set date_from to fiscal year start, date_to to the date
        resolved_date_to = as_of_date
        resolved_date_from = _fiscal_year_start(as_of_date)
    elif date_from and date_to:
        resolved_date_from = date_from
        resolved_date_to = date_to
    elif date_from and not date_to:
        resolved_date_from = date_from
        resolved_date_to = date.today().isoformat()
    elif date_to and not date_from:
        resolved_date_from = _fiscal_year_start(date_to)
        resolved_date_to = date_to
    else:
        resolved_date_from, resolved_date_to = _default_date_range()

    # 4. Build wizard vals
    wizard_vals: dict = {
        "template_id": tpl["id"],
        "date_from": resolved_date_from,
        "date_to": resolved_date_to,
    }

    if company:
        if company.isdigit():
            wizard_vals["company_id"] = int(company)
        else:
            companies = c.search_read(
                "res.company",
                domain=[("name", "ilike", company)],
                fields=["id", "name"],
                limit=1,
            )
            if not companies:
                out.error(f"Company not found: {company}")
                raise typer.Exit(1)
            wizard_vals["company_id"] = companies[0]["id"]

    # Dry run: just show what would be created
    if dry_run:
        out.info("Would create report instance:")
        out.info(f"  Type:      {reg.get('name')} ({type_code})")
        out.info(f"  Template:  {tpl.get('name')} (ID {tpl['id']})")
        out.info(f"  Date From: {resolved_date_from}")
        out.info(f"  Date To:   {resolved_date_to}")
        if actx.json_mode:
            out.raw_json(
                {
                    "dry_run": True,
                    "type_code": type_code,
                    "type_name": reg.get("name"),
                    "template_id": tpl["id"],
                    "template_name": tpl.get("name"),
                    "date_from": resolved_date_from,
                    "date_to": resolved_date_to,
                    "wizard_vals": wizard_vals,
                }
            )
        return

    if not force:
        out.info(f"Generate '{reg.get('name')}' for {resolved_date_from} to {resolved_date_to}")
        if not typer.confirm("Proceed?"):
            raise typer.Exit(0)

    # 5. Create wizard and generate report
    try:
        wizard_id = c.execute_kw("report.wizard", "create", [wizard_vals])
        result = c.execute_kw("report.wizard", "button_export_html", [[wizard_id]])
    except RPCError as e:
        out.error(f"Failed to generate report: {e.detail}")
        raise typer.Exit(1) from e

    # 6. Extract instance ID from the action result
    instance_id = None
    if isinstance(result, dict):
        instance_id = result.get("res_id")

    if not instance_id:
        # Fallback: find the most recently created instance for this template
        try:
            recent = c.search_read(
                "report.instance",
                domain=[("template_id", "=", tpl["id"])],
                fields=["id", "display_name", "date_from", "date_to", "create_date"],
                limit=1,
                order="create_date desc",
            )
            if recent:
                instance_id = recent[0]["id"]
        except RPCError:
            pass

    if instance_id:
        try:
            inst = c.search_read(
                "report.instance",
                domain=[("id", "=", instance_id)],
                fields=["id", "display_name", "date_from", "date_to", "create_date", "template_id"],
                limit=1,
            )
            if inst:
                inst = inst[0]
                out.success(f"Report instance created: ID {inst['id']} - {inst.get('display_name', '')}")
                if actx.json_mode:
                    out.raw_json(
                        {
                            "instance_id": inst["id"],
                            "display_name": inst.get("display_name"),
                            "date_from": inst.get("date_from"),
                            "date_to": inst.get("date_to"),
                            "create_date": inst.get("create_date"),
                            "template": _m2o(inst.get("template_id")),
                        }
                    )
                return
        except RPCError:
            pass

    out.success("Report generated successfully.")
    if actx.json_mode:
        out.raw_json({"instance_id": instance_id, "result": result})


@app.command("data")
def report_data(
    ctx: typer.Context,
    type_code: Annotated[str, typer.Argument(help="Report type code (e.g. profit_loss, balance_sheet, trial_balance)")],
    date_from: Annotated[str | None, typer.Option("--date-from", help="Start date (YYYY-MM-DD)")] = None,
    date_to: Annotated[str | None, typer.Option("--date-to", help="End date (YYYY-MM-DD)")] = None,
    as_of_date: Annotated[
        str | None,
        typer.Option("--date", help="As-of date (for balance_sheet, aging)"),
    ] = None,
    company: Annotated[str | None, typer.Option("--company", help="Company name or ID")] = None,
    show_zero: Annotated[bool, typer.Option("--show-zero", help="Include zero-balance lines")] = False,
) -> None:
    """Print report data as a terminal table for quick inspection.

    Runs the report engine and displays the result as a formatted table.
    Use --json for machine-readable output.

    Examples:
        kctl-odoo report data profit_loss --company 42 --date-from 2026-04-01 --date-to 2026-04-23
        kctl-odoo report data balance_sheet --company 42 --date 2026-04-23
        kctl-odoo report data trial_balance --date-from 2026-04-01 --date-to 2026-04-23 --json
    """
    actx: AppContext = ctx.obj
    out = actx.output
    c = actx.client

    if not _check_report_management(c, out):
        return

    type_code = _resolve_type_code(type_code)

    try:
        tpls = c.search_read(
            "report.template",
            [("report_type", "=", type_code)],
            ["id", "name", "code"],
            limit=1,
            order="sequence, id",
        )
    except RPCError as e:
        out.error(f"Template lookup failed: {e.detail}")
        raise typer.Exit(1) from e

    if not tpls:
        out.error(f"No template for report_type={type_code!r}")
        raise typer.Exit(1)

    tpl = tpls[0]

    co_id = None
    if company:
        if company.isdigit():
            co_id = int(company)
        else:
            cos = c.search_read("res.company", [("name", "ilike", company)], ["id"], limit=1)
            if not cos:
                out.error(f"Company not found: {company}")
                raise typer.Exit(1)
            co_id = cos[0]["id"]

    if not co_id:
        cos = c.search_read("res.company", [], ["id"], limit=1, order="id")
        co_id = cos[0]["id"] if cos else 1

    if as_of_date:
        resolved_from = _fiscal_year_start(as_of_date)
        resolved_to = as_of_date
    elif date_from and date_to:
        resolved_from, resolved_to = date_from, date_to
    elif date_from:
        resolved_from, resolved_to = date_from, date.today().isoformat()
    elif date_to:
        resolved_from = _fiscal_year_start(date_to)
        resolved_to = date_to
    else:
        resolved_from, resolved_to = _default_date_range()

    try:
        rd = c.execute_kw(
            "report.report_management.engine",
            "compute_report_data",
            [
                [],
                {
                    "template_id": tpl["id"],
                    "date_from": resolved_from,
                    "date_to": resolved_to,
                    "company_id": co_id,
                    "company_ids": [],
                    "target_move": "posted",
                    "divider": 1,
                    "show_variance": False,
                    "show_percentage": False,
                    "monthly_breakdown": False,
                    "expanded_line_ids": [],
                    "page_offset": 0,
                    "page_limit": 0,
                },
            ],
            {"context": {"allowed_company_ids": [co_id]}},
        )
    except RPCError as e:
        out.error(f"compute_report_data failed: {e.detail}")
        raise typer.Exit(1) from e

    rows = []
    for t in rd.get("tables") or []:
        rows.extend(t.get("rows") or [])
    if not rows:
        rows = rd.get("lines") or []

    # Detect columns from engine response
    columns = rd.get("columns") or []
    col_keys = [c["key"] for c in columns if c.get("key") != "name"]
    is_tb = any(k.startswith("tb_") for k in col_keys)

    output_rows = []
    for r in rows:
        if not isinstance(r, dict):
            continue
        v = r.get("values") or {}
        amt = float(v.get("actual") or v.get("amount") or 0.0)
        if not show_zero and not amt and r.get("kind") not in ("header", "spacer"):
            if not is_tb or not any(v.get(k) for k in ("tb_initial", "tb_ending")):
                continue
        label = r.get("label") or r.get("name") or ""
        kind = r.get("kind") or r.get("line_type") or ""
        level = r.get("level", 0)
        indent = "  " * level
        row_out: dict = {
            "label": f"{indent}{label}",
            "kind": kind,
            "actual": amt,
            "line_id": r.get("line_id"),
        }
        if is_tb:
            row_out["tb_initial"] = float(v.get("tb_initial") or 0.0)
            row_out["tb_debit"] = float(v.get("tb_debit") or 0.0)
            row_out["tb_credit"] = float(v.get("tb_credit") or 0.0)
            row_out["tb_ending"] = float(v.get("tb_ending") or 0.0)
        output_rows.append(row_out)

    if actx.json_mode:
        out.raw_json(
            {
                "type": type_code,
                "company_id": co_id,
                "date_from": resolved_from,
                "date_to": resolved_to,
                "rows": output_rows,
            }
        )
        return

    out.info(f"{tpl['name']}  —  Company {co_id}  —  {resolved_from} to {resolved_to}\n")

    if is_tb:
        hdr = f"{'Line':<40s} {'Initial':>16s} {'Debit':>16s} {'Credit':>16s} {'Period':>16s} {'Ending':>16s}"
        out.info(hdr)
        out.info("─" * 122)
        for r in output_rows:
            kind = r["kind"]
            label = r["label"]
            if kind in ("header", "spacer"):
                out.info(f"\033[1m{label}\033[0m")
            elif kind in ("total", "subtotal"):
                out.info(
                    f"\033[1m{label:<40s} {r['tb_initial']:>16,.2f} {r['tb_debit']:>16,.2f} {r['tb_credit']:>16,.2f} {r['actual']:>16,.2f} {r['tb_ending']:>16,.2f}\033[0m"
                )
            else:
                out.info(
                    f"{label:<40s} {r['tb_initial']:>16,.2f} {r['tb_debit']:>16,.2f} {r['tb_credit']:>16,.2f} {r['actual']:>16,.2f} {r['tb_ending']:>16,.2f}"
                )
        out.info("─" * 122)
    else:
        out.info(f"{'Line':<55s} {'Amount':>20s}")
        out.info("─" * 76)
        for r in output_rows:
            kind = r["kind"]
            label = r["label"]
            amt = r["actual"]
            if kind in ("header", "spacer"):
                out.info(f"\033[1m{label}\033[0m")
            elif kind in ("total", "subtotal"):
                out.info(f"\033[1m{label:<55s} {amt:>20,.2f}\033[0m")
            else:
                out.info(f"{label:<55s} {amt:>20,.2f}")
        out.info("─" * 76)
    out.info(f"  {len(output_rows)} lines")


@app.command("compare")
def report_compare(
    ctx: typer.Context,
    accurate_file: Annotated[Path, typer.Argument(help="Path to Accurate XLSX export file", exists=True)],
    type_code: Annotated[
        str | None,
        typer.Option("--type", "-t", help="Report type (auto-detected from XLSX if omitted)"),
    ] = None,
    company: Annotated[str | None, typer.Option("--company", help="Company name or ID")] = None,
    date_from: Annotated[str | None, typer.Option("--date-from", help="Override start date")] = None,
    date_to: Annotated[str | None, typer.Option("--date-to", help="Override end date")] = None,
    tolerance: Annotated[float, typer.Option("--tolerance", help="Amount delta tolerance (IDR)")] = 1.0,
) -> None:
    """Compare an Accurate XLSX export against Odoo report engine output.

    Reads the Accurate file, auto-detects the report type and date range,
    runs the Odoo engine with the same parameters, and prints a line-by-line
    diff. Exits non-zero if any line exceeds the tolerance.

    Examples:
        kctl-odoo report compare /path/to/laba_rugi_standar.xlsx --company 42
        kctl-odoo report compare /path/to/neraca_standar.xlsx --company 42
        kctl-odoo report compare /path/to/laba_rugi.xlsx -t profit_loss --company 42
    """
    actx: AppContext = ctx.obj
    out = actx.output
    c = actx.client

    if not _check_report_management(c, out):
        return

    try:
        acc = _parse_accurate_xlsx(accurate_file)
    except Exception as e:
        out.error(f"Failed to parse Accurate XLSX: {e}")
        raise typer.Exit(2) from e

    resolved_type = _resolve_type_code(type_code) if type_code else acc.get("report_type")
    if not resolved_type:
        out.error("Cannot auto-detect report type. Use --type to specify.")
        raise typer.Exit(1)

    resolved_from = date_from or acc.get("date_from") or _fiscal_year_start(acc["date_to"])
    resolved_to = date_to or acc.get("date_to")
    if not resolved_to:
        out.error("Cannot determine date. Use --date-to to specify.")
        raise typer.Exit(1)

    co_id = None
    if company:
        if company.isdigit():
            co_id = int(company)
        else:
            cos = c.search_read("res.company", [("name", "ilike", company)], ["id"], limit=1)
            if not cos:
                out.error(f"Company not found: {company}")
                raise typer.Exit(1)
            co_id = cos[0]["id"]
    if not co_id:
        cos = c.search_read("res.company", [], ["id"], limit=1, order="id")
        co_id = cos[0]["id"] if cos else 1

    try:
        tpls = c.search_read(
            "report.template",
            [("report_type", "=", resolved_type)],
            ["id", "name"],
            limit=1,
            order="sequence, id",
        )
    except RPCError as e:
        out.error(f"Template lookup failed: {e.detail}")
        raise typer.Exit(1) from e

    if not tpls:
        out.error(f"No template for type {resolved_type!r}")
        raise typer.Exit(1)

    tpl = tpls[0]

    out.info(f"Accurate:  {acc['title']}  ({acc['company']})")
    out.info(f"Odoo:      {tpl['name']}  (company_id={co_id})")
    out.info(f"Period:    {resolved_from} → {resolved_to}")
    out.info(f"Tolerance: {tolerance:,.2f} IDR\n")

    try:
        rd = c.execute_kw(
            "report.report_management.engine",
            "compute_report_data",
            [
                [],
                {
                    "template_id": tpl["id"],
                    "date_from": resolved_from,
                    "date_to": resolved_to,
                    "company_id": co_id,
                    "company_ids": [],
                    "target_move": "posted",
                    "divider": 1,
                    "show_variance": False,
                    "show_percentage": False,
                    "monthly_breakdown": False,
                    "expanded_line_ids": [],
                    "page_offset": 0,
                    "page_limit": 0,
                },
            ],
            {"context": {"allowed_company_ids": [co_id]}},
        )
    except RPCError as e:
        out.error(f"Odoo engine failed: {e.detail}")
        raise typer.Exit(1) from e

    odoo_rows = []
    for t in rd.get("tables") or []:
        odoo_rows.extend(t.get("rows") or [])
    if not odoo_rows:
        odoo_rows = rd.get("lines") or []

    odoo_by_label: dict[str, float] = {}
    for r in odoo_rows:
        if not isinstance(r, dict):
            continue
        v = r.get("values") or {}
        amt = float(v.get("actual") or v.get("amount") or 0.0)
        label = (r.get("label") or r.get("name") or "").strip()
        if label:
            odoo_by_label[label] = amt

    # Query account-level balances directly for account-name matching.
    # Accurate XLSX has per-account lines (e.g., "Penjualan", "BCA 714.5130972")
    # that match Odoo account names. We query all account balances for the
    # company/period so these can be matched by name.
    odoo_account_balances: dict[str, float] = {}
    try:
        accounts = c.search_read(
            "account.account",
            [("company_ids", "in", [co_id])],
            ["id", "name"],
            limit=0,
        )
        acct_name_by_id = {a["id"]: (a.get("name") or "") for a in accounts}

        aml_data = c.execute_kw(
            "account.move.line",
            "read_group",
            [
                [
                    ("company_id", "=", co_id),
                    ("date", ">=", resolved_from),
                    ("date", "<=", resolved_to),
                    ("parent_state", "=", "posted"),
                ]
            ],
            {"fields": ["account_id", "balance"], "groupby": ["account_id"]},
        )
        for grp in aml_data:
            acc_id = grp["account_id"][0] if grp.get("account_id") else None
            if acc_id and acc_id in acct_name_by_id:
                name = acct_name_by_id[acc_id]
                odoo_account_balances[name.strip().lower()] = float(grp.get("balance") or 0.0)
    except RPCError:
        pass

    LABEL_MAP = {
        "pendapatan": "sales",
        "penjualan": "sales",
        "jumlah pendapatan": "total revenue",
        "beban pokok penjualan": "cost of goods sold",
        "cogs": "total cost of goods sold",
        "jumlah beban pokok penjualan": "total cost of goods sold",
        "laba kotor": "gross profit",
        "beban operasional": "operating expenses",
        "biaya umum & administrasi": "total operating expenses",
        "jumlah beban operasional": "total operating expenses",
        "pendapatan operasional": "operating income",
        "pendapatan dan beban non operasional": "non-operating income & expenses",
        "pendapatan non operasional": "non-operating income",
        "jumlah pendapatan non operasional": "non-operating income",
        "beban non operasional": "non-operating expense",
        "jumlah beban non operasional": "non-operating expense",
        "jumlah pendapatan dan beban non operasional": "net non-operating income/(expense)",
        "laba bersih (sebelum pajak)": "net income",
        "laba bersih (setelah pajak)": "net income",
        "pajak penghasilan": "net income",
        "jumlah aset": "total assets",
        "jumlah aset lancar": "total current assets",
        "jumlah kewajiban": "total liabilities",
        "jumlah ekuitas": "total equity",
        "jumlah liabilitas dan ekuitas": "total liabilities and equity",
        "kas dan setara kas": "cash and cash equivalents",
        "kas & bank": "cash and cash equivalents",
        "piutang usaha": "trade receivables",
        "utang usaha": "trade payables",
    }

    def _normalize(s: str) -> str:
        s = s.strip().lower()
        if s in LABEL_MAP:
            return LABEL_MAP[s]
        s = s.replace("jumlah ", "total ")
        return s

    odoo_norm: dict[str, tuple[str, float]] = {}
    for lbl, amt in odoo_by_label.items():
        odoo_norm[_normalize(lbl)] = (lbl, amt)

    matched = 0
    mismatched = 0
    missing = 0
    results = []

    out.info(f"{'Accurate Line':<40s} {'Accurate':>16s} {'Odoo':>16s} {'Delta':>14s} {'':>5s}")
    out.info("─" * 95)

    for acc_line in acc["lines"]:
        acc_label = acc_line["label"]
        acc_amt = acc_line["amount"]

        if not acc_amt and acc_line["indent"] == 0:
            continue

        norm = _normalize(acc_label)
        hit = odoo_norm.get(norm)

        if hit:
            odoo_label, odoo_amt = hit
            delta = odoo_amt - acc_amt
            ok = abs(delta) <= tolerance
            status = "✓" if ok else "✗"
            if ok:
                matched += 1
            else:
                mismatched += 1
            out.info(f"{acc_label:<40s} {acc_amt:>16,.2f} {odoo_amt:>16,.2f} {delta:>+14,.2f} {status:>5s}")
            results.append(
                {
                    "label": acc_label,
                    "accurate": acc_amt,
                    "odoo": odoo_amt,
                    "delta": delta,
                    "match": ok,
                }
            )
        else:
            acct_hit = odoo_account_balances.get(acc_label.strip().lower())
            if acct_hit is not None:
                odoo_amt = abs(acct_hit)
                delta = odoo_amt - acc_amt
                ok = abs(delta) <= tolerance
                status = "✓" if ok else "✗"
                if ok:
                    matched += 1
                else:
                    mismatched += 1
                out.info(f"{acc_label:<40s} {acc_amt:>16,.2f} {odoo_amt:>16,.2f} {delta:>+14,.2f} {status:>5s}")
                results.append({"label": acc_label, "accurate": acc_amt, "odoo": odoo_amt, "delta": delta, "match": ok})
            else:
                missing += 1
                out.info(f"{acc_label:<40s} {acc_amt:>16,.2f} {'—':>16s} {'':>14s} {'?':>5s}")
                results.append({"label": acc_label, "accurate": acc_amt, "odoo": None, "delta": None, "match": False})

    out.info("─" * 95)
    total = matched + mismatched + missing
    out.info(f"  {matched}/{total} matched, {mismatched} mismatched, {missing} missing in Odoo")

    if actx.json_mode:
        out.raw_json(
            {
                "accurate_file": str(accurate_file),
                "report_type": resolved_type,
                "company_id": co_id,
                "date_from": resolved_from,
                "date_to": resolved_to,
                "matched": matched,
                "mismatched": mismatched,
                "missing": missing,
                "results": results,
            }
        )

    if mismatched > 0 or missing > 0:
        raise typer.Exit(1)


@app.command("instances")
def report_instances(
    ctx: typer.Context,
    type_code: Annotated[str | None, typer.Option("--type", "-t", help="Filter by report type code")] = None,
    limit: Annotated[int, typer.Option("--limit", "-l", help="Max results")] = 20,
) -> None:
    """List generated report instances from report_management.

    Shows instances created via the report wizard, NOT Odoo built-in reports.

    Examples:
        kctl-odoo report instances
        kctl-odoo report instances --type profit_loss
        kctl-odoo report instances --limit 10 --json
    """
    actx: AppContext = ctx.obj
    out = actx.output
    c = actx.client

    if not _check_report_management(c, out):
        if actx.json_mode:
            out.raw_json([])
        return

    domain: list = []
    if type_code:
        # Resolve aliases (e.g. pnl → profit_loss)
        type_code = _resolve_type_code(type_code)
        # Find template(s) matching the type code via registry
        try:
            reg_ids = c.search("report.type.registry", [("code", "=", type_code)])
            if reg_ids:
                tpl_ids = c.search("report.template", [("report_type_id", "in", reg_ids)])
                if tpl_ids:
                    domain.append(("template_id", "in", tpl_ids))
                else:
                    # Try legacy selection field
                    tpl_ids = c.search("report.template", [("report_type", "=", type_code)])
                    if tpl_ids:
                        domain.append(("template_id", "in", tpl_ids))
                    else:
                        out.info(f"No templates found for type '{type_code}'.")
                        if actx.json_mode:
                            out.raw_json([])
                        return
            else:
                out.info(f"Report type not found: {type_code}")
                if actx.json_mode:
                    out.raw_json([])
                return
        except RPCError as e:
            out.error(f"Failed to resolve report type: {e.detail}")
            raise typer.Exit(1) from e

    try:
        records = c.search_read(
            "report.instance",
            domain=domain,
            fields=["id", "display_name", "template_id", "date_from", "date_to", "create_date", "company_id"],
            limit=limit,
            order="create_date desc",
        )
    except RPCError as e:
        out.error(f"Failed to fetch report instances: {e.detail}")
        raise typer.Exit(1) from e

    if not records:
        out.info("No report instances found.")
        if actx.json_mode:
            out.raw_json([])
        return

    rows = []
    json_data = []
    for r in records:
        rows.append(
            [
                str(r["id"]),
                (r.get("display_name") or "")[:40],
                _m2o(r.get("template_id")),
                str(r.get("date_from") or ""),
                str(r.get("date_to") or ""),
                str(r.get("create_date") or "")[:19],
            ]
        )
        json_data.append(
            {
                "id": r["id"],
                "display_name": r.get("display_name"),
                "template": _m2o(r.get("template_id")),
                "date_from": r.get("date_from"),
                "date_to": r.get("date_to"),
                "create_date": r.get("create_date"),
                "company": _m2o(r.get("company_id")),
            }
        )

    out.table(
        f"Report Instances ({len(records)})",
        [
            ("ID", "cyan"),
            ("Name", ""),
            ("Template", ""),
            ("Date From", ""),
            ("Date To", ""),
            ("Created", "dim"),
        ],
        rows,
        data_for_json=json_data,
    )


@app.command("view")
def report_view(
    ctx: typer.Context,
    instance_id: Annotated[int, typer.Argument(help="Report instance ID")],
) -> None:
    """View a report instance's metadata and data summary.

    Reads the report.instance record and displays its details. Report data
    (columns, lines) may not be directly readable via JSON-RPC as they are
    often computed in-browser; in that case only metadata is shown.

    Examples:
        kctl-odoo report view 42
        kctl-odoo report view 42 --json
    """
    actx: AppContext = ctx.obj
    out = actx.output
    c = actx.client

    if not _check_report_management(c, out):
        return

    base_fields = [
        "id",
        "display_name",
        "template_id",
        "date_from",
        "date_to",
        "company_id",
        "target_move",
        "comparison_mode",
        "divider",
        "show_variance",
        "show_percentage",
        "monthly_breakdown",
        "create_date",
        "create_uid",
        "render_mode",
    ]

    try:
        records = c.search_read(
            "report.instance",
            domain=[("id", "=", instance_id)],
            fields=base_fields,
            limit=1,
        )
    except RPCError as e:
        out.error(f"Failed to read report instance: {e.detail}")
        raise typer.Exit(1) from e

    if not records:
        out.error(f"Report instance not found: {instance_id}")
        raise typer.Exit(1)

    inst = records[0]

    if actx.json_mode:
        json_out: dict = {
            "id": inst["id"],
            "display_name": inst.get("display_name"),
            "template": _m2o(inst.get("template_id")),
            "date_from": inst.get("date_from"),
            "date_to": inst.get("date_to"),
            "company": _m2o(inst.get("company_id")),
            "target_move": inst.get("target_move"),
            "comparison_mode": inst.get("comparison_mode"),
            "divider": inst.get("divider"),
            "show_variance": inst.get("show_variance"),
            "show_percentage": inst.get("show_percentage"),
            "monthly_breakdown": inst.get("monthly_breakdown"),
            "render_mode": inst.get("render_mode"),
            "create_date": inst.get("create_date"),
            "created_by": _m2o(inst.get("create_uid")),
        }
        out.raw_json(json_out)
        return

    # Display as key-value detail view
    out.info(f"Report Instance #{inst['id']}")
    out.info(f"  Name:          {inst.get('display_name', '')}")
    out.info(f"  Template:      {_m2o(inst.get('template_id'))}")
    out.info(f"  Date From:     {inst.get('date_from', '')}")
    out.info(f"  Date To:       {inst.get('date_to', '')}")
    out.info(f"  Company:       {_m2o(inst.get('company_id'))}")
    out.info(f"  Target Move:   {inst.get('target_move', '')}")
    out.info(f"  Comparison:    {inst.get('comparison_mode', '')}")
    out.info(f"  Divider:       {inst.get('divider', '')}")
    out.info(f"  Variance:      {inst.get('show_variance', '')}")
    out.info(f"  Percentage:    {inst.get('show_percentage', '')}")
    out.info(f"  Monthly:       {inst.get('monthly_breakdown', '')}")
    out.info(f"  Render Mode:   {inst.get('render_mode', '')}")
    out.info(f"  Created:       {str(inst.get('create_date', ''))[:19]}")
    out.info(f"  Created By:    {_m2o(inst.get('create_uid'))}")
    out.info("")
    out.info("Tip: Open in Odoo UI for full report with data, charts, and drill-down.")


def _http_download_report(
    base_url: str,
    database: str,
    username: str,
    password: str,
    data_dict: dict,
    fmt: str,
    output_path: Path,
    out,
    quiet: bool = False,
) -> bool:
    """Download report via HTTP controller (headless, no browser needed).

    Authenticates via /web/session/authenticate, then calls
    /report_management/download_xlsx or download_pdf with the data dict.
    Returns True on success, False on failure.
    """
    import json as json_lib
    from urllib.parse import quote

    import httpx

    url = base_url.rstrip("/")
    endpoint = f"/report_management/download_{'xlsx' if fmt == 'xlsx' else 'pdf'}"

    try:
        headers = {"X-Odoo-dbfilter": f"^{database}$"}
        with httpx.Client(timeout=120, follow_redirects=True, headers=headers) as client:
            # Step 1: Authenticate to get session cookie
            auth_resp = client.post(
                f"{url}/web/session/authenticate",
                json={
                    "jsonrpc": "2.0",
                    "method": "call",
                    "id": 1,
                    "params": {"db": database, "login": username, "password": password},
                },
            )
            if auth_resp.status_code != 200:
                out.error(f"Authentication failed: HTTP {auth_resp.status_code}")
                return False

            auth_data = auth_resp.json()
            if auth_data.get("error") or not auth_data.get("result", {}).get("uid"):
                out.error("Authentication failed: invalid credentials")
                return False

            # Step 2: Download report via controller
            encoded_data = quote(json_lib.dumps(data_dict))
            download_resp = client.get(f"{url}{endpoint}?data={encoded_data}")

            if download_resp.status_code != 200:
                out.error(f"Download failed: HTTP {download_resp.status_code}")
                return False

            content_type = download_resp.headers.get("content-type", "")
            if "html" in content_type and download_resp.status_code == 200 and len(download_resp.content) < 500:
                out.error("Download returned HTML (likely redirect to login). Check credentials.")
                return False

            output_path.write_bytes(download_resp.content)
            size_kb = len(download_resp.content) / 1024
            if not quiet:
                out.success(f"Report saved: {output_path} ({size_kb:.1f} KB)")
            return True

    except httpx.ConnectError as exc:
        out.error(f"Connection failed: {exc}")
        return False
    except Exception as exc:
        out.error(f"Download failed: {exc}")
        return False


@app.command("export")
def report_export(
    ctx: typer.Context,
    instance_id: Annotated[int, typer.Argument(help="Report instance ID")],
    fmt: Annotated[str, typer.Option("--format", "-f", help="Export format: xlsx or pdf")] = "xlsx",
    output: Annotated[str | None, typer.Option("--output", "-o", help="Output file path")] = None,
) -> None:
    """Export a report instance to XLSX or PDF file.

    Downloads the report via Odoo's HTTP controller (headless, no browser).

    Examples:
        kctl-odoo report export 1 -o pnl.xlsx
        kctl-odoo report export 1 --format pdf -o pnl.pdf
    """
    actx: AppContext = ctx.obj
    out = actx.output
    c = actx.client

    if not _check_report_management(c, out):
        return

    if fmt not in ("xlsx", "pdf"):
        out.error("Format must be 'xlsx' or 'pdf'.")
        raise typer.Exit(1)

    # Read instance to get template_id and dates
    try:
        inst_records = c.search_read(
            "report.instance",
            domain=[("id", "=", instance_id)],
            fields=["id", "display_name", "template_id", "date_from", "date_to", "company_id", "target_move"],
            limit=1,
        )
    except RPCError as e:
        out.error(f"Failed to read report instance: {e.detail}")
        raise typer.Exit(1) from e

    if not inst_records:
        out.error(f"Report instance not found: {instance_id}")
        raise typer.Exit(1)

    inst = inst_records[0]
    tpl_id = inst["template_id"][0] if isinstance(inst.get("template_id"), list) else inst.get("template_id")
    company_id = inst["company_id"][0] if isinstance(inst.get("company_id"), list) else inst.get("company_id")

    # Build data dict for the controller — company_id is required
    if not company_id:
        try:
            user_data = c.read("res.users", [c.uid], ["company_id"])
            if user_data:
                co = user_data[0].get("company_id")
                company_id = co[0] if isinstance(co, list) else co
        except Exception:
            company_id = 1

    data_dict = {
        "template_id": tpl_id,
        "date_from": str(inst.get("date_from") or ""),
        "date_to": str(inst.get("date_to") or ""),
        "target_move": inst.get("target_move", "posted"),
        "company_id": company_id or 1,
    }

    # Resolve output path
    if output:
        output_path = Path(output)
    else:
        safe_name = (inst.get("display_name") or f"report_{instance_id}").replace(" ", "_").replace("/", "-")
        output_path = Path(f"{safe_name}.{fmt}")

    # Get connection details
    from kctl_odoo.core.config import resolve_connection

    url, database, username, api_key = resolve_connection(
        profile_name=actx.profile,
        url_override=actx.url_override,
        api_key_override=actx.api_key_override,
        database_override=actx.database_override,
        username_override=actx.username_override,
    )

    out.info(f"Downloading {fmt.upper()} for instance #{instance_id}...")
    success = _http_download_report(url, database, username, api_key, data_dict, fmt, output_path, out)

    if success and actx.json_mode:
        out.raw_json(
            {
                "instance_id": instance_id,
                "format": fmt,
                "file": str(output_path),
                "size": output_path.stat().st_size,
            }
        )
    elif not success:
        if actx.json_mode:
            out.raw_json(
                {
                    "instance_id": instance_id,
                    "format": fmt,
                    "hint": "Download from Odoo UI",
                }
            )
        raise typer.Exit(1)


@app.command("download")
def report_download(
    ctx: typer.Context,
    type_code: Annotated[
        str, typer.Argument(help="Report type code (e.g. profit_loss, balance_sheet, receivable_aging)")
    ],
    date_from: Annotated[str | None, typer.Option("--date-from", help="Start date (YYYY-MM-DD)")] = None,
    date_to: Annotated[str | None, typer.Option("--date-to", help="End date (YYYY-MM-DD)")] = None,
    as_of_date: Annotated[
        str | None,
        typer.Option("--date", help="As-of / cumulative date (for balance_sheet, receivable_aging, payable_aging)"),
    ] = None,
    fmt: Annotated[str, typer.Option("--format", "-f", help="Export format: xlsx or pdf")] = "xlsx",
    output: Annotated[str | None, typer.Option("--output", "-o", help="Output file path")] = None,
    force: Annotated[bool, typer.Option("--force", "-y", help="Skip confirmation")] = False,
) -> None:
    """Generate a report and export it in one step.

    Combines 'report run' + 'report export': creates a report.instance via
    the wizard, then exports to XLSX or PDF.

    Old short codes (pnl, bs, cf, tb, cma, ar_aging, ap_aging) are still
    accepted via backward-compatible aliases.

    Examples:
        kctl-odoo report download profit_loss --date-from 2026-01-01 --date-to 2026-01-31
        kctl-odoo report download balance_sheet --date 2026-03-31 --format pdf
        kctl-odoo report download receivable_aging --date 2026-03-27 -o aging.xlsx -y
        kctl-odoo report download pnl --date-from 2026-01-01 --date-to 2026-01-31  # alias
    """
    actx: AppContext = ctx.obj
    out = actx.output
    c = actx.client

    if not _check_report_management(c, out):
        return

    if fmt not in ("xlsx", "pdf"):
        out.error("Format must be 'xlsx' or 'pdf'.")
        raise typer.Exit(1)

    # Resolve aliases (e.g. pnl → profit_loss)
    type_code = _resolve_type_code(type_code)

    # 1. Resolve report type
    try:
        reg_records = c.search_read(
            "report.type.registry",
            domain=[("code", "=", type_code)],
            fields=["id", "name", "code", "date_mode"],
            limit=1,
        )
    except RPCError as e:
        out.error(f"Failed to query report type registry: {e.detail}")
        raise typer.Exit(1) from e

    if not reg_records:
        out.error(f"Report type not found: {type_code}")
        raise typer.Exit(1)

    reg = reg_records[0]

    # 2. Find template
    try:
        tpl_records = c.search_read(
            "report.template",
            domain=[("report_type_id", "=", reg["id"])],
            fields=["id", "name"],
            limit=1,
            order="sequence, id",
        )
    except RPCError as e:
        out.error(f"Failed to query report templates: {e.detail}")
        raise typer.Exit(1) from e

    if not tpl_records:
        with contextlib.suppress(RPCError):
            tpl_records = c.search_read(
                "report.template",
                domain=[("report_type", "=", type_code)],
                fields=["id", "name"],
                limit=1,
                order="sequence, id",
            )

    if not tpl_records:
        out.error(f"No report template found for type '{type_code}'.")
        raise typer.Exit(1)

    tpl = tpl_records[0]

    # 3. Resolve dates
    if as_of_date:
        resolved_date_to = as_of_date
        resolved_date_from = _fiscal_year_start(as_of_date)
    elif date_from and date_to:
        resolved_date_from = date_from
        resolved_date_to = date_to
    elif date_from and not date_to:
        resolved_date_from = date_from
        resolved_date_to = date.today().isoformat()
    elif date_to and not date_from:
        resolved_date_from = _fiscal_year_start(date_to)
        resolved_date_to = date_to
    else:
        resolved_date_from, resolved_date_to = _default_date_range()

    if not force:
        out.info(
            f"Generate + export '{reg.get('name')}' ({fmt.upper()}) for {resolved_date_from} to {resolved_date_to}"
        )
        if not typer.confirm("Proceed?"):
            raise typer.Exit(0)

    # 4. Create wizard and generate
    wizard_vals: dict = {
        "template_id": tpl["id"],
        "date_from": resolved_date_from,
        "date_to": resolved_date_to,
    }

    try:
        wizard_id = c.execute_kw("report.wizard", "create", [wizard_vals])
        result = c.execute_kw("report.wizard", "button_export_html", [[wizard_id]])
    except RPCError as e:
        out.error(f"Failed to generate report: {e.detail}")
        raise typer.Exit(1) from e

    # Extract instance ID
    instance_id = None
    if isinstance(result, dict):
        instance_id = result.get("res_id")

    if not instance_id:
        try:
            recent = c.search_read(
                "report.instance",
                domain=[("template_id", "=", tpl["id"])],
                fields=["id"],
                limit=1,
                order="create_date desc",
            )
            if recent:
                instance_id = recent[0]["id"]
        except RPCError:
            pass

    if not instance_id:
        out.error("Report was generated but instance ID could not be determined.")
        raise typer.Exit(1)

    out.success(f"Report instance #{instance_id} created.")

    # 5. Export via HTTP controller (headless download)
    # Resolve company_id (required by controller)
    dl_company_id = None
    try:
        user_data = c.read("res.users", [c.uid], ["company_id"])
        if user_data:
            co = user_data[0].get("company_id")
            dl_company_id = co[0] if isinstance(co, list) else co
    except Exception:
        dl_company_id = 1

    data_dict = {
        "template_id": tpl["id"],
        "date_from": resolved_date_from,
        "date_to": resolved_date_to,
        "company_id": dl_company_id or 1,
        "target_move": "posted",
    }

    output_path = Path(output) if output else Path(f"{type_code}_{resolved_date_from}_{resolved_date_to}.{fmt}")

    from kctl_odoo.core.config import resolve_connection

    url, database_name, username, api_key = resolve_connection(
        profile_name=actx.profile,
        url_override=actx.url_override,
        api_key_override=actx.api_key_override,
        database_override=actx.database_override,
        username_override=actx.username_override,
    )

    out.info(f"Downloading {fmt.upper()}...")
    success = _http_download_report(url, database_name, username, api_key, data_dict, fmt, output_path, out)

    if success and actx.json_mode:
        out.raw_json(
            {
                "instance_id": instance_id,
                "format": fmt,
                "file": str(output_path),
                "size": output_path.stat().st_size,
            }
        )
    elif not success:
        raise typer.Exit(1)


# ---------------------------------------------------------------------------
# Validate
# ---------------------------------------------------------------------------


@app.command("validate")
def report_validate(
    ctx: typer.Context,
    type_code: Annotated[
        str | None,
        typer.Argument(help="Report type code to validate, or 'all'"),
    ] = "all",
    date_from: Annotated[str | None, typer.Option("--date-from")] = None,
    date_to: Annotated[str | None, typer.Option("--date-to")] = None,
    company: Annotated[
        int | None,
        typer.Option("--company", help="Target company id (default: first active)"),
    ] = None,
    category: Annotated[
        str | None,
        typer.Option("--category", help="Filter registry by category"),
    ] = None,
) -> None:
    """Validate report quality — check reports generate with data.

    Downloads each report and checks for errors, empty files, and data rows.

    Examples:
        kctl-odoo report validate
        kctl-odoo report validate profit_loss
        kctl-odoo report validate all --date-from 2026-01-01 --date-to 2026-03-31
    """
    import tempfile

    actx: AppContext = ctx.obj
    out = actx.output
    c = actx.client

    if not _check_report_management(c, out):
        return

    if not date_to:
        date_to = date.today().strftime("%Y-%m-%d")
    if not date_from:
        date_from = date.today().replace(month=1, day=1).strftime("%Y-%m-%d")

    resolved = _resolve_type_code(type_code) if type_code and type_code != "all" else None

    try:
        domain: list = []
        if resolved:
            domain.append(("code", "=", resolved))
        if category:
            domain.append(("category", "=", category))
        types = c.search_read(
            "report.type.registry",
            domain=domain,
            fields=["id", "code", "name", "category"],
            order="category,sequence",
        )
    except RPCError as e:
        out.error(f"Failed to query: {e.detail}")
        raise typer.Exit(1) from e

    if not types:
        out.error(f"Not found: {type_code}")
        raise typer.Exit(1)

    # Resolve connection for HTTP download
    from kctl_odoo.core.config import resolve_connection

    url, database_name, username, api_key = resolve_connection(
        profile_name=actx.profile,
        url_override=actx.url_override,
        api_key_override=actx.api_key_override,
        database_override=actx.database_override,
        username_override=actx.username_override,
    )

    # Resolve company
    if company is not None:
        company_id = company
    else:
        company_id = 1
        try:
            cos = c.search_read("res.company", [("active", "=", True)], ["id"], limit=1)
            if cos:
                company_id = cos[0]["id"]
        except RPCError:
            pass

    out.info(f"Validating {len(types)} reports (company={company_id}, {date_from} to {date_to})")

    ok_count = warn_count = fail_count = 0
    results = []
    current_cat = ""

    for rt in types:
        code, category = rt["code"], rt["category"]

        if category != current_cat:
            current_cat = category
            out.info(f"\n  [{category}]")

        try:
            tpls = c.search_read("report.template", [("report_type", "=", code)], ["id"], limit=1)
        except RPCError:
            out.error(f"    FAIL  {code:40s} template query error")
            fail_count += 1
            results.append({"code": code, "status": "fail"})
            continue
        if not tpls:
            out.warn(f"    SKIP  {code:40s} no template")
            warn_count += 1
            results.append({"code": code, "status": "skip"})
            continue

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            tmp_path = Path(tmp.name)

        try:
            data_dict = {
                "template_id": tpls[0]["id"],
                "date_from": date_from,
                "date_to": date_to,
                "company_id": company_id,
                "target_move": "posted",
            }

            ok = _http_download_report(
                url, database_name, username, api_key, data_dict, "xlsx", tmp_path, out, quiet=True
            )

            if not ok:
                out.error(f"    FAIL  {code:40s} download error")
                fail_count += 1
                results.append({"code": code, "status": "fail"})
                continue

            fsize = tmp_path.stat().st_size
            if fsize == 0:
                out.error(f"    FAIL  {code:40s} empty")
                fail_count += 1
                results.append({"code": code, "status": "fail"})
                continue

            # Count data rows
            data_rows = 0
            try:
                import openpyxl

                wb = openpyxl.load_workbook(tmp_path)
                ws = wb.active
                for row in ws.iter_rows(min_row=5, max_row=200, values_only=True):
                    if not row or not row[0]:
                        continue
                    lbl = str(row[0]).strip()
                    if lbl in ("", "Generated by Report Management"):
                        continue
                    for v in row[1:]:
                        if v is not None and v != "" and v != 0:
                            data_rows += 1
                            break
                    else:
                        if len(lbl) > 2:
                            data_rows += 1
                wb.close()
            except Exception:
                data_rows = -1

            if data_rows > 0:
                out.success(f"    OK    {code:40s} {fsize:>6}B {data_rows:>3} rows")
                ok_count += 1
            elif data_rows == 0:
                out.warn(f"    WARN  {code:40s} {fsize:>6}B no data")
                warn_count += 1
            else:
                out.success(f"    OK    {code:40s} {fsize:>6}B")
                ok_count += 1

            results.append(
                {"code": code, "status": "ok" if data_rows != 0 else "warn", "size": fsize, "rows": data_rows}
            )

        except Exception as e:
            out.error(f"    FAIL  {code:40s} {str(e)[:50]}")
            fail_count += 1
            results.append({"code": code, "status": "fail"})
        finally:
            tmp_path.unlink(missing_ok=True)

    total = ok_count + warn_count + fail_count
    out.info(f"\n{'=' * 60}")

    if fail_count == 0 and warn_count == 0:
        out.success(f"  ALL {total} REPORTS PASS")
    elif fail_count == 0:
        out.success(f"  {ok_count} pass, {warn_count} warnings")
    else:
        out.error(f"  {fail_count} FAILURES | {ok_count} ok | {warn_count} warn")

    if actx.json_mode:
        import json as _json

        print(
            _json.dumps(
                {"total": total, "ok": ok_count, "warn": warn_count, "fail": fail_count, "results": results}, indent=2
            )
        )


# ===================================================================
# VERIFY-ACCURATE — row-by-row compare to a stored Accurate reference
# ===================================================================


@app.command("verify-accurate")
def report_verify_accurate(
    ctx: typer.Context,
    reference: Annotated[
        Path,
        typer.Option("--reference", "-r", help="Path to the Accurate reference JSON", exists=True),
    ],
    company: Annotated[
        int | None,
        typer.Option("--company", help="Override company_id (default: from reference)"),
    ] = None,
    date_from: Annotated[
        str | None,
        typer.Option("--date-from", help="Override start date (default: from reference)"),
    ] = None,
    date_to: Annotated[
        str | None,
        typer.Option("--date-to", help="Override end date (default: from reference)"),
    ] = None,
    tolerance: Annotated[
        float, typer.Option("--tolerance", help="Absolute amount delta tolerated (default 1.0 rupiah)")
    ] = 1.0,
    pct_tolerance: Annotated[float, typer.Option("--pct-tolerance", help="% delta tolerated (default 0.05)")] = 0.05,
) -> None:
    """Compare live engine output row-by-row against a stored Accurate XLSX
    reference (exported as JSON). Exits non-zero on any mismatch; prints a
    structured per-row diff naming the line_id, expected vs actual, and delta.

    The reference JSON is produced once by hand from an Accurate XLSX export
    and lives in `report_management/tests/accurate_refs/`.

    Examples:
        kctl-odoo --profile tpp-odoo-erp report verify-accurate \\
          -r tests/accurate_refs/laba_rugi_cv_sumber__2026-04-01_2026-04-15.json
    """
    import json as _json

    actx: AppContext = ctx.obj
    out = actx.output
    c = actx.client

    try:
        expected = _json.loads(reference.read_text())
    except Exception as e:
        out.error(f"Failed to parse reference: {e}")
        raise typer.Exit(2) from e

    co_id = company or expected.get("company_id_in_odoo")
    df = date_from or expected["period"]["date_from"]
    dt = date_to or expected["period"]["date_to"]
    report_type = expected["report_type"]

    if not co_id:
        out.error("--company required when reference has no company_id_in_odoo")
        raise typer.Exit(2)

    try:
        tpls = c.search_read(
            "report.template",
            [("report_type", "=", report_type), ("code", "=", report_type)],
            ["id", "name"],
            limit=1,
        )
        if not tpls:
            tpls = c.search_read(
                "report.template",
                [("report_type", "=", report_type)],
                ["id", "name"],
                limit=1,
            )
    except RPCError as e:
        out.error(f"Template lookup failed: {e.detail}")
        raise typer.Exit(1) from e

    if not tpls:
        out.error(f"No template for report_type={report_type!r}")
        raise typer.Exit(1)

    tpl = tpls[0]
    out.info(f"Template: {tpl['name']!r} (id={tpl['id']})")
    out.info(f"Company: {co_id}  Period: {df}..{dt}  Source: {expected.get('source', '?')}")
    out.info(f"Reference: {reference.name}\n")

    try:
        rd = c.execute_kw(
            "report.report_management.engine",
            "compute_report_data",
            [
                [],
                {
                    "template_id": tpl["id"],
                    "date_from": df,
                    "date_to": dt,
                    "company_id": co_id,
                    "company_ids": [],
                    "target_move": "posted",
                    "divider": 1,
                    "show_variance": False,
                    "show_percentage": False,
                    "monthly_breakdown": False,
                    "expanded_line_ids": [],
                    "page_offset": 0,
                    "page_limit": 0,
                },
            ],
            {"context": {"allowed_company_ids": [co_id]}},
        )
    except RPCError as e:
        out.error(f"compute_report_data failed: {e.detail}")
        raise typer.Exit(1) from e

    rows = []
    for t in rd.get("tables") or []:
        rows.extend(t.get("rows") or [])
    if not rows:
        rows = rd.get("lines") or []

    # Build an id lookup — line_id may be numeric (template.line.id) or
    # the xmlid suffix. We also index by the template line's external id
    # if resolvable, and by the label as a fallback.
    by_line_id = {}
    line_ids_numeric = [r.get("line_id") for r in rows if isinstance(r.get("line_id"), int)]
    xmlid_by_id = {}
    if line_ids_numeric:
        try:
            data = c.execute_kw(
                "ir.model.data",
                "search_read",
                [
                    [("model", "=", "report.template.line"), ("res_id", "in", line_ids_numeric)],
                    ["res_id", "name"],
                ],
            )
            xmlid_by_id = {d["res_id"]: d["name"] for d in data}
        except RPCError:
            pass
    by_label = {}
    for r in rows:
        lid = r.get("line_id")
        if lid is not None:
            by_line_id[str(lid)] = r
        if isinstance(lid, int) and lid in xmlid_by_id:
            by_line_id[xmlid_by_id[lid]] = r
        lbl = (r.get("label") or r.get("name") or "").strip()
        if lbl:
            by_label.setdefault(lbl, r)

    fails = 0
    results = []
    for exp in expected["totals"]:
        lid = exp["label_id"]
        got = by_line_id.get(lid)
        if got is None:
            got = by_label.get(exp.get("label_accurate", ""))
        if got is None:
            out.error(f"  MISSING  {lid:30s} (Accurate: {exp.get('label_accurate', '?')!r})")
            fails += 1
            results.append({"label_id": lid, "status": "missing"})
            continue
        v = got.get("values") or {}
        amt = float(v.get("amount") or v.get("actual") or 0.0)
        pct = float(v.get("pct") or 0.0)
        delta_amt = abs(amt - exp["amount"])
        delta_pct = abs(pct - exp["pct"])
        if delta_amt > tolerance or delta_pct > pct_tolerance:
            out.error(
                f"  FAIL     {lid:30s} "
                f"amt={amt:>18,.2f} (expected {exp['amount']:>18,.2f}, Δ={amt - exp['amount']:>+14,.2f})  "
                f"pct={pct:>6.2f} (expected {exp['pct']:>6.2f})"
            )
            fails += 1
            results.append(
                {
                    "label_id": lid,
                    "status": "fail",
                    "amount": amt,
                    "expected_amount": exp["amount"],
                    "delta_amount": amt - exp["amount"],
                    "pct": pct,
                    "expected_pct": exp["pct"],
                }
            )
        else:
            out.success(f"  OK       {lid:30s} amt={amt:>18,.2f}  pct={pct:>6.2f}")
            results.append({"label_id": lid, "status": "ok", "amount": amt, "pct": pct})

    out.info(f"\n{'=' * 70}")
    if fails == 0:
        out.success(f"  All {len(expected['totals'])} totals match Accurate within tolerance")
    else:
        out.error(f"  {fails}/{len(expected['totals'])} totals DIFFER from Accurate")

    if actx.json_mode:
        print(
            _json.dumps(
                {
                    "reference": str(reference),
                    "company_id": co_id,
                    "period": {"date_from": df, "date_to": dt},
                    "tolerance": tolerance,
                    "pct_tolerance": pct_tolerance,
                    "fails": fails,
                    "results": results,
                },
                indent=2,
            )
        )

    if fails:
        raise typer.Exit(1)


# ===================================================================
# VERIFY-MATH — consolidation math check across reports
# ===================================================================


def _extract_totals(rd: dict) -> dict[str, float]:
    """Pull amount-shaped values from every `total`/`subtotal` row in the
    returned report data. Returns a `{row_label: amount}` map so we can
    compare the same logical totals across company sets."""
    out: dict[str, float] = {}
    AMOUNT_KEYS = ("actual", "balance", "debit", "total", "amount", "subtotal", "net")
    tables = rd.get("tables") or []
    rows: list = []
    for t in tables:
        rows.extend(t.get("rows") or [])
    if not rows:
        rows = rd.get("lines") or []
    for r in rows:
        if r.get("kind") not in ("total", "subtotal"):
            continue
        label = r.get("label") or r.get("name") or r.get("line_id") or "?"
        vals = r.get("values") or {}
        amt = None
        for k in AMOUNT_KEYS:
            v = vals.get(k)
            if isinstance(v, (int, float)):
                amt = float(v)
                break
        if amt is None:
            amt = float(r.get("actual") or 0)
        out[label] = out.get(label, 0.0) + amt
    return out


@app.command("verify-math")
def report_verify_math(
    ctx: typer.Context,
    type_code: Annotated[
        str | None,
        typer.Argument(help="Report type code to verify, or 'all'"),
    ] = "all",
    companies: Annotated[
        str | None,
        typer.Option("--companies", help="Comma-separated company ids for consolidation check (e.g. '1,3,4')"),
    ] = None,
    category: Annotated[
        str | None,
        typer.Option("--category", "-c", help="Filter by registry category (financial, aging, sales, ...)"),
    ] = None,
    date_from: Annotated[str | None, typer.Option("--date-from")] = None,
    date_to: Annotated[str | None, typer.Option("--date-to")] = None,
    tolerance: Annotated[float, typer.Option("--tolerance", help="Absolute delta tolerated (default 1.0)")] = 1.0,
) -> None:
    """Verify consolidation math: consolidated totals should equal the sum
    of per-company totals. Catches silent data-leakage regressions across
    handlers (account resolution, company filters, SQL predicates).

    Examples:
        kctl-odoo report verify-math --companies 1,3
        kctl-odoo report verify-math --companies 1,3,4 --category financial
        kctl-odoo report verify-math profit_loss --companies 1,4
    """
    actx: AppContext = ctx.obj
    out = actx.output
    c = actx.client

    if not _check_report_management(c, out):
        return

    if not date_to:
        date_to = date.today().strftime("%Y-%m-%d")
    if not date_from:
        date_from = date.today().replace(month=1, day=1).strftime("%Y-%m-%d")

    # Parse company list
    if not companies:
        out.error("--companies is required (e.g. --companies 1,3)")
        raise typer.Exit(2)
    try:
        co_ids = [int(x.strip()) for x in companies.split(",") if x.strip()]
    except ValueError as e:
        out.error(f"Invalid --companies value: {e}")
        raise typer.Exit(2) from e
    if len(co_ids) < 2:
        out.error("--companies must include at least 2 ids to verify consolidation math")
        raise typer.Exit(2)

    # Pre-flight: RPC user must have access to all target companies
    try:
        uid = c.uid
        users = c.execute_kw(
            "res.users",
            "search_read",
            [[("id", "=", uid)], ["login", "company_ids"]],
            {"context": {"active_test": False}, "limit": 1},
        )
        if users:
            user_co_ids = set(users[0].get("company_ids") or [])
            missing = [cid for cid in co_ids if cid not in user_co_ids]
            if missing:
                out.error(
                    f"RPC user '{users[0]['login']}' lacks access to companies {missing}. "
                    f"User company_ids={sorted(user_co_ids)}. "
                    f"Either grant access via `res_company_users_rel` or run the "
                    f"verification via `kctl-odoo local shell` instead."
                )
                raise typer.Exit(2)
    except RPCError:
        pass  # pre-flight is best-effort

    # Filter report types
    resolved = _resolve_type_code(type_code) if type_code and type_code != "all" else None
    domain: list = []
    if resolved:
        domain.append(("code", "=", resolved))
    if category:
        domain.append(("category", "=", category))
    try:
        types = c.search_read(
            "report.type.registry",
            domain=domain,
            fields=["id", "code", "name", "category"],
            order="category,sequence",
        )
    except RPCError as e:
        out.error(f"Failed to query registry: {e.detail}")
        raise typer.Exit(1) from e
    if not types:
        out.error("No matching report types")
        raise typer.Exit(1)

    _last_error: str = ""

    def _run(template_id: int, target_co_ids: list[int]) -> dict | None:
        data = {
            "instance_id": 0,
            "template_id": template_id,
            "date_from": date_from,
            "date_to": date_to,
            "company_id": target_co_ids[0],
            "company_ids": target_co_ids if len(target_co_ids) > 1 else [],
            "target_move": "posted",
            "divider": 1,
            "show_variance": False,
            "show_percentage": False,
            "monthly_breakdown": False,
            "expanded_line_ids": [],
            "page_offset": 0,
            "page_limit": 0,
        }
        try:
            return c.execute_kw(
                "report.report_management.engine",
                "compute_report_data",
                [[], data],
                {"context": {"allowed_company_ids": target_co_ids, "active_test": False}},
            )
        except RPCError as e:
            nonlocal _last_error
            _last_error = str(e)[:120]
            return None

    out.info(f"Verifying consolidation math: companies={co_ids} period={date_from}..{date_to} tolerance={tolerance}")
    out.info(f"Checking {len(types)} report types\n")

    ok_count = fail_count = skip_count = 0
    results: list[dict] = []
    current_cat = ""

    for rt in types:
        code, cat_name = rt["code"], rt["category"]
        if cat_name != current_cat:
            current_cat = cat_name
            out.info(f"  [{cat_name}]")
        try:
            tpls = c.search_read("report.template", [("report_type", "=", code)], ["id"], limit=1)
        except RPCError:
            out.warn(f"    SKIP  {code:30s} template query error")
            skip_count += 1
            continue
        if not tpls:
            skip_count += 1
            continue
        tpl_id = tpls[0]["id"]

        per_co_totals: dict[int, dict[str, float]] = {}
        crashed = False
        for cid in co_ids:
            rd = _run(tpl_id, [cid])
            if rd is None:
                crashed = True
                break
            per_co_totals[cid] = _extract_totals(rd)
        if crashed:
            out.error(f"    FAIL  {code:30s} per-company call crashed: {_last_error}")
            fail_count += 1
            results.append({"code": code, "status": "fail", "reason": _last_error or "crash"})
            continue

        rd_cons = _run(tpl_id, co_ids)
        if rd_cons is None:
            out.error(f"    FAIL  {code:30s} consolidated call crashed")
            fail_count += 1
            results.append({"code": code, "status": "fail", "reason": "consolidated call crashed"})
            continue
        cons_totals = _extract_totals(rd_cons)

        # Compare: consolidated[label] should equal sum(per_co[label])
        all_labels = set(cons_totals) | {lbl for m in per_co_totals.values() for lbl in m}
        max_delta = 0.0
        worst_label = ""
        for lbl in all_labels:
            expected = sum(per_co_totals[cid].get(lbl, 0.0) for cid in co_ids)
            got = cons_totals.get(lbl, 0.0)
            delta = abs(got - expected)
            if delta > max_delta:
                max_delta = delta
                worst_label = lbl

        if max_delta <= tolerance:
            out.success(f"    OK    {code:30s} Δ_max={max_delta:>12,.2f} ({len(all_labels)} rows)")
            ok_count += 1
            results.append({"code": code, "status": "ok", "max_delta": max_delta, "rows": len(all_labels)})
        else:
            out.error(f"    FAIL  {code:30s} Δ={max_delta:>12,.2f} at {worst_label!r}")
            fail_count += 1
            results.append(
                {
                    "code": code,
                    "status": "fail",
                    "max_delta": max_delta,
                    "worst_label": worst_label,
                    "rows": len(all_labels),
                }
            )

    total = ok_count + fail_count + skip_count
    out.info(f"\n{'=' * 60}")
    if fail_count == 0:
        out.success(f"  {ok_count}/{total} reports consolidation-math OK  ({skip_count} no-template)")
    else:
        out.error(f"  {fail_count} FAILURES | {ok_count} ok | {skip_count} skipped")

    if actx.json_mode:
        import json as _json

        print(
            _json.dumps(
                {
                    "companies": co_ids,
                    "date_from": date_from,
                    "date_to": date_to,
                    "tolerance": tolerance,
                    "ok": ok_count,
                    "fail": fail_count,
                    "skipped": skip_count,
                    "results": results,
                },
                indent=2,
            )
        )

    if fail_count > 0:
        raise typer.Exit(1)


# ===================================================================
# CATALOG — report type overview with template counts
# ===================================================================


@app.command("catalog")
def report_catalog(
    ctx: typer.Context,
) -> None:
    """Show all report types with template count and status.

    Queries report.type.registry and counts associated templates per type.

    Examples:
        kctl-odoo report catalog
        kctl-odoo report catalog --json
    """
    actx: AppContext = ctx.obj
    out = actx.output
    c = actx.client

    if not _check_report_management(c, out):
        if actx.json_mode:
            out.raw_json([])
        return

    try:
        records = c.search_read(
            "report.type.registry",
            domain=[],
            fields=["id", "code", "name", "category", "active"],
            limit=500,
            order="category, sequence, name",
        )
    except RPCError as e:
        out.error(f"Failed to fetch report types: {e}")
        raise typer.Exit(1) from e

    if not records:
        out.info("No report types found.")
        if actx.json_mode:
            out.raw_json([])
        return

    # Count templates per report type
    template_counts: dict[str, int] = {}
    for r in records:
        code = r.get("code", "")
        try:
            count = c.search_count(
                "report.template",
                [("report_type", "=", code)],
            )
            template_counts[code] = count
        except RPCError:
            template_counts[code] = 0

    rows = []
    json_data = []
    for r in records:
        code = r.get("code", "")
        cat = r.get("category", "") or ""
        cat_display = {
            "financial": "[green]financial[/green]",
            "aging": "[yellow]aging[/yellow]",
            "management": "[cyan]management[/cyan]",
            "custom": "[dim]custom[/dim]",
        }.get(cat, cat)
        status = "[green]active[/green]" if r.get("active") else "[dim]inactive[/dim]"
        tpl_count = template_counts.get(code, 0)

        rows.append(
            [
                code,
                (r.get("name") or "")[:40],
                cat_display,
                str(tpl_count),
                status,
            ]
        )
        json_data.append(
            {
                "code": code,
                "name": r.get("name"),
                "category": cat,
                "templates": tpl_count,
                "active": r.get("active"),
            }
        )

    out.table(
        f"Report Catalog ({len(records)} types)",
        [
            ("Code", "cyan"),
            ("Name", ""),
            ("Category", ""),
            ("Templates", ""),
            ("Status", ""),
        ],
        rows,
        data_for_json=json_data,
    )


@app.command("schedule")
def schedule_report(
    ctx: typer.Context,
    type_code: Annotated[str, typer.Argument(help="Report type code (e.g. profit_loss)")],
    cron: Annotated[str, typer.Option("--cron", help="Cron schedule expression")] = "0 8 * * 1",
    email_to: Annotated[str | None, typer.Option("--email-to", help="Email address to send report to")] = None,
) -> None:
    """Schedule a report to run periodically.

    Shows instructions for setting up a cron job to generate and email reports.
    """
    actx: AppContext = ctx.obj
    out = actx.output

    type_code = _resolve_type_code(type_code)
    run_cmd = f"kctl-odoo report run {type_code}"
    export_cmd = f"kctl-odoo report export {type_code} --format xlsx"

    sections: list[tuple[str, list[tuple[str, str]]]] = [
        (
            "Report Details",
            [
                ("Type Code", type_code),
                ("Cron Schedule", cron),
                ("Email To", email_to or "(not set)"),
            ],
        ),
        (
            "Crontab Entry",
            [
                ("Generate & export", f"{cron}  {export_cmd}"),
            ],
        ),
    ]

    if email_to:
        mail_cmd = f"{cron}  {export_cmd} && mail -s 'Report: {type_code}' {email_to} < /dev/null"
        sections.append(
            (
                "With Email",
                [
                    ("Send via mail", mail_cmd),
                    (
                        "Send via kctl",
                        f"{cron}  {run_cmd} --email {email_to}",
                    ),
                ],
            )
        )

    sections.append(
        (
            "Setup Instructions",
            [
                ("1. Edit crontab", "crontab -e"),
                ("2. Add entry", f"{cron}  {export_cmd}"),
                ("3. Verify cron", "crontab -l"),
                (
                    "4. Odoo ir.cron",
                    "Or use Odoo Technical > Automation > Scheduled Actions",
                ),
            ],
        )
    )

    json_data = {
        "type_code": type_code,
        "cron": cron,
        "email_to": email_to,
        "run_command": run_cmd,
        "export_command": export_cmd,
    }

    out.detail(f"Schedule Report: {type_code}", sections, data_for_json=json_data)


@app.command("batch-download")
def batch_download(
    ctx: typer.Context,
    category: Annotated[str | None, typer.Option("--category", "-c", help="Filter by report category")] = None,
    date_from: Annotated[str | None, typer.Option("--date-from", help="Start date (YYYY-MM-DD)")] = None,
    date_to: Annotated[str | None, typer.Option("--date-to", help="End date (YYYY-MM-DD)")] = None,
    output_dir: Annotated[str, typer.Option("--output-dir", "-o", help="Output directory")] = ".",
) -> None:
    """Bulk export multiple reports to files."""
    actx: AppContext = ctx.obj
    out = actx.output
    c = actx.client

    output_path = Path(output_dir)
    output_path.mkdir(parents=True, exist_ok=True)

    # Default date range: last 30 days
    if not date_to:
        date_to = date.today().isoformat()
    if not date_from:
        date_from = (date.today() - timedelta(days=30)).isoformat()

    # Try to list report types from report_management module
    domain: list = [("active", "=", True)]
    if category:
        domain.append(("category", "=", category))

    try:
        report_types = c.search_read(
            "report.type.registry",
            domain=domain,
            fields=["id", "code", "name", "category"],
            limit=50,
            order="category, name",
        )
    except RPCError:
        out.warn("report.type.registry model not available. Is report_management installed?")
        out.info("Falling back to ir.actions.report...")
        try:
            report_types_raw = c.search_read(
                "ir.actions.report",
                domain=[("report_type", "in", ["qweb-pdf", "qweb-xlsx"])],
                fields=["id", "name", "report_name", "report_type"],
                limit=50,
                order="name",
            )
            report_types = [
                {
                    "id": r["id"],
                    "code": r["report_name"],
                    "name": r["name"],
                    "category": r.get("report_type", ""),
                }
                for r in report_types_raw
            ]
        except RPCError as e:
            out.error(f"Could not list reports: {e.detail}")
            raise typer.Exit(1) from e

    if not report_types:
        out.info("No report types found" + (f" for category '{category}'" if category else "") + ".")
        if actx.json_mode:
            out.raw_json([])
        return

    out.info(f"Found {len(report_types)} report type(s) to download.")
    out.info(f"Date range: {date_from} to {date_to}")
    out.info(f"Output directory: {output_path.resolve()}")

    results = []
    for rt in report_types:
        code = rt.get("code") or rt.get("report_name") or str(rt["id"])
        name = rt.get("name", code)
        file_name = f"{code.replace('.', '_').replace('/', '_')}.xlsx"
        file_path = output_path / file_name

        try:
            # Try to use report_management export if available
            instance_domain: list = [
                ("type_code", "=", code),
                ("date_from", ">=", date_from),
                ("date_to", "<=", date_to),
            ]
            instances = c.search_read(
                "report.instance",
                domain=instance_domain,
                fields=["id", "name", "state"],
                limit=1,
                order="create_date desc",
            )
            if instances:
                inst = instances[0]
                raw = c.execute_kw("report.instance", "action_export_xlsx", [[inst["id"]]])
                if raw and isinstance(raw, dict):
                    data = raw.get("data") or raw.get("file_data")
                    if data:
                        file_path.write_bytes(base64.b64decode(data))
                        results.append({"code": code, "name": name, "file": str(file_path), "status": "ok"})
                        out.info(f"  [green]OK[/green] {name} -> {file_name}")
                        continue

            # Fallback: generate a placeholder file
            file_path.write_text(
                f"Report: {name}\nCode: {code}\n"
                f"Date range: {date_from} to {date_to}\n"
                f"Generated at: {date.today().isoformat()}\n"
            )
            results.append({"code": code, "name": name, "file": str(file_path), "status": "placeholder"})
            out.info(f"  [yellow]PLACEHOLDER[/yellow] {name} -> {file_name}")

        except (RPCError, Exception) as e:
            results.append({"code": code, "name": name, "file": None, "status": "error", "error": str(e)})
            out.warn(f"  [red]FAILED[/red] {name}: {e}")

    ok_count = sum(1 for r in results if r["status"] == "ok")
    placeholder_count = sum(1 for r in results if r["status"] == "placeholder")
    err_count = sum(1 for r in results if r["status"] == "error")

    out.info(f"\nBatch download complete: {ok_count} OK, {placeholder_count} placeholder, {err_count} failed")
    if actx.json_mode:
        out.raw_json(results)
