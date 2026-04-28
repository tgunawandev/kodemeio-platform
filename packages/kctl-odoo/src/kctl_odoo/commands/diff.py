"""Instance diff/comparison commands."""

from __future__ import annotations

from pathlib import Path
from typing import Annotated

import typer

from kctl_odoo.core.callbacks import AppContext
from kctl_odoo.core.client import OdooClient
from kctl_odoo.core.config import resolve_connection

app = typer.Typer(help="Compare Odoo instances across profiles.")


def _get_client_for_profile(profile_name: str) -> OdooClient:
    """Build an OdooClient from a named config profile."""
    url, db, user, key = resolve_connection(
        profile_name=profile_name,
        url_override=None,
        api_key_override=None,
        database_override=None,
        username_override=None,
    )
    return OdooClient(base_url=url, database=db, username=user, api_key=key)


def _fetch_installed_modules(client: OdooClient) -> dict[str, dict]:
    """Return {technical_name: {version, state}} for all non-uninstalled modules."""
    records = client.search_read(
        "ir.module.module",
        domain=[("state", "!=", "uninstalled")],
        fields=["name", "installed_version", "state"],
        limit=0,
        order="name",
    )
    return {r["name"]: {"version": r.get("installed_version") or "", "state": r.get("state", "")} for r in records}


def _fetch_config_params(client: OdooClient, pattern: str | None = None) -> dict[str, str]:
    """Return {key: value} for system parameters."""
    domain: list = []
    if pattern:
        domain.append(("key", "ilike", pattern))
    records = client.search_read(
        "ir.config_parameter",
        domain=domain,
        fields=["key", "value"],
        limit=0,
        order="key",
    )
    return {r["key"]: str(r.get("value", "")) for r in records}


def _fetch_users(client: OdooClient) -> dict[str, dict]:
    """Return {login: {name, active}} for all users."""
    records = client.search_read(
        "res.users",
        domain=[],
        fields=["login", "name", "active"],
        limit=0,
        order="login",
    )
    return {r["login"]: {"name": r.get("name", ""), "active": r.get("active", False)} for r in records}


@app.command()
def modules(
    ctx: typer.Context,
    profile_a: Annotated[str, typer.Argument(help="First profile name")],
    profile_b: Annotated[str, typer.Argument(help="Second profile name")],
    output: Annotated[
        Path | None,
        typer.Option("--output", "-o", help="Export comparison to Excel (.xlsx)"),
    ] = None,
) -> None:
    """Compare installed modules between two Odoo instances.

    Shows modules only in A, only in B, and those with differing versions.

    Examples:
      kctl-odoo diff modules production staging
      kctl-odoo diff modules mac-odoo-erp tpp-odoo-erp -o diff.xlsx
    """
    actx: AppContext = ctx.obj
    out = actx.output

    out.info(f"Connecting to profile '{profile_a}'...")
    client_a = _get_client_for_profile(profile_a)
    mods_a = _fetch_installed_modules(client_a)
    client_a.close()

    out.info(f"Connecting to profile '{profile_b}'...")
    client_b = _get_client_for_profile(profile_b)
    mods_b = _fetch_installed_modules(client_b)
    client_b.close()

    keys_a = set(mods_a.keys())
    keys_b = set(mods_b.keys())

    only_a = sorted(keys_a - keys_b)
    only_b = sorted(keys_b - keys_a)
    common = sorted(keys_a & keys_b)

    different: list[tuple[str, str, str]] = []
    for name in common:
        va = mods_a[name]["version"]
        vb = mods_b[name]["version"]
        if va != vb:
            different.append((name, va, vb))

    # JSON output
    if actx.json_mode:
        out.raw_json(
            {
                "profile_a": profile_a,
                "profile_b": profile_b,
                "summary": {
                    "total_a": len(mods_a),
                    "total_b": len(mods_b),
                    "only_a": len(only_a),
                    "only_b": len(only_b),
                    "different_version": len(different),
                },
                "only_a": only_a,
                "only_b": only_b,
                "different_version": [
                    {"name": n, f"{profile_a}_version": va, f"{profile_b}_version": vb} for n, va, vb in different
                ],
            }
        )
        return

    # Summary
    out.detail(
        "Module Comparison",
        [
            (
                "Summary",
                [
                    (profile_a, f"{len(mods_a)} modules"),
                    (profile_b, f"{len(mods_b)} modules"),
                    (f"Only in {profile_a}", str(len(only_a))),
                    (f"Only in {profile_b}", str(len(only_b))),
                    ("Version differs", str(len(different))),
                ],
            ),
        ],
    )

    if only_a:
        rows = [[name, mods_a[name]["state"]] for name in only_a]
        out.table(
            f"Only in {profile_a} ({len(only_a)})",
            [("Module", ""), ("State", "dim")],
            rows,
        )

    if only_b:
        rows = [[name, mods_b[name]["state"]] for name in only_b]
        out.table(
            f"Only in {profile_b} ({len(only_b)})",
            [("Module", ""), ("State", "dim")],
            rows,
        )

    if different:
        rows = [[n, va, vb] for n, va, vb in different]
        out.table(
            f"Version Differences ({len(different)})",
            [("Module", ""), (f"{profile_a} Version", "cyan"), (f"{profile_b} Version", "yellow")],
            rows,
        )

    if not only_a and not only_b and not different:
        out.success("Instances are identical in module state and versions.")

    if output:
        _write_diff_xlsx(output, profile_a, profile_b, mods_a, mods_b, only_a, only_b, common, different)
        out.success(f"Wrote {output}")


def _write_diff_xlsx(
    output_path: Path,
    label_a: str,
    label_b: str,
    mods_a: dict[str, dict],
    mods_b: dict[str, dict],
    only_a: list[str],
    only_b: list[str],
    common: list[str],
    different: list[tuple[str, str, str]],
) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    hdr_font = Font(bold=True, color="FFFFFF", size=11)
    hdr_fill = PatternFill("solid", fgColor="2F5496")
    only_a_fill = PatternFill("solid", fgColor="DAEEF3")
    only_b_fill = PatternFill("solid", fgColor="FDE9D9")
    diff_fill = PatternFill("solid", fgColor="FFEB9C")
    same_fill = PatternFill("solid", fgColor="C6EFCE")
    border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )

    def write_headers(ws, headers, widths=None):
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = hdr_font
            cell.fill = hdr_fill
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
        if widths:
            for i, w in enumerate(widths, 1):
                ws.column_dimensions[get_column_letter(i)].width = w
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    # Build diff set for quick lookup
    diff_names = {n for n, _, _ in different}

    # ── Sheet 1: Side-by-Side ──
    ws1 = wb.active
    ws1.title = "Side-by-Side"
    headers1 = [
        "Module",
        f"{label_a} Status",
        f"{label_a} Version",
        f"{label_b} Status",
        f"{label_b} Version",
        "Comparison",
    ]
    write_headers(ws1, headers1, [35, 14, 16, 14, 16, 18])

    all_names = sorted(set(mods_a.keys()) | set(mods_b.keys()))
    row = 2
    for name in all_names:
        in_a = name in mods_a
        in_b = name in mods_b
        a_state = mods_a[name]["state"] if in_a else ""
        a_ver = mods_a[name]["version"] if in_a else ""
        b_state = mods_b[name]["state"] if in_b else ""
        b_ver = mods_b[name]["version"] if in_b else ""

        if not in_b:
            comparison = f"Only {label_a}"
            fill = only_a_fill
        elif not in_a:
            comparison = f"Only {label_b}"
            fill = only_b_fill
        elif name in diff_names:
            comparison = "Version differs"
            fill = diff_fill
        else:
            comparison = "Same"
            fill = same_fill

        values = [name, a_state, a_ver, b_state, b_ver, comparison]
        for col_idx, val in enumerate(values, 1):
            cell = ws1.cell(row=row, column=col_idx, value=val)
            cell.border = border
            cell.fill = fill
        row += 1

    # ── Sheet 2: Only in A ──
    ws2 = wb.create_sheet(f"Only {label_a}")
    write_headers(ws2, ["Module", "State", "Version"], [35, 14, 16])
    for r_idx, name in enumerate(only_a, 2):
        for c_idx, val in enumerate([name, mods_a[name]["state"], mods_a[name]["version"]], 1):
            cell = ws2.cell(row=r_idx, column=c_idx, value=val)
            cell.border = border
            cell.fill = only_a_fill

    # ── Sheet 3: Only in B ──
    ws3 = wb.create_sheet(f"Only {label_b}")
    write_headers(ws3, ["Module", "State", "Version"], [35, 14, 16])
    for r_idx, name in enumerate(only_b, 2):
        for c_idx, val in enumerate([name, mods_b[name]["state"], mods_b[name]["version"]], 1):
            cell = ws3.cell(row=r_idx, column=c_idx, value=val)
            cell.border = border
            cell.fill = only_b_fill

    # ── Sheet 4: Version Differences ──
    ws4 = wb.create_sheet("Version Differs")
    write_headers(ws4, ["Module", f"{label_a} Version", f"{label_b} Version"], [35, 20, 20])
    for r_idx, (name, va, vb) in enumerate(different, 2):
        for c_idx, val in enumerate([name, va, vb], 1):
            cell = ws4.cell(row=r_idx, column=c_idx, value=val)
            cell.border = border
            cell.fill = diff_fill

    wb.save(output_path)


@app.command()
def config(
    ctx: typer.Context,
    profile_a: Annotated[str, typer.Argument(help="First profile name")],
    profile_b: Annotated[str, typer.Argument(help="Second profile name")],
    key: Annotated[str | None, typer.Option("--key", "-k", help="Filter keys by pattern (ilike)")] = None,
) -> None:
    """Compare ir.config_parameter values between two instances.

    Examples:
      kctl-odoo diff config production staging
      kctl-odoo diff config production staging --key web.base
    """
    actx: AppContext = ctx.obj
    out = actx.output

    out.info(f"Connecting to profile '{profile_a}'...")
    client_a = _get_client_for_profile(profile_a)
    params_a = _fetch_config_params(client_a, key)
    client_a.close()

    out.info(f"Connecting to profile '{profile_b}'...")
    client_b = _get_client_for_profile(profile_b)
    params_b = _fetch_config_params(client_b, key)
    client_b.close()

    keys_a = set(params_a.keys())
    keys_b = set(params_b.keys())

    only_a = sorted(keys_a - keys_b)
    only_b = sorted(keys_b - keys_a)
    common = sorted(keys_a & keys_b)

    different: list[tuple[str, str, str]] = []
    for k in common:
        if params_a[k] != params_b[k]:
            different.append((k, params_a[k], params_b[k]))

    if actx.json_mode:
        out.raw_json(
            {
                "profile_a": profile_a,
                "profile_b": profile_b,
                "filter": key,
                "summary": {
                    "total_a": len(params_a),
                    "total_b": len(params_b),
                    "only_a": len(only_a),
                    "only_b": len(only_b),
                    "different_value": len(different),
                },
                "only_a": {k: params_a[k] for k in only_a},
                "only_b": {k: params_b[k] for k in only_b},
                "different_value": [
                    {"key": k, f"{profile_a}_value": va, f"{profile_b}_value": vb} for k, va, vb in different
                ],
            }
        )
        return

    out.detail(
        "Config Parameter Comparison",
        [
            (
                "Summary",
                [
                    (profile_a, f"{len(params_a)} params"),
                    (profile_b, f"{len(params_b)} params"),
                    (f"Only in {profile_a}", str(len(only_a))),
                    (f"Only in {profile_b}", str(len(only_b))),
                    ("Value differs", str(len(different))),
                ],
            ),
        ],
    )

    if only_a:
        rows = [[k, _truncate(params_a[k])] for k in only_a]
        out.table(
            f"Only in {profile_a} ({len(only_a)})",
            [("Key", ""), ("Value", "dim")],
            rows,
        )

    if only_b:
        rows = [[k, _truncate(params_b[k])] for k in only_b]
        out.table(
            f"Only in {profile_b} ({len(only_b)})",
            [("Key", ""), ("Value", "dim")],
            rows,
        )

    if different:
        rows = [[k, _truncate(va), _truncate(vb)] for k, va, vb in different]
        out.table(
            f"Value Differences ({len(different)})",
            [("Key", ""), (f"{profile_a}", "cyan"), (f"{profile_b}", "yellow")],
            rows,
        )

    if not only_a and not only_b and not different:
        out.success("Config parameters are identical.")


@app.command()
def users(
    ctx: typer.Context,
    profile_a: Annotated[str, typer.Argument(help="First profile name")],
    profile_b: Annotated[str, typer.Argument(help="Second profile name")],
) -> None:
    """Compare user lists between two Odoo instances.

    Examples:
      kctl-odoo diff users production staging
    """
    actx: AppContext = ctx.obj
    out = actx.output

    out.info(f"Connecting to profile '{profile_a}'...")
    client_a = _get_client_for_profile(profile_a)
    users_a = _fetch_users(client_a)
    client_a.close()

    out.info(f"Connecting to profile '{profile_b}'...")
    client_b = _get_client_for_profile(profile_b)
    users_b = _fetch_users(client_b)
    client_b.close()

    logins_a = set(users_a.keys())
    logins_b = set(users_b.keys())

    only_a = sorted(logins_a - logins_b)
    only_b = sorted(logins_b - logins_a)
    common = sorted(logins_a & logins_b)

    different: list[tuple[str, str, str, str, str]] = []
    for login in common:
        ua = users_a[login]
        ub = users_b[login]
        if ua["name"] != ub["name"] or ua["active"] != ub["active"]:
            different.append(
                (
                    login,
                    ua["name"],
                    ub["name"],
                    str(ua["active"]),
                    str(ub["active"]),
                )
            )

    if actx.json_mode:
        out.raw_json(
            {
                "profile_a": profile_a,
                "profile_b": profile_b,
                "summary": {
                    "total_a": len(users_a),
                    "total_b": len(users_b),
                    "only_a": len(only_a),
                    "only_b": len(only_b),
                    "different": len(different),
                },
                "only_a": [{"login": lg, **users_a[lg]} for lg in only_a],
                "only_b": [{"login": lg, **users_b[lg]} for lg in only_b],
                "different": [
                    {
                        "login": login,
                        f"{profile_a}_name": na,
                        f"{profile_b}_name": nb,
                        f"{profile_a}_active": aa,
                        f"{profile_b}_active": ab,
                    }
                    for login, na, nb, aa, ab in different
                ],
            }
        )
        return

    out.detail(
        "User Comparison",
        [
            (
                "Summary",
                [
                    (profile_a, f"{len(users_a)} users"),
                    (profile_b, f"{len(users_b)} users"),
                    (f"Only in {profile_a}", str(len(only_a))),
                    (f"Only in {profile_b}", str(len(only_b))),
                    ("Different", str(len(different))),
                ],
            ),
        ],
    )

    if only_a:
        rows = [[lg, users_a[lg]["name"], str(users_a[lg]["active"])] for lg in only_a]
        out.table(
            f"Only in {profile_a} ({len(only_a)})",
            [("Login", ""), ("Name", ""), ("Active", "dim")],
            rows,
        )

    if only_b:
        rows = [[lg, users_b[lg]["name"], str(users_b[lg]["active"])] for lg in only_b]
        out.table(
            f"Only in {profile_b} ({len(only_b)})",
            [("Login", ""), ("Name", ""), ("Active", "dim")],
            rows,
        )

    if different:
        rows = [[login, na, nb, aa, ab] for login, na, nb, aa, ab in different]
        out.table(
            f"Differences ({len(different)})",
            [
                ("Login", ""),
                (f"{profile_a} Name", ""),
                (f"{profile_b} Name", ""),
                (f"{profile_a} Active", "dim"),
                (f"{profile_b} Active", "dim"),
            ],
            rows,
        )

    if not only_a and not only_b and not different:
        out.success("User lists are identical.")


def _truncate(value: str, max_len: int = 50) -> str:
    """Truncate a string for display."""
    if len(value) > max_len:
        return value[: max_len - 3] + "..."
    return value
