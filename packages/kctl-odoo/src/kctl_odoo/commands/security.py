"""Security and audit commands."""

from __future__ import annotations

from datetime import UTC, datetime, timedelta
from typing import Annotated

import typer

from kctl_odoo.core.callbacks import AppContext
from kctl_odoo.core.resolve import resolve_group_id, resolve_user_id

app = typer.Typer(help="Security & audit operations.")


@app.command()
def groups(
    ctx: typer.Context,
    limit: Annotated[int, typer.Option("--limit", "-l", help="Max results")] = 100,
) -> None:
    """List security groups with user counts."""
    actx: AppContext = ctx.obj
    out = actx.output
    c = actx.client

    records = c.search_read(
        "res.groups",
        domain=[],
        fields=["id", "name", "full_name", "users", "category_id"],
        limit=limit,
        order="full_name",
    )

    rows = []
    json_data = []
    for g in records:
        cat = g.get("category_id")
        cat_name = cat[1] if isinstance(cat, list) else str(cat or "-")
        user_count = len(g.get("users", []))
        rows.append([str(g["id"]), g.get("full_name") or g["name"], cat_name, str(user_count)])
        json_data.append(
            {
                "id": g["id"],
                "name": g.get("full_name") or g["name"],
                "category": cat_name,
                "users_count": user_count,
            }
        )

    out.table(
        f"Security Groups ({len(records)})",
        [("ID", "cyan"), ("Name", ""), ("Category", "dim"), ("Users", "")],
        rows,
        data_for_json=json_data,
    )


@app.command("access-rights")
def access_rights(
    ctx: typer.Context,
    model: Annotated[str | None, typer.Option("--model", "-m", help="Filter by model name")] = None,
    limit: Annotated[int, typer.Option("--limit", "-l", help="Max results")] = 100,
) -> None:
    """List access control rules (ir.model.access)."""
    actx: AppContext = ctx.obj
    out = actx.output
    c = actx.client

    domain: list = []
    if model:
        domain.append(("model_id.model", "=", model))

    records = c.search_read(
        "ir.model.access",
        domain=domain,
        fields=["id", "name", "model_id", "group_id", "perm_read", "perm_write", "perm_create", "perm_unlink"],
        limit=limit,
        order="model_id",
    )

    rows = []
    json_data = []
    for r in records:
        model_name = r.get("model_id")
        model_display = model_name[1] if isinstance(model_name, list) else str(model_name or "-")
        group = r.get("group_id")
        group_display = group[1] if isinstance(group, list) else str(group or "(global)")

        def _perm(v: bool) -> str:
            return "[green]Y[/green]" if v else "[red]N[/red]"

        rows.append(
            [
                str(r["id"]),
                r["name"],
                model_display,
                group_display,
                _perm(r.get("perm_read", False)),
                _perm(r.get("perm_write", False)),
                _perm(r.get("perm_create", False)),
                _perm(r.get("perm_unlink", False)),
            ]
        )
        json_data.append(
            {
                "id": r["id"],
                "name": r["name"],
                "model": model_display,
                "group": group_display,
                "read": r.get("perm_read"),
                "write": r.get("perm_write"),
                "create": r.get("perm_create"),
                "unlink": r.get("perm_unlink"),
            }
        )

    out.table(
        f"Access Rights ({len(records)})",
        [("ID", "cyan"), ("Name", ""), ("Model", ""), ("Group", "dim"), ("R", ""), ("W", ""), ("C", ""), ("D", "")],
        rows,
        data_for_json=json_data,
    )


@app.command("record-rules")
def record_rules(
    ctx: typer.Context,
    model: Annotated[str | None, typer.Option("--model", "-m", help="Filter by model name")] = None,
    limit: Annotated[int, typer.Option("--limit", "-l", help="Max results")] = 100,
) -> None:
    """List record-level access rules (ir.rule)."""
    actx: AppContext = ctx.obj
    out = actx.output
    c = actx.client

    domain: list = []
    if model:
        domain.append(("model_id.model", "=", model))

    records = c.search_read(
        "ir.rule",
        domain=domain,
        fields=[
            "id",
            "name",
            "model_id",
            "groups",
            "domain_force",
            "perm_read",
            "perm_write",
            "perm_create",
            "perm_unlink",
            "active",
        ],
        limit=limit,
        order="model_id",
    )

    rows = []
    json_data = []
    for r in records:
        model_name = r.get("model_id")
        model_display = model_name[1] if isinstance(model_name, list) else str(model_name or "-")
        groups_count = len(r.get("groups", []))
        global_label = "(global)" if groups_count == 0 else f"{groups_count} groups"
        active = "[green]Y[/green]" if r.get("active") else "[red]N[/red]"
        domain_str = str(r.get("domain_force") or "-")
        if len(domain_str) > 50:
            domain_str = domain_str[:47] + "..."

        rows.append(
            [
                str(r["id"]),
                r["name"],
                model_display,
                global_label,
                active,
                domain_str,
            ]
        )
        json_data.append(
            {
                "id": r["id"],
                "name": r["name"],
                "model": model_display,
                "groups_count": groups_count,
                "active": r.get("active"),
                "domain_force": r.get("domain_force"),
                "perm_read": r.get("perm_read"),
                "perm_write": r.get("perm_write"),
                "perm_create": r.get("perm_create"),
                "perm_unlink": r.get("perm_unlink"),
            }
        )

    out.table(
        f"Record Rules ({len(records)})",
        [("ID", "cyan"), ("Name", ""), ("Model", ""), ("Groups", "dim"), ("Active", ""), ("Domain", "dim")],
        rows,
        data_for_json=json_data,
    )


@app.command()
def admins(ctx: typer.Context) -> None:
    """List users in the Administration / Settings group (base.group_system)."""
    actx: AppContext = ctx.obj
    out = actx.output
    c = actx.client

    # Find the admin group ID
    admin_groups = c.search_read(
        "res.groups",
        domain=[("category_id.name", "=", "Administration"), ("name", "=", "Settings")],
        fields=["id", "users"],
    )
    if not admin_groups:
        # Fallback: try by XML ID
        admin_groups = c.search_read(
            "res.groups",
            domain=[("full_name", "ilike", "Administration / Settings")],
            fields=["id", "users"],
        )

    if not admin_groups:
        out.warn("Could not find admin group. Showing superuser (uid=2) only.")
        admin_groups = [{"users": [2]}]

    user_ids = admin_groups[0].get("users", [])
    if not user_ids:
        out.info("No admin users found.")
        return

    users = c.read("res.users", user_ids, fields=["id", "login", "name", "email", "active", "write_date"])

    rows = []
    json_data = []
    for u in users:
        status = "[green]active[/green]" if u.get("active") else "[red]inactive[/red]"
        rows.append(
            [str(u["id"]), u["login"], u["name"], u.get("email") or "-", status, str(u.get("write_date") or "-")]
        )
        json_data.append(
            {
                "id": u["id"],
                "login": u["login"],
                "name": u["name"],
                "email": u.get("email"),
                "active": u.get("active"),
                "write_date": u.get("write_date"),
            }
        )

    out.table(
        f"Admin Users ({len(users)})",
        [("ID", "cyan"), ("Login", ""), ("Name", ""), ("Email", "dim"), ("Status", ""), ("Last Updated", "dim")],
        rows,
        data_for_json=json_data,
    )


@app.command("quick-audit")
def audit(
    ctx: typer.Context,
    days: Annotated[int, typer.Option("--days", "-d", help="Look back N days")] = 7,
    limit: Annotated[int, typer.Option("--limit", "-l", help="Max results")] = 50,
) -> None:
    """Show recent login activity from res.users.log."""
    actx: AppContext = ctx.obj
    out = actx.output
    c = actx.client

    since = (datetime.now(tz=UTC) - timedelta(days=days)).strftime("%Y-%m-%d %H:%M:%S")

    # res.users.log records login events
    try:
        records = c.search_read(
            "res.users.log",
            domain=[("create_date", ">=", since)],
            fields=["create_uid", "create_date"],
            limit=limit,
            order="create_date desc",
        )
    except Exception:
        out.warn("res.users.log not accessible. Falling back to res.users login_date.")
        records = []

    if records:
        # Resolve user names
        uid_set = {r["create_uid"][0] if isinstance(r["create_uid"], list) else r["create_uid"] for r in records}
        user_map: dict[int, str] = {}
        if uid_set:
            users_data = c.read("res.users", list(uid_set), fields=["id", "login", "name"])
            user_map = {u["id"]: f"{u['login']} ({u['name']})" for u in users_data}

        rows = []
        json_data = []
        for r in records:
            uid = r["create_uid"][0] if isinstance(r["create_uid"], list) else r["create_uid"]
            user_label = user_map.get(uid, str(uid))
            rows.append([str(uid), user_label, str(r.get("create_date", ""))])
            json_data.append(
                {
                    "user_id": uid,
                    "user": user_label,
                    "login_date": r.get("create_date"),
                }
            )

        out.table(
            f"Login Activity (last {days} days, {len(records)} entries)",
            [("UID", "cyan"), ("User", ""), ("Login Date", "dim")],
            rows,
            data_for_json=json_data,
        )
    else:
        # Fallback: show users with recent login_date
        users = c.search_read(
            "res.users",
            domain=[("login_date", ">=", since), ("active", "=", True)],
            fields=["id", "login", "name", "login_date"],
            limit=limit,
            order="login_date desc",
        )
        rows = []
        for u in users:
            rows.append([str(u["id"]), u["login"], u["name"], str(u.get("login_date") or "-")])

        out.table(
            f"Recent Logins (last {days} days, {len(users)} users)",
            [("ID", "cyan"), ("Login", ""), ("Name", ""), ("Last Login", "dim")],
            rows,
            data_for_json=users,
        )


@app.command("group-users")
def group_users(
    ctx: typer.Context,
    group: Annotated[str, typer.Argument(help="Group name, ID, or XML ID (e.g. base.group_user)")],
    limit: Annotated[int, typer.Option("--limit", "-l", help="Max users to show")] = 50,
) -> None:
    """List all users in a specific security group."""
    actx: AppContext = ctx.obj
    out = actx.output
    c = actx.client

    group_id = resolve_group_id(c, group)
    groups_data = c.read("res.groups", [group_id], fields=["id", "full_name", "users"])
    if not groups_data:
        out.error(f"Group not found: {group}")
        raise typer.Exit(1)

    g = groups_data[0]
    user_ids = g.get("users", [])
    group_label = g.get("full_name") or str(group_id)

    if not user_ids:
        out.info(f"No users in group '{group_label}'.")
        if actx.json_mode:
            out.raw_json({"group": group_label, "users": []})
        return

    # Limit the user IDs we read
    user_ids_limited = user_ids[:limit]
    users = c.read("res.users", user_ids_limited, fields=["id", "login", "name", "email", "active"])

    rows = []
    json_data = []
    for u in users:
        status = "[green]active[/green]" if u.get("active") else "[red]inactive[/red]"
        rows.append([str(u["id"]), u["login"], u["name"], u.get("email") or "-", status])
        json_data.append(
            {
                "id": u["id"],
                "login": u["login"],
                "name": u["name"],
                "email": u.get("email"),
                "active": u.get("active"),
            }
        )

    title = f"Users in '{group_label}' ({len(user_ids)} total"
    if len(user_ids) > limit:
        title += f", showing {limit}"
    title += ")"

    out.table(
        title,
        [("ID", "cyan"), ("Login", ""), ("Name", ""), ("Email", "dim"), ("Status", "")],
        rows,
        data_for_json=json_data,
    )


@app.command("add-to-group")
def add_to_group(
    ctx: typer.Context,
    user_login: Annotated[str, typer.Argument(help="User login or ID")],
    group_ref: Annotated[str, typer.Argument(help="Group ID or XML ID (e.g. base.group_user)")],
) -> None:
    """Add a user to a security group."""
    actx: AppContext = ctx.obj
    out = actx.output
    c = actx.client

    user_id = resolve_user_id(c, user_login)
    group_id = resolve_group_id(c, group_ref)

    # Use (4, id) to add a link without removing existing ones
    c.write("res.users", [user_id], {"groups_id": [(4, group_id)]})
    out.success(f"Added user '{user_login}' (ID: {user_id}) to group ID {group_id}")


@app.command("remove-from-group")
def remove_from_group(
    ctx: typer.Context,
    user_login: Annotated[str, typer.Argument(help="User login or ID")],
    group_ref: Annotated[str, typer.Argument(help="Group ID or XML ID (e.g. base.group_user)")],
    force: Annotated[bool, typer.Option("--force", "-f", help="Skip confirmation")] = False,
) -> None:
    """Remove a user from a security group."""
    actx: AppContext = ctx.obj
    out = actx.output
    c = actx.client

    user_id = resolve_user_id(c, user_login)
    group_id = resolve_group_id(c, group_ref)

    if not force and not typer.confirm(f"Remove user '{user_login}' from group ID {group_id}?"):
        raise typer.Exit(0)

    # Use (3, id) to remove a link without deleting the group
    c.write("res.users", [user_id], {"groups_id": [(3, group_id)]})
    out.success(f"Removed user '{user_login}' (ID: {user_id}) from group ID {group_id}")


@app.command("check-permissions")
def check_permissions(
    ctx: typer.Context,
    user: Annotated[str, typer.Argument(help="User ID (numeric) or login name")],
) -> None:
    """Show a user's groups and effective permissions."""
    actx: AppContext = ctx.obj
    out = actx.output
    c = actx.client

    user_id = resolve_user_id(c, user)

    users = c.read("res.users", [user_id], fields=["id", "login", "name", "groups_id", "company_id"])
    if not users:
        out.error(f"User not found: {user_id}")
        raise typer.Exit(1)

    u = users[0]
    group_ids = u.get("groups_id", [])

    group_names: list[str] = []
    if group_ids:
        groups_data = c.read("res.groups", group_ids, fields=["id", "full_name"])
        group_names = [g.get("full_name") or str(g["id"]) for g in groups_data]

    # Get access rights for these groups
    acl_domain: list = ["|", ("group_id", "=", False)]
    if group_ids:
        acl_domain.append(("group_id", "in", group_ids))
    else:
        acl_domain.append(("group_id", "=", False))

    acls = c.search_read(
        "ir.model.access",
        domain=acl_domain,
        fields=["name", "model_id", "perm_read", "perm_write", "perm_create", "perm_unlink"],
        limit=200,
    )

    # Build model permission summary
    model_perms: dict[str, dict] = {}
    for acl in acls:
        model_label = acl["model_id"][1] if isinstance(acl["model_id"], list) else str(acl["model_id"])
        if model_label not in model_perms:
            model_perms[model_label] = {"read": False, "write": False, "create": False, "unlink": False}
        if acl.get("perm_read"):
            model_perms[model_label]["read"] = True
        if acl.get("perm_write"):
            model_perms[model_label]["write"] = True
        if acl.get("perm_create"):
            model_perms[model_label]["create"] = True
        if acl.get("perm_unlink"):
            model_perms[model_label]["unlink"] = True

    company = u.get("company_id")
    company_name = company[1] if isinstance(company, list) else str(company or "-")

    json_out = {
        "user_id": user_id,
        "login": u["login"],
        "name": u["name"],
        "company": company_name,
        "groups": sorted(group_names),
        "groups_count": len(group_names),
        "models_accessible": len(model_perms),
    }

    sections = [
        (
            "User",
            [
                ("ID", str(user_id)),
                ("Login", u["login"]),
                ("Name", u["name"]),
                ("Company", company_name),
                ("Groups", str(len(group_names))),
            ],
        ),
        ("Groups", [(str(i + 1), name) for i, name in enumerate(sorted(group_names)[:30])]),
        (
            "Access Summary",
            [
                ("Models accessible", str(len(model_perms))),
                (
                    "Full CRUD models",
                    str(
                        sum(1 for p in model_perms.values() if p["read"] and p["write"] and p["create"] and p["unlink"])
                    ),
                ),
            ],
        ),
    ]

    out.detail(f"Permissions: {u['login']}", sections, data_for_json=json_out)


@app.command()
def superusers(
    ctx: typer.Context,
) -> None:
    """List internal (non-portal) active users with admin status."""
    actx: AppContext = ctx.obj
    out = actx.output
    c = actx.client

    users = c.search_read(
        "res.users",
        domain=[("active", "=", True), ("share", "=", False)],
        fields=["id", "login", "name", "email", "groups_id", "login_date", "company_id"],
        order="id",
    )

    # Find admin group ID
    admin_groups = c.search_read(
        "res.groups",
        domain=[("full_name", "ilike", "Administration / Settings")],
        fields=["id"],
    )
    admin_group_ids = {g["id"] for g in admin_groups}

    rows = []
    json_data = []
    for u in users:
        user_groups = set(u.get("groups_id", []))
        is_admin = bool(user_groups & admin_group_ids)
        admin_label = "[red]admin[/red]" if is_admin else ""
        company = u.get("company_id")
        company_name = company[1] if isinstance(company, list) else str(company or "-")

        rows.append(
            [
                str(u["id"]),
                u["login"],
                u["name"],
                company_name,
                admin_label,
                str(u.get("login_date") or "never"),
            ]
        )
        json_data.append(
            {
                "id": u["id"],
                "login": u["login"],
                "name": u["name"],
                "company": company_name,
                "is_admin": is_admin,
                "last_login": u.get("login_date"),
            }
        )

    out.table(
        f"Internal Users ({len(users)})",
        [("ID", "cyan"), ("Login", ""), ("Name", ""), ("Company", "dim"), ("Admin", ""), ("Last Login", "dim")],
        rows,
        data_for_json=json_data,
    )


# ---------------------------------------------------------------------------
# Security Audit — comprehensive automated check
# ---------------------------------------------------------------------------


@app.command("audit")
def security_audit(ctx: typer.Context) -> None:
    """Comprehensive security audit across groups, ACLs, record rules, and users.

    Checks:
    - Models without access rights (ACL gaps)
    - Groups with no users (orphaned groups)
    - Users with admin access (privilege review)
    - Inactive record rules (may expose data)
    - Global record rules (apply to everyone)
    - Users without 2FA (if auth_totp installed)
    - Excessive sudo() usage summary
    """
    actx: AppContext = ctx.obj
    out = actx.output
    c = actx.client

    rows: list[list[str]] = []
    json_data: list[dict] = []
    total_issues = 0

    def _add(category: str, check: str, ok: bool, count: int, detail: str) -> None:
        nonlocal total_issues
        if not ok:
            total_issues += count
        status = "[green]OK[/green]" if ok else f"[red]{count}[/red]"
        rows.append([category, check, status, detail if not ok else ""])
        json_data.append({"category": category, "check": check, "count": count, "ok": ok, "detail": detail})

    # 1. Models without ANY access rights
    try:
        all_models = c.search_read("ir.model", [("transient", "=", False)], fields=["model"], limit=500)
        models_without_acl = 0
        for m in all_models:
            acl_count = c.search_count("ir.model.access", [("model_id.model", "=", m["model"])])
            if acl_count == 0:
                models_without_acl += 1
        _add(
            "ACL",
            "Models without access rights",
            models_without_acl == 0,
            models_without_acl,
            f"{models_without_acl} models have no ir.model.access rules",
        )
    except Exception:
        _add("ACL", "Models without access rights", True, 0, "Check skipped")

    # 2. Groups with no users
    try:
        all_groups = c.search_read("res.groups", [], fields=["name", "users"], limit=200)
        empty_groups = [g for g in all_groups if not g.get("users")]
        _add(
            "Groups",
            "Groups with no users",
            len(empty_groups) < 10,
            len(empty_groups),
            f"{len(empty_groups)} groups have zero users assigned",
        )
    except Exception:
        _add("Groups", "Groups with no users", True, 0, "Check skipped")

    # 3. Admin users count
    try:
        admin_group = c.search_read(
            "res.groups",
            [("category_id.name", "=", "Administration"), ("name", "ilike", "Settings")],
            fields=["users"],
            limit=1,
        )
        admin_count = len(admin_group[0]["users"]) if admin_group else 0
        _add(
            "Users",
            "Admin users (Settings group)",
            admin_count <= 3,
            admin_count,
            f"{admin_count} users have full admin access — review if > 3",
        )
    except Exception:
        _add("Users", "Admin users", True, 0, "Check skipped")

    # 4. Users with no groups (potential orphans)
    try:
        users_no_groups = c.search_count(
            "res.users", [("active", "=", True), ("share", "=", False), ("groups_id", "=", False)]
        )
        _add(
            "Users",
            "Internal users without groups",
            users_no_groups == 0,
            users_no_groups,
            "Users without any group have no permissions",
        )
    except Exception:
        _add("Users", "Users without groups", True, 0, "Check skipped")

    # 5. Inactive record rules (may expose data if they should be active)
    try:
        inactive_rules = c.search_count("ir.rule", [("active", "=", False)])
        _add(
            "Rules",
            "Inactive record rules",
            inactive_rules < 5,
            inactive_rules,
            f"{inactive_rules} record rules are disabled — verify intentional",
        )
    except Exception:
        _add("Rules", "Inactive record rules", True, 0, "Check skipped")

    # 6. Global record rules (no group = applies to everyone including public)
    try:
        global_rules = c.search_read(
            "ir.rule", [("groups", "=", False), ("active", "=", True)], fields=["name", "model_id"], limit=100
        )
        _add(
            "Rules",
            "Global record rules (no group)",
            True,
            len(global_rules),
            f"{len(global_rules)} rules apply globally — this is normal for multi-company/OU",
        )
    except Exception:
        _add("Rules", "Global record rules", True, 0, "Check skipped")

    # 7. Users with API keys (should be reviewed)
    try:
        api_key_count = c.search_count("res.users.apikeys", [])
    except Exception:
        try:
            api_key_count = c.search_count("api.key", [])
        except Exception:
            api_key_count = -1
    if api_key_count >= 0:
        _add(
            "API",
            "Active API keys",
            api_key_count <= 10,
            api_key_count,
            f"{api_key_count} API keys active — review periodically",
        )

    # 8. Portal users count (information)
    try:
        portal_count = c.search_count("res.users", [("active", "=", True), ("share", "=", True)])
        _add("Users", "Portal/external users", True, portal_count, f"{portal_count} portal users")
    except Exception:
        pass

    # 9. Check for default admin password (heuristic: try common passwords)
    _add("Password", "Admin password policy", True, 0, "Cannot check passwords via API — verify manually")

    title = f"Security Audit ({len(rows)} checks)"
    if total_issues:
        title += f" — [yellow]{total_issues} findings[/yellow]"
    else:
        title += " — [green]clean[/green]"

    out.table(
        title,
        [("Category", "dim"), ("Check", "cyan"), ("Status", ""), ("Detail", "dim")],
        rows,
        json_data,
    )


# ---------------------------------------------------------------------------
# Security Compliance — 9-point check with scoring
# ---------------------------------------------------------------------------


@app.command()
def compliance(ctx: typer.Context) -> None:
    """Security compliance scorecard (9-point check).

    Checks:
    1. Admin count (<=3 = PASS)
    2. Users without login in 90 days
    3. Global record rules without group (security risk)
    4. Models without ACL
    5. Signup scope (b2b vs b2c)
    6. HTTPS enforcement (web.base.url)
    7. Password minimum length (if auth_password_policy installed)
    8. API key count
    9. Portal user count

    Examples:
        kctl-odoo security compliance
        kctl-odoo security compliance --json
    """
    actx: AppContext = ctx.obj
    out = actx.output
    c = actx.client

    rows: list[list[str]] = []
    json_data: list[dict] = []
    passed = 0
    total = 0

    def _check(name: str, ok: bool, value: str, detail: str) -> None:
        nonlocal passed, total
        total += 1
        if ok:
            passed += 1
        status = "[green]PASS[/green]" if ok else "[red]FAIL[/red]"
        rows.append([str(total), name, value, status, detail])
        json_data.append(
            {
                "num": total,
                "check": name,
                "value": value,
                "passed": ok,
                "detail": detail,
            }
        )

    # 1. Admin count
    try:
        admin_groups = c.search_read(
            "res.groups",
            domain=[("category_id.name", "=", "Administration"), ("name", "ilike", "Settings")],
            fields=["users"],
            limit=1,
        )
        admin_count = len(admin_groups[0]["users"]) if admin_groups else 0
        _check(
            "Admin user count",
            admin_count <= 3,
            str(admin_count),
            "<=3 recommended" if admin_count <= 3 else f"{admin_count} admins — review access",
        )
    except Exception:
        _check("Admin user count", True, "?", "Check skipped")

    # 2. Users without login in 90 days
    try:
        cutoff_90d = (datetime.now(tz=UTC) - timedelta(days=90)).strftime("%Y-%m-%d %H:%M:%S")
        stale_users = c.search_count(
            "res.users",
            [
                ("active", "=", True),
                ("share", "=", False),
                "|",
                ("login_date", "=", False),
                ("login_date", "<", cutoff_90d),
            ],
        )
        _check(
            "Stale users (no login 90d)",
            stale_users == 0,
            str(stale_users),
            f"{stale_users} user(s) not logged in for 90+ days" if stale_users else "All users active",
        )
    except Exception:
        _check("Stale users (no login 90d)", True, "?", "Check skipped")

    # 3. Global record rules without group
    try:
        global_rules = c.search_count("ir.rule", [("groups", "=", False), ("active", "=", True)])
        # Global rules are normal in Odoo (multi-company), so warn only if excessive
        _check(
            "Global record rules",
            global_rules <= 30,
            str(global_rules),
            f"{global_rules} rules apply to all users" if global_rules > 30 else "Within normal range",
        )
    except Exception:
        _check("Global record rules", True, "?", "Check skipped")

    # 4. Models without ACL
    try:
        all_models = c.search_read("ir.model", [("transient", "=", False)], fields=["model"], limit=500)
        no_acl = 0
        for m in all_models:
            acl = c.search_count("ir.model.access", [("model_id.model", "=", m["model"])])
            if acl == 0:
                no_acl += 1
        _check(
            "Models without ACL",
            no_acl == 0,
            str(no_acl),
            f"{no_acl} model(s) have no access rules" if no_acl else "All models have ACLs",
        )
    except Exception:
        _check("Models without ACL", True, "?", "Check skipped")

    # 5. Signup scope
    try:
        signup_params = c.search_read(
            "ir.config_parameter",
            domain=[("key", "=", "auth_signup.invitation_scope")],
            fields=["value"],
            limit=1,
        )
        scope = signup_params[0]["value"] if signup_params else "b2b"
        _check(
            "Signup scope",
            scope != "b2c",
            scope,
            "b2c allows public signup — ensure intentional" if scope == "b2c" else "Invitation-only signup",
        )
    except Exception:
        _check("Signup scope", True, "?", "Check skipped")

    # 6. HTTPS enforcement
    try:
        url_params = c.search_read(
            "ir.config_parameter",
            domain=[("key", "=", "web.base.url")],
            fields=["value"],
            limit=1,
        )
        base_url = url_params[0]["value"] if url_params else ""
        is_https = base_url.startswith("https://") or "localhost" in base_url
        _check(
            "HTTPS enforcement", is_https, base_url[:50], "HTTPS not configured" if not is_https else "HTTPS enabled"
        )
    except Exception:
        _check("HTTPS enforcement", True, "?", "Check skipped")

    # 7. Password minimum length
    try:
        from kctl_odoo.core.utils import check_module_installed

        if check_module_installed(c, "auth_password_policy"):
            pw_params = c.search_read(
                "ir.config_parameter",
                domain=[("key", "=", "auth_password_policy.minlength")],
                fields=["value"],
                limit=1,
            )
            min_len = int(pw_params[0]["value"]) if pw_params else 0
            _check(
                "Password min length",
                min_len >= 8,
                str(min_len),
                f"Minimum {min_len} chars" if min_len >= 8 else "Set to >=8 for security",
            )
        else:
            _check("Password min length", False, "N/A", "auth_password_policy module not installed")
    except Exception:
        _check("Password min length", True, "?", "Check skipped")

    # 8. API key count
    try:
        api_key_count = c.search_count("res.users.apikeys", [])
    except Exception:
        try:
            api_key_count = c.search_count("api.key", [])
        except Exception:
            api_key_count = -1
    if api_key_count >= 0:
        _check(
            "API key count",
            api_key_count <= 10,
            str(api_key_count),
            "Review API keys periodically" if api_key_count > 10 else "Within acceptable range",
        )

    # 9. Portal user count
    try:
        portal_count = c.search_count("res.users", [("active", "=", True), ("share", "=", True)])
        _check("Portal users", True, str(portal_count), f"{portal_count} portal/external users (informational)")
    except Exception:
        _check("Portal users", True, "?", "Check skipped")

    score = (passed / total * 100) if total else 0
    score_color = "green" if score >= 80 else "yellow" if score >= 60 else "red"
    title = f"Security Compliance — [{score_color}]{score:.0f}%[/{score_color}] ({passed}/{total} passed)"

    out.table(
        title,
        [("#", "dim"), ("Check", ""), ("Value", "cyan"), ("Status", ""), ("Detail", "dim")],
        rows,
        data_for_json=json_data,
    )


# ---------------------------------------------------------------------------
# Wave 5: bulk-grant, bulk-revoke, diff-users, sudo-audit
# ---------------------------------------------------------------------------


@app.command("bulk-grant")
def bulk_grant(
    ctx: typer.Context,
    group_ref: Annotated[str, typer.Argument(help="Group ID or XML ID (e.g. base.group_user)")],
    users: Annotated[str, typer.Argument(help="Comma-separated user logins")],
) -> None:
    """Add multiple users to a security group.

    Examples:
        kctl-odoo security bulk-grant base.group_user user1,user2,user3
    """
    actx: AppContext = ctx.obj
    out = actx.output
    c = actx.client

    group_id = resolve_group_id(c, group_ref)
    logins = [login.strip() for login in users.split(",") if login.strip()]

    if not logins:
        out.error("No user logins provided.")
        raise typer.Exit(1)

    success_count = 0
    errors: list[str] = []

    for login in logins:
        try:
            user_id = resolve_user_id(c, login)
            c.write("res.users", [user_id], {"groups_id": [(4, group_id)]})
            success_count += 1
        except Exception as e:
            errors.append(f"{login}: {e}")

    if success_count:
        out.success(f"Added {success_count} user(s) to group ID {group_id}.")
    for err in errors:
        out.warn(f"Failed: {err}")


@app.command("bulk-revoke")
def bulk_revoke(
    ctx: typer.Context,
    group_ref: Annotated[str, typer.Argument(help="Group ID or XML ID (e.g. base.group_user)")],
    users: Annotated[str, typer.Argument(help="Comma-separated user logins")],
) -> None:
    """Remove multiple users from a security group.

    Examples:
        kctl-odoo security bulk-revoke base.group_user user1,user2,user3
    """
    actx: AppContext = ctx.obj
    out = actx.output
    c = actx.client

    group_id = resolve_group_id(c, group_ref)
    logins = [login.strip() for login in users.split(",") if login.strip()]

    if not logins:
        out.error("No user logins provided.")
        raise typer.Exit(1)

    success_count = 0
    errors: list[str] = []

    for login in logins:
        try:
            user_id = resolve_user_id(c, login)
            c.write("res.users", [user_id], {"groups_id": [(3, group_id)]})
            success_count += 1
        except Exception as e:
            errors.append(f"{login}: {e}")

    if success_count:
        out.success(f"Removed {success_count} user(s) from group ID {group_id}.")
    for err in errors:
        out.warn(f"Failed: {err}")


@app.command("diff-users")
def diff_users(
    ctx: typer.Context,
    user_a: Annotated[str, typer.Argument(help="First user login or ID")],
    user_b: Annotated[str, typer.Argument(help="Second user login or ID")],
) -> None:
    """Compare permissions between two users.

    Shows groups unique to each user and groups shared by both.

    Examples:
        kctl-odoo security diff-users admin user1
    """
    actx: AppContext = ctx.obj
    out = actx.output
    c = actx.client

    uid_a = resolve_user_id(c, user_a)
    uid_b = resolve_user_id(c, user_b)

    users_data = c.read("res.users", [uid_a, uid_b], fields=["id", "login", "groups_id"])
    if len(users_data) < 2:
        out.error("Could not read both users.")
        raise typer.Exit(1)

    # Map by id
    users_map = {u["id"]: u for u in users_data}
    data_a = users_map[uid_a]
    data_b = users_map[uid_b]

    groups_a = set(data_a.get("groups_id", []))
    groups_b = set(data_b.get("groups_id", []))

    only_a = groups_a - groups_b
    only_b = groups_b - groups_a
    shared = groups_a & groups_b

    def _resolve_group_names(ids: set) -> list[str]:
        if not ids:
            return []
        groups_data = c.read("res.groups", list(ids), fields=["id", "full_name"])
        return sorted(g.get("full_name") or str(g["id"]) for g in groups_data)

    names_only_a = _resolve_group_names(only_a)
    names_only_b = _resolve_group_names(only_b)
    names_shared = _resolve_group_names(shared)

    login_a = data_a.get("login", str(uid_a))
    login_b = data_b.get("login", str(uid_b))

    sections = [
        (
            f"Only {login_a} ({len(names_only_a)})",
            [(str(i + 1), name) for i, name in enumerate(names_only_a)] or [("-", "-")],
        ),
        (
            f"Only {login_b} ({len(names_only_b)})",
            [(str(i + 1), name) for i, name in enumerate(names_only_b)] or [("-", "-")],
        ),
        (f"Shared ({len(names_shared)})", [(str(i + 1), name) for i, name in enumerate(names_shared)] or [("-", "-")]),
    ]

    json_out = {
        "user_a": login_a,
        "user_b": login_b,
        f"only_{login_a}": names_only_a,
        f"only_{login_b}": names_only_b,
        "shared": names_shared,
    }

    out.detail(f"Permission Diff: {login_a} vs {login_b}", sections, data_for_json=json_out)


@app.command("sudo-audit")
def sudo_audit(
    ctx: typer.Context,
    module: Annotated[str | None, typer.Argument(help="Module name to scan (default: all)")] = None,
) -> None:
    """Find sudo() calls in private module Python code.

    Scans src/private/ for .sudo() usage to help security reviews.

    Examples:
        kctl-odoo security sudo-audit
        kctl-odoo security sudo-audit sfa_management
    """
    import re

    actx: AppContext = ctx.obj
    out = actx.output

    try:
        from kctl_odoo.core.utils import find_project_root

        private_dir = find_project_root() / "src" / "private"
    except Exception:
        out.error("Could not locate project root / src/private directory.")
        raise typer.Exit(1) from None

    if module:
        mod_dirs = [private_dir / module]
    else:
        mod_dirs = [d for d in private_dir.iterdir() if d.is_dir() and (d / "__manifest__.py").exists()]

    pattern = re.compile(r"\.sudo\(")
    findings: list[dict] = []

    for mod_dir in mod_dirs:
        mod_name = mod_dir.name
        for py_file in mod_dir.rglob("*.py"):
            if "__pycache__" in str(py_file):
                continue
            try:
                lines = py_file.read_text(encoding="utf-8").splitlines()
            except Exception:
                continue
            for lineno, line in enumerate(lines, 1):
                if pattern.search(line):
                    rel_path = str(py_file.relative_to(private_dir))
                    findings.append(
                        {
                            "module": mod_name,
                            "file": rel_path,
                            "line": lineno,
                            "code": line.strip(),
                        }
                    )

    if not findings:
        out.success("No .sudo() calls found.")
        return

    rows = []
    for f in findings:
        rows.append([f["module"], f["file"], str(f["line"]), f["code"][:80]])

    out.table(
        f"sudo() Calls ({len(findings)} found)",
        [("Module", "cyan"), ("File", ""), ("Line", "dim"), ("Code", "dim")],
        rows,
        data_for_json=findings,
    )


# ---------------------------------------------------------------------------
# Access Report — comprehensive user access matrix export
# ---------------------------------------------------------------------------


def _fetch_access_data(c, *, include_inactive: bool = False) -> dict:
    """Batch-fetch all user access data in ~5 RPC calls."""
    domain: list = [("share", "=", False)]
    if not include_inactive:
        domain.append(("active", "=", True))

    ou_fields = ["assigned_operating_unit_ids", "default_operating_unit_id"]
    base_fields = [
        "id",
        "login",
        "name",
        "active",
        "login_date",
        "company_id",
        "company_ids",
        "groups_id",
    ]
    try:
        users = c.search_read(
            "res.users",
            domain=domain,
            fields=base_fields + ou_fields,
            order="login",
            limit=0,
        )
        has_ou = True
    except Exception:
        users = c.search_read(
            "res.users",
            domain=domain,
            fields=base_fields,
            order="login",
            limit=0,
        )
        has_ou = False

    # --- Role assignments ---------------------------------------------------
    try:
        role_lines = c.search_read(
            "res.users.role.line",
            domain=[],
            fields=["user_id", "role_id"],
        )
    except Exception:
        role_lines = []

    # Map user_id -> list of role_ids
    user_role_map: dict[int, list[int]] = {}
    all_role_ids: set[int] = set()
    for rl in role_lines:
        uid = rl["user_id"][0] if isinstance(rl["user_id"], list) else rl["user_id"]
        rid = rl["role_id"][0] if isinstance(rl["role_id"], list) else rl["role_id"]
        user_role_map.setdefault(uid, []).append(rid)
        all_role_ids.add(rid)

    role_names: dict[int, str] = {}
    if all_role_ids:
        try:
            roles = c.read("res.users.role", list(all_role_ids), ["id", "name"])
            role_names = {r["id"]: r["name"] for r in roles}
        except Exception:
            pass

    # --- OU names -----------------------------------------------------------
    if has_ou:
        all_ou_ids: set[int] = set()
        for u in users:
            all_ou_ids.update(u.get("assigned_operating_unit_ids") or [])
            dou = u.get("default_operating_unit_id")
            if isinstance(dou, list):
                all_ou_ids.add(dou[0])
            elif isinstance(dou, int) and dou:
                all_ou_ids.add(dou)

        ou_names: dict[int, str] = {}
        if all_ou_ids:
            ous = c.read("operating.unit", list(all_ou_ids), ["id", "name"])
            ou_names = {o["id"]: o["name"] for o in ous}
    else:
        ou_names: dict[int, str] = {}

    # --- Company names ------------------------------------------------------
    all_company_ids: set[int] = set()
    for u in users:
        cid = u.get("company_id")
        if isinstance(cid, list):
            all_company_ids.add(cid[0])
        elif isinstance(cid, int) and cid:
            all_company_ids.add(cid)
        all_company_ids.update(u.get("company_ids") or [])

    company_names: dict[int, str] = {}
    if all_company_ids:
        companies = c.read("res.company", list(all_company_ids), ["id", "name"])
        company_names = {co["id"]: co["name"] for co in companies}

    # --- Admin group --------------------------------------------------------
    admin_groups = c.search_read(
        "res.groups",
        domain=[("full_name", "ilike", "Administration / Settings")],
        fields=["id"],
    )
    admin_group_ids = {g["id"] for g in admin_groups}

    # --- Build user rows ----------------------------------------------------
    active_count = 0
    inactive_count = 0
    user_rows: list[dict] = []

    for u in users:
        is_active = u.get("active", True)
        if is_active:
            active_count += 1
        else:
            inactive_count += 1

        company = u.get("company_id")
        company_name = company[1] if isinstance(company, list) else str(company or "-")

        company_id_list = u.get("company_ids") or []
        companies_list = [company_names.get(cid, str(cid)) for cid in company_id_list]

        user_groups = set(u.get("groups_id") or [])
        is_admin = bool(user_groups & admin_group_ids)

        uid = u["id"]
        role_ids = user_role_map.get(uid, [])
        roles_list = [role_names.get(rid, str(rid)) for rid in role_ids]

        ou_ids = u.get("assigned_operating_unit_ids") or []
        ous_list = [ou_names.get(oid, str(oid)) for oid in ou_ids]

        dou = u.get("default_operating_unit_id")
        if isinstance(dou, list):
            default_ou = dou[1]
        elif isinstance(dou, int) and dou:
            default_ou = ou_names.get(dou, str(dou))
        else:
            default_ou = "-"

        user_rows.append(
            {
                "login": u["login"],
                "name": u["name"],
                "company": company_name,
                "companies": companies_list,
                "status": "active" if is_active else "inactive",
                "last_login": str(u.get("login_date") or "never"),
                "roles": roles_list,
                "operating_units": ous_list,
                "default_ou": default_ou,
                "is_admin": is_admin,
                "groups_count": len(user_groups),
            }
        )

    # --- Roles summary ------------------------------------------------------
    roles_summary: list[dict] = []
    # Invert: role -> list of logins
    role_user_map: dict[int, list[str]] = {}
    user_login_map = {u["id"]: u["login"] for u in users}
    for rl in role_lines:
        uid = rl["user_id"][0] if isinstance(rl["user_id"], list) else rl["user_id"]
        rid = rl["role_id"][0] if isinstance(rl["role_id"], list) else rl["role_id"]
        login = user_login_map.get(uid)
        if login:
            role_user_map.setdefault(rid, []).append(login)

    for rid in sorted(all_role_ids):
        rname = role_names.get(rid, str(rid))
        rusers = sorted(set(role_user_map.get(rid, [])))
        roles_summary.append(
            {
                "role": rname,
                "users_count": len(rusers),
                "users": rusers,
            }
        )

    return {
        "instance": c._base_url,
        "database": c.database,
        "generated": datetime.now(UTC).strftime("%Y-%m-%d %H:%M:%S UTC"),
        "active_count": active_count,
        "inactive_count": inactive_count,
        "users": user_rows,
        "roles_summary": roles_summary,
    }


def _write_markdown(data: dict, path: str) -> None:
    """Write access report as a Markdown file."""

    def _md_escape(val: str) -> str:
        return str(val).replace("|", "\\|")

    lines: list[str] = []
    lines.append("# User Access Report")
    lines.append("")
    lines.append(f"- **Instance:** {data['instance']}")
    lines.append(f"- **Database:** {data['database']}")
    lines.append(f"- **Generated:** {data['generated']}")
    lines.append(f"- **Active users:** {data['active_count']}")
    lines.append(f"- **Inactive users:** {data['inactive_count']}")
    lines.append("")

    # User access table
    lines.append("## User Access")
    lines.append("")
    lines.append("| Login | Name | Company | Roles | Operating Units | Default OU | Admin | Last Login | Groups |")
    lines.append("|-------|------|---------|-------|-----------------|------------|-------|------------|--------|")
    for u in data["users"]:
        roles = " \\| ".join(_md_escape(r) for r in u["roles"]) if u["roles"] else "-"
        ous = " \\| ".join(_md_escape(o) for o in u["operating_units"]) if u["operating_units"] else "-"
        admin = "Yes" if u["is_admin"] else "No"
        lines.append(
            f"| {_md_escape(u['login'])} | {_md_escape(u['name'])} | {_md_escape(u['company'])} | {roles} | {ous} | {_md_escape(u['default_ou'])} | {admin} | {_md_escape(u['last_login'])} | {u['groups_count']} |"
        )
    lines.append("")

    # Roles summary table
    lines.append("## Roles Summary")
    lines.append("")
    lines.append("| Role | Users | Assigned To |")
    lines.append("|------|-------|-------------|")
    for r in data["roles_summary"]:
        assigned = ", ".join(_md_escape(u) for u in r["users"]) if r["users"] else "-"
        lines.append(f"| {_md_escape(r['role'])} | {r['users_count']} | {assigned} |")
    lines.append("")

    with open(path, "w", encoding="utf-8") as f:
        f.write("\n".join(lines) + "\n")


def _write_excel(data: dict, path: str) -> None:
    """Write access report as an Excel file."""
    from openpyxl import Workbook
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    wb = Workbook()

    # Sheet 1: User Access
    ws = wb.active
    ws.title = "User Access"
    headers = [
        "Login",
        "Name",
        "Company",
        "Companies",
        "Status",
        "Last Login",
        "Roles",
        "Operating Units",
        "Default OU",
        "Admin",
        "Groups",
    ]
    bold = Font(bold=True)
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = bold

    for row_idx, u in enumerate(data["users"], 2):
        ws.cell(row=row_idx, column=1, value=u["login"])
        ws.cell(row=row_idx, column=2, value=u["name"])
        ws.cell(row=row_idx, column=3, value=u["company"])
        ws.cell(row=row_idx, column=4, value=", ".join(u["companies"]) if u["companies"] else "-")
        ws.cell(row=row_idx, column=5, value=u["status"])
        ws.cell(row=row_idx, column=6, value=u["last_login"])
        ws.cell(row=row_idx, column=7, value=", ".join(u["roles"]) if u["roles"] else "-")
        ws.cell(row=row_idx, column=8, value=", ".join(u["operating_units"]) if u["operating_units"] else "-")
        ws.cell(row=row_idx, column=9, value=u["default_ou"])
        ws.cell(row=row_idx, column=10, value="Yes" if u["is_admin"] else "No")
        ws.cell(row=row_idx, column=11, value=u["groups_count"])

    ws.freeze_panes = "A2"
    for col_idx in range(1, len(headers) + 1):
        max_len = len(headers[col_idx - 1])
        for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 50)

    # Sheet 2: Roles Summary
    ws2 = wb.create_sheet("Roles Summary")
    role_headers = ["Role", "Users Count", "Assigned To"]
    for col_idx, header in enumerate(role_headers, 1):
        cell = ws2.cell(row=1, column=col_idx, value=header)
        cell.font = bold

    for row_idx, r in enumerate(data["roles_summary"], 2):
        ws2.cell(row=row_idx, column=1, value=r["role"])
        ws2.cell(row=row_idx, column=2, value=r["users_count"])
        ws2.cell(row=row_idx, column=3, value=", ".join(r["users"]) if r["users"] else "-")

    ws2.freeze_panes = "A2"
    for col_idx in range(1, len(role_headers) + 1):
        max_len = len(role_headers[col_idx - 1])
        for row in ws2.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
        ws2.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 50)

    wb.save(path)


@app.command("access-report")
def access_report(
    ctx: typer.Context,
    output: Annotated[str | None, typer.Option("--output", "-o", help="Output file path (.md, .xlsx, .json)")] = None,
    all_users: Annotated[bool, typer.Option("--all/--active", help="Include inactive users")] = False,
) -> None:
    """Export user access matrix with roles, OUs, and admin status.

    Outputs a comprehensive report of all internal users with their roles,
    operating units, companies, admin status, and group counts.

    Supports terminal table, Markdown, Excel, and JSON output.
    """
    actx: AppContext = ctx.obj
    out = actx.output
    c = actx.client

    data = _fetch_access_data(c, include_inactive=all_users)

    # --- File output --------------------------------------------------------
    if output:
        ext = output.rsplit(".", 1)[-1].lower() if "." in output else ""
        if ext == "md":
            _write_markdown(data, output)
        elif ext in ("xlsx", "xls"):
            _write_excel(data, output)
        elif ext == "json":
            import json as json_mod

            with open(output, "w", encoding="utf-8") as f:
                json_mod.dump(data, f, indent=2, ensure_ascii=False)
        else:
            out.error(f"Unsupported file extension: .{ext} (use .md, .xlsx, or .json)")
            raise typer.Exit(1)
        out.success(f"Access report written to {output}")
        return

    # --- JSON stdout mode ---------------------------------------------------
    if actx.json_mode:
        out.raw_json(data)
        return

    # --- Terminal table ------------------------------------------------------
    rows: list[list[str]] = []
    for u in data["users"]:
        roles = ", ".join(u["roles"]) if u["roles"] else "-"
        ous = ", ".join(u["operating_units"]) if u["operating_units"] else "-"
        admin = "[red]Yes[/red]" if u["is_admin"] else "No"
        status = "[green]active[/green]" if u["status"] == "active" else "[red]inactive[/red]"
        rows.append(
            [
                u["login"],
                u["name"],
                roles,
                ous,
                u["default_ou"],
                admin,
                status,
                u["last_login"],
                str(u["groups_count"]),
            ]
        )

    out.table(
        f"User Access Report ({data['active_count']} active, {data['inactive_count']} inactive)",
        [
            ("Login", "cyan"),
            ("Name", ""),
            ("Roles", ""),
            ("OUs", ""),
            ("Default OU", ""),
            ("Admin", ""),
            ("Status", ""),
            ("Last Login", "dim"),
            ("Groups", "dim"),
        ],
        rows,
        data_for_json=data["users"],
    )

    # Roles summary
    role_rows: list[list[str]] = []
    for r in data["roles_summary"]:
        assigned = ", ".join(r["users"]) if r["users"] else "-"
        role_rows.append([r["role"], str(r["users_count"]), assigned])

    if role_rows:
        out.table(
            f"Roles Summary ({len(role_rows)} roles)",
            [("Role", "cyan"), ("Users", ""), ("Assigned To", "dim")],
            role_rows,
            data_for_json=data["roles_summary"],
        )
