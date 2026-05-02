"""Direct-Accurate credit-limit-proposal command (Bahasa, multi-tenant).

Pulls customer + invoice + receipt data straight from Accurate Online via
the SDK (no Odoo migration required) for ONE OR MORE tenants and writes
one combined Excel workbook with Bahasa Indonesia labels.

Sheets (combined):

* Ringkasan — totals across all tenants
* Duplikat Antar-CV — customers whose Identity Key appears in 2+ CVs
* Usulan Limit Kredit — one row per (tenant × customer) with proposed limit
* Master Pelanggan — full customer master rows, all tenants
* Faktur — every invoice across all tenants, with payment state + DSO

For 100% data accuracy (shareholder review), invoice detail is fetched
per record — no estimation for partial-payment residuals.

Spec: docs/superpowers/specs/2026-04-29-credit-limit-proposal-design.md
"""

from __future__ import annotations

from datetime import date
from pathlib import Path
from typing import Annotated, Any

import typer
import yaml
from accurate_sdk import AccurateClient

from kctl_odoo.commands._credit_limit.accurate_pull import (
    pull_customer_details,
    pull_customers,
    pull_sales_invoices_and_payments_accurate,
)
from kctl_odoo.commands._credit_limit.compute import (
    Params,
    ProposalRow,
    compute_per_pair,
)
from kctl_odoo.core.callbacks import AppContext

DEFAULT_ENV_FILE = Path("/home/tgunawan/project/00-new-projects/kodemeio-workspace/kodemeio-accurate/.env")
DEFAULT_CONFIG_PATH = Path("~/.config/kodemeio/config.yaml").expanduser()

# Display-name → sub-group mapping (TPP1 / TPP2 / TPP3 / TPP4 / Tier 1).
# Display name = env_key suffix with underscores → spaces.
# Used to colour-code and sub-total the report.
TENANT_SUBGROUP: dict[str, str] = {
    "26 BU TRD CV CAKRAWALA SARANA PRIORITAS": "TPP1",  # CSP
    "CV TUNGGAL PRAWIRA PAKERTI": "TPP1",  # TPRP
    "CV TUNGGAL PANGAN PAKERTI": "TPP1",  # TPK
    "26 BU TRD CV CAKRAWALA INTERNUSA PRIORITAS": "TPP2",  # CIP
    "CV TUNGGAL PUTRA PAKERTI": "TPP2",  # TPTR
    "26 BP TRD CV MAKMUR PANGAN JAYA": "TPP3",  # MPJ
    "26 BU TRD CV LINTAS FRESH INTERNUSA": "TPP3",  # LFI
    "CV LESTARI FRESH INTERNUSA": "TPP4",  # LSFI
    "26 BP TRD CV LUMBUNG YASA DAGANG": "TPP4",  # LYD
    "25 BP TRD CV YATA SIKHA UDAYA": "Tier 1",  # YSU
    "25 BP TRD CV YATA SIKHA UTAMA": "Tier 1",  # YSUT
    "25 BP CV YATA SIKHA ULTIMA": "Tier 1",  # YSUL
}


def _subgroup_of(tenant_name: str) -> str:
    return TENANT_SUBGROUP.get(tenant_name, "Lain-lain")


# 12 trading-type CVs in the TPP group (per Odoo res.company "Trading" tag).
# Used by --trading-only filter. The "25 KT IMP TRD" entry is excluded — it's
# a mixed import-trading entity, not pure trading.
TRADING_TENANT_KEYS: frozenset[str] = frozenset(
    {
        "ACCURATE_TOKEN_26_BU_TRD_CV_CAKRAWALA_SARANA_PRIORITAS",  # CSP — TPP1
        "ACCURATE_TOKEN_CV_TUNGGAL_PRAWIRA_PAKERTI",  # TPRP — TPP1
        "ACCURATE_TOKEN_CV_TUNGGAL_PANGAN_PAKERTI",  # TPK — TPP1
        "ACCURATE_TOKEN_26_BU_TRD_CV_CAKRAWALA_INTERNUSA_PRIORITAS",  # CIP — TPP2
        "ACCURATE_TOKEN_CV_TUNGGAL_PUTRA_PAKERTI",  # TPTR — TPP2
        "ACCURATE_TOKEN_26_BP_TRD_CV_MAKMUR_PANGAN_JAYA",  # MPJ — TPP3
        "ACCURATE_TOKEN_26_BU_TRD_CV_LINTAS_FRESH_INTERNUSA",  # LFI — TPP3
        "ACCURATE_TOKEN_CV_LESTARI_FRESH_INTERNUSA",  # LSFI — TPP4
        "ACCURATE_TOKEN_26_BP_TRD_CV_LUMBUNG_YASA_DAGANG",  # LYD — TPP4
        "ACCURATE_TOKEN_25_BP_TRD_CV_YATA_SIKHA_UDAYA",  # YSU — Tier 1
        "ACCURATE_TOKEN_25_BP_TRD_CV_YATA_SIKHA_UTAMA",  # YSUT — Tier 1
        "ACCURATE_TOKEN_25_BP_CV_YATA_SIKHA_ULTIMA",  # YSUL — Tier 1
    }
)


# ---------------------------------------------------------------------------
# Token + secret resolution
# ---------------------------------------------------------------------------


def _read_env_file(path: Path) -> dict[str, str]:
    """Parse a ``.env`` file into a plain dict."""
    env: dict[str, str] = {}
    if not path.exists():
        return env
    for raw_line in path.read_text().splitlines():
        line = raw_line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, _, value = line.partition("=")
        key = key.strip()
        value = value.strip()
        if (value.startswith('"') and value.endswith('"')) or (value.startswith("'") and value.endswith("'")):
            value = value[1:-1]
        env[key] = value
    return env


def _resolve_token(env_key: str, env_file: Path) -> str:
    env = _read_env_file(env_file)
    if env_key in env and env[env_key]:
        return env[env_key]
    full_key = env_key if env_key.startswith("ACCURATE_TOKEN_") else f"ACCURATE_TOKEN_{env_key}"
    if full_key in env and env[full_key]:
        return env[full_key]
    raise typer.BadParameter(f"Token not found in {env_file} for key '{env_key}' or '{full_key}'.")


def _resolve_signature_secret(config_path: Path = DEFAULT_CONFIG_PATH) -> str:
    if not config_path.exists():
        raise typer.BadParameter(f"kctl config not found at {config_path}")
    cfg = yaml.safe_load(config_path.read_text()) or {}
    try:
        secret = cfg["profiles"]["idtpp"]["accurate"]["signature_secret"]
    except (KeyError, TypeError) as exc:
        raise typer.BadParameter(
            f"signature_secret missing in {config_path} (profiles.idtpp.accurate.signature_secret)"
        ) from exc
    if not secret:
        raise typer.BadParameter(f"signature_secret is empty in {config_path}")
    return secret


def _discover_tenants_from_env(env_file: Path) -> list[tuple[str, str]]:
    """Find every ``ACCURATE_TOKEN_<KEY>`` in the env file.

    Returns list of (display_name, env_var_name) tuples. Display name is
    derived from the env var suffix: ``ACCURATE_TOKEN_CV_TUNGGAL_PANGAN_PAKERTI``
    → ``"CV TUNGGAL PANGAN PAKERTI"``.
    """
    env = _read_env_file(env_file)
    tenants: list[tuple[str, str]] = []
    for key in sorted(env.keys()):
        if not key.startswith("ACCURATE_TOKEN_") or not env[key]:
            continue
        suffix = key[len("ACCURATE_TOKEN_") :]
        display = suffix.replace("_", " ")
        tenants.append((display, key))
    return tenants


# ---------------------------------------------------------------------------
# Per-tenant data pull
# ---------------------------------------------------------------------------


def _identity_keys(cust: dict[str, Any]) -> list[tuple[str, str]]:
    """Build a list of cross-CV match candidates for one customer.

    Returns ``[(match_type, key_value), ...]`` covering up to 4 signals:

    * NPWP — most authoritative for B2B
    * KTP / idCard — most authoritative for individuals
    * WP (Wajib Pajak name) — when NPWP not entered, the WP name on the
      tax identity is still a useful match signal
    * Address — street + city as a soft fallback

    Two customers are flagged as candidate duplicates if **any** of these
    keys match across CVs. This is a SUGGESTION for human review — the
    final decision (real duplicate vs coincidence) belongs to the
    reviewer.
    """
    out: list[tuple[str, str]] = []
    npwp = (cust.get("npwp") or "").strip()
    if npwp:
        out.append(("NPWP", npwp))
    id_card = (cust.get("id_card") or "").strip()
    if id_card:
        out.append(("KTP", id_card))
    wp_name = (cust.get("wp_name") or "").strip().upper()
    name = (cust.get("name") or "").strip().upper()
    if wp_name and wp_name != name:
        out.append(("WP", wp_name))
    street = (cust.get("billing_street") or "").strip().upper()
    city = (cust.get("billing_city") or "").strip().upper()
    if street and city:
        out.append(("Alamat", f"{street}|{city}"))
    return out


def _identity_key(cust: dict[str, Any]) -> str:
    """Compatibility wrapper — return the highest-priority match key as a string.

    Used by sheets that show a single key column. Match-by-any logic in
    the duplicate-detection sheet uses :func:`_identity_keys` instead.
    """
    keys = _identity_keys(cust)
    if keys:
        kt, kv = keys[0]
        return f"{kt}:{kv}"
    name = (cust.get("name") or "").strip().upper()
    city = (cust.get("billing_city") or "").strip().upper()
    return f"NAMA:{name}|{city}"


def _pull_one_tenant(
    *,
    tenant_name: str,
    token: str,
    secret: str,
    today: date,
    params: Params,
    out: Any,
    limit_customers: int | None = None,
) -> tuple[list[dict[str, Any]], dict[int, list[dict[str, Any]]], list[tuple[dict[str, Any], ProposalRow]]]:
    """Pull customers + invoices + receipts for one tenant; return enriched data."""
    out.info(f"  → {tenant_name}: koneksi ke Accurate Online...")
    client = AccurateClient(api_token=token, signature_secret=secret)

    customers = pull_customers(client)
    if limit_customers is not None and limit_customers > 0:
        customers = customers[:limit_customers]
    out.info(f"     {len(customers)} pelanggan (list)")

    out.info(f"     Mengambil faktur + penerimaan dengan detail (akurasi 100%)...")
    by_customer_all = pull_sales_invoices_and_payments_accurate(client, today=today)
    customer_ids = {c["id"] for c in customers}
    by_customer = {cid: invs for cid, invs in by_customer_all.items() if cid in customer_ids}
    n_invoices = sum(len(v) for v in by_customer.values())
    out.info(f"     {n_invoices} faktur, {len(by_customer)} pelanggan dengan transaksi")

    all_ids = sorted({c["id"] for c in customers})
    if all_ids:
        out.info(f"     Mengambil detail master ({len(all_ids)} pelanggan, ~{len(all_ids) // 8 + 1}d)...")
        detail_map = pull_customer_details(client, all_ids)
        for cust in customers:
            cid = cust["id"]
            if cid in detail_map:
                d = detail_map[cid]
                cust["current_credit_limit"] = d["current_credit_limit"]
                cust["billing_street"] = d["billing_street"] or cust.get("billing_street", "")
                cust["billing_city"] = d["billing_city"] or cust.get("billing_city", "")
                cust["billing_province"] = d["billing_province"] or cust.get("billing_province", "")
                cust["billing_zip"] = d["billing_zip"] or cust.get("billing_zip", "")
                cust["billing_country"] = d["billing_country"] or cust.get("billing_country", "")
                cust["shipping_street"] = d["shipping_street"]
                cust["shipping_city"] = d["shipping_city"]
                cust["shipping_province"] = d["shipping_province"]
                cust["id_card"] = d["id_card"]
                cust["npwp"] = d["npwp"] or cust.get("npwp", "")
                cust["wp_name"] = d["wp_name"] or cust.get("wp_name", "")
                cust["term_name"] = d.get("term_name", "")
                cust["term_net_days"] = d.get("term_net_days", 0)
                cust["max_invoice_age"] = d.get("max_invoice_age", 0)
                cust["_raw_accurate"] = d.get("_raw_accurate") or {}
            cust["tenant"] = tenant_name
            cust["identity_key"] = _identity_key(cust)
        # Tag invoices with tenant too
        for invs in by_customer.values():
            for inv in invs:
                inv["tenant"] = tenant_name

    proposals: list[tuple[dict[str, Any], ProposalRow]] = []
    for cust in customers:
        cust_id = cust["id"]
        invoices = by_customer.get(cust_id, [])
        current_limit = cust.get("current_credit_limit", 0.0) or 0.0
        row = compute_per_pair(
            invoices,
            current_limit=current_limit,
            partner_id=cust_id,
            partner_name=cust["name"],
            trading_companies=tenant_name,
            params=params,
        )
        proposals.append((cust, row))
    return customers, by_customer, proposals


# ---------------------------------------------------------------------------
# Workbook writer (Bahasa, multi-tenant combined)
# ---------------------------------------------------------------------------


def _money_fmt() -> str:
    return "#,##0"


def _slugify(name: str) -> str:
    out = name.lower().strip()
    for ch in (" ", "."):
        out = out.replace(ch, "-")
    out = "".join(c for c in out if c.isalnum() or c == "-")
    while "--" in out:
        out = out.replace("--", "-")
    return out.strip("-") or "tenant"


# Display-only translation maps (compute-layer values stay English in code)
_BEHAVIOR_ID = {
    "Good": "Bagus",
    "Slow": "Lambat",
    "Bad": "Buruk",
    "Unknown": "Belum Diketahui",
}
_FLAG_ID = {
    "REVIEW_REQUIRED": "PERLU_REVIEW",
    "": "",
}
_PROPOSED_ID = {
    # Display INSUFFICIENT_DATA as blank — the row is still highlighted yellow
    # so reviewers can see "no proposal" at a glance without literal label text.
    "INSUFFICIENT_DATA": "",
}


_PROPOSED_ROUND_TO = 1_000_000  # Round proposed limits to nearest Rp 1,000,000


def _round_to(value: float, step: float) -> float:
    """Round UP to nearest ``step`` (e.g. 1,000,000) — produces clean credit-ops numbers.

    Ceiling rounding gives a small buffer for normal monthly sales variance
    and yields communicable round numbers. With coverage=1.0 the formula is
    still conservative; the ceiling adds at most ``step - 1`` IDR per row.
    A Rp 12,345,678 result becomes Rp 13,000,000.
    """
    if step <= 0:
        return value
    import math

    return float(math.ceil(value / step) * step)


def _bahasa_proposed(value: Any) -> Any:
    """Map proposed_limit to display value: blank for marker, floor-rounded for float."""
    if isinstance(value, str):
        return _PROPOSED_ID.get(value, value)
    if isinstance(value, (int, float)):
        if value < 0:
            return 0.0  # net-negative-sales case stays at 0
        return _round_to(float(value), _PROPOSED_ROUND_TO)
    return value


def _write_combined_workbook(
    output: Path,
    *,
    tenant_results: list[
        tuple[str, list[dict[str, Any]], dict[int, list[dict[str, Any]]], list[tuple[dict[str, Any], ProposalRow]]]
    ],
    params: Params,
    coverage: float,
) -> None:
    """Render the combined-tenant workbook with Bahasa Indonesia labels."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    hdr_font = Font(bold=True, color="FFFFFF", size=11)
    hdr_fill = PatternFill("solid", fgColor="2F5496")
    review_fill = PatternFill("solid", fgColor="FFC7CE")
    insufficient_fill = PatternFill("solid", fgColor="FFEB9C")
    duplicate_fill = PatternFill("solid", fgColor="FFD580")  # orange

    def _write_headers(ws: Any, headers: list[str], widths: list[int] | None = None) -> None:
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = hdr_font
            cell.fill = hdr_fill
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
        if widths:
            for i, w in enumerate(widths, 1):
                ws.column_dimensions[get_column_letter(i)].width = w
        ws.freeze_panes = "A2"

    # Aggregate stats across tenants
    n_tenants = len(tenant_results)
    total_customers = sum(len(custs) for _, custs, _, _ in tenant_results)
    total_invoiced = sum(len(by_c) for _, _, by_c, _ in tenant_results)
    total_invoices = sum(sum(len(v) for v in by_c.values()) for _, _, by_c, _ in tenant_results)
    sum_current = 0.0
    sum_proposed = 0.0
    sum_open_ar = 0.0
    n_review = 0
    n_insufficient = 0
    n_proposal = 0
    for _name, custs, _by_c, props in tenant_results:
        sum_current += sum(c.get("current_credit_limit", 0.0) or 0.0 for c in custs)
        for _cust, row in props:
            sum_open_ar += float(row.total_open_AR_idr)
            if isinstance(row.proposed_limit, (int, float)):
                sum_proposed += float(row.proposed_limit)
                n_proposal += 1
            if row.proposed_limit == "INSUFFICIENT_DATA":
                n_insufficient += 1
            if row.review_flag:
                n_review += 1

    # ── Sheet 1: Ringkasan (executive summary, 9/10 quality) ───────
    from openpyxl.styles import Border, Side

    ws_sum = wb.active
    ws_sum.title = "Ringkasan"

    # Distribution by payment behavior + total sales by behavior
    behavior_dist: dict[str, dict[str, float]] = {
        "Bagus": {"count": 0, "sales": 0.0},
        "Lambat": {"count": 0, "sales": 0.0},
        "Buruk": {"count": 0, "sales": 0.0},
        "Belum Diketahui": {"count": 0, "sales": 0.0},
    }
    for _name, _custs, _by_c, props in tenant_results:
        for _cust, row in props:
            if row.invoice_count == 0:
                continue
            label = _BEHAVIOR_ID.get(row.payment_behavior, row.payment_behavior)
            if label not in behavior_dist:
                behavior_dist[label] = {"count": 0, "sales": 0.0}
            behavior_dist[label]["count"] += 1
            behavior_dist[label]["sales"] += row.total_sales_idr

    # Reusable styles
    title_font = Font(bold=True, size=18, color="1F4E79")
    subtitle_font = Font(italic=True, size=10, color="555555")
    section_font = Font(bold=True, size=12, color="FFFFFF")
    section_fill = PatternFill("solid", fgColor="1F4E79")
    label_font = Font(bold=True, size=10)
    big_value_font = Font(bold=True, size=14, color="1F4E79")
    money_value_font = Font(bold=True, size=11)
    grand_fill = PatternFill("solid", fgColor="DDEBF7")
    thin_border = Border(
        left=Side(style="thin", color="BFBFBF"),
        right=Side(style="thin", color="BFBFBF"),
        top=Side(style="thin", color="BFBFBF"),
        bottom=Side(style="thin", color="BFBFBF"),
    )

    def _section(row: int, text: str, span: int = 7) -> None:
        ws_sum.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
        cell = ws_sum.cell(row=row, column=1, value=text)
        cell.font = section_font
        cell.fill = section_fill
        cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws_sum.row_dimensions[row].height = 22

    # Column widths
    widths_map = {"A": 38, "B": 22, "C": 22, "D": 18, "E": 22, "F": 22, "G": 22}
    for col, w in widths_map.items():
        ws_sum.column_dimensions[col].width = w

    # Title banner (rows 1-3)
    ws_sum.merge_cells(start_row=1, start_column=1, end_row=1, end_column=7)
    c1 = ws_sum.cell(row=1, column=1, value="ANALISA KREDIT LIMIT")
    c1.font = title_font
    c1.alignment = Alignment(horizontal="center", vertical="center")
    ws_sum.row_dimensions[1].height = 32
    ws_sum.merge_cells(start_row=2, start_column=1, end_row=2, end_column=7)
    c2 = ws_sum.cell(row=2, column=1, value="Konsolidasi Pelanggan Lintas-CV: Group TPP")
    c2.font = Font(italic=True, size=11, color="555555")
    c2.alignment = Alignment(horizontal="center")
    ws_sum.merge_cells(start_row=3, start_column=1, end_row=3, end_column=7)
    c3 = ws_sum.cell(
        row=3,
        column=1,
        value=f"Tanggal: {params.today.isoformat()}    |    Sumber: Accurate Online (live)    "
        f"|    Akurasi Data: 100% (detail per faktur)",
    )
    c3.font = subtitle_font
    c3.alignment = Alignment(horizontal="center")

    # ── Section: KPI BLOCK ──
    # Compute richer metrics for shareholder review.
    sum_total_sales = 0.0
    proposed_limits_list: list[float] = []
    n_overdue_30: int = 0
    n_overdue_120: int = 0
    for _name, _custs, _by_c, props in tenant_results:
        for _cust, row in props:
            sum_total_sales += float(row.total_sales_idr)
            if isinstance(row.proposed_limit, (int, float)) and row.proposed_limit > 0:
                # Use the same ceiling-rounded value users see in the report
                rounded = _round_to(float(row.proposed_limit), _PROPOSED_ROUND_TO)
                proposed_limits_list.append(rounded)
            if row.oldest_overdue_days > 30:
                n_overdue_30 += 1
            if row.has_overdue_120d:
                n_overdue_120 += 1
    proposed_limits_list.sort()
    avg_proposed = sum(proposed_limits_list) / len(proposed_limits_list) if proposed_limits_list else 0.0
    median_proposed = proposed_limits_list[len(proposed_limits_list) // 2] if proposed_limits_list else 0.0
    max_proposed = proposed_limits_list[-1] if proposed_limits_list else 0.0
    sum_proposed_rounded = sum(proposed_limits_list)

    _section(5, "RINGKASAN UTAMA")
    kpi_pairs: list[tuple[str, Any, str]] = [
        # Cakupan Data
        ("Jumlah CV / Tenant Diproses", n_tenants, ""),
        ("Total Pelanggan (semua CV)", total_customers, ""),
        (
            "Pelanggan dengan Transaksi",
            f"{total_invoiced}  ({total_invoiced / total_customers * 100:.0f}% dari total)" if total_customers else 0,
            "",
        ),
        ("Pelanggan Tanpa Transaksi (DATA_KOSONG)", n_insufficient, ""),
        ("", "", ""),
        # Volume Transaksi
        ("Total Faktur Tercatat", total_invoices, ""),
        ("Total Penjualan / Sales (IDR)", sum_total_sales, "money"),
        ("Total AR Berjalan / Outstanding (IDR)", sum_open_ar, "money"),
        ("", "", ""),
        # Usulan Limit Kredit
        ("Total Usulan Limit Kredit (IDR)", sum_proposed_rounded, "money"),
        ("  ↳ Rata-rata Usulan per Pelanggan (IDR)", avg_proposed, "money"),
        ("  ↳ Median Usulan per Pelanggan (IDR)", median_proposed, "money"),
        ("  ↳ Usulan Tertinggi (IDR)", max_proposed, "money"),
        ("", "", ""),
        # Indikator Risiko
        ("Pelanggan Perlu Review (PERLU_REVIEW)", n_review, ""),
        ("Pelanggan dengan Tunggakan > 30 hari", n_overdue_30, ""),
        ("Pelanggan dengan Tunggakan > 120 hari (KRITIS)", n_overdue_120, ""),
        ("", "", ""),
        # Konteks tambahan (de-emphasized)
        (
            "Konteks: Limit Saat Ini di Accurate",
            "Sebagian besar pelanggan belum di-set limit; perbandingan langsung tidak relevan",
            "",
        ),
        ("  ↳ Total Limit Saat Ini (IDR, parsial)", sum_current, "money"),
    ]
    base_row = 6
    for i, (label, value, kind) in enumerate(kpi_pairs):
        r = base_row + i
        if not label and not value:
            continue
        cl = ws_sum.cell(row=r, column=1, value=label)
        cl.font = label_font
        cv = ws_sum.cell(row=r, column=2, value=value)
        if kind == "money" and isinstance(value, (int, float)):
            cv.number_format = _money_fmt()
            cv.font = money_value_font
        elif label.startswith(("Total ", "Jumlah ", "Pelanggan ")):
            cv.font = big_value_font if isinstance(value, (int, float)) else Font(bold=True, size=11)
        cv.alignment = Alignment(horizontal="right")

    # ── Section: Distribusi Perilaku Pembayaran ──
    behavior_section_row = base_row + len(kpi_pairs) + 1
    _section(behavior_section_row, "DISTRIBUSI PERILAKU PEMBAYARAN")
    bh_hdr_row = behavior_section_row + 1
    bh_headers = ["Perilaku", "Jumlah Pelanggan", "Total Penjualan (IDR)", "% dari Total"]
    for c_idx, h in enumerate(bh_headers, start=1):
        cell = ws_sum.cell(row=bh_hdr_row, column=c_idx, value=h)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = thin_border
    behavior_total_sales = sum(d["sales"] for d in behavior_dist.values())
    behavior_total_count = sum(d["count"] for d in behavior_dist.values())
    behavior_color = {"Bagus": "C6EFCE", "Lambat": "FFEB9C", "Buruk": "FFC7CE", "Belum Diketahui": "F2F2F2"}
    for i, label in enumerate(["Bagus", "Lambat", "Buruk", "Belum Diketahui"], start=1):
        r = bh_hdr_row + i
        d = behavior_dist[label]
        pct = (d["sales"] / behavior_total_sales * 100.0) if behavior_total_sales else 0.0
        for c_idx, val in enumerate([label, d["count"], d["sales"], f"{pct:.1f}%"], start=1):
            cell = ws_sum.cell(row=r, column=c_idx, value=val)
            cell.border = thin_border
            cell.fill = PatternFill("solid", fgColor=behavior_color.get(label, "FFFFFF"))
            if c_idx == 3:
                cell.number_format = _money_fmt()
            if c_idx == 1:
                cell.font = label_font
            cell.alignment = Alignment(horizontal="left" if c_idx == 1 else "right")
    # Behavior grand total
    bg_row = bh_hdr_row + 5
    for c_idx, val in enumerate(["TOTAL", behavior_total_count, behavior_total_sales, "100.0%"], start=1):
        cell = ws_sum.cell(row=bg_row, column=c_idx, value=val)
        cell.fill = grand_fill
        cell.border = thin_border
        cell.font = Font(bold=True)
        if c_idx == 3:
            cell.number_format = _money_fmt()
        cell.alignment = Alignment(horizontal="left" if c_idx == 1 else "right")

    # ── Section: DAFTAR CV YANG DIPROSES ──
    cv_section_row = bg_row + 2
    _section(cv_section_row, "DAFTAR CV YANG DIPROSES (per CV)")
    cv_hdr_row = cv_section_row + 1
    cv_headers = [
        "Sub-group",
        "CV / Tenant",
        "Total Pelanggan",
        "Pelanggan Aktif",
        "Total Faktur",
        "Total Penjualan (IDR)",
        "Total AR Berjalan (IDR)",
        "Pelanggan Perlu Review",
    ]
    for c_idx, h in enumerate(cv_headers, start=1):
        cell = ws_sum.cell(row=cv_hdr_row, column=c_idx, value=h)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = thin_border
    # Sort by sub-group order (TPP1→TPP4→Tier 1→Lain-lain), then by AR desc
    subgroup_order = {"TPP1": 1, "TPP2": 2, "TPP3": 3, "TPP4": 4, "Tier 1": 5, "Lain-lain": 9}
    sorted_tenants = sorted(
        tenant_results,
        key=lambda t: (
            subgroup_order.get(_subgroup_of(t[0]), 9),
            -sum(float(r.total_open_AR_idr) for _c, r in t[3]),
        ),
    )
    # Sub-group color map
    subgroup_color = {
        "TPP1": "E8F4FD",  # light blue
        "TPP2": "FFF4E0",  # light orange
        "TPP3": "EAF7E8",  # light green
        "TPP4": "FAEBF4",  # light pink
        "Tier 1": "F5F0FA",  # light purple
        "Lain-lain": "F2F2F2",
    }
    cv_grand: dict[str, float] = {
        "customers": 0,
        "invoiced": 0,
        "invoices": 0,
        "sales": 0.0,
        "open_ar": 0.0,
        "review": 0,
    }
    subgroup_totals: dict[str, dict[str, float]] = {}
    for i, (tname, custs, by_c, props) in enumerate(sorted_tenants, start=1):
        cv_invoiced = len(by_c)
        cv_invoices = sum(len(v) for v in by_c.values())
        cv_sales = sum(float(r.total_sales_idr) for _c, r in props)
        cv_open_ar = sum(float(r.total_open_AR_idr) for _c, r in props)
        cv_review = sum(1 for _c, r in props if r.review_flag)
        cv_grand["customers"] += len(custs)
        cv_grand["invoiced"] += cv_invoiced
        cv_grand["invoices"] += cv_invoices
        cv_grand["sales"] += cv_sales
        cv_grand["open_ar"] += cv_open_ar
        cv_grand["review"] += cv_review
        sg = _subgroup_of(tname)
        sgt = subgroup_totals.setdefault(
            sg, {"customers": 0, "invoiced": 0, "invoices": 0, "sales": 0.0, "open_ar": 0.0, "review": 0}
        )
        sgt["customers"] += len(custs)
        sgt["invoiced"] += cv_invoiced
        sgt["invoices"] += cv_invoices
        sgt["sales"] += cv_sales
        sgt["open_ar"] += cv_open_ar
        sgt["review"] += cv_review
        r = cv_hdr_row + i
        cells = [
            (1, sg, "center"),
            (2, tname, "left"),
            (3, len(custs), "right"),
            (4, cv_invoiced, "right"),
            (5, cv_invoices, "right"),
            (6, cv_sales, "right"),
            (7, cv_open_ar, "right"),
            (8, cv_review, "right"),
        ]
        sg_fill = PatternFill("solid", fgColor=subgroup_color.get(sg, "FFFFFF"))
        for c_idx, val, align in cells:
            cell = ws_sum.cell(row=r, column=c_idx, value=val)
            cell.alignment = Alignment(horizontal=align)
            cell.border = thin_border
            cell.fill = sg_fill
            if c_idx in (6, 7):
                cell.number_format = _money_fmt()
            if c_idx == 1:
                cell.font = Font(bold=True, size=10)
            if c_idx == 8 and val > 0:
                cell.font = Font(bold=True, color="C00000")
    # Grand total row
    gt_row = cv_hdr_row + len(sorted_tenants) + 1
    grand_cells = [
        (1, "", "left"),
        (2, "TOTAL GROUP TPP", "left"),
        (3, cv_grand["customers"], "right"),
        (4, cv_grand["invoiced"], "right"),
        (5, cv_grand["invoices"], "right"),
        (6, cv_grand["sales"], "right"),
        (7, cv_grand["open_ar"], "right"),
        (8, cv_grand["review"], "right"),
    ]
    for c_idx, val, align in grand_cells:
        cell = ws_sum.cell(row=gt_row, column=c_idx, value=val)
        cell.fill = grand_fill
        cell.font = Font(bold=True, size=11)
        cell.border = thin_border
        cell.alignment = Alignment(horizontal=align)
        if c_idx in (6, 7):
            cell.number_format = _money_fmt()

    # ── Section: PARAMETER & METODOLOGI ──
    param_row = gt_row + 2
    _section(param_row, "PARAMETER & METODOLOGI")
    pe_rows = [
        ("Formula Usulan Limit", f"Usulan = Rata-rata Penjualan Bulanan × {coverage}"),
        ("Pengali Coverage", coverage),
        ("Ambang Faktur Minimum", f"{params.insufficient_min_invoices} faktur"),
        ("Ambang Tunggakan Perlu Review", f"> {params.overdue_review_days} hari"),
        (
            "Pencocokan Duplikat Antar-CV",
            "Multi-signal: NPWP / KTP / Nama Wajib Pajak / Alamat (saran, perlu verifikasi manual)",
        ),
        ("Sumber Data", "Accurate Online (endpoint detail per record, akurasi 100%)"),
        ("Hari Perhitungan", params.today.isoformat()),
    ]
    for i, (k, v) in enumerate(pe_rows, start=1):
        r = param_row + i
        ws_sum.cell(row=r, column=1, value=k).font = label_font
        ws_sum.merge_cells(start_row=r, start_column=2, end_row=r, end_column=7)
        cell = ws_sum.cell(row=r, column=2, value=v)
        cell.alignment = Alignment(horizontal="left", indent=1)
        if isinstance(v, (int, float)):
            cell.number_format = "0.00"

    # ── Section: DAFTAR ISTILAH (Glossary) ──
    glossary_row = param_row + len(pe_rows) + 2
    _section(glossary_row, "DAFTAR ISTILAH (GLOSSARY)")
    glos_hdr_row = glossary_row + 1
    glos_headers = ["Istilah", "Singkatan / Inggris", "Penjelasan"]
    for c_idx, h in enumerate(glos_headers, start=1):
        cell = ws_sum.cell(row=glos_hdr_row, column=c_idx, value=h)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = thin_border
    # Merge "Penjelasan" across columns C-G for readability
    glos_terms: list[tuple[str, str, str]] = [
        (
            "AR Berjalan / Piutang Usaha",
            "AR (Accounts Receivable) / Outstanding",
            "Sisa tagihan yang belum dibayar pelanggan dari faktur yang sudah dikirim. Sumber risiko utama yang harus dijaga di bawah limit kredit.",
        ),
        (
            "Faktur Penjualan",
            "Sales Invoice",
            "Tagihan yang dikeluarkan saat barang/jasa diserahkan. Total tagihan = pokok penjualan + PPN.",
        ),
        (
            "Hari Rata-rata Penagihan",
            "DSO (Days Sales Outstanding)",
            "Rata-rata jumlah hari dari tanggal faktur sampai pembayaran masuk. DSO ≤ 14 hari = Bagus. 14–30 hari = Lambat. >30 hari = Buruk.",
        ),
        (
            "Tunggakan / Overdue",
            "Past-Due Days",
            "Selisih hari antara tanggal jatuh tempo dengan tanggal hari ini, untuk faktur yang masih ada sisa (residual) > 0.",
        ),
        (
            "Limit Kredit",
            "Credit Limit",
            "Plafon maksimum tagihan yang boleh berjalan untuk satu pelanggan di satu CV. Diatur per CV (per perusahaan) di Accurate / Odoo.",
        ),
        (
            "Usulan Limit",
            "Proposed Credit Limit",
            f"Hasil perhitungan otomatis: Rata² Penjualan Bulanan × {coverage}. Bersifat usulan; keputusan akhir tetap di tangan reviewer.",
        ),
        (
            "Pengali Coverage",
            "Coverage Multiplier",
            "Faktor pengali pada formula usulan limit. 1.5 = limit menutupi 1,5 bulan penjualan rata-rata. Nilai standar konservatif.",
        ),
        (
            "Rata² Penjualan Bulanan",
            "Average Monthly Sales",
            "Total penjualan tercatat ÷ jumlah bulan aktif (sejak faktur pertama). Refund dihitung sebagai negatif (net).",
        ),
        (
            "Bulan Aktif",
            "Months Active",
            "Selisih hari dari faktur pertama sampai tanggal laporan, dibagi 30,44. Floor 1.0 untuk pelanggan baru.",
        ),
        (
            "Perilaku Pembayaran: Bagus",
            "Good payer",
            "Pelanggan dengan rata-rata DSO ≤ 14 hari. Indikator pembayaran sangat tepat waktu (cocok untuk produk segar).",
        ),
        (
            "Perilaku Pembayaran: Lambat",
            "Slow payer",
            "Pelanggan dengan rata-rata DSO 14–30 hari. Pembayaran lebih lama dari termin standar fresh-produce tetapi belum kritis.",
        ),
        (
            "Perilaku Pembayaran: Buruk",
            "Bad payer",
            "Pelanggan dengan rata-rata DSO > 30 hari. Risiko tinggi; otomatis di-flag PERLU_REVIEW.",
        ),
        (
            "Belum Diketahui",
            "Unknown",
            "Pelanggan belum punya faktur lunas, sehingga DSO tidak bisa dihitung. Bukan negatif, hanya belum cukup data.",
        ),
        (
            "PERLU_REVIEW",
            "REVIEW_REQUIRED",
            "Tanda bahwa baris perlu peninjauan manual sebelum limit diterima. Dipicu oleh: (a) perilaku Buruk, atau (b) ada faktur dengan tunggakan > 120 hari.",
        ),
        (
            "DATA TIDAK CUKUP (cell kosong + tint kuning)",
            "INSUFFICIENT_DATA",
            f"Pelanggan punya < {params.insufficient_min_invoices} faktur sehingga tidak ada usulan limit yang dihitung. Pertahankan limit saat ini.",
        ),
        (
            "NPWP",
            "Tax ID (Nomor Pokok Wajib Pajak)",
            "Identitas pajak resmi badan/perorangan di Indonesia. Sinyal pencocokan duplikat antar-CV paling otoritatif untuk B2B.",
        ),
        (
            "KTP",
            "National ID (Kartu Tanda Penduduk)",
            "Identitas penduduk Indonesia. Sinyal pencocokan duplikat untuk pelanggan perorangan.",
        ),
        (
            "WP / Wajib Pajak",
            "WP (Tax-bearer Name)",
            "Nama pemilik NPWP. Sinyal pencocokan duplikat ketika NPWP tidak terisi tapi nama wajib pajak ada.",
        ),
        (
            "Kunci Identitas",
            "Identity Key",
            "Kunci unik per pelanggan yang dipakai untuk pencocokan antar-CV. Prioritas: NPWP > KTP > Nama Wajib Pajak > Nama+Kota.",
        ),
        (
            "Duplikat Antar-CV",
            "Cross-CV Duplicate",
            "Pelanggan yang muncul di ≥2 CV dengan NPWP/KTP/WP/Alamat yang sama. Saran otomatis, wajib diverifikasi reviewer.",
        ),
        (
            "Eksposur Grup",
            "Group Exposure",
            "Total tagihan/AR satu pelanggan dijumlahkan dari semua CV di group. Jumlah inilah yang harus dipantau untuk risiko default.",
        ),
        ("Selisih (Δ)", "Delta", "Selisih antara Usulan Limit dan Limit Saat Ini. Positif = naik, negatif = turun."),
        (
            "CV / Tenant",
            "Company / Tenant",
            "Satu badan hukum (CV/PT) yang punya database Accurate sendiri. Group TPP terdiri dari 27 CV.",
        ),
    ]
    for i, (term_id, term_en, expl) in enumerate(glos_terms, start=1):
        r = glos_hdr_row + i
        c1 = ws_sum.cell(row=r, column=1, value=term_id)
        c1.font = label_font
        c1.alignment = Alignment(horizontal="left", wrap_text=True, vertical="top")
        c1.border = thin_border
        c2 = ws_sum.cell(row=r, column=2, value=term_en)
        c2.font = Font(italic=True)
        c2.alignment = Alignment(horizontal="left", wrap_text=True, vertical="top")
        c2.border = thin_border
        ws_sum.merge_cells(start_row=r, start_column=3, end_row=r, end_column=7)
        c3 = ws_sum.cell(row=r, column=3, value=expl)
        c3.alignment = Alignment(horizontal="left", wrap_text=True, vertical="top", indent=1)
        c3.border = thin_border
        # Lighter striping for readability
        if i % 2 == 1:
            for c_idx in range(1, 8):
                ws_sum.cell(row=r, column=c_idx).fill = PatternFill("solid", fgColor="FAFAFA")
        ws_sum.row_dimensions[r].height = 32

    # Freeze top 4 rows (title banner)
    ws_sum.freeze_panes = "A5"

    # ── Sheet 2: Duplikat Antar-CV (multi-signal: NPWP / KTP / WP / Address) ──
    # For each customer, generate up to 4 match keys (NPWP, KTP, WP name,
    # Address). Two customers in DIFFERENT CVs sharing ANY key are flagged
    # as candidate duplicates. This is a SUGGESTION for human review.
    #
    # Map: (match_type, key_value) -> list of (tenant_name, customer, row)
    key_to_rows: dict[tuple[str, str], list[tuple[str, dict[str, Any], ProposalRow]]] = {}
    for tname, custs, _by_c, props in tenant_results:
        prop_by_cust = {cust["id"]: row for cust, row in props}
        for cust in custs:
            row = prop_by_cust.get(cust["id"])
            if row is None:
                continue
            for match_type, key_value in _identity_keys(cust):
                key_to_rows.setdefault((match_type, key_value), []).append((tname, cust, row))

    # Filter to keys that appear in at least 2 DISTINCT CVs (cross-CV)
    duplicate_keys: dict[tuple[str, str], list[tuple[str, dict[str, Any], ProposalRow]]] = {}
    for k, entries in key_to_rows.items():
        cvs_in_match = {tname for tname, _c, _r in entries}
        if len(cvs_in_match) >= 2:
            duplicate_keys[k] = entries

    # Track customers flagged as duplicate (any key match) for highlighting
    flagged_cust_ids: set[tuple[str, int]] = set()
    for entries in duplicate_keys.values():
        for tname, cust, _row in entries:
            flagged_cust_ids.add((tname, cust["id"]))

    ws_dup = wb.create_sheet("Duplikat Antar-CV")
    dup_headers = [
        "Tipe Pencocokan",
        "Nilai Pencocokan",
        "Jumlah CV",
        "CV",
        "ID Pelanggan",
        "No Pelanggan",
        "Nama Pelanggan",
        "NPWP",
        "No KTP",
        "Nama Wajib Pajak",
        "Alamat",
        "Kota",
        "Provinsi",
        "Jumlah Faktur",
        "Total Penjualan (IDR)",
        "Usulan Limit (IDR)",
        "Limit Saat Ini (IDR)",
    ]
    dup_widths = [16, 36, 10, 32, 12, 14, 38, 22, 18, 28, 32, 18, 20, 12, 18, 18, 18]
    _write_headers(ws_dup, dup_headers, dup_widths)
    dup_money_cols = {15, 16, 17}
    # Add a note row at the top explaining the suggestion nature
    ws_dup.cell(
        row=1,
        column=len(dup_headers) + 2,
        value="Catatan: Ini adalah saran pencocokan otomatis berdasarkan NPWP / KTP / Nama WP / Alamat. Wajib diverifikasi manual.",
    ).font = Font(italic=True, color="555555")

    r_offset = 2
    # Sort: most-matches-first (most suspicious), then by match type
    sorted_keys = sorted(
        duplicate_keys.keys(),
        key=lambda k: (-len(duplicate_keys[k]), k[0], k[1]),
    )
    for match_type, key_value in sorted_keys:
        entries = duplicate_keys[(match_type, key_value)]
        for tname, cust, row in entries:
            proposed = _bahasa_proposed(row.proposed_limit)
            proposed_value: Any = float(proposed) if isinstance(proposed, (int, float)) else proposed
            values = [
                match_type,
                key_value,
                len(entries),
                tname,
                cust["id"],
                cust.get("customer_no", ""),
                cust.get("name", ""),
                cust.get("npwp", ""),
                cust.get("id_card", ""),
                cust.get("wp_name", ""),
                cust.get("billing_street", ""),
                cust.get("billing_city", ""),
                cust.get("billing_province", ""),
                row.invoice_count,
                row.total_sales_idr,
                proposed_value,
                row.current_limit_sum_idr,
            ]
            for c_idx, val in enumerate(values, start=1):
                cell = ws_dup.cell(row=r_offset, column=c_idx, value=val)
                if c_idx in dup_money_cols and isinstance(val, (int, float)):
                    cell.number_format = _money_fmt()
                cell.fill = duplicate_fill
            r_offset += 1
    if not duplicate_keys:
        ws_dup.cell(row=2, column=1, value="(Tidak ditemukan pelanggan duplikat antar-CV)").font = Font(italic=True)

    # ── Sheet 3: Usulan Limit Kredit (combined) ─────────────────────
    ws_prop = wb.create_sheet("Usulan Limit Kredit")
    proposed_header = f"Usulan Limit (IDR)\n[= Rata² Penjualan × {coverage}]"
    headers = [
        "CV",
        "ID Pelanggan",
        "No Pelanggan",
        "Nama Pelanggan",
        "NPWP",
        "No KTP",
        "Nama Wajib Pajak",
        "No Telp",
        "Kota",
        "Provinsi",
        "ToP (Term Of Payment)",
        "ToP Hari (netDays)",
        "Kunci Identitas",
        "Kandidat Duplikat",
        "Jumlah Faktur",
        "Faktur Lunas",
        "Faktur Pertama",
        "Bulan Aktif",
        "Total Penjualan (IDR)",
        "Rata-rata Penjualan Bulanan (IDR)",
        proposed_header,
        "Limit Saat Ini (IDR)",
        "Selisih (IDR)",
        "Selisih %",
        "Rata-rata DSO (hari)",
        "Perilaku Pembayaran",
        "Tunggakan >120 hari",
        "Tunggakan Terlama (hari)",
        "Total AR Berjalan (IDR)",
        "Flag Review",
        "Catatan",
    ]
    widths = [
        28,
        12,
        14,
        38,
        22,
        18,
        28,
        18,
        18,
        20,
        18,
        10,  # ToP, ToP Hari (new)
        38,
        22,
        14,
        14,
        14,
        14,
        18,
        18,
        22,
        18,
        16,
        10,
        12,
        18,
        16,
        16,
        18,
        18,
        30,
    ]
    _write_headers(ws_prop, headers, widths)
    # Money columns shifted by +2 due to new ToP columns at positions 11, 12
    money_cols = {19, 20, 21, 22, 23, 29}
    r_offset = 2
    for tname, custs, by_c, props in tenant_results:
        # Only customers with at least one invoice belong on the Proposal sheet.
        # Customers with no transactions still appear in Master Pelanggan.
        props_with_invoices = [(cust, row) for cust, row in props if row.invoice_count > 0]
        props_sorted = sorted(props_with_invoices, key=lambda p: -p[1].total_sales_idr)
        for cust, row in props_sorted:
            proposed = _bahasa_proposed(row.proposed_limit)
            proposed_value: Any = float(proposed) if isinstance(proposed, (int, float)) else proposed
            # Build "Kandidat Duplikat" cell — list which match types this customer hits
            cust_match_types: set[str] = set()
            for match_type, key_value in _identity_keys(cust):
                if (match_type, key_value) in duplicate_keys:
                    cust_match_types.add(match_type)
            duplicate_label = ", ".join(sorted(cust_match_types)) if cust_match_types else ""
            values = [
                tname,
                cust["id"],
                cust.get("customer_no", ""),
                cust.get("name", ""),
                cust.get("npwp", ""),
                cust.get("id_card", ""),
                cust.get("wp_name", ""),
                cust.get("mobile_phone", ""),
                cust.get("billing_city", ""),
                cust.get("billing_province", ""),
                cust.get("term_name", "") or "",
                cust.get("term_net_days", 0),
                cust.get("identity_key", ""),
                duplicate_label,
                row.invoice_count,
                row.settled_invoice_count,
                row.first_invoice_date.isoformat() if row.first_invoice_date else "",
                row.months_active,
                row.total_sales_idr,
                row.avg_monthly_sales_idr,
                proposed_value,
                row.current_limit_sum_idr,
                row.delta_idr,
                row.delta_pct,
                row.avg_DSO_days if row.avg_DSO_days is not None else "",
                _BEHAVIOR_ID.get(row.payment_behavior, row.payment_behavior),
                "Y" if row.has_overdue_120d else "",
                row.oldest_overdue_days,
                row.total_open_AR_idr,
                _FLAG_ID.get(row.review_flag, row.review_flag),
                row.note,
            ]
            for c_idx, val in enumerate(values, start=1):
                cell = ws_prop.cell(row=r_offset, column=c_idx, value=val)
                if c_idx in money_cols and isinstance(val, (int, float)):
                    cell.number_format = _money_fmt()
            if row.review_flag == "REVIEW_REQUIRED":
                for c_idx in range(1, len(headers) + 1):
                    ws_prop.cell(row=r_offset, column=c_idx).fill = review_fill
            elif row.proposed_limit == "INSUFFICIENT_DATA":
                for c_idx in range(1, len(headers) + 1):
                    ws_prop.cell(row=r_offset, column=c_idx).fill = insufficient_fill
            # Highlight cross-CV duplicates with orange (any match-type hit)
            # Identity Key is now col 13, Kandidat Duplikat is col 14 (shifted +2 by ToP cols)
            if (tname, cust["id"]) in flagged_cust_ids:
                ws_prop.cell(row=r_offset, column=13).fill = duplicate_fill
                ws_prop.cell(row=r_offset, column=14).fill = duplicate_fill
            # Yellow highlight: customer has no current credit limit set in Accurate.
            # Limit Saat Ini is now col 22 (shifted from 20 by +2 ToP cols).
            if not isinstance(row.current_limit_sum_idr, (int, float)) or row.current_limit_sum_idr <= 0:
                no_limit_fill = PatternFill("solid", fgColor="FFEB9C")  # Excel "yellow"
                ws_prop.cell(row=r_offset, column=22).fill = no_limit_fill
            r_offset += 1

    # ── Sheet 3b: Top 20 Pelanggan (group-level concentration) ───────
    ws_top = wb.create_sheet("Top 20 Pelanggan")

    # Compact title banner only (rows 1-2). Full explainer goes BELOW the table.
    ws_top.merge_cells(start_row=1, start_column=1, end_row=1, end_column=14)
    ct = ws_top.cell(row=1, column=1, value="TOP 20 PELANGGAN: KONSENTRASI RISIKO KREDIT GRUP")
    ct.font = Font(bold=True, size=14, color="1F4E79")
    ct.alignment = Alignment(horizontal="center", vertical="center")
    ws_top.row_dimensions[1].height = 24

    ws_top.merge_cells(start_row=2, start_column=1, end_row=2, end_column=14)
    ct2 = ws_top.cell(
        row=2,
        column=1,
        value=(
            "Peringkat 20 pelanggan teratas berdasarkan total Usulan Limit Kredit "
            "(dijumlah lintas-CV bila pelanggan terdaftar di beberapa CV)."
        ),
    )
    ct2.font = Font(italic=True, size=10, color="555555")
    ct2.alignment = Alignment(horizontal="center", wrap_text=True)
    ws_top.row_dimensions[2].height = 18

    # Table starts at row 4 (right after title + subtitle)
    table_start = 4
    top_headers = [
        "Peringkat",
        "Nama Pelanggan",
        "NPWP",
        "No KTP",
        "Jumlah CV",
        "Daftar CV",
        "Total Faktur",
        "Total Penjualan (IDR)",
        "Total AR Berjalan (IDR)",
        "Total Usulan Limit (IDR)",
        "% dari Total Usulan",
        "Rata-rata DSO (hari)",
        "Perilaku Pembayaran",
        "Catatan",
    ]
    top_widths = [8, 38, 22, 18, 8, 28, 12, 22, 22, 22, 12, 14, 18, 28]
    # Write headers at table_start (row 11) — _write_headers hardcodes row 1
    # which collides with the explainer banner above.
    for col, h in enumerate(top_headers, 1):
        cell = ws_top.cell(row=table_start, column=col, value=h)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
    for i, w in enumerate(top_widths, 1):
        ws_top.column_dimensions[get_column_letter(i)].width = w
    ws_top.freeze_panes = f"A{table_start + 1}"

    # Aggregate per partner across tenants (group-level)
    by_partner_group: dict[int, dict[str, Any]] = {}
    for tname, custs, _by_c, props in tenant_results:
        for cust, row in props:
            if row.invoice_count == 0:
                continue
            pid = cust["id"]
            agg = by_partner_group.setdefault(
                pid,
                {
                    "name": cust.get("name", ""),
                    "npwp": cust.get("npwp", ""),
                    "id_card": cust.get("id_card", ""),
                    "cvs": set(),
                    "invoice_count": 0,
                    "total_sales": 0.0,
                    "total_ar": 0.0,
                    "total_proposed": 0.0,
                    "dso_weighted_num": 0.0,
                    "dso_weighted_den": 0,
                    "behaviors": [],
                },
            )
            agg["cvs"].add(tname)
            agg["invoice_count"] += row.invoice_count
            agg["total_sales"] += row.total_sales_idr
            agg["total_ar"] += row.total_open_AR_idr
            if isinstance(row.proposed_limit, (int, float)):
                agg["total_proposed"] += _round_to(float(row.proposed_limit), _PROPOSED_ROUND_TO)
            if row.avg_DSO_days is not None and row.settled_invoice_count > 0:
                agg["dso_weighted_num"] += row.avg_DSO_days * row.settled_invoice_count
                agg["dso_weighted_den"] += row.settled_invoice_count
            agg["behaviors"].append(row.payment_behavior)

    grand_total_proposed = sum(a["total_proposed"] for a in by_partner_group.values()) or 1.0
    sorted_partners = sorted(by_partner_group.values(), key=lambda a: -a["total_proposed"])
    top_n = sorted_partners[:20]
    for rank, agg in enumerate(top_n, start=1):
        avg_dso = (agg["dso_weighted_num"] / agg["dso_weighted_den"]) if agg["dso_weighted_den"] else None
        from collections import Counter

        behavior_counter = Counter(agg["behaviors"])
        top_behavior = behavior_counter.most_common(1)[0][0]
        notes = []
        if len(agg["cvs"]) >= 3:
            notes.append(f"Eksposur di {len(agg['cvs'])} CV")
        if agg["total_ar"] > 1_000_000_000:
            notes.append("AR > 1 Miliar")
        pct = agg["total_proposed"] / grand_total_proposed * 100.0
        values = [
            rank,
            agg["name"],
            agg["npwp"] or "",
            agg["id_card"] or "",
            len(agg["cvs"]),
            ", ".join(sorted(agg["cvs"]))[:60],
            agg["invoice_count"],
            agg["total_sales"],
            agg["total_ar"],
            agg["total_proposed"],
            f"{pct:.1f}%",
            f"{avg_dso:.1f}" if avg_dso is not None else "",
            _BEHAVIOR_ID.get(top_behavior, top_behavior),
            "; ".join(notes),
        ]
        money_top_cols = {8, 9, 10}
        for c_idx, val in enumerate(values, start=1):
            cell = ws_top.cell(row=table_start + rank, column=c_idx, value=val)
            if c_idx in money_top_cols and isinstance(val, (int, float)):
                cell.number_format = _money_fmt()
            cell.alignment = Alignment(horizontal="right" if c_idx in (1, 5, 7, 8, 9, 10, 11, 12) else "left")
            cell.border = thin_border
            if rank % 2 == 1:
                cell.fill = PatternFill("solid", fgColor="F7F9FC")
            if rank <= 5 and c_idx == 10:
                cell.font = Font(bold=True, size=11)

    # Explainer block AFTER the data table (rows table_start + len(top_n) + 2 onwards)
    explainer_lines = [
        ("Cara membaca tabel:", ""),
        ("Peringkat 1-5", "Pelanggan paling berkontribusi terhadap eksposur kredit grup. Wajib review oleh komite."),
        (
            "Jumlah CV >= 3",
            "Pelanggan terdaftar di banyak CV. Kemungkinan duplikasi atau eksposur ganda. Lihat sheet 'Duplikat Antar-CV'.",
        ),
        (
            "% dari Total Usulan",
            "Porsi pelanggan dari total Usulan Limit grup. Di atas 5% = konsentrasi tinggi (perlu mitigasi).",
        ),
        (
            "Total AR Berjalan",
            "Sisa tagihan yang belum dibayar saat ini. Di atas Rp 1 Miliar di-flag pada kolom Catatan.",
        ),
        (
            "Rata-rata DSO",
            "Hari rata-rata pembayaran. Sampai 14 hari = Bagus (target untuk fresh produce). Di atas 30 hari = Buruk.",
        ),
        (
            "Catatan",
            "Auto-flag dari sistem: 'Eksposur di N CV' bila lebih dari 2 CV; 'AR > 1 Miliar' bila open AR signifikan.",
        ),
    ]
    explainer_start = table_start + len(top_n) + 2
    for i, (label, expl) in enumerate(explainer_lines):
        r = explainer_start + i
        ws_top.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
        c1 = ws_top.cell(row=r, column=1, value=label)
        c1.font = Font(bold=True, size=10) if i == 0 else Font(size=10)
        c1.alignment = Alignment(horizontal="left", indent=1, vertical="center")
        ws_top.merge_cells(start_row=r, start_column=3, end_row=r, end_column=14)
        c2 = ws_top.cell(row=r, column=3, value=expl)
        c2.font = Font(italic=True, size=10, color="555555")
        c2.alignment = Alignment(horizontal="left", wrap_text=True, vertical="center", indent=1)
        ws_top.row_dimensions[r].height = 20 if expl else 14

    # ── Sheet 3c: Aging AR (per customer × CV, days past due buckets) ──
    ws_aging = wb.create_sheet("Aging AR")
    aging_headers = [
        "CV",
        "Sub-group",
        "ID Pelanggan",
        "Nama Pelanggan",
        "NPWP",
        "Jumlah Faktur Terbuka",
        "Total AR (IDR)",
        "Belum Jatuh Tempo (IDR)",
        "1-7 hari (IDR)",
        "8-14 hari (IDR)",
        "15-30 hari (IDR)",
        "31-60 hari (IDR)",
        "61-120 hari (IDR)",
        ">120 hari KRITIS (IDR)",
        "Tunggakan Terlama (hari)",
    ]
    aging_widths = [28, 10, 12, 38, 22, 12, 18, 18, 14, 14, 14, 14, 14, 18, 14]
    _write_headers(ws_aging, aging_headers, aging_widths)

    def _aging_bucket(days_past_due: int) -> str:
        if days_past_due <= 0:
            return "current"
        if days_past_due <= 7:
            return "1-7"
        if days_past_due <= 14:
            return "8-14"
        if days_past_due <= 30:
            return "15-30"
        if days_past_due <= 60:
            return "31-60"
        if days_past_due <= 120:
            return "61-120"
        return ">120"

    aging_money_cols = {7, 8, 9, 10, 11, 12, 13, 14}
    aging_rows: list[dict[str, Any]] = []
    for tname, custs, by_c, _props in tenant_results:
        sg = _subgroup_of(tname)
        # Customer lookup for this tenant — id → customer dict
        cust_lookup: dict[int, dict[str, Any]] = {c["id"]: c for c in custs}
        for cust_id, invoices in by_c.items():
            buckets = {"current": 0.0, "1-7": 0.0, "8-14": 0.0, "15-30": 0.0, "31-60": 0.0, "61-120": 0.0, ">120": 0.0}
            n_open = 0
            total_ar = 0.0
            oldest = 0
            for inv in invoices:
                residual = float(inv.get("amount_residual") or 0.0)
                if residual <= 0:
                    continue
                n_open += 1
                total_ar += residual
                due = inv.get("invoice_date_due")
                if due:
                    days_late = (params.today - due).days
                    if days_late > oldest:
                        oldest = days_late
                else:
                    days_late = 0
                buckets[_aging_bucket(days_late)] += residual
            if n_open == 0:
                continue
            cust = cust_lookup.get(cust_id, {})
            aging_rows.append(
                {
                    "CV": tname,
                    "sg": sg,
                    "id": cust_id,
                    "name": cust.get("name", ""),
                    "npwp": cust.get("npwp", ""),
                    "n_open": n_open,
                    "total_ar": total_ar,
                    "buckets": buckets,
                    "oldest": oldest,
                }
            )
    # Sort: most-overdue first (>120 desc, then 61-120, etc.)
    aging_rows.sort(key=lambda r: (-r["buckets"][">120"], -r["buckets"]["61-120"], -r["total_ar"]))
    for i, r in enumerate(aging_rows, start=2):
        b = r["buckets"]
        values = [
            r["CV"],
            r["sg"],
            r["id"],
            r["name"],
            r["npwp"],
            r["n_open"],
            r["total_ar"],
            b["current"],
            b["1-7"],
            b["8-14"],
            b["15-30"],
            b["31-60"],
            b["61-120"],
            b[">120"],
            r["oldest"],
        ]
        for c_idx, val in enumerate(values, start=1):
            cell = ws_aging.cell(row=i, column=c_idx, value=val)
            if c_idx in aging_money_cols and isinstance(val, (int, float)):
                cell.number_format = _money_fmt()
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="right" if c_idx >= 6 else "left")
        # Highlight critical (>120d) rows in red
        if b[">120"] > 0:
            for c_idx in range(1, len(aging_headers) + 1):
                ws_aging.cell(row=i, column=c_idx).fill = PatternFill("solid", fgColor="FFC7CE")
        elif b["61-120"] > 0:
            for c_idx in range(1, len(aging_headers) + 1):
                ws_aging.cell(row=i, column=c_idx).fill = PatternFill("solid", fgColor="FFEB9C")
    # Aging grand total row
    if aging_rows:
        gtot_row = len(aging_rows) + 2
        bucket_totals = {
            k: sum(r["buckets"][k] for r in aging_rows)
            for k in ("current", "1-7", "8-14", "15-30", "31-60", "61-120", ">120")
        }
        total_total = sum(r["total_ar"] for r in aging_rows)
        total_open = sum(r["n_open"] for r in aging_rows)
        gt_cells = [
            "TOTAL",
            "",
            "",
            "",
            "",
            total_open,
            total_total,
            bucket_totals["current"],
            bucket_totals["1-7"],
            bucket_totals["8-14"],
            bucket_totals["15-30"],
            bucket_totals["31-60"],
            bucket_totals["61-120"],
            bucket_totals[">120"],
            "",
        ]
        for c_idx, val in enumerate(gt_cells, start=1):
            cell = ws_aging.cell(row=gtot_row, column=c_idx, value=val)
            cell.fill = grand_fill
            cell.font = Font(bold=True)
            cell.border = thin_border
            if c_idx in aging_money_cols and isinstance(val, (int, float)):
                cell.number_format = _money_fmt()
            cell.alignment = Alignment(horizontal="right" if c_idx >= 6 else "left")

    # ── Sheet 3d: Strategi & Rekomendasi ────────────────────────────
    ws_strat = wb.create_sheet("Strategi & Rekomendasi")
    # Title row
    ws_strat.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)
    cs = ws_strat.cell(row=1, column=1, value="STRATEGI MITIGASI RISIKO KREDIT LINTAS-CV")
    cs.font = title_font
    cs.alignment = Alignment(horizontal="center", vertical="center")
    ws_strat.row_dimensions[1].height = 32
    ws_strat.merge_cells(start_row=2, start_column=1, end_row=2, end_column=4)
    cs2 = ws_strat.cell(
        row=2,
        column=1,
        value="Cara mencegah eksposur ganda (double credit) terhadap pelanggan yang sama di banyak CV",
    )
    cs2.font = subtitle_font
    cs2.alignment = Alignment(horizontal="center")

    ws_strat.column_dimensions["A"].width = 6
    ws_strat.column_dimensions["B"].width = 38
    ws_strat.column_dimensions["C"].width = 22
    ws_strat.column_dimensions["D"].width = 70

    def _strat_section(row: int, text: str) -> int:
        ws_strat.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        cell = ws_strat.cell(row=row, column=1, value=text)
        cell.font = section_font
        cell.fill = section_fill
        cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws_strat.row_dimensions[row].height = 22
        return row + 1

    def _strat_table_header(row: int) -> int:
        for c_idx, h in enumerate(["No", "Tindakan", "Pemilik / PIC", "Detail"], start=1):
            cell = ws_strat.cell(row=row, column=c_idx, value=h)
            cell.font = hdr_font
            cell.fill = hdr_fill
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
            cell.border = thin_border
        return row + 1

    def _strat_row(row: int, no: int, action: str, owner: str, detail: str) -> int:
        cells = [
            (1, no, "center"),
            (2, action, "left"),
            (3, owner, "center"),
            (4, detail, "left"),
        ]
        for c_idx, val, align in cells:
            cell = ws_strat.cell(row=row, column=c_idx, value=val)
            cell.alignment = Alignment(horizontal=align, wrap_text=True, vertical="top")
            cell.border = thin_border
            if c_idx == 2:
                cell.font = label_font
        ws_strat.row_dimensions[row].height = 48
        return row + 1

    # ── STRATEGI SEMENTARA ──
    r = _strat_section(4, "STRATEGI SEMENTARA: Saat Ini (Masih di Accurate, sebelum migrasi Odoo)")
    r = _strat_table_header(r)
    r = _strat_row(
        r,
        1,
        "Laporan Konsolidasi Mingguan",
        "Finance",
        "Jalankan kctl-odoo accurate reports credit-limit --all-tenants setiap Senin pagi. "
        "Review sheet 'Duplikat Antar-CV' + 'Top 20 Pelanggan' untuk identifikasi eksposur ganda.",
    )
    r = _strat_row(
        r,
        2,
        "Master Customer Registry (Manual)",
        "Finance + IT",
        "Buat satu spreadsheet konsolidasi (atau dokumen Outline) berisi 1 baris per pelanggan riil. Kunci utama: "
        "NPWP > KTP > Nama+Kota. Kolom: NPWP, KTP, daftar CV tempat terdaftar, limit per CV, total eksposur grup. "
        "Owner: 1 orang di Finance.",
    )
    r = _strat_row(
        r,
        3,
        "Wajib Lengkapi NPWP / KTP saat Onboard",
        "Sales + Operations (per CV)",
        "Kebijakan baru: setiap pelanggan baru di Accurate WAJIB dilengkapi NPWP (untuk B2B) atau KTP (untuk perorangan). "
        "Alamat lengkap (jalan + kota + provinsi) wajib. Saat ini ~0% pelanggan TPK punya NPWP, sehingga risiko fraud detection sangat lemah.",
    )
    r = _strat_row(
        r,
        4,
        "Cross-CV Approval Gate",
        "CFO + Direktur",
        "Sebelum CV mana pun memberi limit > Rp 500 juta ke satu pelanggan: jalankan laporan, cek apakah pelanggan sudah ada di CV lain. "
        "Jika ya, wajib persetujuan komite (CFO + minimal 1 direktur). Dokumentasikan keputusan.",
    )
    r = _strat_row(
        r,
        5,
        "Group Exposure Ceiling",
        "CFO",
        "Tetapkan batas total eksposur per pelanggan di seluruh grup (mis. Rp 3 Miliar untuk perorangan, Rp 10 Miliar untuk PT). "
        "Saat laporan menunjukkan pelanggan melewati batas: bekukan penjualan baru sampai dilunasi.",
    )
    r = _strat_row(
        r,
        6,
        "Bulanan: Review Komite Kredit",
        "CFO + Komite",
        "Review Top 20 Pelanggan + Duplikat Antar-CV bersama. Tetapkan ulang limit pelanggan dengan eksposur > Rp 5 Miliar. "
        "Tanda-tangan pada laporan ini sebagai bukti review.",
    )

    # ── STRATEGI JANGKA PANJANG ──
    r += 1
    r = _strat_section(r, "STRATEGI JANGKA PANJANG: Setelah Migrasi ke Odoo")
    r = _strat_table_header(r)
    r = _strat_row(
        r,
        1,
        "Satu res.partner Kanonik per Pelanggan",
        "Odoo",
        "Multi-company Odoo: pelanggan yang sama di CV A, B, C = SATU record res.partner. "
        "Diberikan akses ke beberapa company via res.partner.company_ids (Many2many). Hilangkan duplikat secara struktural.",
    )
    r = _strat_row(
        r,
        2,
        "Limit Kredit per Company (per CV) di Odoo",
        "Odoo (built-in)",
        "Field res.partner.credit_limit di Odoo 18 sudah per-company. Satu pelanggan punya 12 limit (1 per CV). "
        "Total eksposur grup = SUM(credit_limit + AR) live tanpa konsolidasi manual.",
    )
    r = _strat_row(
        r,
        3,
        "Duplicate Prevention saat Pembuatan",
        "Odoo OCA partner_unique",
        "Modul OCA partner_unique memblokir pembuatan partner dengan NPWP/KTP yang sudah ada. "
        "partner_management (custom) tambahan: required-field rules per tipe pelanggan.",
    )
    r = _strat_row(
        r,
        4,
        "Intercompany Partner Mirroring",
        "intercompany_integration (sudah ada)",
        "Modul intercompany yang akan dibangun. Pelanggan yang dibuat di CV A otomatis dimirror ke CV B/C. "
        "Hilangkan duplikat-by-accident sepenuhnya.",
    )
    r = _strat_row(
        r,
        5,
        "Tier Validation pada Perubahan Limit",
        "OCA base_tier_validation",
        "Setiap perubahan credit_limit > 10% wajib approval Tier-1. Perubahan > Rp 1 Miliar wajib approval CFO. "
        "Audit log tersimpan otomatis.",
    )
    r = _strat_row(
        r,
        6,
        "Dashboard Komite Kredit (Custom)",
        "Custom (perlu dibangun)",
        "Per pelanggan: 12 eksposur CV, total AR, payment history, DSO, group limit utilization. "
        "Alert otomatis saat eksposur > threshold. Email digest bulanan ke CFO.",
    )
    r = _strat_row(
        r,
        7,
        "Onboarding Governance (Wajib NPWP/KTP)",
        "Custom partner_management",
        "Saat create partner di Odoo, NPWP/KTP wajib diisi. Validasi format. Soft-block dengan justifikasi tertulis untuk pengecualian.",
    )
    r = _strat_row(
        r,
        8,
        "Ganti Laporan Ini dengan Versi Odoo",
        "kctl-odoo partners credit-limit-proposal",
        "Phase 1-3 sudah ada di main (compute layer + readiness layer). Phase 4-6 (pull dari Odoo + workbook) "
        "tinggal dibangun. Setelah migrasi, laporan ini menggunakan data Odoo, bukan Accurate langsung.",
    )

    # ── ROADMAP MIGRASI ──
    r += 1
    r = _strat_section(r, "ROADMAP MIGRASI YANG DIREKOMENDASIKAN")
    r = _strat_row(
        r,
        1,
        "Minggu 1-2: Aktifkan Strategi Sementara",
        "Finance + IT",
        "Roll-out laporan mingguan. Setup Master Customer Registry. Sosialisasi kebijakan NPWP/KTP wajib.",
    )
    r = _strat_row(
        r,
        2,
        "Bulan 1: Pilot Migrasi 1 CV ke Odoo",
        "IT + Finance",
        "Pilih 1 CV trading kecil (mis. YSU). Jalankan migrasi penuh ke Odoo. Verify intercompany mirroring berjalan.",
    )
    r = _strat_row(
        r,
        3,
        "Bulan 2-3: Migrasi 11 CV Trading Sisanya",
        "IT + Finance",
        "Batch migrate sesuai sub-group (TPP1 → TPP2 → TPP3 → TPP4 → Tier 1). Validasi parity 100% per CV.",
    )
    r = _strat_row(
        r,
        4,
        "Bulan 4: Beralih ke Laporan Versi Odoo",
        "IT",
        "Build kctl-odoo partners credit-limit-proposal Phase 4-6. Bandingkan output dengan versi Accurate ini. "
        "Saat parity tercapai, decommission versi Accurate (atau pertahankan sebagai fallback).",
    )
    r = _strat_row(
        r,
        5,
        "Berkelanjutan: Review Komite Bulanan",
        "CFO",
        "Tetap jalankan review bulanan. Modul Odoo akan menyederhanakan (single source of truth, real-time).",
    )

    ws_strat.freeze_panes = "A4"

    # ── Sheet 4: Master Pelanggan (combined) ────────────────────────
    ws_cust = wb.create_sheet("Master Pelanggan")
    cust_headers = [
        "CV",
        "ID",
        "No Pelanggan",
        "Nama",
        "NPWP",
        "No KTP",
        "Nama Wajib Pajak",
        "Email",
        "No HP",
        "Alamat Tagihan",
        "Kota Tagihan",
        "Provinsi Tagihan",
        "Kode Pos",
        "Negara",
        "Alamat Kirim",
        "Kota Kirim",
        "Provinsi Kirim",
        "ToP (Term Of Payment)",
        "ToP Hari (netDays)",
        "Kunci Identitas",
        "Limit Kredit Saat Ini (IDR)",
        "Suspended",
    ]
    cust_widths = [28, 10, 14, 38, 22, 18, 28, 28, 18, 32, 18, 18, 12, 14, 32, 18, 18, 18, 10, 38, 22, 10]
    _write_headers(ws_cust, cust_headers, cust_widths)
    cust_money_cols = {21}
    r_offset = 2
    for tname, custs, _by_c, _props in tenant_results:
        for c in custs:
            ikey = c.get("identity_key") or _identity_key(c)
            values = [
                tname,
                c.get("id"),
                c.get("customer_no", ""),
                c.get("name", ""),
                c.get("npwp", ""),
                c.get("id_card", ""),
                c.get("wp_name", ""),
                c.get("email", ""),
                c.get("mobile_phone", ""),
                c.get("billing_street", ""),
                c.get("billing_city", ""),
                c.get("billing_province", ""),
                c.get("billing_zip", ""),
                c.get("billing_country", ""),
                c.get("shipping_street", ""),
                c.get("shipping_city", ""),
                c.get("shipping_province", ""),
                c.get("term_name", "") or "",
                c.get("term_net_days", 0),
                ikey,
                c.get("current_credit_limit", 0.0) or 0.0,
                "Y" if c.get("suspended") else "",
            ]
            for c_idx, val in enumerate(values, start=1):
                cell = ws_cust.cell(row=r_offset, column=c_idx, value=val)
                if c_idx in cust_money_cols and isinstance(val, (int, float)):
                    cell.number_format = _money_fmt()
            # Yellow highlight if no current credit limit set in Accurate (col 21)
            curr_limit = c.get("current_credit_limit", 0.0) or 0.0
            if curr_limit <= 0:
                ws_cust.cell(row=r_offset, column=21).fill = PatternFill("solid", fgColor="FFEB9C")
            # Identity key column shifted to col 20 (was 18 before ToP cols added)
            if (tname, c.get("id")) in flagged_cust_ids:
                ws_cust.cell(row=r_offset, column=20).fill = duplicate_fill
            r_offset += 1

    # ── Sheet 5: Pelanggan - Atribut Lengkap (raw Accurate fields) ──
    # Every Accurate field for every customer (with detail fetched). This
    # is the authoritative dump for shareholder review — no curation.
    ws_raw = wb.create_sheet("Pelanggan - Atribut Lengkap")

    # Build the union of all Accurate field names across all customers.
    # Skip nested objects/lists (we render only scalar fields here; keep
    # complex fields as JSON-serialised strings so nothing is lost).
    all_keys: set[str] = set()
    for tname, custs, _by_c, _props in tenant_results:
        for c in custs:
            raw = c.get("_raw_accurate") or {}
            for k in raw.keys():
                all_keys.add(k)
    sorted_keys = sorted(all_keys)
    raw_headers = ["CV", "ID Pelanggan (Odoo-side)"] + sorted_keys
    raw_widths = [28, 18] + [22] * len(sorted_keys)
    _write_headers(ws_raw, raw_headers, raw_widths)

    import json as _json

    def _scalarize(v: Any) -> Any:
        if v is None or isinstance(v, (str, int, float, bool)):
            return v
        try:
            return _json.dumps(v, ensure_ascii=False, default=str)
        except Exception:  # noqa: BLE001
            return str(v)

    r_offset = 2
    for tname, custs, _by_c, _props in tenant_results:
        for c in custs:
            raw = c.get("_raw_accurate") or {}
            row_values: list[Any] = [tname, c.get("id")]
            for k in sorted_keys:
                row_values.append(_scalarize(raw.get(k)))
            for c_idx, val in enumerate(row_values, start=1):
                ws_raw.cell(row=r_offset, column=c_idx, value=val)
            r_offset += 1

    # ── Sheet 6: Faktur (combined) ──────────────────────────────────
    ws_inv = wb.create_sheet("Faktur")
    inv_headers = [
        "CV",
        "ID Faktur",
        "No Faktur",
        "ID Pelanggan",
        "No Pelanggan",
        "Nama Pelanggan",
        "Tanggal Faktur",
        "Jatuh Tempo",
        "Total (IDR)",
        "Sisa (IDR)",
        "Status Bayar",
        "Status Accurate",
        "Tanggal Bayar Terakhir",
        "Hari ke Bayar",
    ]
    inv_widths = [28, 12, 22, 12, 14, 38, 14, 14, 16, 16, 14, 16, 18, 14]
    _write_headers(ws_inv, inv_headers, inv_widths)
    inv_money_cols = {9, 10}
    r_offset = 2
    for tname, _custs, by_c, _props in tenant_results:
        for cust_id in sorted(by_c.keys()):
            for inv in by_c[cust_id]:
                inv_date = inv.get("invoice_date")
                last_pay = inv.get("last_payment_date")
                days_to_pay: Any = ""
                if inv_date and last_pay:
                    days_to_pay = (last_pay - inv_date).days
                values = [
                    tname,
                    inv.get("id"),
                    inv.get("doc_number", ""),
                    inv.get("customer_id"),
                    inv.get("customer_no", ""),
                    inv.get("customer_name", ""),
                    inv_date.isoformat() if inv_date else "",
                    inv["invoice_date_due"].isoformat() if inv.get("invoice_date_due") else "",
                    inv.get("amount_total", 0.0),
                    inv.get("amount_residual", 0.0),
                    inv.get("payment_state", ""),
                    inv.get("status", ""),
                    last_pay.isoformat() if last_pay else "",
                    days_to_pay,
                ]
                for c_idx, val in enumerate(values, start=1):
                    cell = ws_inv.cell(row=r_offset, column=c_idx, value=val)
                    if c_idx in inv_money_cols and isinstance(val, (int, float)):
                        cell.number_format = _money_fmt()
                r_offset += 1

    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)


# ---------------------------------------------------------------------------
# Typer command
# ---------------------------------------------------------------------------


def credit_limit_report(
    ctx: typer.Context,
    tenant_name: Annotated[
        str | None,
        typer.Option("--tenant-name", help="(single-tenant mode) Display name on the cover sheet"),
    ] = None,
    env_key: Annotated[
        str | None,
        typer.Option("--env-key", help="(single-tenant mode) Token variable name in --env-file"),
    ] = None,
    all_tenants: Annotated[
        bool,
        typer.Option("--all-tenants/--no-all-tenants", help="Process every ACCURATE_TOKEN_* in the env file"),
    ] = False,
    trading_only: Annotated[
        bool,
        typer.Option(
            "--trading-only/--no-trading-only",
            help=(
                "When --all-tenants, process ONLY the 12 trading-type CVs "
                "(default ON; matches Odoo res.company 'Trading' tag). Use "
                "--no-trading-only to include import + holding entities too."
            ),
        ),
    ] = True,
    env_file: Annotated[
        Path,
        typer.Option("--env-file", help="Path to .env containing ACCURATE_TOKEN_* vars"),
    ] = DEFAULT_ENV_FILE,
    output: Annotated[
        Path | None,
        typer.Option("--output", "-o", help="Output xlsx path"),
    ] = None,
    coverage: Annotated[
        float,
        typer.Option(
            "--coverage",
            help=(
                "Multiplier on avg monthly sales. Default 1.25 (= 1 month of sales + "
                "25% buffer, rounded UP to nearest Rp 1M)."
            ),
        ),
    ] = 1.25,
    limit_customers: Annotated[
        int | None,
        typer.Option("--limit-customers", help="Pilot debug: only process the first N customers per tenant"),
    ] = None,
    insufficient_min_invoices: Annotated[
        int,
        typer.Option("--insufficient-min-invoices", help="Minimum total invoices required to compute a proposal"),
    ] = 3,
    overdue_review_days: Annotated[
        int,
        typer.Option("--overdue-review-days", help="Days-past-due threshold to flag PERLU_REVIEW"),
    ] = 120,
) -> None:
    """Pull customer + invoice data from Accurate Online and propose credit limits.

    Uses the Accurate API directly (no Odoo migration required) for ONE
    or MORE tenants. With --all-tenants, processes every
    ``ACCURATE_TOKEN_*`` in the env file and produces ONE combined Excel
    workbook with cross-CV duplicate detection.

    For 100% data accuracy (shareholder review), invoice detail is fetched
    per record — no estimation for partial-payment residuals.
    """
    actx: AppContext = ctx.obj
    out = actx.output
    secret = _resolve_signature_secret()
    today = date.today()

    # Resolve tenants
    tenants: list[tuple[str, str]] = []
    if all_tenants:
        tenants = _discover_tenants_from_env(env_file)
        if trading_only:
            before = len(tenants)
            tenants = [(name, key) for name, key in tenants if key in TRADING_TENANT_KEYS]
            out.info(f"Filter --trading-only: {before} tenant ditemukan → {len(tenants)} CV trading")
        if not tenants:
            out.error(f"Tidak ditemukan tenant cocok di {env_file}")
            raise typer.Exit(code=1)
    else:
        if not tenant_name or not env_key:
            out.error("Berikan --tenant-name + --env-key, atau gunakan --all-tenants")
            raise typer.Exit(code=1)
        tenants = [(tenant_name, env_key)]

    if output is None:
        if all_tenants or len(tenants) > 1:
            output = Path(
                f"./laporan-analisa-kredit-trading-all-tpp.xlsx"
                if trading_only
                else f"./laporan-analisa-kredit-{today.strftime('%Y%m%d')}.xlsx"
            )
        else:
            slug = _slugify(tenants[0][0])
            output = Path(f"./laporan-analisa-kredit-{slug}-{today.strftime('%Y%m%d')}.xlsx")

    params = Params(
        coverage=coverage,
        insufficient_min_invoices=insufficient_min_invoices,
        overdue_review_days=overdue_review_days,
        today=today,
    )

    out.info(f"Memproses {len(tenants)} CV / tenant...")
    tenant_results: list[
        tuple[str, list[dict[str, Any]], dict[int, list[dict[str, Any]]], list[tuple[dict[str, Any], ProposalRow]]]
    ] = []
    skipped_tenants: list[tuple[str, str]] = []
    for tname, ekey in tenants:
        token = _resolve_token(ekey, env_file)
        try:
            customers, by_customer, proposals = _pull_one_tenant(
                tenant_name=tname,
                token=token,
                secret=secret,
                today=today,
                params=params,
                out=out,
                limit_customers=limit_customers,
            )
            tenant_results.append((tname, customers, by_customer, proposals))
        except Exception as exc:  # noqa: BLE001
            # Skip and continue — token expired / Accurate API hiccup / 500 etc.
            # Shareholder review needs partial result with clear "skipped" list,
            # not an aborted run.
            short = str(exc).splitlines()[0] if str(exc) else type(exc).__name__
            out.error(f"  ✗ {tname}: SKIP (alasan: {short})")
            skipped_tenants.append((tname, short))
            continue

    if not tenant_results:
        out.error("Semua CV gagal diproses; tidak ada data untuk ditulis.")
        for tn, why in skipped_tenants:
            out.error(f"  ✗ {tn}: {why}")
        raise typer.Exit(code=2)

    out.info(f"Menulis laporan ke {output}...")
    _write_combined_workbook(
        output,
        tenant_results=tenant_results,
        params=params,
        coverage=coverage,
    )
    out.success(f"Selesai: {output}")

    # Per-sub-group split files (only when --all-tenants + --trading-only)
    if all_tenants and trading_only and len(tenant_results) > 1:
        out_dir = output.parent
        for sg in ["TPP1", "TPP2", "TPP3", "TPP4", "Tier 1"]:
            subset = [t for t in tenant_results if _subgroup_of(t[0]) == sg]
            if not subset:
                continue
            sg_slug = _slugify(sg)
            sg_path = out_dir / f"laporan-analisa-kredit-{sg_slug}.xlsx"
            out.info(f"Menulis split per sub-group: {sg} ({len(subset)} CV) -> {sg_path}")
            try:
                _write_combined_workbook(sg_path, tenant_results=subset, params=params, coverage=coverage)
            except Exception as exc:  # noqa: BLE001
                out.error(f"  gagal menulis {sg_path}: {exc}")
        out.info("Split per sub-group selesai.")

    if skipped_tenants:
        out.warn(f"{len(skipped_tenants)} CV dilewati (token error / API issue):")
        for tn, why in skipped_tenants:
            out.warn(f"  ✗ {tn}: {why}")
