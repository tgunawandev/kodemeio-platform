"""Access control report builder.

Generates a flat user x application access matrix by cross-referencing
Authentik policy bindings with user group memberships.
"""

from __future__ import annotations

import json
import sys
from datetime import UTC, datetime
from pathlib import Path
from typing import Any

from pydantic import BaseModel

_SERVICE_ACCOUNT_TYPES = frozenset({"internal_service_account", "service_account"})

_COLUMNS = [
    "username",
    "email",
    "name",
    "user_type",
    "is_active",
    "application",
    "app_slug",
    "has_access",
    "access_via",
    "binding_negate",
]


class ReportFilters(BaseModel):
    active_only: bool = False
    include_service_accounts: bool = False
    include_deactivated: bool = False
    app_slug: str | None = None


class AccessRow(BaseModel):
    username: str
    email: str
    name: str
    user_type: str
    is_active: str
    application: str
    app_slug: str
    has_access: str
    access_via: str
    binding_negate: str


class ReportMeta(BaseModel):
    profile: str
    generated_at: str
    user_count: int
    app_count: int
    row_count: int


class ReportData(BaseModel):
    users: list[dict[str, Any]]
    apps: list[dict[str, Any]]
    bindings: list[dict[str, Any]]


def _classify_user_type(user: dict[str, Any]) -> str:
    return "service" if user.get("type", "internal") in _SERVICE_ACCOUNT_TYPES else "human"


def _user_group_pks(user: dict[str, Any]) -> set[str]:
    return {str(g.get("pk", "")) for g in user.get("groups_obj", [])}


def _build_access_map(
    apps: list[dict[str, Any]],
    bindings: list[dict[str, Any]],
) -> dict[str, list[dict[str, Any]]]:
    by_app: dict[str, list[dict[str, Any]]] = {str(a["pk"]): [] for a in apps}
    for b in bindings:
        if not b.get("enabled", True):
            continue
        target = str(b.get("target", ""))
        if target in by_app:
            by_app[target].append(b)
    return by_app


def _resolve_access(
    user: dict[str, Any],
    app_bindings: list[dict[str, Any]],
) -> tuple[str, str, str]:
    user_pk = user.get("pk")
    user_groups = _user_group_pks(user)

    for b in app_bindings:
        group_obj = b.get("group_obj") or {}
        user_obj = b.get("user_obj") or {}
        negate = b.get("negate", False)

        if user_obj and user_obj.get("pk") == user_pk:
            access = "no" if negate else "yes"
            return access, "user:direct", "yes" if negate else "no"

        if group_obj and str(group_obj.get("pk", "")) in user_groups:
            group_name = group_obj.get("name", "unknown")
            access = "no" if negate else "yes"
            return access, f"group:{group_name}", "yes" if negate else "no"

    return "no", "none", "no"


def build_rows(data: ReportData, filters: ReportFilters) -> list[AccessRow]:
    users = data.users
    apps = data.apps

    if filters.app_slug:
        apps = [a for a in apps if a.get("slug") == filters.app_slug]

    filtered_users: list[dict[str, Any]] = []
    for u in users:
        is_service = u.get("type", "internal") in _SERVICE_ACCOUNT_TYPES
        is_active = u.get("is_active", False)

        if is_service and not filters.include_service_accounts:
            continue
        if not is_active and not filters.include_deactivated:
            continue
        if filters.active_only and not is_active:
            continue

        filtered_users.append(u)

    access_map = _build_access_map(data.apps, data.bindings)

    rows: list[AccessRow] = []
    for u in filtered_users:
        for a in apps:
            app_pk = str(a["pk"])
            app_bindings = access_map.get(app_pk, [])
            has_access, access_via, binding_negate = _resolve_access(u, app_bindings)

            rows.append(
                AccessRow(
                    username=u.get("username", ""),
                    email=u.get("email", ""),
                    name=u.get("name", ""),
                    user_type=_classify_user_type(u),
                    is_active="yes" if u.get("is_active") else "no",
                    application=a.get("name", ""),
                    app_slug=a.get("slug", ""),
                    has_access=has_access,
                    access_via=access_via,
                    binding_negate=binding_negate,
                )
            )

    rows.sort(key=lambda r: (r.username, r.app_slug))
    return rows


def write_markdown(rows: list[AccessRow], meta: ReportMeta, dest: Any) -> None:
    """Write Markdown pipe-table report to a file-like object."""
    dest.write("# Access Control Report\n\n")
    dest.write(f"- **Profile:** {meta.profile}\n")
    dest.write(f"- **Generated:** {meta.generated_at}\n")
    dest.write(f"- **Users:** {meta.user_count} | **Applications:** {meta.app_count} | **Rows:** {meta.row_count}\n\n")

    headers = ["Username", "Email", "Name", "Type", "Active", "Application", "Slug", "Access", "Via", "Negate"]
    dest.write("| " + " | ".join(headers) + " |\n")
    dest.write("| " + " | ".join("---" for _ in headers) + " |\n")

    for r in rows:
        vals = [
            r.username,
            r.email,
            r.name,
            r.user_type,
            r.is_active,
            r.application,
            r.app_slug,
            r.has_access,
            r.access_via,
            r.binding_negate,
        ]
        dest.write("| " + " | ".join(vals) + " |\n")


def write_xlsx(rows: list[AccessRow], meta: ReportMeta, dest: Path) -> None:
    """Write Excel report with Access Control data sheet and Summary sheet."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise ImportError(
            "openpyxl is required for Excel output. Install it with: uv pip install kctl-ak[reports]"
        ) from None

    wb = Workbook()

    ws = wb.active
    ws.title = "Access Control"

    headers = ["Username", "Email", "Name", "Type", "Active", "Application", "Slug", "Access", "Via", "Negate"]
    ws.append(headers)

    bold = Font(bold=True)
    for col_idx in range(1, len(headers) + 1):
        ws.cell(1, col_idx).font = bold

    for r in rows:
        ws.append(
            [
                r.username,
                r.email,
                r.name,
                r.user_type,
                r.is_active,
                r.application,
                r.app_slug,
                r.has_access,
                r.access_via,
                r.binding_negate,
            ]
        )

    last_col = get_column_letter(len(headers))
    ws.auto_filter.ref = f"A1:{last_col}{len(rows) + 1}"

    ws.freeze_panes = "A2"

    for col_idx, header in enumerate(headers, 1):
        max_len = len(header)
        for row_idx in range(2, len(rows) + 2):
            val = str(ws.cell(row_idx, col_idx).value or "")
            if len(val) > max_len:
                max_len = len(val)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 50)

    summary = wb.create_sheet("Summary")
    summary_data = [
        ("Profile", meta.profile),
        ("Generated", meta.generated_at),
        ("Users", meta.user_count),
        ("Applications", meta.app_count),
        ("Rows", meta.row_count),
    ]
    for label, value in summary_data:
        summary.append([label, value])
    summary.cell(1, 1).font = bold

    wb.save(dest)


def to_json(rows: list[AccessRow], meta: ReportMeta) -> dict[str, Any]:
    """Build JSON-serializable dict with meta envelope."""
    return {
        "meta": meta.model_dump(),
        "rows": [r.model_dump() for r in rows],
    }


def fetch_report_data(client: Any, filters: ReportFilters) -> ReportData:
    """Fetch users, apps, and policy bindings from Authentik API."""
    users = client.get_all("core/users/")
    params: dict[str, Any] = {"superuser_full_list": "true"}
    if filters.app_slug:
        params["slug"] = filters.app_slug
    apps = client.get_all("core/applications/", params=params)
    bindings = client.get_all("policies/bindings/")
    return ReportData(users=users, apps=apps, bindings=bindings)
