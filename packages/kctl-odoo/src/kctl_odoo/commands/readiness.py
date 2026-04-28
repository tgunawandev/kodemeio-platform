"""Production readiness audit command for Odoo 18.

Compares a target Odoo instance (typically a staging/local profile) against a
reference instance (typically MAC production) and emits a comprehensive
multi-sheet Excel workbook (18 sheets) showing what exists, what is missing
and what differs at the company-configuration level.

Without ``--reference``, falls back to existence-only checks (legacy behaviour).
"""

from __future__ import annotations

from datetime import date
from pathlib import Path
from typing import Annotated, Any

import typer
import yaml

from kctl_odoo.core.callbacks import AppContext


app = typer.Typer(help="Production readiness audit.")


# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

ACCOUNT_TYPES = [
    "asset_cash",
    "asset_receivable",
    "asset_current",
    "asset_non_current",
    "asset_prepayments",
    "asset_fixed",
    "liability_payable",
    "liability_credit_card",
    "liability_current",
    "liability_non_current",
    "equity",
    "equity_unaffected",
    "income",
    "income_other",
    "expense",
    "expense_direct_cost",
    "expense_depreciation",
    "off_balance",
]

CRITICAL_ACCOUNT_TYPES = {
    "asset_cash",
    "asset_receivable",
    "liability_payable",
    "income",
    "expense_direct_cost",
    "equity",
}

REQUIRED_MODULES = ["account", "stock", "mail"]
OPTIONAL_MODULES = [
    "l10n_id",
    "accurate_integration",
    "purchase",
    "sale",
    "point_of_sale",
    "hr",
    "project",
    "report_layout",
    "report_management",
    "report_management_financial",
]
TRACKED_MODULES = REQUIRED_MODULES + OPTIONAL_MODULES

CURRENCIES = ["IDR", "USD", "CNY", "THB"]

CRITICAL_PARAMS = {
    "web.base.url",
    "mail.catchall.domain",
    "mail.bounce.alias",
    "database.uuid",
    "report.url",
    "account.use_anglo_saxon",
    "accurate.signature_secret",
}

COMPANY_FIELDS = [
    "id",
    "name",
    "currency_id",
    "account_fiscal_country_id",
    "account_sale_tax_id",
    "account_purchase_tax_id",
    "income_currency_exchange_account_id",
    "expense_currency_exchange_account_id",
    "transfer_account_id",
    "anglo_saxon_accounting",
    "fiscalyear_lock_date",
    "tax_lock_date",
    "account_journal_suspense_account_id",
    "account_journal_payment_debit_account_id",
    "account_journal_payment_credit_account_id",
    "default_cash_difference_income_account_id",
    "default_cash_difference_expense_account_id",
    "po_lock",
    "po_double_validation",
    "po_double_validation_amount",
    "po_lead",
    "quotation_validity_days",
    "portal_confirmation_sign",
    "portal_confirmation_pay",
]


# ---------------------------------------------------------------------------
# Profile + RPC helpers
# ---------------------------------------------------------------------------


def _load_profile(profile_name: str) -> dict:
    cfg_path = Path.home() / ".config" / "kodemeio" / "config.yaml"
    if not cfg_path.exists():
        raise FileNotFoundError(f"Config file not found: {cfg_path}")
    cfg = yaml.safe_load(cfg_path.read_text()) or {}
    profiles = cfg.get("profiles", {})
    p = (profiles.get(profile_name) or {}).get("odoo") or {}
    if not p:
        raise ValueError(f"Profile '{profile_name}' not found or has no 'odoo' section in {cfg_path}")
    missing = [k for k in ("url", "database", "username", "api_key") if not p.get(k)]
    if missing:
        raise ValueError(f"Profile '{profile_name}' is missing required fields: {', '.join(missing)}")
    return {
        "url": p["url"],
        "database": p["database"],
        "username": p["username"],
        "api_key": p["api_key"],
    }


class XmlRpcClient:
    """Thin xmlrpc wrapper exposing search_read/search_count/read."""

    def __init__(self, profile: dict, *, timeout: float = 600.0):
        import xmlrpc.client

        url = profile["url"].rstrip("/")
        self.url = url
        self.database = profile["database"]
        self._pw = profile["api_key"]
        # xmlrpc transport timeout is socket-level; expose for use_builtin_types
        common = xmlrpc.client.ServerProxy(f"{url}/xmlrpc/2/common", allow_none=True, use_builtin_types=True)
        uid = common.authenticate(self.database, profile["username"], self._pw, {})
        if not uid:
            raise RuntimeError(f"Authentication failed for db={self.database} user={profile['username']}")
        self.uid = uid
        self._models = xmlrpc.client.ServerProxy(f"{url}/xmlrpc/2/object", allow_none=True, use_builtin_types=True)

    def execute(self, model: str, method: str, *args, **kwargs):
        return self._models.execute_kw(self.database, self.uid, self._pw, model, method, list(args), kwargs)

    def search_read(
        self,
        model: str,
        domain: list | None = None,
        fields: list | None = None,
        limit: int = 0,
        order: str | None = None,
    ) -> list[dict]:
        kw: dict[str, Any] = {}
        if fields:
            kw["fields"] = fields
        if limit:
            kw["limit"] = limit
        if order:
            kw["order"] = order
        return self.execute(model, "search_read", domain or [], **kw)

    def search_count(self, model: str, domain: list | None = None) -> int:
        return self.execute(model, "search_count", domain or [])

    def read(self, model: str, ids: list[int], fields: list | None = None) -> list[dict]:
        kw: dict[str, Any] = {}
        if fields:
            kw["fields"] = fields
        return self.execute(model, "read", ids, **kw)


def _bump_timeout(actx: AppContext, seconds: float = 600.0) -> None:
    try:
        import httpx

        actx.client._client.timeout = httpx.Timeout(seconds)
    except Exception:
        pass


def _safe(value: Any, default: str = "") -> str:
    if value is False or value is None:
        return default
    if isinstance(value, list) and len(value) == 2:
        return str(value[1])
    return str(value)


def _safe_id(value: Any) -> int | None:
    if isinstance(value, list) and len(value) >= 1:
        return value[0]
    if isinstance(value, int):
        return value
    return None


def _mask(value: str) -> str:
    if not value:
        return ""
    s = str(value)
    if len(s) <= 8:
        return "****"
    return f"{s[:4]}****{s[-4:]}"


def _truncate(value: Any, length: int = 100) -> str:
    if value in (False, None):
        return ""
    s = str(value)
    return s if len(s) <= length else s[: length - 3] + "..."


# ---------------------------------------------------------------------------
# Data pulls
# ---------------------------------------------------------------------------


class _ClientShim:
    """Adapter exposing the same minimal API on AppContext.client and XmlRpcClient.

    The platform's AppContext.client implements search_read(model, domain, fields=...)
    so we just pass through. XmlRpcClient also matches.
    """

    def __init__(self, inner: Any):
        self._inner = inner

    @property
    def database(self) -> str:
        if hasattr(self._inner, "database"):
            db = self._inner.database
            return db() if callable(db) else db
        return ""

    def search_read(self, model, domain=None, fields=None, limit=0, order=None):
        kwargs: dict[str, Any] = {}
        if fields is not None:
            kwargs["fields"] = fields
        if limit:
            kwargs["limit"] = limit
        if order:
            kwargs["order"] = order
        return self._inner.search_read(model, domain or [], **kwargs)

    def search_count(self, model, domain=None):
        return self._inner.search_count(model, domain or [])


def _try(fn, default=None):
    try:
        return fn()
    except Exception:
        return default


def _pull_company_data(client: _ClientShim, company_id: int, company_name: str) -> dict:
    """Pull every piece of configuration we audit, scoped to one company.

    Returns a dict that both target and reference share the shape of, so the
    comparison logic can be symmetric.
    """
    data: dict[str, Any] = {
        "company_id": company_id,
        "company_name": company_name,
    }

    # Company settings
    co = _try(
        lambda: client.search_read("res.company", [("id", "=", company_id)], fields=COMPANY_FIELDS + ["logo"]),
        default=[],
    )
    data["company"] = co[0] if co else {}

    # Accounts: full listing for this company (Odoo 18 = company_ids many2many)
    accounts = _try(
        lambda: client.search_read(
            "account.account",
            [("company_ids", "in", company_id)],
            fields=["id", "code", "name", "account_type", "deprecated"],
            order="code",
        ),
        default=[],
    )
    data["accounts"] = accounts

    # Account ext-ids (ir.model.data) for accounts in this company
    ext_ids: dict[int, str] = {}
    if accounts:
        acc_ids = [a["id"] for a in accounts]
        # batch in chunks of 500 to avoid huge domains
        for i in range(0, len(acc_ids), 500):
            chunk = acc_ids[i : i + 500]
            rows = (
                _try(
                    lambda c=chunk: client.search_read(
                        "ir.model.data",
                        [("model", "=", "account.account"), ("res_id", "in", c)],
                        fields=["res_id", "module", "name"],
                    ),
                    default=[],
                )
                or []
            )
            for r in rows:
                ext_ids[r["res_id"]] = f"{r['module']}.{r['name']}"
    data["account_ext_ids"] = ext_ids

    # Journals (full)
    data["journals"] = _try(
        lambda: client.search_read(
            "account.journal",
            [("company_id", "=", company_id)],
            fields=[
                "id",
                "name",
                "code",
                "type",
                "default_account_id",
                "suspense_account_id",
                "currency_id",
                "active",
            ],
            order="type, code",
        ),
        default=[],
    )

    # Taxes (full)
    taxes = _try(
        lambda: client.search_read(
            "account.tax",
            [("company_id", "=", company_id)],
            fields=[
                "id",
                "name",
                "type_tax_use",
                "amount",
                "amount_type",
                "tax_group_id",
                "active",
            ],
            order="type_tax_use, name",
        ),
        default=[],
    )
    data["taxes"] = taxes

    # Tax repartition lines (so we can map tax -> GL accounts)
    tax_accounts: dict[int, list[str]] = {}
    if taxes:
        rep = (
            _try(
                lambda: client.search_read(
                    "account.tax.repartition.line",
                    [("company_id", "=", company_id), ("tax_id", "in", [t["id"] for t in taxes])],
                    fields=["tax_id", "account_id", "repartition_type", "factor_percent"],
                ),
                default=[],
            )
            or []
        )
        for r in rep:
            tid = _safe_id(r.get("tax_id"))
            acc = _safe(r.get("account_id"))
            if tid and acc:
                tax_accounts.setdefault(tid, []).append(acc)
    data["tax_accounts"] = tax_accounts

    # Tax groups
    data["tax_groups"] = _try(
        lambda: client.search_read(
            "account.tax.group",
            [],
            fields=[
                "id",
                "name",
                "tax_payable_account_id",
                "tax_receivable_account_id",
                "advance_tax_payment_account_id",
            ],
            order="name",
        ),
        default=[],
    )

    # Fiscal positions
    data["fiscal_positions"] = _try(
        lambda: client.search_read(
            "account.fiscal.position",
            [("company_id", "=", company_id)],
            fields=[
                "id",
                "name",
                "country_id",
                "auto_apply",
                "account_ids",
                "tax_ids",
            ],
            order="name",
        ),
        default=[],
    )

    # Product categories (NOT company-scoped — shared)
    data["categories"] = _try(
        lambda: client.search_read(
            "product.category",
            [],
            fields=[
                "id",
                "name",
                "property_cost_method",
                "property_valuation",
                "property_account_income_categ_id",
                "property_account_expense_categ_id",
                "property_stock_account_input_categ_id",
                "property_stock_account_output_categ_id",
                "property_stock_valuation_account_id",
                "property_stock_journal",
            ],
            order="name",
        ),
        default=[],
    )

    # Warehouses
    data["warehouses"] = _try(
        lambda: client.search_read(
            "stock.warehouse",
            [("company_id", "=", company_id)],
            fields=[
                "id",
                "name",
                "code",
                "lot_stock_id",
                "reception_steps",
                "delivery_steps",
                "active",
            ],
            order="code",
        ),
        default=[],
    )

    # Operation types
    data["picking_types"] = _try(
        lambda: client.search_read(
            "stock.picking.type",
            [("company_id", "=", company_id)],
            fields=[
                "id",
                "name",
                "sequence_code",
                "code",
                "warehouse_id",
                "default_location_src_id",
                "default_location_dest_id",
            ],
            order="warehouse_id, sequence_code",
        ),
        default=[],
    )

    # Stock locations
    data["locations"] = _try(
        lambda: client.search_read(
            "stock.location",
            [("company_id", "=", company_id)],
            fields=["id", "name", "complete_name", "usage", "active"],
            order="usage, complete_name",
        ),
        default=[],
    )

    # Stock summary
    data["stock_summary"] = {
        "quants": _try(lambda: client.search_count("stock.quant", [("company_id", "=", company_id)]), default=0),
        "products_with_stock": _try(
            lambda: len(
                {
                    _safe_id(q["product_id"])
                    for q in client.search_read(
                        "stock.quant",
                        [("company_id", "=", company_id), ("quantity", ">", 0)],
                        fields=["product_id"],
                        limit=20000,
                    )
                    if _safe_id(q.get("product_id"))
                }
            ),
            default=0,
        ),
        "svl_count": _try(
            lambda: client.search_count("stock.valuation.layer", [("company_id", "=", company_id)]), default=0
        ),
    }

    # Operating units
    ou_data: list = _try(
        lambda: client.search_read(
            "operating.unit",
            [("company_id", "=", company_id)],
            fields=["id", "code", "name", "partner_id", "active"],
            order="code",
        ),
        default=None,
    )
    data["operating_units"] = ou_data if ou_data is not None else "MODULE_NOT_INSTALLED"

    # User counts per OU
    ou_user_counts: dict[int, int] = {}
    if isinstance(ou_data, list) and ou_data:
        for ou in ou_data:
            cnt = _try(
                lambda oid=ou["id"]: client.search_count(
                    "res.users", [("default_operating_unit_id", "=", oid), ("active", "=", True)]
                ),
                default=0,
            )
            ou_user_counts[ou["id"]] = cnt
    data["ou_user_counts"] = ou_user_counts

    # Report formats (report_layout module)
    rfmt = _try(
        lambda: client.search_read(
            "report.format",
            [("company_id", "=", company_id)],
            fields=[
                "id",
                "report_type",
                "name",
                "is_default",
                "active",
                "logo",
                "band_ids",
                "signature_line_ids",
            ],
            order="report_type, name",
        ),
        default=None,
    )
    data["report_formats"] = rfmt if rfmt is not None else "MODULE_NOT_INSTALLED"

    # Report management instances (report_management_financial)
    rinst = _try(
        lambda: client.search_read(
            "report.instance",
            [("company_id", "=", company_id)],
            fields=["id", "name", "template_id", "report_type"],
            order="name",
        ),
        default=None,
    )
    data["report_instances"] = rinst if rinst is not None else "MODULE_NOT_INSTALLED"

    # Report types registry (global)
    rtypes = _try(
        lambda: client.search_read(
            "report.type",
            [],
            fields=["id", "name", "code", "category", "handler"],
            order="category, name",
        ),
        default=None,
    )
    data["report_types"] = rtypes if rtypes is not None else "MODULE_NOT_INSTALLED"

    # Sequences (per-company)
    data["sequences"] = _try(
        lambda: client.search_read(
            "ir.sequence",
            ["|", ("company_id", "=", company_id), ("company_id", "=", False)],
            fields=["id", "code", "name", "number_next", "padding", "implementation", "active"],
            order="code",
        ),
        default=[],
    )

    # ir.property (company-dependent)
    data["properties"] = _try(
        lambda: client.search_read(
            "ir.property",
            [("company_id", "=", company_id)],
            fields=["id", "name", "fields_id", "res_id", "value_text", "value_reference", "type"],
            order="name",
            limit=300,
        ),
        default=[],
    )

    return data


def _pull_global_data(client: _ClientShim) -> dict:
    """Pull data not scoped to a single company (modules, currencies, params, mail)."""
    data: dict[str, Any] = {}

    # Modules
    data["modules"] = _try(
        lambda: client.search_read(
            "ir.module.module",
            [("name", "in", TRACKED_MODULES)],
            fields=["name", "state", "installed_version", "shortdesc"],
            order="name",
        ),
        default=[],
    )

    # Currencies
    data["currencies"] = _try(
        lambda: client.search_read(
            "res.currency",
            [("name", "in", CURRENCIES)],
            fields=["name", "active", "rate", "rounding", "symbol"],
            order="name",
        ),
        default=[],
    )

    # Mail servers
    data["mail_servers"] = _try(
        lambda: client.search_read(
            "ir.mail_server",
            [],
            fields=["id", "name", "smtp_host", "smtp_port", "smtp_encryption", "active"],
            order="sequence, id",
        ),
        default=[],
    )

    # ir.config_parameter (filter to a sane size)
    params = _try(
        lambda: client.search_read(
            "ir.config_parameter",
            [],
            fields=["key", "value"],
            order="key",
            limit=2000,
        ),
        default=[],
    )
    data["config_params"] = params

    return data


# ---------------------------------------------------------------------------
# Per-company comparison/check builder
# ---------------------------------------------------------------------------


def _row(company: str, **fields: Any) -> dict:
    out = {"Company": company}
    out.update(fields)
    return out


def _check_company(target: dict, ref: dict | None, company_name: str) -> dict[str, list[dict]]:
    """Compare one target company against the reference (if provided).

    Returns dict[sheet_name -> list[row_dict]].
    """
    sheets: dict[str, list[dict]] = {
        "Company Settings": [],
        "CoA - Account Listing": [],
        "Account Types Coverage": [],
        "Journals": [],
        "Taxes": [],
        "Tax Groups": [],
        "Fiscal Positions": [],
        "Product Categories": [],
        "Warehouses + Op Types": [],
        "Stock Locations": [],
        "Operating Units": [],
        "Report Layouts": [],
        "Report Management": [],
        "Sequences": [],
        "ir.property": [],
    }

    # ── Company Settings ────────────────────────────────────────────────
    co_t = target.get("company") or {}
    co_r = (ref or {}).get("company") or {}

    def cmp_setting(label: str, tval: Any, rval: Any, status: str, note: str = "") -> dict:
        return _row(
            company_name,
            Setting=label,
            Value=tval,
            **{"Reference Value": rval if ref else "(no reference)"},
            Status=status,
            Note=note,
        )

    # currency
    t = _safe(co_t.get("currency_id"), "Not set")
    r = _safe(co_r.get("currency_id"), "")
    sheets["Company Settings"].append(cmp_setting("Currency", t, r, "PASS" if t == "IDR" else "WARN"))

    # Fiscal country
    t = _safe(co_t.get("account_fiscal_country_id"), "Not set")
    r = _safe(co_r.get("account_fiscal_country_id"), "")
    sheets["Company Settings"].append(
        cmp_setting(
            "Fiscal Country",
            t,
            r,
            "PASS" if "Indonesia" in t else "WARN",
            note="Should be Indonesia",
        )
    )

    # Default sale tax / purchase tax
    for fld, label in [
        ("account_sale_tax_id", "Default Sale Tax"),
        ("account_purchase_tax_id", "Default Purchase Tax"),
    ]:
        t = _safe(co_t.get(fld), "Not set")
        r = _safe(co_r.get(fld), "")
        sheets["Company Settings"].append(cmp_setting(label, t, r, "PASS" if co_t.get(fld) else "FAIL"))

    # FX accounts + transfer + suspense + cash diff + payment debit/credit
    for fld, label, level in [
        ("income_currency_exchange_account_id", "FX Income Account", "WARN"),
        ("expense_currency_exchange_account_id", "FX Expense Account", "WARN"),
        ("transfer_account_id", "Transfer Account", "WARN"),
        ("account_journal_suspense_account_id", "Journal Suspense Account", "WARN"),
        ("account_journal_payment_debit_account_id", "Outstanding Receipts Account", "WARN"),
        ("account_journal_payment_credit_account_id", "Outstanding Payments Account", "WARN"),
        ("default_cash_difference_income_account_id", "Cash Diff Income Account", "INFO"),
        ("default_cash_difference_expense_account_id", "Cash Diff Expense Account", "INFO"),
    ]:
        t = _safe(co_t.get(fld), "Not set")
        r = _safe(co_r.get(fld), "")
        ok = bool(co_t.get(fld))
        status = "PASS" if ok else level
        sheets["Company Settings"].append(cmp_setting(label, t, r, status))

    # Booleans + scalars
    for fld, label in [
        ("anglo_saxon_accounting", "Anglo-Saxon Accounting"),
        ("fiscalyear_lock_date", "Fiscal Year Lock"),
        ("tax_lock_date", "Tax Lock"),
        ("po_lock", "PO Lock"),
        ("po_double_validation", "PO Double Validation"),
        ("po_double_validation_amount", "PO Double Validation Amount"),
        ("po_lead", "PO Lead Time"),
        ("quotation_validity_days", "Quotation Validity (days)"),
        ("portal_confirmation_sign", "Portal Sign Confirmation"),
        ("portal_confirmation_pay", "Portal Pay Confirmation"),
    ]:
        t = co_t.get(fld)
        r = co_r.get(fld)
        sheets["Company Settings"].append(
            cmp_setting(
                label,
                "" if t in (False, None) else str(t),
                "" if r in (False, None) else str(r),
                "INFO",
            )
        )

    # Logo
    has_logo = bool(co_t.get("logo"))
    has_logo_ref = bool(co_r.get("logo"))
    sheets["Company Settings"].append(
        cmp_setting(
            "Logo",
            "Set" if has_logo else "Missing",
            "Set" if has_logo_ref else "Missing" if ref else "",
            "PASS" if has_logo else "WARN",
        )
    )

    # ── CoA - Account Listing ───────────────────────────────────────────
    accounts = target.get("accounts") or []
    ext = target.get("account_ext_ids") or {}
    for a in accounts:
        sheets["CoA - Account Listing"].append(
            _row(
                company_name,
                Code=a.get("code") or "",
                Name=a.get("name") or "",
                Type=a.get("account_type") or "",
                **{"Has Ext-ID": "Yes" if ext.get(a["id"]) else "No"},
                **{"Ext-ID": ext.get(a["id"], "")},
                Deprecated="Yes" if a.get("deprecated") else "No",
                Status="PASS" if a.get("code") else "FAIL",
                **{"Vs Reference": ""},
            )
        )

    # ── Account Types Coverage ──────────────────────────────────────────
    by_type_t: dict[str, int] = {}
    for a in accounts:
        by_type_t[a.get("account_type") or ""] = by_type_t.get(a.get("account_type") or "", 0) + 1

    by_type_r: dict[str, int] = {}
    if ref:
        for a in ref.get("accounts") or []:
            by_type_r[a.get("account_type") or ""] = by_type_r.get(a.get("account_type") or "", 0) + 1

    for at in ACCOUNT_TYPES:
        tcnt = by_type_t.get(at, 0)
        rcnt = by_type_r.get(at, 0)
        if at in CRITICAL_ACCOUNT_TYPES:
            status = "PASS" if tcnt >= 1 else "FAIL"
        elif tcnt == 0 and rcnt > 0:
            status = "WARN"
        else:
            status = "INFO"
        sheets["Account Types Coverage"].append(
            _row(
                company_name,
                **{"Account Type": at},
                **{"Target Count": tcnt},
                **{"Reference Count": rcnt if ref else "-"},
                Status=status,
                Note="Critical type" if at in CRITICAL_ACCOUNT_TYPES else "",
            )
        )

    # ── Journals ────────────────────────────────────────────────────────
    journals = target.get("journals") or []
    ref_journals = (ref or {}).get("journals") or []
    ref_codes = {j.get("code") for j in ref_journals}
    target_codes = {j.get("code") for j in journals}

    for j in journals:
        active = j.get("active")
        has_default = bool(j.get("default_account_id"))
        jtype = j.get("type") or ""
        if active and not has_default and jtype != "general":
            status = "FAIL"
            note = "Active journal missing default account"
        else:
            status = "PASS"
            note = ""
        sheets["Journals"].append(
            _row(
                company_name,
                Code=j.get("code") or "",
                Type=jtype,
                Name=j.get("name") or "",
                **{"Default Account": _safe(j.get("default_account_id"))},
                **{"Suspense Account": _safe(j.get("suspense_account_id"))},
                Currency=_safe(j.get("currency_id")),
                Active="Yes" if active else "No",
                Status=status,
                Note=note,
                **{"Vs Reference": "in MAC" if j.get("code") in ref_codes else "(target only)"},
            )
        )

    # Missing journal codes (in MAC but not in target)
    if ref:
        missing = ref_codes - target_codes
        for code in sorted(c or "" for c in missing if c):
            ref_j = next((j for j in ref_journals if j.get("code") == code), {})
            sheets["Journals"].append(
                _row(
                    company_name,
                    Code=code,
                    Type=ref_j.get("type") or "",
                    Name=f"(missing — present in MAC)",
                    **{"Default Account": ""},
                    **{"Suspense Account": ""},
                    Currency="",
                    Active="No",
                    Status="WARN",
                    Note="Journal in reference but missing in target",
                    **{"Vs Reference": "MAC only"},
                )
            )

    # ── Taxes ───────────────────────────────────────────────────────────
    taxes = target.get("taxes") or []
    tax_accounts = target.get("tax_accounts") or {}
    ref_taxes = (ref or {}).get("taxes") or []
    ref_tax_names = {(t.get("name"), t.get("type_tax_use")) for t in ref_taxes}

    for t in taxes:
        gl_accs = tax_accounts.get(t["id"], [])
        gl_str = ", ".join(gl_accs) if gl_accs else ""
        active = t.get("active")
        if active and not gl_accs:
            status = "FAIL"
            note = "Active tax missing GL account on repartition"
        else:
            status = "PASS"
            note = ""
        in_ref = (t.get("name"), t.get("type_tax_use")) in ref_tax_names
        sheets["Taxes"].append(
            _row(
                company_name,
                Name=t.get("name") or "",
                **{"Type Use": t.get("type_tax_use") or ""},
                Amount=t.get("amount"),
                **{"Amount Type": t.get("amount_type") or ""},
                **{"Tax Group": _safe(t.get("tax_group_id"))},
                **{"GL Account(s)": gl_str},
                Active="Yes" if active else "No",
                Status=status,
                Note=note,
                **{"Vs Reference": "in MAC" if in_ref else "(target only)"} if ref else {"Vs Reference": ""},
            )
        )

    # ── Tax Groups ──────────────────────────────────────────────────────
    tgroups = target.get("tax_groups") or []
    used_groups = {_safe_id(t.get("tax_group_id")) for t in taxes if t.get("active")}
    for g in tgroups:
        used = g["id"] in used_groups
        has_payable = bool(g.get("tax_payable_account_id"))
        has_recv = bool(g.get("tax_receivable_account_id"))
        if used and not (has_payable and has_recv):
            status = "WARN"
            note = "Used by active taxes but missing payable/receivable account"
        else:
            status = "PASS"
            note = ""
        sheets["Tax Groups"].append(
            _row(
                company_name,
                **{"Group Name": g.get("name") or ""},
                **{"Payable Account": _safe(g.get("tax_payable_account_id"))},
                **{"Receivable Account": _safe(g.get("tax_receivable_account_id"))},
                **{"Advance Account": _safe(g.get("advance_tax_payment_account_id"))},
                **{"Used By Active Tax": "Yes" if used else "No"},
                Status=status,
                Note=note,
                **{"Vs Reference": ""},
            )
        )

    # ── Fiscal Positions ────────────────────────────────────────────────
    fps = target.get("fiscal_positions") or []
    for fp in fps:
        sheets["Fiscal Positions"].append(
            _row(
                company_name,
                Name=fp.get("name") or "",
                Country=_safe(fp.get("country_id")),
                **{"Auto Apply": "Yes" if fp.get("auto_apply") else "No"},
                **{"# Account Mappings": len(fp.get("account_ids") or [])},
                **{"# Tax Mappings": len(fp.get("tax_ids") or [])},
                Status="PASS",
                **{"Vs Reference": ""},
            )
        )
    if not fps:
        sheets["Fiscal Positions"].append(
            _row(
                company_name,
                Name="(none)",
                Country="",
                **{"Auto Apply": ""},
                **{"# Account Mappings": 0},
                **{"# Tax Mappings": 0},
                Status="WARN",
                **{"Vs Reference": ""},
            )
        )

    # ── Product Categories ──────────────────────────────────────────────
    cats = target.get("categories") or []
    for c in cats:
        valuation = c.get("property_valuation") or ""
        has_in = bool(c.get("property_account_income_categ_id"))
        has_ex = bool(c.get("property_account_expense_categ_id"))
        has_stock = (
            bool(c.get("property_stock_account_input_categ_id"))
            and bool(c.get("property_stock_account_output_categ_id"))
            and bool(c.get("property_stock_valuation_account_id"))
        )
        if not (has_in and has_ex):
            status = "FAIL"
            note = "Missing income or expense account"
        elif valuation == "real_time" and not has_stock:
            status = "WARN"
            note = "Real-time valuation missing stock accounts"
        else:
            status = "PASS"
            note = ""
        sheets["Product Categories"].append(
            _row(
                company_name,
                Name=c.get("name") or "",
                **{"Cost Method": c.get("property_cost_method") or ""},
                Valuation=valuation,
                Income=_safe(c.get("property_account_income_categ_id")),
                Expense=_safe(c.get("property_account_expense_categ_id")),
                **{"Stock Input": _safe(c.get("property_stock_account_input_categ_id"))},
                **{"Stock Output": _safe(c.get("property_stock_account_output_categ_id"))},
                **{"Stock Valuation": _safe(c.get("property_stock_valuation_account_id"))},
                **{"Stock Journal": _safe(c.get("property_stock_journal"))},
                Status=status,
                Note=note,
            )
        )

    # ── Warehouses + Operation Types ────────────────────────────────────
    whs = target.get("warehouses") or []
    if whs:
        for w in whs:
            sheets["Warehouses + Op Types"].append(
                _row(
                    company_name,
                    Section="Warehouse",
                    Code=w.get("code") or "",
                    Name=w.get("name") or "",
                    **{"Stock Loc / Source": _safe(w.get("lot_stock_id"))},
                    **{"Reception Steps / Dest": w.get("reception_steps") or ""},
                    **{"Delivery Steps / Type": w.get("delivery_steps") or ""},
                    Active="Yes" if w.get("active") else "No",
                    Status="PASS",
                )
            )
    else:
        sheets["Warehouses + Op Types"].append(
            _row(
                company_name,
                Section="Warehouse",
                Code="",
                Name="(none)",
                **{"Stock Loc / Source": ""},
                **{"Reception Steps / Dest": ""},
                **{"Delivery Steps / Type": ""},
                Active="",
                Status="FAIL",
            )
        )

    pts = target.get("picking_types") or []
    for p in pts:
        sheets["Warehouses + Op Types"].append(
            _row(
                company_name,
                Section="Operation Type",
                Code=p.get("sequence_code") or "",
                Name=p.get("name") or "",
                **{"Stock Loc / Source": _safe(p.get("default_location_src_id"))},
                **{"Reception Steps / Dest": _safe(p.get("default_location_dest_id"))},
                **{"Delivery Steps / Type": p.get("code") or ""},
                Active="Yes",
                Status="PASS",
            )
        )

    # ── Stock Locations ─────────────────────────────────────────────────
    locs = target.get("locations") or []
    by_usage: dict[str, int] = {}
    for l in locs:
        by_usage[l.get("usage") or ""] = by_usage.get(l.get("usage") or "", 0) + 1
    for usage, count in sorted(by_usage.items()):
        sheets["Stock Locations"].append(
            _row(
                company_name,
                **{"Location Type": usage},
                Count=count,
                Note="",
                Status="PASS" if count >= 1 else "WARN",
            )
        )
    ss = target.get("stock_summary") or {}
    sheets["Stock Locations"].append(
        _row(
            company_name,
            **{"Location Type": "(summary)"},
            Count=ss.get("quants") or 0,
            Note=f"Quants={ss.get('quants', 0)}, Distinct products with qty>0={ss.get('products_with_stock', 0)}, SVL records={ss.get('svl_count', 0)}",
            Status="INFO",
        )
    )

    # ── Operating Units ─────────────────────────────────────────────────
    ous = target.get("operating_units")
    counts = target.get("ou_user_counts") or {}
    if ous == "MODULE_NOT_INSTALLED":
        sheets["Operating Units"].append(
            _row(
                company_name,
                Code="",
                **{"OU Name": "(operating_unit module not installed)"},
                Partner="",
                **{"# Users Assigned": 0},
                Active="",
                Status="WARN",
            )
        )
    elif not ous:
        sheets["Operating Units"].append(
            _row(
                company_name,
                Code="",
                **{"OU Name": "(no operating units configured)"},
                Partner="",
                **{"# Users Assigned": 0},
                Active="",
                Status="WARN",
            )
        )
    else:
        for ou in ous:
            sheets["Operating Units"].append(
                _row(
                    company_name,
                    Code=ou.get("code") or "",
                    **{"OU Name": ou.get("name") or ""},
                    Partner=_safe(ou.get("partner_id")),
                    **{"# Users Assigned": counts.get(ou["id"], 0)},
                    Active="Yes" if ou.get("active") else "No",
                    Status="PASS" if ou.get("active") else "WARN",
                )
            )

    # ── Report Layouts ──────────────────────────────────────────────────
    rfmt = target.get("report_formats")
    ref_rfmt = (ref or {}).get("report_formats") if ref else None
    if rfmt == "MODULE_NOT_INSTALLED":
        sheets["Report Layouts"].append(
            _row(
                company_name,
                **{"Report Type": "(report_layout module not installed)"},
                **{"Format Name": ""},
                **{"Is Default": ""},
                **{"# Bands": 0},
                **{"# Signature Lines": 0},
                **{"Has Logo": ""},
                Status="N/A",
                Note="Module not installed",
            )
        )
    else:
        target_default_types = {f["report_type"] for f in (rfmt or []) if f.get("is_default")}
        for f in rfmt or []:
            sheets["Report Layouts"].append(
                _row(
                    company_name,
                    **{"Report Type": f.get("report_type") or ""},
                    **{"Format Name": f.get("name") or ""},
                    **{"Is Default": "Yes" if f.get("is_default") else "No"},
                    **{"# Bands": len(f.get("band_ids") or [])},
                    **{"# Signature Lines": len(f.get("signature_line_ids") or [])},
                    **{"Has Logo": "Yes" if f.get("logo") else "No"},
                    Status="PASS" if f.get("active") else "WARN",
                    Note="",
                )
            )
        # Compare report types
        if isinstance(ref_rfmt, list):
            ref_default_types = {f["report_type"] for f in ref_rfmt if f.get("is_default")}
            missing = ref_default_types - target_default_types
            for rt in sorted(missing):
                sheets["Report Layouts"].append(
                    _row(
                        company_name,
                        **{"Report Type": rt},
                        **{"Format Name": "(missing default)"},
                        **{"Is Default": "No"},
                        **{"# Bands": 0},
                        **{"# Signature Lines": 0},
                        **{"Has Logo": ""},
                        Status="FAIL",
                        Note="MAC has a default for this type — target has none",
                    )
                )

    # ── Report Management Instances ─────────────────────────────────────
    rinst = target.get("report_instances")
    rtypes = target.get("report_types")
    if rinst == "MODULE_NOT_INSTALLED":
        sheets["Report Management"].append(
            _row(
                company_name,
                Section="Instance",
                Name="(report_management_financial not installed)",
                Template="",
                **{"Report Type / Handler": ""},
                Status="N/A",
            )
        )
    else:
        if not rinst:
            sheets["Report Management"].append(
                _row(
                    company_name,
                    Section="Instance",
                    Name="(no instances configured)",
                    Template="",
                    **{"Report Type / Handler": ""},
                    Status="WARN",
                )
            )
        for inst in rinst or []:
            sheets["Report Management"].append(
                _row(
                    company_name,
                    Section="Instance",
                    Name=inst.get("name") or "",
                    Template=_safe(inst.get("template_id")),
                    **{"Report Type / Handler": inst.get("report_type") or ""},
                    Status="PASS",
                )
            )
    if isinstance(rtypes, list):
        for rt in rtypes:
            sheets["Report Management"].append(
                _row(
                    company_name,
                    Section="Registered Type",
                    Name=rt.get("name") or "",
                    Template=rt.get("category") or "",
                    **{"Report Type / Handler": rt.get("handler") or rt.get("code") or ""},
                    Status="INFO",
                )
            )

    # ── Sequences ───────────────────────────────────────────────────────
    seqs = target.get("sequences") or []
    for s in seqs:
        sheets["Sequences"].append(
            _row(
                company_name,
                Code=s.get("code") or "",
                Name=s.get("name") or "",
                **{"Number Next": s.get("number_next") or 0},
                Padding=s.get("padding") or 0,
                Implementation=s.get("implementation") or "",
                Active="Yes" if s.get("active") else "No",
                Status="PASS" if s.get("active") else "WARN",
            )
        )

    # ── ir.property ─────────────────────────────────────────────────────
    props = target.get("properties") or []
    for p in props[:200]:
        sheets["ir.property"].append(
            _row(
                company_name,
                Field=_safe(p.get("fields_id")) or p.get("name") or "",
                Resource=_truncate(p.get("res_id"), 60),
                Type=p.get("type") or "",
                Value=_truncate(p.get("value_text") or p.get("value_reference"), 80),
                Status="INFO",
            )
        )

    return sheets


# ---------------------------------------------------------------------------
# Excel writer
# ---------------------------------------------------------------------------


SHEET_ORDER = [
    "Summary",
    "Company Settings",
    "CoA - Account Listing",
    "Account Types Coverage",
    "Journals",
    "Taxes",
    "Tax Groups",
    "Fiscal Positions",
    "Product Categories",
    "Warehouses + Op Types",
    "Stock Locations",
    "Operating Units",
    "Report Layouts",
    "Report Management",
    "ir.config_parameter",
    "ir.property",
    "Modules + Currencies",
    "Sequences",
]


def _write_excel(
    company_results: list[tuple[str, dict[str, list[dict]]]],
    target_global: dict,
    ref_global: dict | None,
    output: Path,
    target_label: str,
    reference_label: str | None,
) -> None:
    """Write the multi-sheet workbook."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    hdr_font = Font(bold=True, color="FFFFFF", size=11)
    hdr_fill = PatternFill("solid", fgColor="2F5496")
    pass_fill = PatternFill("solid", fgColor="C6EFCE")
    warn_fill = PatternFill("solid", fgColor="FFEB9C")
    fail_fill = PatternFill("solid", fgColor="FFC7CE")
    info_fill = PatternFill("solid", fgColor="DDEBF7")
    border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )

    def status_fill(s: str):
        return {
            "PASS": pass_fill,
            "WARN": warn_fill,
            "FAIL": fail_fill,
            "INFO": info_fill,
            "N/A": info_fill,
        }.get(s)

    def write_table(ws, headers: list[str], rows: list[dict], widths: dict[str, int] | None = None):
        # Header row
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = hdr_font
            cell.fill = hdr_fill
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
            cell.border = border
        # Default widths
        for col, h in enumerate(headers, 1):
            w = (widths or {}).get(h)
            if w is None:
                if h in ("Status", "Active", "Has Ext-ID", "Is Default", "Has Logo", "Auto Apply"):
                    w = 12
                elif h in ("Code", "Type", "Type Use", "Amount", "Padding", "Number Next"):
                    w = 14
                elif h in ("Note", "Vs Reference", "Reference Value"):
                    w = 50
                elif h == "Name" or h == "Setting" or h == "Field" or h == "OU Name":
                    w = 35
                elif h == "Company":
                    w = 32
                else:
                    w = 22
            ws.column_dimensions[get_column_letter(col)].width = w

        for r_idx, row in enumerate(rows, 2):
            for c_idx, h in enumerate(headers, 1):
                val = row.get(h, "")
                cell = ws.cell(row=r_idx, column=c_idx, value=val if val != "" else None)
                cell.font = Font(size=10)
                cell.border = border
                if h == "Status":
                    fill = status_fill(str(val))
                    if fill:
                        cell.fill = fill
                    cell.font = Font(size=10, bold=True)
                    cell.alignment = Alignment(horizontal="center")

        ws.freeze_panes = "A2"
        if rows:
            ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(rows) + 1}"

    # ── Sheet 1: Summary ──────────────────────────────────────────────
    ws = wb.active
    ws.title = "Summary"
    sections = [
        "Company Settings",
        "CoA - Account Listing",
        "Account Types Coverage",
        "Journals",
        "Taxes",
        "Tax Groups",
        "Fiscal Positions",
        "Product Categories",
        "Warehouses + Op Types",
        "Stock Locations",
        "Operating Units",
        "Report Layouts",
    ]
    headers = ["Company"] + sections + ["Total Pass", "Total Fail", "Total Warn", "Ready?"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = border
    ws.column_dimensions["A"].width = 35
    for col in range(2, 2 + len(sections)):
        ws.column_dimensions[get_column_letter(col)].width = 13
    ws.column_dimensions[get_column_letter(2 + len(sections))].width = 12
    ws.column_dimensions[get_column_letter(3 + len(sections))].width = 12
    ws.column_dimensions[get_column_letter(4 + len(sections))].width = 12
    ws.column_dimensions[get_column_letter(5 + len(sections))].width = 13

    row = 2
    for company_name, sheets in company_results:
        ws.cell(row=row, column=1, value=company_name).font = Font(size=10, bold=True)
        total_pass = total_fail = total_warn = 0
        for col_off, sec in enumerate(sections):
            checks = sheets.get(sec, [])
            counted = [c for c in checks if c.get("Status") in ("PASS", "WARN", "FAIL")]
            passes = sum(1 for c in counted if c["Status"] == "PASS")
            fails = sum(1 for c in counted if c["Status"] == "FAIL")
            warns = sum(1 for c in counted if c["Status"] == "WARN")
            total_pass += passes
            total_fail += fails
            total_warn += warns
            cell = ws.cell(row=row, column=2 + col_off, value=f"{passes}/{len(counted)}")
            cell.alignment = Alignment(horizontal="center")
            cell.border = border
            if fails > 0:
                cell.fill = fail_fill
            elif warns > 0:
                cell.fill = warn_fill
            elif passes > 0:
                cell.fill = pass_fill
            else:
                cell.fill = info_fill
        ws.cell(row=row, column=2 + len(sections), value=total_pass).font = Font(bold=True)
        ws.cell(row=row, column=3 + len(sections), value=total_fail).font = Font(bold=True)
        ws.cell(row=row, column=4 + len(sections), value=total_warn).font = Font(bold=True)
        ready = "READY" if total_fail == 0 else "NOT READY"
        ready_cell = ws.cell(row=row, column=5 + len(sections), value=ready)
        ready_cell.font = Font(bold=True)
        ready_cell.fill = pass_fill if total_fail == 0 else fail_fill
        ready_cell.alignment = Alignment(horizontal="center")
        for col in range(1, len(headers) + 1):
            ws.cell(row=row, column=col).border = border
        row += 1
    ws.freeze_panes = "B2"
    if company_results:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(company_results) + 1}"

    # Add header banner row at top describing target/reference
    ws_meta = wb.create_sheet("README", 0)
    wb.move_sheet("README", offset=-wb.index(ws_meta))
    ws_meta.column_dimensions["A"].width = 24
    ws_meta.column_dimensions["B"].width = 60
    meta_rows = [
        ("Generated", date.today().isoformat()),
        ("Target", target_label),
        ("Reference", reference_label or "(none — existence-only check)"),
        ("# Companies", str(len(company_results))),
        ("Sheets", ", ".join(SHEET_ORDER)),
    ]
    for r, (k, v) in enumerate(meta_rows, 1):
        c1 = ws_meta.cell(row=r, column=1, value=k)
        c1.font = Font(bold=True)
        c1.fill = hdr_fill
        c1.font = hdr_font
        ws_meta.cell(row=r, column=2, value=v).font = Font(size=10)

    # ── Detail sheets driven by per-company checks ──────────────────────
    detail_sections: list[tuple[str, list[str]]] = [
        ("Company Settings", ["Company", "Setting", "Value", "Reference Value", "Status", "Note"]),
        (
            "CoA - Account Listing",
            ["Company", "Code", "Name", "Type", "Has Ext-ID", "Ext-ID", "Deprecated", "Status", "Vs Reference"],
        ),
        (
            "Account Types Coverage",
            ["Company", "Account Type", "Target Count", "Reference Count", "Status", "Note"],
        ),
        (
            "Journals",
            [
                "Company",
                "Code",
                "Type",
                "Name",
                "Default Account",
                "Suspense Account",
                "Currency",
                "Active",
                "Status",
                "Note",
                "Vs Reference",
            ],
        ),
        (
            "Taxes",
            [
                "Company",
                "Name",
                "Type Use",
                "Amount",
                "Amount Type",
                "Tax Group",
                "GL Account(s)",
                "Active",
                "Status",
                "Note",
                "Vs Reference",
            ],
        ),
        (
            "Tax Groups",
            [
                "Company",
                "Group Name",
                "Payable Account",
                "Receivable Account",
                "Advance Account",
                "Used By Active Tax",
                "Status",
                "Note",
                "Vs Reference",
            ],
        ),
        (
            "Fiscal Positions",
            [
                "Company",
                "Name",
                "Country",
                "Auto Apply",
                "# Account Mappings",
                "# Tax Mappings",
                "Status",
                "Vs Reference",
            ],
        ),
        (
            "Product Categories",
            [
                "Company",
                "Name",
                "Cost Method",
                "Valuation",
                "Income",
                "Expense",
                "Stock Input",
                "Stock Output",
                "Stock Valuation",
                "Stock Journal",
                "Status",
                "Note",
            ],
        ),
        (
            "Warehouses + Op Types",
            [
                "Company",
                "Section",
                "Code",
                "Name",
                "Stock Loc / Source",
                "Reception Steps / Dest",
                "Delivery Steps / Type",
                "Active",
                "Status",
            ],
        ),
        ("Stock Locations", ["Company", "Location Type", "Count", "Note", "Status"]),
        (
            "Operating Units",
            ["Company", "Code", "OU Name", "Partner", "# Users Assigned", "Active", "Status"],
        ),
        (
            "Report Layouts",
            [
                "Company",
                "Report Type",
                "Format Name",
                "Is Default",
                "# Bands",
                "# Signature Lines",
                "Has Logo",
                "Status",
                "Note",
            ],
        ),
        (
            "Report Management",
            ["Company", "Section", "Name", "Template", "Report Type / Handler", "Status"],
        ),
        ("Sequences", ["Company", "Code", "Name", "Number Next", "Padding", "Implementation", "Active", "Status"]),
        ("ir.property", ["Company", "Field", "Resource", "Type", "Value", "Status"]),
    ]

    for title, hdrs in detail_sections:
        ws = wb.create_sheet(title)
        all_rows: list[dict] = []
        for company_name, sheets in company_results:
            all_rows.extend(sheets.get(title, []))
        write_table(ws, hdrs, all_rows)

    # ── ir.config_parameter (global) ────────────────────────────────────
    ws = wb.create_sheet("ir.config_parameter")
    headers = ["Key", "Value", "Reference Value", "Status", "Note"]
    rows: list[dict] = []
    target_params = {p["key"]: p["value"] for p in (target_global.get("config_params") or [])}
    ref_params = {p["key"]: p["value"] for p in ((ref_global or {}).get("config_params") or [])}
    all_keys = sorted(set(target_params) | set(ref_params))
    for key in all_keys:
        tval = target_params.get(key, "")
        rval = ref_params.get(key, "")
        is_critical = key in CRITICAL_PARAMS
        is_secret = any(s in key.lower() for s in ("secret", "password", "token", "api_key"))
        tdisplay = _mask(tval) if is_secret else _truncate(tval, 100)
        rdisplay = _mask(rval) if is_secret else _truncate(rval, 100)
        if is_critical and not tval:
            status = "FAIL"
            note = "Critical parameter missing"
        elif is_critical:
            status = "PASS"
            note = "Critical parameter"
        else:
            status = "INFO"
            note = ""
        rows.append(
            {
                "Key": key,
                "Value": tdisplay,
                "Reference Value": rdisplay if ref_global else "",
                "Status": status,
                "Note": note,
            }
        )
    write_table(ws, headers, rows, widths={"Key": 50, "Value": 50, "Reference Value": 50, "Note": 40})

    # ── Modules + Currencies + Mail (global) ────────────────────────────
    ws = wb.create_sheet("Modules + Currencies")
    headers = ["Section", "Item", "Target State / Active", "Reference", "Status", "Note"]
    rows = []
    target_modules = {m["name"]: m for m in (target_global.get("modules") or [])}
    ref_modules = {m["name"]: m for m in ((ref_global or {}).get("modules") or [])}
    for mod in TRACKED_MODULES:
        tm = target_modules.get(mod, {})
        rm = ref_modules.get(mod, {})
        tstate = tm.get("state", "not found")
        rstate = rm.get("state", "not found") if ref_global else "-"
        if mod in REQUIRED_MODULES:
            status = "PASS" if tstate == "installed" else "FAIL"
        else:
            if tstate == "installed":
                status = "PASS"
            elif rstate == "installed":
                status = "WARN"
            else:
                status = "INFO"
        rows.append(
            {
                "Section": "Module",
                "Item": mod,
                "Target State / Active": f"{tstate} ({tm.get('installed_version', '')})" if tm else tstate,
                "Reference": rstate,
                "Status": status,
                "Note": "Required" if mod in REQUIRED_MODULES else "",
            }
        )

    # Currencies
    target_curr = {c["name"]: c for c in (target_global.get("currencies") or [])}
    ref_curr = {c["name"]: c for c in ((ref_global or {}).get("currencies") or [])}
    for ccode in CURRENCIES:
        t = target_curr.get(ccode, {})
        r = ref_curr.get(ccode, {})
        active = t.get("active")
        rate = t.get("rate", "")
        rows.append(
            {
                "Section": "Currency",
                "Item": ccode,
                "Target State / Active": f"active={active}, rate={rate}",
                "Reference": f"active={r.get('active')}, rate={r.get('rate', '')}" if ref_global else "-",
                "Status": "PASS" if active else "WARN",
                "Note": "",
            }
        )

    # Mail servers
    msrvs = target_global.get("mail_servers") or []
    rmsrvs = (ref_global or {}).get("mail_servers") or []
    if msrvs:
        for s in msrvs:
            rows.append(
                {
                    "Section": "Mail Server",
                    "Item": s.get("name") or "",
                    "Target State / Active": (
                        f"{s.get('smtp_host')}:{s.get('smtp_port')} ({s.get('smtp_encryption', '-')})"
                        f" {'active' if s.get('active') else 'inactive'}"
                    ),
                    "Reference": "",
                    "Status": "PASS" if s.get("active") else "WARN",
                    "Note": "",
                }
            )
    else:
        rows.append(
            {
                "Section": "Mail Server",
                "Item": "(none configured)",
                "Target State / Active": "",
                "Reference": f"{len(rmsrvs)} server(s) in MAC" if ref_global else "-",
                "Status": "WARN",
                "Note": "No outgoing mail server configured",
            }
        )
    write_table(
        ws, headers, rows, widths={"Section": 14, "Item": 32, "Target State / Active": 50, "Reference": 30, "Note": 30}
    )

    # Reorder sheets to match SHEET_ORDER (README first, then ordered)
    desired = ["README"] + SHEET_ORDER
    for idx, name in enumerate(desired):
        if name in wb.sheetnames:
            sheet = wb[name]
            current = wb.index(sheet)
            offset = idx - current
            if offset:
                wb.move_sheet(sheet, offset=offset)

    wb.save(str(output))


# ---------------------------------------------------------------------------
# Main typer command
# ---------------------------------------------------------------------------


@app.command("readiness-check")
def readiness_check(
    ctx: typer.Context,
    output: Annotated[
        Path | None,
        typer.Option("-o", "--output", help="Output Excel path (default: readiness_<db>_<date>.xlsx)"),
    ] = None,
    company_id: Annotated[
        int,
        typer.Option("--company-id", help="Limit to one company id (default: all)"),
    ] = 0,
    reference: Annotated[
        str | None,
        typer.Option("--reference", help="Reference profile name (e.g. idtpp-mac-odoo-erp)"),
    ] = None,
    reference_company_id: Annotated[
        int,
        typer.Option(
            "--reference-company-id",
            help="Reference company id to use as the gold standard (default: first company on reference)",
        ),
    ] = 0,
) -> None:
    """Generate a multi-sheet production readiness audit Excel.

    Without ``--reference``, runs existence-only checks against the target
    instance. With ``--reference``, also pulls config from a second profile
    (e.g. MAC production) and compares each target company against it.
    """
    actx: AppContext = ctx.obj
    _bump_timeout(actx, seconds=600.0)
    target_inner = actx.client
    out = actx.output

    try:
        from openpyxl import Workbook  # noqa: F401
    except ImportError:
        out.error("openpyxl not installed. Run: pip install openpyxl")
        raise typer.Exit(1)

    target_client = _ClientShim(target_inner)
    target_label = f"{getattr(actx, 'profile', 'target')} ({target_client.database})"

    # Reference setup
    ref_client: _ClientShim | None = None
    reference_label: str | None = None
    if reference:
        try:
            ref_profile = _load_profile(reference)
            xrpc = XmlRpcClient(ref_profile)
            ref_client = _ClientShim(xrpc)
            reference_label = f"{reference} ({ref_profile['database']})"
            out.info(f"Reference: {reference_label}")
        except Exception as exc:
            out.error(f"Failed to load reference profile '{reference}': {exc}")
            raise typer.Exit(1)

    # Target companies
    domain = [("id", "=", company_id)] if company_id else []
    companies = target_client.search_read("res.company", domain, fields=["id", "name"], order="id")
    if not companies:
        out.error("No companies found on target")
        raise typer.Exit(1)
    out.info(f"Target: {target_label} ({len(companies)} company/companies)")

    # Pull reference data once (one company = the gold standard)
    ref_company_data: dict | None = None
    ref_global: dict | None = None
    if ref_client:
        ref_companies = ref_client.search_read("res.company", [], fields=["id", "name"], order="id")
        if not ref_companies:
            out.error("Reference instance has no companies")
            raise typer.Exit(1)
        chosen = None
        if reference_company_id:
            chosen = next((c for c in ref_companies if c["id"] == reference_company_id), None)
            if not chosen:
                out.error(f"Reference company id {reference_company_id} not found")
                raise typer.Exit(1)
        else:
            chosen = ref_companies[0]
        out.info(f"  Reference company: {chosen['name']} (id={chosen['id']})")
        ref_company_data = _pull_company_data(ref_client, chosen["id"], chosen["name"])
        ref_global = _pull_global_data(ref_client)

    # Pull target global once
    target_global = _pull_global_data(target_client)

    # Per-company pull + check
    company_results: list[tuple[str, dict[str, list[dict]]]] = []
    for co in companies:
        out.info(f"  Auditing {co['name']}...")
        try:
            tdata = _pull_company_data(target_client, co["id"], co["name"])
            sheets = _check_company(tdata, ref_company_data, co["name"])
            company_results.append((co["name"], sheets))
        except Exception as exc:
            out.error(f"  {co['name']}: {exc}")

    # Default output path
    if not output:
        db = target_client.database or "odoo"
        output = Path(f"readiness_{db}_{date.today().isoformat()}.xlsx")

    _write_excel(
        company_results,
        target_global,
        ref_global,
        output,
        target_label=target_label,
        reference_label=reference_label,
    )

    # Console summary
    out.info(f"\nReport written to {output}")
    for company_name, sheets in company_results:
        all_rows = [r for rows in sheets.values() for r in rows if r.get("Status") in ("PASS", "WARN", "FAIL")]
        fails = sum(1 for r in all_rows if r["Status"] == "FAIL")
        warns = sum(1 for r in all_rows if r["Status"] == "WARN")
        passes = sum(1 for r in all_rows if r["Status"] == "PASS")
        status = "READY" if fails == 0 else "NOT READY"
        out.info(f"  {company_name}: {status}  ({passes} pass / {fails} fail / {warns} warn)")
