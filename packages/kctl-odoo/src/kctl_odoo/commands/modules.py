"""Module management commands."""

from __future__ import annotations

from datetime import UTC, datetime
from pathlib import Path
from typing import Annotated

import typer

from kctl_odoo.core.callbacks import AppContext
from kctl_odoo.core.snapshots import (
    ModuleSnapshot,
    SnapshotEntry,
    SnapshotMetadata,
    diff_snapshots,
    load_snapshot,
    save_snapshot,
)
from kctl_odoo.core.utils import module_state_color

app = typer.Typer(help="Manage Odoo modules.")


@app.command("list")
def list_(
    ctx: typer.Context,
    state: Annotated[
        str | None, typer.Option("--state", "-s", help="Filter by state (installed, uninstalled, to upgrade)")
    ] = None,
    limit: Annotated[int, typer.Option("--limit", "-l", help="Max results")] = 200,
) -> None:
    """List modules."""
    actx: AppContext = ctx.obj
    out = actx.output
    c = actx.client

    domain: list = []
    if state:
        domain.append(("state", "=", state))

    modules = c.search_read(
        "ir.module.module",
        domain=domain,
        fields=["id", "name", "shortdesc", "state", "installed_version", "author"],
        limit=limit,
        order="name",
    )

    rows = []
    json_data = []
    for m in modules:
        state_display = module_state_color(m.get("state", ""))
        rows.append(
            [
                str(m["id"]),
                m["name"],
                m.get("shortdesc") or "",
                state_display,
                m.get("installed_version") or "-",
                m.get("author") or "",
            ]
        )
        json_data.append(
            {
                "id": m["id"],
                "name": m["name"],
                "summary": m.get("shortdesc"),
                "state": m.get("state"),
                "version": m.get("installed_version"),
                "author": m.get("author"),
            }
        )

    out.table(
        f"Modules ({len(modules)})",
        [("ID", "cyan"), ("Technical Name", ""), ("Summary", ""), ("State", ""), ("Version", "dim"), ("Author", "dim")],
        rows,
        data_for_json=json_data,
    )


@app.command()
def install(
    ctx: typer.Context,
    names: Annotated[str, typer.Argument(help="Comma-separated module names")],
) -> None:
    """Install modules."""
    actx: AppContext = ctx.obj
    out = actx.output
    c = actx.client

    module_names = [n.strip() for n in names.split(",")]
    for name in module_names:
        ids = c.search("ir.module.module", [("name", "=", name)])
        if not ids:
            out.error(f"Module not found: {name}")
            continue
        c.execute_kw("ir.module.module", "button_immediate_install", [ids])
        out.success(f"Installed: {name}")


@app.command()
def upgrade(
    ctx: typer.Context,
    names: Annotated[str, typer.Argument(help="Comma-separated module names")],
) -> None:
    """Upgrade modules."""
    actx: AppContext = ctx.obj
    out = actx.output
    c = actx.client

    module_names = [n.strip() for n in names.split(",")]
    for name in module_names:
        ids = c.search("ir.module.module", [("name", "=", name), ("state", "=", "installed")])
        if not ids:
            out.error(f"Module not found or not installed: {name}")
            continue
        c.execute_kw("ir.module.module", "button_immediate_upgrade", [ids])
        out.success(f"Upgraded: {name}")


@app.command()
def uninstall(
    ctx: typer.Context,
    names: Annotated[str, typer.Argument(help="Comma-separated module names")],
    force: Annotated[bool, typer.Option("--force", help="Skip confirmation")] = False,
) -> None:
    """Uninstall modules."""
    actx: AppContext = ctx.obj
    out = actx.output
    c = actx.client

    module_names = [n.strip() for n in names.split(",")]
    if not force and not typer.confirm(f"Uninstall {', '.join(module_names)}? This may remove data."):
        raise typer.Exit(0)

    for name in module_names:
        ids = c.search("ir.module.module", [("name", "=", name), ("state", "=", "installed")])
        if not ids:
            out.error(f"Module not found or not installed: {name}")
            continue
        c.execute_kw("ir.module.module", "button_immediate_uninstall", [ids])
        out.success(f"Uninstalled: {name}")


@app.command()
def check(
    ctx: typer.Context,
    name: Annotated[str, typer.Argument(help="Module technical name")],
) -> None:
    """Check if a module is installed (exit 0 if yes, exit 1 if not).

    Useful for scripting: kctl-odoo modules check sfa_management && echo "ready"

    Examples:
        kctl-odoo modules check sfa_management
        kctl-odoo modules check sale --json
    """
    actx: AppContext = ctx.obj
    out = actx.output
    c = actx.client

    modules = c.search_read(
        "ir.module.module",
        [("name", "=", name)],
        ["name", "state", "installed_version", "shortdesc"],
        limit=1,
    )

    if not modules:
        if actx.json_mode:
            out.raw_json({"name": name, "found": False, "installed": False})
        else:
            out.error(f"Module not found: {name}")
        raise typer.Exit(1)

    m = modules[0]
    installed = m.get("state") == "installed"

    if actx.json_mode:
        out.raw_json(
            {
                "name": m["name"],
                "found": True,
                "installed": installed,
                "state": m.get("state"),
                "version": m.get("installed_version") or None,
                "summary": m.get("shortdesc") or None,
            }
        )
    else:
        state = m.get("state", "")
        state_display = module_state_color(state)
        out.detail(
            f"Module: {name}",
            [
                (
                    "Info",
                    [
                        ("Name", m["name"]),
                        ("Summary", m.get("shortdesc") or "-"),
                        ("State", state_display),
                        ("Version", m.get("installed_version") or "-"),
                    ],
                )
            ],
        )

    if not installed:
        raise typer.Exit(1)


@app.command()
def search(
    ctx: typer.Context,
    query: Annotated[str, typer.Argument(help="Search term")],
    limit: Annotated[int, typer.Option("--limit", "-l")] = 20,
) -> None:
    """Search for modules by name or description."""
    actx: AppContext = ctx.obj
    out = actx.output
    c = actx.client

    domain = ["|", ("name", "ilike", query), ("shortdesc", "ilike", query)]
    modules = c.search_read(
        "ir.module.module",
        domain=domain,
        fields=["id", "name", "shortdesc", "state", "installed_version"],
        limit=limit,
        order="name",
    )

    rows = []
    for m in modules:
        rows.append(
            [
                str(m["id"]),
                m["name"],
                m.get("shortdesc") or "",
                module_state_color(m.get("state", "")),
                m.get("installed_version") or "-",
            ]
        )

    out.table(
        f"Search: '{query}' ({len(modules)} results)",
        [("ID", "cyan"), ("Name", ""), ("Summary", ""), ("State", ""), ("Version", "dim")],
        rows,
        data_for_json=modules,
    )


@app.command("scan")
def scan(ctx: typer.Context) -> None:
    """Scan for new or updated modules (update module list)."""
    actx: AppContext = ctx.obj
    out = actx.output
    c = actx.client

    out.info("Scanning for new modules...")
    c.execute_kw("ir.module.module", "update_list", [])
    out.success("Module list updated")


# ---------------------------------------------------------------------------
# Helper: build dep maps
# ---------------------------------------------------------------------------


def _build_dep_maps(
    c: object,
) -> tuple[dict[str, str], dict[str, list[str]], dict[str, list[str]]]:
    """Return (id_to_name, dep_map, rdep_map) from Odoo.

    - ``id_to_name``: module id → technical name
    - ``dep_map``: module_name → list of dependency names
    - ``rdep_map``: dep_name → list of modules that depend on it
    """
    # id→name mapping
    mods = c.search_read(  # type: ignore[union-attr]
        "ir.module.module",
        domain=[],
        fields=["id", "name"],
        limit=0,
        order="name",
    )
    id_to_name: dict[str, str] = {str(m["id"]): m["name"] for m in mods}

    # dependency records
    deps = c.search_read(  # type: ignore[union-attr]
        "ir.module.module.dependency",
        domain=[],
        fields=["module_id", "name"],
        limit=0,
    )

    dep_map: dict[str, list[str]] = {}
    rdep_map: dict[str, list[str]] = {}

    for d in deps:
        mid = d.get("module_id")
        if isinstance(mid, list):
            mid = mid[0]
        parent_name = id_to_name.get(str(mid), str(mid))
        dep_name = d.get("name", "")

        dep_map.setdefault(parent_name, [])
        if dep_name not in dep_map[parent_name]:
            dep_map[parent_name].append(dep_name)

        rdep_map.setdefault(dep_name, [])
        if parent_name not in rdep_map[dep_name]:
            rdep_map[dep_name].append(parent_name)

    return id_to_name, dep_map, rdep_map


def _collect_transitive_deps(
    name: str,
    dep_map: dict[str, list[str]],
    max_depth: int,
    current_depth: int = 0,
    visited: set[str] | None = None,
) -> list[str]:
    """Recursively collect transitive dependency names."""
    if visited is None:
        visited = set()
    if name in visited:
        return []
    visited.add(name)

    result: list[str] = []
    for dep in dep_map.get(name, []):
        if dep not in visited:
            result.append(dep)
            if max_depth == 0 or current_depth < max_depth - 1:
                result.extend(_collect_transitive_deps(dep, dep_map, max_depth, current_depth + 1, visited))
    return result


def _print_dep_tree(
    name: str,
    dep_map: dict[str, list[str]],
    out: object,
    indent: int = 0,
    visited: set[str] | None = None,
    max_depth: int = 0,
) -> None:
    """Print an indented dependency tree to output."""
    if visited is None:
        visited = set()
    if max_depth > 0 and indent >= max_depth:
        return

    children = dep_map.get(name, [])
    for child in sorted(children):
        prefix = "  " * indent + ("└─ " if indent > 0 else "")
        out.info(f"{prefix}{child}")  # type: ignore[union-attr]
        if child not in visited:
            visited.add(child)
            _print_dep_tree(child, dep_map, out, indent + 1, visited, max_depth)


# ---------------------------------------------------------------------------
# deps
# ---------------------------------------------------------------------------


@app.command()
def deps(
    ctx: typer.Context,
    name: Annotated[str, typer.Argument(help="Module technical name")],
    depth: Annotated[int, typer.Option("--depth", "-d", help="Max depth (0 = all)")] = 0,
) -> None:
    """Show module install dependency tree.

    Examples:
        kctl-odoo modules deps sale_management
        kctl-odoo modules deps sale_management --depth 2
        kctl-odoo modules deps sale_management --json
    """
    actx: AppContext = ctx.obj
    out = actx.output
    c = actx.client

    _id_to_name, dep_map, _rdep_map = _build_dep_maps(c)

    if actx.json_mode:
        all_deps = _collect_transitive_deps(name, dep_map, max_depth=depth)
        out.raw_json({"module": name, "depth": depth, "dependencies": sorted(set(all_deps))})
        return

    direct = dep_map.get(name, [])
    if not direct:
        out.info(f"No dependencies found for: {name}")
        return

    out.info(f"Dependencies of [bold]{name}[/bold] (depth={depth or 'all'}):")
    _print_dep_tree(name, dep_map, out, indent=0, max_depth=depth)


# ---------------------------------------------------------------------------
# rdeps
# ---------------------------------------------------------------------------


@app.command()
def rdeps(
    ctx: typer.Context,
    name: Annotated[str, typer.Argument(help="Module technical name")],
    depth: Annotated[int, typer.Option("--depth", "-d", help="Max depth (0 = all, default: 1)")] = 1,
) -> None:
    """Show reverse dependencies (modules that depend on this one).

    Examples:
        kctl-odoo modules rdeps base_management
        kctl-odoo modules rdeps base_management --depth 0
        kctl-odoo modules rdeps base_management --json
    """
    actx: AppContext = ctx.obj
    out = actx.output
    c = actx.client

    _id_to_name, _dep_map, rdep_map = _build_dep_maps(c)

    # Collect reverse deps (with depth control)
    def _collect_rdeps(n: str, remaining: int, seen: set[str]) -> list[str]:
        result: list[str] = []
        for mod in rdep_map.get(n, []):
            if mod not in seen:
                seen.add(mod)
                result.append(mod)
                if remaining != 1:  # 0=unlimited, >1 means go deeper
                    result.extend(_collect_rdeps(mod, max(0, remaining - 1), seen))
        return result

    seen: set[str] = {name}
    all_rdeps = _collect_rdeps(name, depth, seen)

    if actx.json_mode:
        out.raw_json({"module": name, "depth": depth, "reverse_dependencies": sorted(set(all_rdeps))})
        return

    if not all_rdeps:
        out.info(f"No reverse dependencies found for: {name}")
        return

    rows = [[m] for m in sorted(set(all_rdeps))]
    out.table(
        f"Reverse dependencies of '{name}' ({len(rows)})",
        [("Module", "")],
        rows,
        data_for_json=sorted(set(all_rdeps)),
    )


# ---------------------------------------------------------------------------
# snapshot
# ---------------------------------------------------------------------------


@app.command()
def snapshot(
    ctx: typer.Context,
    output: Annotated[str | None, typer.Option("--output", "-o", help="Output file path")] = None,
    state: Annotated[str, typer.Option("--state", "-s", help="Module state filter: installed or all")] = "installed",
) -> None:
    """Export current module state to a JSON snapshot file.

    Examples:
        kctl-odoo modules snapshot
        kctl-odoo modules snapshot --output /tmp/snap.json
        kctl-odoo modules snapshot --state all
    """
    actx: AppContext = ctx.obj
    out = actx.output
    c = actx.client

    domain: list = []
    if state == "installed":
        domain.append(("state", "=", "installed"))

    mods = c.search_read(
        "ir.module.module",
        domain=domain,
        fields=["name", "installed_version", "state", "author", "category_id"],
        limit=0,
        order="name",
    )

    entries: list[SnapshotEntry] = []
    for m in mods:
        cat = m.get("category_id")
        category = cat[1] if isinstance(cat, list) and len(cat) >= 2 else ""
        entries.append(
            SnapshotEntry(
                name=m["name"],
                version=m.get("installed_version") or "",
                state=m.get("state") or "installed",
                author=m.get("author") or "",
                category=category,
            )
        )

    date_str = datetime.now(tz=UTC).strftime("%Y-%m-%d")
    database = c.database
    profile = actx.profile or ""

    snap = ModuleSnapshot(
        metadata=SnapshotMetadata(
            database=database,
            profile=profile,
        ),
        modules=entries,
    )

    dest = Path(output) if output else Path(f"snapshot-{database}-{date_str}.json")

    save_snapshot(snap, dest)
    out.success(f"Snapshot saved: {dest} ({len(entries)} modules)")

    if actx.json_mode:
        out.raw_json({"file": str(dest), "modules": len(entries), "database": database})


# ---------------------------------------------------------------------------
# diff-snapshots
# ---------------------------------------------------------------------------


@app.command("diff-snapshots")
def diff_snapshots_cmd(
    ctx: typer.Context,
    file_a: Annotated[str, typer.Argument(help="Path to snapshot file A (baseline)")],
    file_b: Annotated[str, typer.Argument(help="Path to snapshot file B (comparison)")],
) -> None:
    """Compare two snapshot JSON files.

    Examples:
        kctl-odoo modules diff-snapshots snap-prod.json snap-staging.json
        kctl-odoo modules diff-snapshots snap-prod.json snap-staging.json --json
    """
    actx: AppContext = ctx.obj
    out = actx.output

    path_a = Path(file_a)
    path_b = Path(file_b)

    snap_a = load_snapshot(path_a)
    snap_b = load_snapshot(path_b)

    diff = diff_snapshots(snap_a, snap_b)

    if actx.json_mode:
        out.raw_json(
            {
                "source_a": diff.source_a or file_a,
                "source_b": diff.source_b or file_b,
                "added": diff.added,
                "removed": diff.removed,
                "version_changed": [
                    {"module": n, "version_a": va, "version_b": vb} for n, va, vb in diff.version_changed
                ],
                "state_changed": [{"module": n, "state_a": sa, "state_b": sb} for n, sa, sb in diff.state_changed],
            }
        )
        return

    total = len(diff.added) + len(diff.removed) + len(diff.version_changed) + len(diff.state_changed)
    sections = [
        (
            "Summary",
            [
                ("File A", file_a),
                ("File B", file_b),
                ("Source A", diff.source_a or "-"),
                ("Source B", diff.source_b or "-"),
                ("Added", str(len(diff.added))),
                ("Removed", str(len(diff.removed))),
                ("Version Changed", str(len(diff.version_changed))),
                ("State Changed", str(len(diff.state_changed))),
                ("Total Changes", str(total)),
            ],
        ),
    ]

    if diff.added:
        sections.append(("Added (in B, not in A)", [(m, "") for m in diff.added]))

    if diff.removed:
        sections.append(("Removed (in A, not in B)", [(m, "") for m in diff.removed]))

    if diff.version_changed:
        sections.append(
            (
                "Version Changed",
                [(f"{n}", f"{va} → {vb}") for n, va, vb in diff.version_changed],
            )
        )

    if diff.state_changed:
        sections.append(
            (
                "State Changed",
                [(f"{n}", f"{sa} → {sb}") for n, sa, sb in diff.state_changed],
            )
        )

    out.detail("Snapshot Diff", sections)


# ---------------------------------------------------------------------------
# diff-live
# ---------------------------------------------------------------------------


@app.command("diff-live")
def diff_live(
    ctx: typer.Context,
    profile_a: Annotated[str, typer.Argument(help="Profile name for instance A")],
    profile_b: Annotated[str, typer.Argument(help="Profile name for instance B")],
) -> None:
    """Compare installed modules between two kctl-odoo connection profiles.

    Examples:
        kctl-odoo modules diff-live production staging
        kctl-odoo modules diff-live production staging --json
    """
    from kctl_odoo.core.client import OdooClient
    from kctl_odoo.core.config import resolve_connection

    actx: AppContext = ctx.obj
    out = actx.output

    def _get_installed(profile_name: str) -> tuple[str, set[str]]:
        url, database, username, api_key = resolve_connection(profile_name=profile_name)
        client = OdooClient(base_url=url, database=database, username=username, api_key=api_key)
        try:
            mods = client.search_read(
                "ir.module.module",
                domain=[("state", "=", "installed")],
                fields=["name"],
                limit=0,
                order="name",
            )
            return database, {m["name"] for m in mods}
        finally:
            client.close()

    out.info(f"Connecting to profile '{profile_a}'...")
    db_a, mods_a = _get_installed(profile_a)

    out.info(f"Connecting to profile '{profile_b}'...")
    db_b, mods_b = _get_installed(profile_b)

    only_a = sorted(mods_a - mods_b)
    only_b = sorted(mods_b - mods_a)
    common = len(mods_a & mods_b)

    if actx.json_mode:
        out.raw_json(
            {
                "profile_a": profile_a,
                "profile_b": profile_b,
                "database_a": db_a,
                "database_b": db_b,
                "only_in_a": only_a,
                "only_in_b": only_b,
                "common_count": common,
            }
        )
        return

    sections = [
        (
            "Summary",
            [
                ("Profile A", f"{profile_a} ({db_a})"),
                ("Profile B", f"{profile_b} ({db_b})"),
                ("Installed in A", str(len(mods_a))),
                ("Installed in B", str(len(mods_b))),
                ("Common", str(common)),
                ("Only in A", str(len(only_a))),
                ("Only in B", str(len(only_b))),
            ],
        ),
    ]

    if only_a:
        sections.append((f"Only in A ({profile_a})", [(m, "") for m in only_a]))

    if only_b:
        sections.append((f"Only in B ({profile_b})", [(m, "") for m in only_b]))

    out.detail("Live Module Diff", sections)


# ---------------------------------------------------------------------------
# audit: per-module row-count usage analysis
# ---------------------------------------------------------------------------


_AUDIT_STATUS_ORDER = {"UNUSED": 0, "LOW": 1, "INHERIT": 2, "ACTIVE": 3, "HEAVY": 4}


def _audit_classify(total: int, has_persistent_table: bool) -> str:
    if not has_persistent_table:
        return "INHERIT"
    if total == 0:
        return "UNUSED"
    if total < 10:
        return "LOW"
    if total < 100:
        return "ACTIVE"
    return "HEAVY"


def _audit_collect(
    client,  # OdooClient
    source_filter: str,
    private_names: set[str] | None,
    model_detail: bool,
) -> list[dict]:
    """Walk installed modules + models, return per-module usage record."""
    modules = client.search_read(
        "ir.module.module",
        domain=[("state", "in", ["installed", "to upgrade"])],
        fields=["name", "shortdesc", "state"],
        limit=0,
        order="name",
    )

    # Map module -> list of model.ids via ir.model.data
    mapping_rows = client.search_read(
        "ir.model.data",
        domain=[("model", "=", "ir.model")],
        fields=["module", "res_id"],
        limit=0,
    )
    by_module: dict[str, list[int]] = {}
    for row in mapping_rows:
        by_module.setdefault(row["module"], []).append(row["res_id"])

    # Collect all referenced ir.model records (transient flag + model name)
    all_model_ids = sorted({mid for ids in by_module.values() for mid in ids})
    model_info_map: dict[int, dict] = {}
    if all_model_ids:
        for chunk_start in range(0, len(all_model_ids), 1000):
            chunk = all_model_ids[chunk_start : chunk_start + 1000]
            info = client.search_read(
                "ir.model",
                domain=[("id", "in", chunk)],
                fields=["id", "model", "transient"],
                limit=0,
            )
            for mi in info:
                model_info_map[mi["id"]] = mi

    results = []
    for mod in modules:
        name = mod["name"]
        is_private = private_names is not None and name in private_names
        if source_filter == "private" and not is_private:
            continue
        if source_filter == "oca" and is_private:
            continue

        model_ids = by_module.get(name, [])
        persistent_models = [
            model_info_map[mid]
            for mid in model_ids
            if mid in model_info_map and not model_info_map[mid].get("transient")
        ]

        per_model_rows = []
        total = 0
        for mi in persistent_models:
            model_name = mi["model"]
            try:
                cnt = client.search_count(model_name, [])
            except Exception:
                cnt = None
            if cnt is not None:
                total += cnt
            per_model_rows.append({"model": model_name, "rows": cnt})

        status = _audit_classify(total, bool(persistent_models))

        record = {
            "module": name,
            "shortdesc": mod.get("shortdesc") or "",
            "source": "private" if is_private else "oca/core",
            "persistent_model_count": len(persistent_models),
            "total_rows": total,
            "status": status,
        }
        if model_detail:
            record["models"] = per_model_rows
        results.append(record)

    return results


def _audit_render_markdown(records: list[dict], profile_label: str) -> str:
    lines = [
        f"# Module Usage Audit — `{profile_label}`",
        "",
        "Row counts per persistent model (ignoring transient wizards).",
        "Status buckets: HEAVY (≥100) · ACTIVE (10–99) · LOW (1–9) · "
        "UNUSED (0 with persistent tables) · INHERIT (no persistent tables).",
        "",
        "## Summary",
        "",
        "| Module | Source | Models | Rows | Status |",
        "|--------|--------|-------:|-----:|--------|",
    ]
    ordered = sorted(
        records,
        key=lambda r: (
            _AUDIT_STATUS_ORDER.get(r["status"], 99),
            -r["total_rows"],
            r["module"],
        ),
    )
    for r in ordered:
        lines.append(
            f"| `{r['module']}` | {r['source']} | {r['persistent_model_count']} "
            f"| {r['total_rows']:,} | **{r['status']}** |"
        )

    if any("models" in r for r in ordered):
        lines += ["", "## Per-Model Detail (UNUSED + LOW)", ""]
        for r in ordered:
            if r["status"] not in ("UNUSED", "LOW"):
                continue
            models = r.get("models") or []
            lines += [f"### `{r['module']}` · {r['status']}", ""]
            if not models:
                lines += ["_No persistent models._", ""]
                continue
            lines += ["| Model | Rows |", "|-------|-----:|"]
            for m in sorted(models, key=lambda m: (-(m["rows"] or 0), m["model"])):
                rows_s = f"{m['rows']:,}" if m["rows"] is not None else "err"
                lines.append(f"| `{m['model']}` | {rows_s} |")
            lines.append("")

    return "\n".join(lines)


@app.command("audit")
def audit(
    ctx: typer.Context,
    source: Annotated[
        str,
        typer.Option(
            "--source",
            "-s",
            help="Filter by source: all | private | oca",
            case_sensitive=False,
        ),
    ] = "all",
    status_filter: Annotated[
        str | None,
        typer.Option(
            "--status",
            help="Only show modules matching status (UNUSED/LOW/ACTIVE/HEAVY/INHERIT). Comma-separated for multiple.",
        ),
    ] = None,
    model_detail: Annotated[
        bool,
        typer.Option(
            "--model-detail/--no-model-detail",
            help="Include per-model row counts (slower).",
        ),
    ] = False,
    markdown_out: Annotated[
        Path | None,
        typer.Option(
            "--markdown",
            "-m",
            help="Write full markdown report to this path.",
        ),
    ] = None,
    private_root: Annotated[
        Path | None,
        typer.Option(
            "--private-root",
            help="Path to src/private (auto-detected from project_root config).",
        ),
    ] = None,
) -> None:
    """Audit installed modules by row count — find simplification candidates.

    Walks ``ir.module.module`` + ``ir.model.data`` + ``ir.model`` and runs
    ``search_count`` on each non-transient model to measure real usage.

    Modules classified as:
      - HEAVY   (>=100 rows) — core usage, keep
      - ACTIVE  (10–99)      — in use
      - LOW     (1–9)        — review
      - UNUSED  (0, has tables) — strong uninstall candidate
      - INHERIT (no persistent tables) — bridge / extension only

    Examples:
        kctl-odoo modules audit
        kctl-odoo modules audit --source private --status UNUSED,LOW
        kctl-odoo modules audit -m report.md --model-detail
        kctl-odoo -p local-tpp-odoo-erp modules audit -m tpp.md
    """
    actx: AppContext = ctx.obj
    out = actx.output
    c = actx.client

    source = source.lower()
    if source not in {"all", "private", "oca"}:
        out.error("--source must be all, private, or oca")
        raise typer.Exit(2)

    # Resolve private module names
    private_names: set[str] | None = None
    if source in {"private", "all"}:
        root = private_root
        if root is None:
            try:
                from kctl_odoo.core.utils import find_project_root

                candidate = find_project_root() / "src" / "private"
                if candidate.is_dir():
                    root = candidate
            except Exception:
                pass
        if root is None:
            cwd_candidate = Path.cwd() / "src" / "private"
            if cwd_candidate.is_dir():
                root = cwd_candidate
        if root is None:
            if source == "private":
                out.warn("Could not auto-detect src/private — pass --private-root PATH.")
            private_names = set()
        else:
            private_names = {m.parent.name for m in root.rglob("__manifest__.py")}

    out.info("Collecting module usage (this may take a minute)...")
    records = _audit_collect(c, source, private_names, model_detail)

    # Apply status filter
    if status_filter:
        wanted = {s.strip().upper() for s in status_filter.split(",")}
        records = [r for r in records if r["status"] in wanted]

    # Summary counts
    by_status: dict[str, int] = {}
    for r in records:
        by_status[r["status"]] = by_status.get(r["status"], 0) + 1

    # Sort for display
    records.sort(
        key=lambda r: (
            _AUDIT_STATUS_ORDER.get(r["status"], 99),
            -r["total_rows"],
            r["module"],
        ),
    )

    profile_label = getattr(c, "database", "") or "odoo"

    # Markdown output
    if markdown_out is not None:
        md = _audit_render_markdown(records, profile_label)
        markdown_out.write_text(md)
        out.success(f"Wrote markdown report: {markdown_out}")

    # Table output
    rows = []
    json_data = []
    for r in records:
        rows.append(
            [
                r["module"],
                r["source"],
                str(r["persistent_model_count"]),
                f"{r['total_rows']:,}",
                r["status"],
            ]
        )
        json_data.append(r)

    out.table(
        f"Module Usage Audit — {profile_label} ({len(records)})",
        [
            ("Module", ""),
            ("Source", "dim"),
            ("Models", ""),
            ("Rows", ""),
            ("Status", ""),
        ],
        rows,
        data_for_json=json_data,
    )

    if by_status:
        summary_line = " · ".join(
            f"{k}={by_status[k]}" for k in sorted(by_status, key=lambda s: _AUDIT_STATUS_ORDER.get(s, 99))
        )
        out.info(f"Status counts: {summary_line}")


# ---------------------------------------------------------------------------
# audit-report: comprehensive Excel audit workbook
# ---------------------------------------------------------------------------


def _audit_report_collect(client, source_filter: str, private_names: set[str] | None, out) -> dict:
    """Collect comprehensive audit data for all installed modules."""
    out.info("Collecting installed modules...")
    modules = client.search_read(
        "ir.module.module",
        domain=[("state", "in", ["installed", "to upgrade"])],
        fields=["name", "shortdesc", "state", "installed_version", "author", "dependencies_id"],
        limit=0,
        order="name",
    )

    filtered = []
    for mod in modules:
        name = mod["name"]
        is_private = private_names is not None and name in private_names
        if source_filter == "private" and not is_private:
            continue
        if source_filter == "oca" and is_private:
            continue
        mod["_source"] = "private" if is_private else "oca/core"
        filtered.append(mod)

    mod_names = {m["name"] for m in filtered}

    out.info("Collecting dependencies...")
    all_dep_ids = []
    for mod in filtered:
        all_dep_ids.extend(mod.get("dependencies_id") or [])
    dep_map: dict[str, list[str]] = {m["name"]: [] for m in filtered}
    if all_dep_ids:
        for chunk_start in range(0, len(all_dep_ids), 1000):
            chunk = all_dep_ids[chunk_start : chunk_start + 1000]
            deps = client.search_read(
                "ir.module.module.dependency",
                domain=[("id", "in", chunk)],
                fields=["module_id", "name", "depend_id"],
                limit=0,
            )
            for d in deps:
                parent = d["module_id"][1] if d["module_id"] else None
                if parent and parent in dep_map:
                    dep_map[parent].append(d["name"])

    out.info("Collecting reverse dependencies...")
    rdep_map: dict[str, list[str]] = {m["name"]: [] for m in filtered}
    rdeps_raw = client.search_read(
        "ir.module.module.dependency",
        domain=[("name", "in", list(mod_names))],
        fields=["module_id", "name"],
        limit=0,
    )
    for d in rdeps_raw:
        dep_name = d["name"]
        parent_name = d["module_id"][1] if d["module_id"] else None
        if dep_name in rdep_map and parent_name:
            rdep_map[dep_name].append(parent_name)

    out.info("Collecting model ownership...")
    mapping_rows = client.search_read(
        "ir.model.data",
        domain=[("model", "=", "ir.model")],
        fields=["module", "res_id"],
        limit=0,
    )
    models_by_module: dict[str, list[int]] = {}
    for row in mapping_rows:
        models_by_module.setdefault(row["module"], []).append(row["res_id"])

    all_model_ids = sorted({mid for ids in models_by_module.values() for mid in ids})
    model_info: dict[int, dict] = {}
    if all_model_ids:
        for cs in range(0, len(all_model_ids), 1000):
            chunk = all_model_ids[cs : cs + 1000]
            info = client.search_read(
                "ir.model",
                domain=[("id", "in", chunk)],
                fields=["id", "model", "name", "transient", "field_id"],
                limit=0,
            )
            for mi in info:
                model_info[mi["id"]] = mi

    out.info("Collecting views, menus, reports, groups, ACL, record rules...")
    imd_counts_models = [
        ("ir.ui.view", "views"),
        ("ir.ui.menu", "menus"),
        ("ir.actions.report", "reports"),
        ("res.groups", "groups"),
        ("ir.model.access", "acl_rules"),
        ("ir.rule", "record_rules"),
    ]
    counts_by_module: dict[str, dict[str, int]] = {m["name"]: {} for m in filtered}
    for odoo_model, key in imd_counts_models:
        imd_rows = client.search_read(
            "ir.model.data",
            domain=[("model", "=", odoo_model), ("module", "in", list(mod_names))],
            fields=["module"],
            limit=0,
        )
        per_mod: dict[str, int] = {}
        for row in imd_rows:
            per_mod[row["module"]] = per_mod.get(row["module"], 0) + 1
        for mname in mod_names:
            counts_by_module.setdefault(mname, {})[key] = per_mod.get(mname, 0)

    out.info("Counting rows and last activity per model (this may take a minute)...")
    model_details: list[dict] = []
    module_totals: dict[str, int] = {}
    module_last_activity: dict[str, str] = {}

    for mod in filtered:
        name = mod["name"]
        m_ids = models_by_module.get(name, [])
        total = 0
        for mid in m_ids:
            mi = model_info.get(mid)
            if not mi:
                continue
            model_name = mi["model"]
            is_transient = mi.get("transient", False)
            cnt = 0
            last_write = None
            if not is_transient:
                try:
                    cnt = client.search_count(model_name, [])
                except Exception:
                    cnt = 0
                if cnt > 0:
                    try:
                        latest = client.search_read(
                            model_name,
                            domain=[],
                            fields=["write_date"],
                            limit=1,
                            order="write_date desc",
                        )
                        if latest:
                            last_write = latest[0].get("write_date")
                    except Exception:
                        pass
                total += cnt

            field_count = len(mi.get("field_id") or [])
            model_details.append(
                {
                    "module": name,
                    "model": model_name,
                    "description": mi.get("name") or "",
                    "transient": is_transient,
                    "row_count": cnt,
                    "last_write": last_write or "",
                    "field_count": field_count,
                }
            )

            if last_write:
                prev = module_last_activity.get(name, "")
                if last_write > prev:
                    module_last_activity[name] = last_write

        module_totals[name] = total

    out.info("Collecting group access matrix...")
    group_access: list[dict] = []
    group_imd = client.search_read(
        "ir.model.data",
        domain=[("model", "=", "res.groups"), ("module", "in", list(mod_names))],
        fields=["module", "res_id"],
        limit=0,
    )
    group_ids = [r["res_id"] for r in group_imd]
    group_to_module = {r["res_id"]: r["module"] for r in group_imd}
    if group_ids:
        groups_data = client.search_read(
            "res.groups",
            domain=[("id", "in", group_ids)],
            fields=["id", "full_name", "users"],
            limit=0,
        )
        acl_rows = client.search_read(
            "ir.model.access",
            domain=[("group_id", "in", group_ids)],
            fields=["group_id", "model_id", "perm_read", "perm_write", "perm_create", "perm_unlink"],
            limit=0,
        )
        acl_by_group: dict[int, list[str]] = {}
        for a in acl_rows:
            gid = a["group_id"][0] if a["group_id"] else 0
            model_label = a["model_id"][1] if a["model_id"] else "?"
            perms = []
            if a.get("perm_read"):
                perms.append("R")
            if a.get("perm_write"):
                perms.append("W")
            if a.get("perm_create"):
                perms.append("C")
            if a.get("perm_unlink"):
                perms.append("D")
            acl_by_group.setdefault(gid, []).append(f"{model_label} ({''.join(perms)})")
        for g in groups_data:
            group_access.append(
                {
                    "module": group_to_module.get(g["id"], "?"),
                    "group": g.get("full_name") or "",
                    "users_count": len(g.get("users") or []),
                    "models_access": "; ".join(acl_by_group.get(g["id"], [])),
                }
            )

    summary = []
    for mod in filtered:
        name = mod["name"]
        total = module_totals.get(name, 0)
        persistent = [
            model_info[mid]
            for mid in models_by_module.get(name, [])
            if mid in model_info and not model_info[mid].get("transient")
        ]
        status = _audit_classify(total, bool(persistent))
        if status == "UNUSED":
            rec = "UNINSTALL candidate"
        elif status == "LOW":
            rec = "REVIEW"
        elif status == "INHERIT":
            rec = "KEEP (bridge)"
        else:
            rec = "KEEP"

        c_data = counts_by_module.get(name, {})
        summary.append(
            {
                "module": name,
                "display_name": mod.get("shortdesc") or "",
                "source": mod["_source"],
                "version": mod.get("installed_version") or "",
                "status": status,
                "total_models": len(models_by_module.get(name, [])),
                "total_rows": total,
                "views": c_data.get("views", 0),
                "menus": c_data.get("menus", 0),
                "reports": c_data.get("reports", 0),
                "groups": c_data.get("groups", 0),
                "acl_rules": c_data.get("acl_rules", 0),
                "record_rules": c_data.get("record_rules", 0),
                "dependencies": ", ".join(dep_map.get(name, [])),
                "reverse_deps": ", ".join(rdep_map.get(name, [])),
                "last_activity": module_last_activity.get(name, ""),
                "recommendation": rec,
            }
        )

    summary.sort(
        key=lambda r: (
            _AUDIT_STATUS_ORDER.get(r["status"], 99),
            -r["total_rows"],
            r["module"],
        ),
    )

    return {
        "summary": summary,
        "model_details": model_details,
        "group_access": group_access,
    }


def _write_audit_xlsx(data: dict, output_path: Path, profile_label: str) -> None:
    """Write the audit workbook."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    hdr_font = Font(bold=True, color="FFFFFF", size=11)
    hdr_fill = PatternFill("solid", fgColor="2F5496")
    unused_fill = PatternFill("solid", fgColor="FFC7CE")
    low_fill = PatternFill("solid", fgColor="FFEB9C")
    heavy_fill = PatternFill("solid", fgColor="C6EFCE")
    border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )
    status_fills = {
        "UNUSED": unused_fill,
        "LOW": low_fill,
        "HEAVY": heavy_fill,
    }

    def write_headers(ws, headers, widths=None):
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = hdr_font
            cell.fill = hdr_fill
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
        if widths:
            for i, w in enumerate(widths, 1):
                ws.column_dimensions[get_column_letter(i)].width = w
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    # ── Sheet 1: Module Summary ──
    ws1 = wb.active
    ws1.title = "Module Summary"
    headers1 = [
        "Module",
        "Display Name",
        "Source",
        "Version",
        "Status",
        "Models",
        "Total Rows",
        "Views",
        "Menus",
        "Reports",
        "Groups",
        "ACL Rules",
        "Record Rules",
        "Dependencies",
        "Reverse Deps",
        "Last Activity",
        "Recommendation",
    ]
    widths1 = [30, 30, 12, 14, 12, 8, 12, 8, 8, 8, 8, 10, 12, 40, 30, 20, 18]
    write_headers(ws1, headers1, widths1)

    for row_idx, r in enumerate(data["summary"], 2):
        values = [
            r["module"],
            r["display_name"],
            r["source"],
            r["version"],
            r["status"],
            r["total_models"],
            r["total_rows"],
            r["views"],
            r["menus"],
            r["reports"],
            r["groups"],
            r["acl_rules"],
            r["record_rules"],
            r["dependencies"],
            r["reverse_deps"],
            r["last_activity"],
            r["recommendation"],
        ]
        for col_idx, val in enumerate(values, 1):
            cell = ws1.cell(row=row_idx, column=col_idx, value=val)
            cell.border = border
        fill = status_fills.get(r["status"])
        if fill:
            for col_idx in range(1, len(values) + 1):
                ws1.cell(row=row_idx, column=col_idx).fill = fill

    # ── Sheet 2: Model Detail ──
    ws2 = wb.create_sheet("Model Detail")
    headers2 = ["Module", "Model", "Description", "Transient", "Row Count", "Last Write", "Fields"]
    widths2 = [30, 35, 30, 10, 12, 20, 8]
    write_headers(ws2, headers2, widths2)

    sorted_models = sorted(
        data["model_details"],
        key=lambda m: (m["module"], -m["row_count"], m["model"]),
    )
    for row_idx, m in enumerate(sorted_models, 2):
        values = [
            m["module"],
            m["model"],
            m["description"],
            "Yes" if m["transient"] else "No",
            m["row_count"],
            m["last_write"],
            m["field_count"],
        ]
        for col_idx, val in enumerate(values, 1):
            cell = ws2.cell(row=row_idx, column=col_idx, value=val)
            cell.border = border

    # ── Sheet 3: Group & Access Matrix ──
    ws3 = wb.create_sheet("Group Access Matrix")
    headers3 = ["Module", "Group", "Users", "Models with Access (RWCD)"]
    widths3 = [30, 40, 8, 80]
    write_headers(ws3, headers3, widths3)

    sorted_groups = sorted(
        data["group_access"],
        key=lambda g: (g["module"], g["group"]),
    )
    for row_idx, g in enumerate(sorted_groups, 2):
        values = [g["module"], g["group"], g["users_count"], g["models_access"]]
        for col_idx, val in enumerate(values, 1):
            cell = ws3.cell(row=row_idx, column=col_idx, value=val)
            cell.border = border

    wb.save(output_path)


@app.command("audit-report")
def audit_report(
    ctx: typer.Context,
    source: Annotated[
        str,
        typer.Option(
            "--source",
            "-s",
            help="Filter by source: all | private | oca",
            case_sensitive=False,
        ),
    ] = "all",
    output: Annotated[
        Path,
        typer.Option(
            "--output",
            "-o",
            help="Output Excel file path.",
        ),
    ] = Path("modules-audit-report.xlsx"),
) -> None:
    """Comprehensive module audit report as Excel workbook.

    Produces a multi-sheet .xlsx with:
      - Sheet 1: Module Summary (rows, views, menus, reports, groups, ACL, deps, recommendation)
      - Sheet 2: Model Detail (per-model row counts, last write date, field counts)
      - Sheet 3: Group & Access Matrix (groups per module, user counts, CRUD access)

    Examples:
        kctl-odoo modules audit-report
        kctl-odoo modules audit-report --source private -o tpp-audit.xlsx
        kctl-odoo -p idtpp-tpp-odoo-erp modules audit-report -o tpp-erp-audit.xlsx
    """
    actx: AppContext = ctx.obj
    out = actx.output
    c = actx.client

    source = source.lower()
    if source not in {"all", "private", "oca"}:
        out.error("--source must be all, private, or oca")
        raise typer.Exit(2)

    private_names: set[str] | None = None
    if source in {"private", "all"}:
        root = None
        try:
            from kctl_odoo.core.utils import find_project_root

            candidate = find_project_root() / "src" / "private"
            if candidate.is_dir():
                root = candidate
        except Exception:
            pass
        if root is None:
            cwd_candidate = Path.cwd() / "src" / "private"
            if cwd_candidate.is_dir():
                root = cwd_candidate
        if root is None:
            if source == "private":
                out.warn("Could not auto-detect src/private — pass --private-root PATH.")
            private_names = set()
        else:
            private_names = {m.parent.name for m in root.rglob("__manifest__.py")}

    data = _audit_report_collect(c, source, private_names, out)

    profile_label = getattr(c, "database", "") or "odoo"
    _write_audit_xlsx(data, output, profile_label)

    summary = data["summary"]
    by_status: dict[str, int] = {}
    for r in summary:
        by_status[r["status"]] = by_status.get(r["status"], 0) + 1

    out.success(
        f"Wrote {output} — {len(summary)} modules, "
        f"{len(data['model_details'])} models, "
        f"{len(data['group_access'])} group-access rows"
    )
    if by_status:
        summary_line = " · ".join(
            f"{k}={by_status[k]}" for k in sorted(by_status, key=lambda s: _AUDIT_STATUS_ORDER.get(s, 99))
        )
        out.info(f"Status: {summary_line}")
