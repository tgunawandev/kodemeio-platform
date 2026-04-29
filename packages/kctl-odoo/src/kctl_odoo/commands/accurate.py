"""Accurate → Odoo 1-click migration commands.

Thin wrappers over accurate.company JSON-RPC methods. Business logic
lives in the Odoo addon (`accurate_integration`); this CLI is a skin
over the same durable state machine that the UI drives.

See docs/superpowers/specs/2026-04-19-accurate-1click-migration-spine-design.md
"""

from __future__ import annotations

import json
from datetime import date, datetime
from pathlib import Path
from typing import Annotated, Optional

import httpx
import typer

from kctl_odoo.core.callbacks import AppContext


def _bump_client_timeout(actx: AppContext, seconds: float = 1800.0) -> None:
    """Extend httpx client read-timeout for long-running Accurate RPCs.

    Default kctl-odoo client has a 30s timeout — fine for most RPCs
    but the Accurate-side methods (gap_sync, post_draft_imports,
    run_parity, full-remediate) can fetch hundreds of records from
    Accurate's API in a single JSON-RPC call. Bump to 30 min so the
    client doesn't disconnect mid-import.
    """
    try:
        actx.client._client.timeout = httpx.Timeout(seconds)
    except Exception:  # noqa: BLE001
        pass


app = typer.Typer(help="Accurate → Odoo 1-click migration.")


def _resolve_accurate_company(client: object, identifier: str) -> dict:
    """Find an accurate.company by slug or numeric ID. Returns full record."""
    recs = client.search_read(  # type: ignore[attr-defined]
        "accurate.company",
        domain=[
            "|",
            ("slug", "=", identifier),
            ("id", "=", int(identifier)) if identifier.isdigit() else ("id", "=", 0),
        ],
        fields=[
            "id",
            "name",
            "slug",
            "db_id",
            "state",
            "current_phase",
            "current_phase_state",
            "progress_percent",
            "last_sync_at",
            "sync_enabled",
            "odoo_company_id",
        ],
        limit=1,
    )
    if not recs:
        raise typer.BadParameter(f"Accurate company not found: {identifier}")
    return recs[0]


# --- companies ---------------------------------------------------------


companies_app = typer.Typer(help="Manage Accurate company registrations.")


@companies_app.command("list")
def companies_list(
    ctx: typer.Context,
    state: Annotated[
        str | None,
        typer.Option("--state", "-s", help="Filter by state"),
    ] = None,
    limit: Annotated[int, typer.Option("--limit", "-l")] = 80,
) -> None:
    """List registered Accurate companies."""
    actx: AppContext = ctx.obj
    out = actx.output
    c = actx.client

    domain = []
    if state:
        domain.append(("state", "=", state))

    recs = c.search_read(
        "accurate.company",
        domain=domain,
        fields=[
            "id",
            "slug",
            "name",
            "state",
            "current_phase",
            "progress_percent",
            "last_sync_at",
        ],
        limit=limit,
        order="name",
    )
    rows = [
        [
            str(r["id"]),
            r["slug"],
            r["name"],
            r["state"],
            r.get("current_phase") or "-",
            f"{r.get('progress_percent', 0)}%",
            str(r.get("last_sync_at") or "-"),
        ]
        for r in recs
    ]
    out.table(
        f"Accurate Companies ({len(recs)})",
        [
            ("ID", "cyan"),
            ("Slug", ""),
            ("Name", ""),
            ("State", ""),
            ("Phase", "dim"),
            ("Progress", "dim"),
            ("Last Sync", "dim"),
        ],
        rows,
        data_for_json=recs,
    )


@companies_app.command("show")
def companies_show(
    ctx: typer.Context,
    identifier: Annotated[str, typer.Argument(help="Slug or ID")],
) -> None:
    """Show details for one Accurate company."""
    actx: AppContext = ctx.obj
    out = actx.output
    c = actx.client

    rec = _resolve_accurate_company(c, identifier)
    # phase history
    phases = c.search_read(
        "accurate.company.phase",
        domain=[("company_id", "=", rec["id"])],
        fields=[
            "phase_code",
            "state",
            "started_at",
            "completed_at",
            "duration_s",
            "triggered_by",
        ],
        order="create_date desc",
        limit=20,
    )

    out.header(f"{rec['name']} ({rec['slug']})")
    for k, v in [
        ("State", rec["state"]),
        ("Current Phase", rec.get("current_phase") or "-"),
        ("Phase State", rec.get("current_phase_state") or "-"),
        ("Progress", f"{rec.get('progress_percent', 0)}%"),
        ("Sync Enabled", "yes" if rec.get("sync_enabled") else "no"),
        ("Last Sync", str(rec.get("last_sync_at") or "-")),
    ]:
        out.kv(k, v)
    if phases:
        rows = [
            [
                p["phase_code"],
                p["state"],
                str(p.get("started_at") or "-"),
                str(p.get("completed_at") or "-"),
                f"{p.get('duration_s', 0):.1f}s",
            ]
            for p in phases
        ]
        out.table(
            "Recent Phases",
            [("Phase", ""), ("State", ""), ("Started", "dim"), ("Finished", "dim"), ("Dur", "dim")],
            rows,
        )


# --- phases ------------------------------------------------------------


@app.command("preflight")
def preflight(
    ctx: typer.Context,
    identifier: Annotated[str, typer.Argument(help="Accurate company slug or ID")],
) -> None:
    """Run Phase 1 Preflight checks."""
    actx: AppContext = ctx.obj
    out = actx.output
    c = actx.client

    rec = _resolve_accurate_company(c, identifier)
    out.info(f"Running preflight for {rec['slug']}...")
    report = c.execute_kw("accurate.company", "action_run_preflight", [[rec["id"]]])

    # report is either the return dict or None (if write-only model method)
    # action_run_preflight returns the report dict per our implementation
    if not isinstance(report, dict):
        # Refresh to read the persisted summary
        updated = c.search_read(
            "accurate.company.phase",
            domain=[
                ("company_id", "=", rec["id"]),
                ("phase_code", "=", "preflight"),
            ],
            fields=["state", "summary_json"],
            order="create_date desc",
            limit=1,
        )
        if updated and updated[0].get("summary_json"):
            report = json.loads(updated[0]["summary_json"])
        else:
            report = {"checks": [], "all_blockers_passed": False}

    rows = []
    for chk in report.get("checks", []):
        rows.append(
            [
                chk["name"],
                "✓" if chk["passed"] else "✗",
                chk["severity"],
                (chk.get("message") or "")[:80],
            ]
        )
    out.table(
        "Preflight Results",
        [("Check", ""), ("Result", ""), ("Severity", "dim"), ("Message", "dim")],
        rows,
    )
    if not report.get("all_blockers_passed"):
        out.error("Preflight BLOCKED — resolve issues and re-run.")
        raise typer.Exit(1)
    out.success("Preflight PASSED — ready for Setup phase.")


@app.command("setup")
def setup(
    ctx: typer.Context,
    identifier: Annotated[str, typer.Argument(help="Accurate company slug or ID")],
    coa_template: Annotated[
        str,
        typer.Option(
            "--coa-template",
            help="Odoo CoA template XML ID to install",
        ),
    ] = "l10n_id.l10n_id_chart_template_amd",
) -> None:
    """Run Phase 2 Setup — install CoA + resolve PPh accounts."""
    actx: AppContext = ctx.obj
    out = actx.output
    c = actx.client

    rec = _resolve_accurate_company(c, identifier)
    out.info(f"Running setup for {rec['slug']} (CoA: {coa_template})...")
    summary = c.execute_kw(
        "accurate.company",
        "action_run_setup",
        [[rec["id"]]],
        {"coa_template_xmlid": coa_template},
    )
    out.success("Setup PASSED.")
    for k, v in [
        ("CoA Installed", str(summary.get("coa_installed", "-"))),
        ("Mapping Rows Seeded", str(summary.get("mapping_rows_seeded", 0))),
    ]:
        out.kv(k, v)
    pph = summary.get("pph_accounts") or {}
    if pph:
        rows = []
        for field_name, val in pph.items():
            if val:
                rows.append([field_name, val.get("code", "-"), val.get("name", "-")])
            else:
                rows.append([field_name, "-", "NOT FOUND"])
        out.table(
            "PPh Accounts Resolved",
            [("Type", ""), ("Code", ""), ("Name", "dim")],
            rows,
        )


@app.command("phases")
def phases(
    ctx: typer.Context,
    identifier: Annotated[str, typer.Argument(help="Accurate company slug or ID")],
    phase: Annotated[
        str | None,
        typer.Option("--phase", help="Filter by phase_code"),
    ] = None,
    state: Annotated[
        str | None,
        typer.Option("--state", help="Filter by state"),
    ] = None,
) -> None:
    """Show phase execution history for one Accurate company."""
    actx: AppContext = ctx.obj
    out = actx.output
    c = actx.client

    rec = _resolve_accurate_company(c, identifier)
    domain = [("company_id", "=", rec["id"])]
    if phase:
        domain.append(("phase_code", "=", phase))
    if state:
        domain.append(("state", "=", state))
    phases_ = c.search_read(
        "accurate.company.phase",
        domain=domain,
        fields=[
            "phase_code",
            "state",
            "started_at",
            "completed_at",
            "duration_s",
            "triggered_by",
            "error_log",
        ],
        order="create_date desc",
        limit=50,
    )
    rows = []
    for p in phases_:
        rows.append(
            [
                p["phase_code"],
                p["state"],
                str(p.get("started_at") or "-"),
                f"{p.get('duration_s', 0):.1f}s",
                (p.get("error_log") or "")[:50],
            ]
        )
    out.table(
        f"Phase History — {rec['slug']} ({len(phases_)})",
        [("Phase", ""), ("State", ""), ("Started", "dim"), ("Dur", "dim"), ("Error", "dim")],
        rows,
        data_for_json=phases_,
    )


# --- cutover ------------------------------------------------------------


@app.command("cutover")
def cutover(
    ctx: typer.Context,
    identifier: Annotated[str, typer.Argument(help="Slug or ID")],
    cutover_date: Annotated[str, typer.Option("--date", help="YYYY-MM-DD")],
    mode: Annotated[str, typer.Option("--mode")] = "cutover",
) -> None:
    """Phase 3 Cutover Config — commits the cutover date + mode."""
    actx: AppContext = ctx.obj
    out = actx.output
    c = actx.client
    rec = _resolve_accurate_company(c, identifier)
    out.info(f"Configuring cutover for {rec['slug']} — date={cutover_date} mode={mode}")
    summary = c.execute_kw(
        "accurate.company",
        "action_run_cutover_config",
        [[rec["id"]]],
        {"cutover_date": cutover_date, "mode": mode},
    )
    out.success("Cutover config committed.")
    for k, v in [
        ("Cutover Date", summary.get("cutover_date", "-")),
        ("Current FY Start", summary.get("current_fy_start", "-")),
        ("Mode", summary.get("mode", "-")),
    ]:
        out.kv(k, v)


# --- foundation --------------------------------------------------------


@app.command("foundation")
def foundation(
    ctx: typer.Context,
    identifier: Annotated[str, typer.Argument(help="Slug or ID")],
) -> None:
    """Phase 4 Foundation Sync."""
    actx: AppContext = ctx.obj
    out = actx.output
    c = actx.client
    rec = _resolve_accurate_company(c, identifier)
    out.info(f"Running Foundation sync for {rec['slug']}...")
    try:
        summary = c.execute_kw("accurate.company", "action_run_foundation", [[rec["id"]]])
    except Exception as exc:
        out.error(f"Foundation failed: {exc}")
        raise typer.Exit(1)
    out.success("Foundation PASSED.")
    mods = summary.get("modules", {})
    rows = [[name, str(v)] for name, v in mods.items()]
    out.table("Foundation Modules", [("Module", ""), ("Result", "")], rows)


# --- transactions ------------------------------------------------------


@app.command("transactions")
def transactions(
    ctx: typer.Context,
    identifier: Annotated[str, typer.Argument(help="Slug or ID")],
    phase: Annotated[str, typer.Option("--phase", help="prior|current|both|waves")] = "waves",
    sync: Annotated[
        bool,
        typer.Option(
            "--sync/--async",
            help=(
                "--sync runs the import inline (blocking, results known on "
                "return — good for CI and small tenants). --async enqueues "
                "via queue_job (default; background with OCA queue worker)."
            ),
        ),
    ] = False,
    skip_bs_opening: Annotated[
        bool,
        typer.Option(
            "--skip-bs-opening/--with-bs-opening",
            help="Skip Wave 0 (opening BS JE) for companies that started in the current FY.",
        ),
    ] = True,
    timeout: Annotated[
        int,
        typer.Option(
            "--timeout",
            help=("HTTP client timeout in seconds. Default 900 (15 min). Raise for very large tenants."),
        ),
    ] = 900,
) -> None:
    """Import transactions from Accurate.

    Default mode (--phase waves) uses the wave-based pipeline:
    invoices → payments → bank → JVs, with pre-flight gates,
    auto-reconciliation, and 12-check validation.

    Legacy modes (prior/current/both) call the old phase methods.

    Examples::

        kctl-odoo accurate transactions mandira-copy-frozen
        kctl-odoo accurate transactions mandira-copy-frozen --phase waves --sync
        kctl-odoo accurate transactions my-company --with-bs-opening
    """
    import httpx

    actx: AppContext = ctx.obj
    out = actx.output
    c = actx.client
    c._client.timeout = httpx.Timeout(timeout)
    rec = _resolve_accurate_company(c, identifier)

    if phase == "waves":
        out.info(f"Running wave-based import for {rec['slug']} (timeout={timeout}s)...")
        result = c.execute_kw(
            "accurate.company",
            "action_import_transactions",
            [[rec["id"]]],
            {"skip_bs_opening": skip_bs_opening, "skip_optional": True},
        )
        if isinstance(result, dict):
            validation = result.get("validation", {})
            waves = result.get("waves", [])
            total_moves = sum(w.get("posted", 0) for w in waves)
            tb = next((c for c in validation.get("checks", []) if c["name"] == "tb_parity"), {})
            status = "PASSED ✓" if validation.get("passed") else "FAILED ✗"
            out.success(f"Import complete: {total_moves} moves posted")
            out.info(f"TB: {tb.get('value', '?')}")
            out.info(f"Validation: {status}")
            if not validation.get("passed"):
                for check in validation.get("checks", []):
                    if not check["passed"]:
                        out.warning(f"  ✗ {check['name']}: {check['detail'][:60]}")
        else:
            out.success("Import complete.")
        return

    # Legacy modes
    if phase in ("prior", "both"):
        out.info(f"Running Transactions — Prior-FY Opens ({'sync' if sync else 'async'}, timeout={timeout}s)...")
        c.execute_kw(
            "accurate.company",
            "action_run_transactions_prior",
            [[rec["id"]]],
            {"context": {"accurate_sync_inline": sync}},
        )
        out.success("Prior-FY opens imported.")
    if phase in ("current", "both"):
        out.info(f"Running Transactions — Current FY ({'sync' if sync else 'async'}, timeout={timeout}s)...")
        c.execute_kw(
            "accurate.company",
            "action_run_transactions_current",
            [[rec["id"]]],
            {"context": {"accurate_sync_inline": sync}},
        )
        out.success("Current-FY transactions imported.")


# --- verify / sign-off / go-live ---------------------------------------


@app.command("verify")
def verify(
    ctx: typer.Context,
    identifier: Annotated[str, typer.Argument(help="Slug or ID")],
) -> None:
    """Phase 7 Verification — generates parity report."""
    actx: AppContext = ctx.obj
    out = actx.output
    c = actx.client
    rec = _resolve_accurate_company(c, identifier)
    out.info(f"Generating verification report for {rec['slug']}...")
    report_id = c.execute_kw("accurate.company", "action_run_verification", [[rec["id"]]])
    # Read report back
    if isinstance(report_id, dict):
        report_id = report_id.get("id")
    elif hasattr(report_id, "id"):
        report_id = report_id.id
    if not isinstance(report_id, int):
        # The method returns the report record — some RPC flavours return {}
        rep = c.search_read(
            "accurate.verification.report",
            domain=[("company_id", "=", rec["id"])],
            fields=["id", "all_passed"],
            limit=1,
            order="create_date desc",
        )
        if not rep:
            out.error("No verification report found.")
            raise typer.Exit(1)
        report_id = rep[0]["id"]
    lines = c.search_read(
        "accurate.verification.line",
        domain=[("report_id", "=", report_id)],
        fields=["check_name", "scope", "accurate_value", "odoo_value", "delta", "passed", "message"],
        order="sequence,id",
    )
    rows = [
        [
            l["check_name"],
            l.get("scope") or "-",
            l.get("accurate_value") or "-",
            l.get("odoo_value") or "-",
            l.get("delta") or "-",
            "✓" if l["passed"] else "✗",
            (l.get("message") or "")[:60],
        ]
        for l in lines
    ]
    out.table(
        f"Verification Report ({len(lines)} checks)",
        [("Check", ""), ("Scope", ""), ("Accurate", ""), ("Odoo", ""), ("Delta", ""), ("Pass", ""), ("Message", "dim")],
        rows,
    )
    failed = [l for l in lines if not l["passed"]]
    if failed:
        out.error(f"{len(failed)} checks FAILED — resolve before sign-off.")
        raise typer.Exit(1)
    out.success("All checks PASSED.")


@app.command("sign-off")
def sign_off(
    ctx: typer.Context,
    identifier: Annotated[str, typer.Argument(help="Slug or ID")],
    confirm: Annotated[bool, typer.Option("--confirm")] = False,
) -> None:
    """Sign off verification — required before go-live."""
    if not confirm:
        raise typer.BadParameter("Add --confirm to proceed (destructive gate).")
    actx: AppContext = ctx.obj
    out = actx.output
    c = actx.client
    rec = _resolve_accurate_company(c, identifier)
    c.execute_kw("accurate.company", "action_sign_off_verification", [[rec["id"]]])
    out.success(f"Verification signed off for {rec['slug']} by {actx.username_override or 'current user'}.")


@app.command("go-live")
def go_live(
    ctx: typer.Context,
    identifier: Annotated[str, typer.Argument(help="Slug or ID")],
) -> None:
    """Phase 8 Go-Live — enable cron sync, state=live."""
    actx: AppContext = ctx.obj
    out = actx.output
    c = actx.client
    rec = _resolve_accurate_company(c, identifier)
    c.execute_kw("accurate.company", "action_run_go_live", [[rec["id"]]])
    out.success(f"{rec['slug']} is LIVE — incremental cron sync enabled.")


@app.command("attachments")
def attachments(
    ctx: typer.Context,
    identifier: Annotated[str, typer.Argument(help="Slug or ID")],
) -> None:
    """Phase 9 Attachments — migrate Accurate attachments to ir.attachment.

    Optional phase. Requires accurate.company.attachments_enabled=True.
    """
    actx: AppContext = ctx.obj
    out = actx.output
    c = actx.client
    rec = _resolve_accurate_company(c, identifier)
    out.info(f"Running attachments phase for {rec['slug']}...")
    try:
        summary = c.execute_kw(
            "accurate.company",
            "action_run_attachments",
            [[rec["id"]]],
        )
    except Exception as exc:
        out.error(f"Attachments failed: {exc}")
        raise typer.Exit(1)
    if summary and summary.get("skipped"):
        out.warn("Attachments phase skipped — attachments_enabled=False.")
    else:
        out.success("Attachments phase complete.")
        if isinstance(summary, dict):
            for k, v in [
                ("Migrated", str(summary.get("migrated_count", 0))),
                ("Skipped", str(summary.get("skipped_count", 0))),
            ]:
                out.kv(k, v)


# --- migrate convenience -----------------------------------------------


@app.command("migrate")
def migrate(
    ctx: typer.Context,
    identifier: Annotated[str, typer.Argument(help="Slug or ID")],
    until: Annotated[str, typer.Option("--until", help="Stop at this phase")] = "verify",
) -> None:
    """Run phases in sequence up to --until (default: verify, stopping before sign-off)."""
    actx: AppContext = ctx.obj
    _bump_client_timeout(actx, seconds=1800.0)
    out = actx.output
    c = actx.client
    rec = _resolve_accurate_company(c, identifier)
    sequence = [
        ("preflight", "action_run_preflight"),
        ("setup", "action_run_setup"),
        ("cutover", "action_run_cutover_config"),
        ("foundation", "action_run_foundation"),
        ("import", "action_import_transactions"),
        ("verify", "action_run_verification"),
        ("go-live", "action_run_go_live"),
    ]
    for label, method in sequence:
        if label == "cutover":
            out.warn("`migrate` cannot run cutover non-interactively — call `cutover` manually first.")
            if not rec.get("current_phase") or rec.get("current_phase") == "preflight":
                out.error("Cutover not yet configured; aborting.")
                raise typer.Exit(2)
            continue
        out.info(f"→ {label}")
        try:
            c.execute_kw("accurate.company", method, [[rec["id"]]])
        except Exception as exc:
            out.error(f"{label} FAILED: {exc}")
            raise typer.Exit(1)
        if label == until:
            out.success(f"Stopped at {until} as requested.")
            return
    out.success("All phases complete.")


# --- clean-slate wipe (18.0.15.2.0) ------------------------------------


@app.command("wipe")
def wipe(
    ctx: typer.Context,
    identifier: Annotated[str, typer.Argument(help="Slug or ID, or 'all'")],
    dry_run: Annotated[
        bool, typer.Option("--dry-run/--execute", help="Preview only (default); use --execute to actually delete")
    ] = True,
    confirm: Annotated[bool, typer.Option("--confirm", help="Required with --execute to proceed")] = False,
) -> None:
    """Wipe migrated transactions while preserving all masters/config.

    Idempotent SQL-based cleanup for clean-slate remigration. Preserves:
    accurate.company config, mappings, parity history, foundation ext-ids
    (customer/vendor/item/glaccount/currency/unit/tax/...), CoA accounts,
    partners, products. Deletes: every account.move tagged with
    x_accurate_migration_source, its lines, partial/full reconciles,
    payment ext-ids, and transaction-module ir.model.data.

    Pass ``all`` as identifier to wipe every accurate.company on this DB.
    """
    actx: AppContext = ctx.obj
    out = actx.output
    c = actx.client
    if not dry_run and not confirm:
        raise typer.BadParameter("Add --confirm to proceed with --execute (destructive).")

    if identifier == "all":
        targets = c.search_read("accurate.company", domain=[], fields=["id", "slug"], order="id")
    else:
        rec = _resolve_accurate_company(c, identifier)
        targets = [{"id": rec["id"], "slug": rec["slug"]}]

    totals = {"moves": 0, "move_lines": 0, "partial_reconcile": 0, "account_payment": 0, "ir_model_data": 0}
    rows = []
    for t in targets:
        res = c.execute_kw(
            "accurate.company",
            "action_wipe_migrated_transactions",
            [[t["id"]]],
            {"dry_run": dry_run},
        )
        rows.append(
            [
                t["slug"],
                str(res["moves"]),
                str(res["move_lines"]),
                str(res["partial_reconcile"]),
                str(res["account_payment"]),
                str(res["ir_model_data"]),
                f"{res.get('elapsed_seconds', 0)}s" if not dry_run else "-",
            ]
        )
        for k in totals:
            totals[k] += res[k]
    header = "Wipe DRY-RUN preview" if dry_run else "Wipe EXECUTED"
    out.table(
        f"{header} ({len(targets)} tenant{'s' if len(targets) != 1 else ''})",
        [("Tenant", ""), ("Moves", ""), ("Lines", ""), ("Recon", ""), ("Pays", ""), ("IMD", ""), ("Elapsed", "dim")],
        rows,
    )
    out.kv("Total moves", str(totals["moves"]))
    out.kv("Total ext-ids", str(totals["ir_model_data"]))
    if dry_run:
        out.warn("DRY-RUN — nothing deleted. Add --execute --confirm to run.")
    else:
        out.success(f"Wiped {totals['moves']} moves across {len(targets)} tenant(s).")


@app.command("wipe-validate")
def wipe_validate(
    ctx: typer.Context,
    identifier: Annotated[str, typer.Argument(help="Slug or ID, or 'all'")],
) -> None:
    """Validate a post-wipe tenant — confirm transactions are truly clean.

    Checks: no migrated moves remain, no transaction-module ext-ids
    remain, no orphan move-lines, no orphan partial_reconcile, no
    orphan account.payment. Also reports foundation + config
    preservation counts (positive proof masters survived).

    Exits non-zero if ANY check flags an issue.
    """
    actx: AppContext = ctx.obj
    out = actx.output
    c = actx.client

    if identifier == "all":
        targets = c.search_read("accurate.company", domain=[], fields=["id", "slug"], order="id")
    else:
        rec = _resolve_accurate_company(c, identifier)
        targets = [{"id": rec["id"], "slug": rec["slug"]}]

    any_dirty = False
    for t in targets:
        res = c.execute_kw("accurate.company", "action_validate_wipe_clean", [[t["id"]]])
        status = "✓ CLEAN" if res["clean"] else "✗ DIRTY"
        out.kv(t["slug"], status)
        if not res["clean"]:
            any_dirty = True
            for issue in res["issues"]:
                out.error(f"  - {issue}")
        fp = res["foundation_preserved"]
        out.kv(
            f"  foundation",
            f"customers={fp.get('customer', 0)} vendors={fp.get('vendor', 0)} items={fp.get('item', 0)} glaccounts={fp.get('glaccount', 0)} taxes={fp.get('tax', 0)}",
        )
        cp = res["config_preserved"]
        out.kv(
            f"  config",
            f"company={cp.get('accurate.company', 0)} account_mapping={cp.get('accurate.account.mapping', 0)} parity_reports={cp.get('accurate.parity.report', 0)}",
        )
    if any_dirty:
        out.error("One or more tenants FAILED validation.")
        raise typer.Exit(1)
    out.success("All tenants validated clean.")


# --- synchronous pipeline helpers (18.0.15.2.0) ------------------------


@app.command("gap-sync")
def gap_sync(
    ctx: typer.Context,
    identifier: Annotated[str, typer.Argument(help="Slug or ID, or 'all'")],
    modules: Annotated[
        Optional[str],
        typer.Option(
            "--modules",
            help="Comma-separated list (default = every transaction module)",
        ),
    ] = None,
    dry_run: Annotated[bool, typer.Option("--dry-run/--execute")] = True,
) -> None:
    """Synchronous gap-sync — fetch every missing Accurate record.

    Surgical complement to ``transactions_current`` which queues async
    queue_job work. This helper runs inline + commits per-tenant, so
    records actually land before the CLI returns.

    Uses ``accurate.company.action_sync_missing_records``.
    """
    actx: AppContext = ctx.obj
    out = actx.output
    c = actx.client
    _bump_client_timeout(actx)

    if identifier == "all":
        targets = c.search_read("accurate.company", domain=[], fields=["id", "slug"], order="id")
    else:
        rec = _resolve_accurate_company(c, identifier)
        targets = [{"id": rec["id"], "slug": rec["slug"]}]

    mods = [m.strip() for m in modules.split(",")] if modules else None

    for t in targets:
        kwargs = {"dry_run": dry_run}
        if mods:
            kwargs["modules"] = mods
        res = c.execute_kw("accurate.company", "action_sync_missing_records", [[t["id"]]], kwargs)
        rows = []
        for mod, v in res.items():
            if (v.get("missing") or 0) > 0 or (v.get("mapped_ok") or 0) > 0:
                rows.append(
                    [
                        mod,
                        str(v.get("missing", 0)),
                        str(v.get("mapped_ok", 0)),
                        str(v.get("mapped_skipped", 0)),
                        str(len(v.get("errors") or [])),
                    ]
                )
        if rows:
            out.table(
                f"{t['slug']} — gap-sync {'DRY-RUN' if dry_run else 'EXECUTED'}",
                [("Module", ""), ("Missing", ""), ("OK", ""), ("Skipped", ""), ("Errors", "")],
                rows,
            )
        else:
            out.info(f"{t['slug']}: nothing to sync.")


@app.command("post-drafts")
def post_drafts(
    ctx: typer.Context,
    identifier: Annotated[str, typer.Argument(help="Slug or ID, or 'all'")],
    dry_run: Annotated[bool, typer.Option("--dry-run/--execute")] = True,
) -> None:
    """Post every draft account.move/payment tagged by the migration.

    Wraps ``accurate.company.action_post_draft_imports``.
    """
    actx: AppContext = ctx.obj
    out = actx.output
    c = actx.client
    _bump_client_timeout(actx)

    if identifier == "all":
        targets = c.search_read("accurate.company", domain=[], fields=["id", "slug"], order="id")
    else:
        rec = _resolve_accurate_company(c, identifier)
        targets = [{"id": rec["id"], "slug": rec["slug"]}]

    for t in targets:
        res = c.execute_kw(
            "accurate.company",
            "action_post_draft_imports",
            [[t["id"]]],
            {"dry_run": dry_run},
        )
        out.kv(
            t["slug"],
            f"candidates={res['move_candidates']} posted={res['moves_posted']} "
            f"failed={len(res['moves_failed'])} payments_posted={res['payments_posted']}",
        )
        for f in res.get("moves_failed", [])[:5]:
            out.warn(f"  FAIL {f.get('move_id')}: {(f.get('error', '') or '')[:150]}")


@app.command("reconcile")
def reconcile(
    ctx: typer.Context,
    identifier: Annotated[str, typer.Argument(help="Slug or ID, or 'all'")],
    dry_run: Annotated[bool, typer.Option("--dry-run/--execute")] = True,
) -> None:
    """Reconcile AR/AP lines across migrated payment + invoice moves.

    Wraps ``accurate.company.action_reconcile_migrated_move_lines``.
    """
    actx: AppContext = ctx.obj
    out = actx.output
    c = actx.client
    _bump_client_timeout(actx)

    if identifier == "all":
        targets = c.search_read("accurate.company", domain=[], fields=["id", "slug"], order="id")
    else:
        rec = _resolve_accurate_company(c, identifier)
        targets = [{"id": rec["id"], "slug": rec["slug"]}]

    for t in targets:
        res = c.execute_kw(
            "accurate.company",
            "action_reconcile_migrated_move_lines",
            [[t["id"]]],
            {"dry_run": dry_run},
        )
        out.kv(
            t["slug"],
            f"groups={res.get('groups_found', 0)} reconciled={res.get('groups_reconciled', 0)} "
            f"skipped={len(res.get('groups_skipped') or [])}",
        )


@app.command("parity")
def parity(
    ctx: typer.Context,
    identifier: Annotated[str, typer.Argument(help="Slug or ID, or 'all'")],
    as_of: Annotated[str, typer.Option("--as-of", help="YYYY-MM-DD")] = "2026-03-31",
) -> None:
    """Run the 12-check parity framework and print BS/PL/TB delta.

    Wraps ``accurate.company.action_run_parity``. Writes a new
    ``accurate.parity.report``. Exits non-zero if any key check fails.
    """
    actx: AppContext = ctx.obj
    out = actx.output
    c = actx.client
    _bump_client_timeout(actx)

    if identifier == "all":
        targets = c.search_read("accurate.company", domain=[], fields=["id", "slug"], order="id")
    else:
        rec = _resolve_accurate_company(c, identifier)
        targets = [{"id": rec["id"], "slug": rec["slug"]}]

    any_fail = False
    for t in targets:
        c.execute_kw(
            "accurate.company",
            "action_run_parity",
            [[t["id"]]],
            {"as_of": as_of},
        )
        reps = c.search_read(
            "accurate.parity.report",
            domain=[("company_id", "=", t["id"])],
            fields=["id"],
            order="id desc",
            limit=1,
        )
        if not reps:
            out.error(f"{t['slug']}: no parity report emitted")
            any_fail = True
            continue
        rep_id = reps[0]["id"]
        lines = c.search_read(
            "accurate.parity.line",
            domain=[
                ("report_id", "=", rep_id),
                ("check_name", "in", ("trial_balance", "bs", "pnl", "ar_aging", "ap_aging")),
            ],
            fields=["check_name", "accurate_total", "odoo_total", "passed"],
        )
        rows = []
        for ln in lines:
            A = ln.get("accurate_total") or 0
            O = ln.get("odoo_total") or 0
            ratio = f"{(O / A):.2f}x" if A else "n/a"
            rows.append(
                [
                    ln["check_name"],
                    f"{A:,.0f}",
                    f"{O:,.0f}",
                    f"{O - A:+,.0f}",
                    ratio,
                    "✓" if ln["passed"] else "✗",
                ]
            )
            if not ln["passed"]:
                any_fail = True
        out.table(
            f"{t['slug']} parity (as_of={as_of}, report={rep_id})",
            [("Check", ""), ("Accurate", ""), ("Odoo", ""), ("Δ", ""), ("Ratio", ""), ("Pass", "")],
            rows,
        )

    if any_fail:
        raise typer.Exit(1)


@app.command("full-remediate")
def full_remediate(
    ctx: typer.Context,
    identifier: Annotated[str, typer.Argument(help="Slug or ID, or 'all'")],
    as_of: Annotated[str, typer.Option("--as-of")] = "2026-03-31",
    skip_wipe: Annotated[bool, typer.Option("--skip-wipe", help="Skip wipe — start from current state")] = False,
    confirm: Annotated[bool, typer.Option("--confirm", help="Required (destructive)")] = False,
) -> None:
    """End-to-end clean-slate remediation: wipe → gap-sync → post → reconcile → parity.

    This is the one-shot operator command. Per tenant, runs the whole
    sequence with per-step commits so container restarts don't lose
    progress. Uses the fixed SDK 0.5.2 mappers (DP-entry pivot + tax_ids
    explicit empty list).
    """
    if not confirm:
        raise typer.BadParameter("Add --confirm to proceed (destructive wipe).")

    actx: AppContext = ctx.obj
    out = actx.output
    c = actx.client
    _bump_client_timeout(actx, seconds=3600.0)  # 1 hour — full pipeline

    if identifier == "all":
        targets = c.search_read("accurate.company", domain=[], fields=["id", "slug"], order="id")
    else:
        rec = _resolve_accurate_company(c, identifier)
        targets = [{"id": rec["id"], "slug": rec["slug"]}]

    for t in targets:
        out.info(f"=== {t['slug']} ===")
        if not skip_wipe:
            r = c.execute_kw(
                "accurate.company",
                "action_wipe_migrated_transactions",
                [[t["id"]]],
                {"dry_run": False},
            )
            out.kv("wipe", f"moves={r['moves']} imd={r['ir_model_data']} ({r['elapsed_seconds']}s)")
            v = c.execute_kw("accurate.company", "action_validate_wipe_clean", [[t["id"]]])
            if not v["clean"]:
                out.error(f"wipe-validate FAILED: {v['issues']}")
                raise typer.Exit(1)
        # gap-sync per-module — each module is a separate RPC call so
        # Odoo's statement_timeout (10 min) can't guillotine the big
        # tenants mid-flight. Larger tenants (SPS 1700+, GDA 1900+
        # records) would exceed the cursor limit on a single-call sync.
        _GAP_MODULES = (
            "sales-invoice",
            "sales-receipt",
            "purchase-invoice",
            "purchase-payment",
            "bank-receipt",
            "bank-transfer",
            "other-payment",
            "other-deposit",
            "expense",
            "journal-voucher",
        )
        total_ok = 0
        total_err = 0
        for mod in _GAP_MODULES:
            try:
                r2 = c.execute_kw(
                    "accurate.company",
                    "action_sync_missing_records",
                    [[t["id"]]],
                    {"dry_run": False, "modules": [mod]},
                )
                s = r2.get(mod, {})
                total_ok += s.get("mapped_ok", 0) or 0
                total_err += len(s.get("errors") or [])
                if (s.get("missing") or 0) > 0 or (s.get("mapped_ok") or 0) > 0:
                    out.info(
                        f"  {mod}: missing={s.get('missing')} ok={s.get('mapped_ok')} "
                        f"skipped={s.get('mapped_skipped')} errors={len(s.get('errors') or [])}"
                    )
            except Exception as exc:  # noqa: BLE001
                out.warn(f"  {mod}: {str(exc)[:180]}")
        out.kv("gap-sync", f"mapped_ok={total_ok} errors={total_err}")
        # post
        r3 = c.execute_kw(
            "accurate.company",
            "action_post_draft_imports",
            [[t["id"]]],
            {"dry_run": False},
        )
        out.kv(
            "post",
            f"posted={r3['moves_posted']}/{r3['move_candidates']} failed={len(r3['moves_failed'])}",
        )
        # reconcile
        r4 = c.execute_kw(
            "accurate.company",
            "action_reconcile_migrated_move_lines",
            [[t["id"]]],
            {"dry_run": False},
        )
        out.kv(
            "reconcile",
            f"{r4.get('groups_reconciled', 0)}/{r4.get('groups_found', 0)} groups",
        )
        # parity
        c.execute_kw(
            "accurate.company",
            "action_run_parity",
            [[t["id"]]],
            {"as_of": as_of},
        )
        reps = c.search_read(
            "accurate.parity.report",
            domain=[("company_id", "=", t["id"])],
            fields=["id"],
            order="id desc",
            limit=1,
        )
        if reps:
            lines = c.search_read(
                "accurate.parity.line",
                domain=[
                    ("report_id", "=", reps[0]["id"]),
                    ("check_name", "in", ("trial_balance", "bs", "pnl")),
                ],
                fields=["check_name", "accurate_total", "odoo_total", "passed"],
            )
            for ln in lines:
                A = ln.get("accurate_total") or 0
                O = ln.get("odoo_total") or 0
                ratio = f"{(O / A):.2f}x" if A else "n/a"
                status = "✓" if ln["passed"] else "✗"
                out.kv(
                    f"  {ln['check_name']}",
                    f"A={A:,.0f} O={O:,.0f} Δ={O - A:+,.0f} {ratio} {status}",
                )


# --- errors ------------------------------------------------------------


errors_app = typer.Typer(help="Migration-error inspection and bulk retry.")


@errors_app.command("list")
def errors_list(
    ctx: typer.Context,
    identifier: Annotated[str, typer.Argument(help="Company slug or ID")],
    error_kind: Annotated[
        str | None,
        typer.Option("--error-kind", help="Filter by error_kind"),
    ] = None,
    error_class: Annotated[
        str | None,
        typer.Option(
            "--error-class",
            help="transient|rate_limited|data_integrity|config_broken|unknown",
        ),
    ] = None,
    retry_status: Annotated[
        str | None,
        typer.Option("--retry-status", help="pending|ignored|retried|succeeded|abandoned"),
    ] = None,
    limit: Annotated[int, typer.Option("--limit", "-l")] = 50,
) -> None:
    """List migration errors for a company."""
    actx: AppContext = ctx.obj
    out = actx.output
    c = actx.client
    rec = _resolve_accurate_company(c, identifier)
    domain = [("company_id", "=", rec["id"])]
    if error_kind:
        domain.append(("error_kind", "=", error_kind))
    if error_class:
        domain.append(("error_class", "=", error_class))
    if retry_status:
        domain.append(("retry_status", "=", retry_status))
    errors = c.search_read(
        "accurate.migration.error",
        domain=domain,
        fields=[
            "id",
            "module",
            "accurate_id",
            "error_class",
            "error_kind",
            "error_message",
            "retry_status",
            "retry_count",
            "create_date",
        ],
        limit=limit,
        order="create_date desc",
    )
    rows = [
        [
            str(e["id"]),
            e["module"],
            str(e.get("accurate_id") or "-"),
            e["error_class"],
            e.get("error_kind") or "-",
            (e.get("error_message") or "")[:40],
            e["retry_status"],
            str(e.get("retry_count") or 0),
        ]
        for e in errors
    ]
    out.table(
        f"Migration Errors — {rec['slug']} ({len(errors)})",
        [
            ("ID", "cyan"),
            ("Module", ""),
            ("Accurate ID", "dim"),
            ("Class", ""),
            ("Kind", "dim"),
            ("Message", "dim"),
            ("Retry Status", ""),
            ("Retries", "dim"),
        ],
        rows,
        data_for_json=errors,
    )


@errors_app.command("retry")
def errors_retry(
    ctx: typer.Context,
    identifier: Annotated[str, typer.Argument(help="Company slug or ID")],
    ids: Annotated[
        str | None,
        typer.Option("--ids", help="Comma-separated error IDs"),
    ] = None,
    error_kind: Annotated[
        str | None,
        typer.Option("--error-kind", help="Retry all errors matching this kind"),
    ] = None,
) -> None:
    """Retry selected errors (by IDs or filtered by error_kind)."""
    actx: AppContext = ctx.obj
    out = actx.output
    c = actx.client
    rec = _resolve_accurate_company(c, identifier)
    domain = [("company_id", "=", rec["id"]), ("retry_status", "=", "pending")]
    if ids:
        id_list = [int(x) for x in ids.split(",") if x.strip()]
        domain.append(("id", "in", id_list))
    elif error_kind:
        domain.append(("error_kind", "=", error_kind))
    else:
        out.error("Provide either --ids or --error-kind to select errors to retry.")
        raise typer.Exit(2)
    error_ids = c.search("accurate.migration.error", domain=domain, limit=500)
    if not error_ids:
        out.warn("No matching errors found.")
        return
    c.execute_kw(
        "accurate.migration.error",
        "action_retry_selected",
        [error_ids],
    )
    out.success(f"Queued retry of {len(error_ids)} error(s) for {rec['slug']}.")


@errors_app.command("ignore")
def errors_ignore(
    ctx: typer.Context,
    identifier: Annotated[str, typer.Argument(help="Company slug or ID")],
    ids: Annotated[str, typer.Option("--ids", help="Comma-separated error IDs")],
    reason: Annotated[str, typer.Option("--reason", help="Ignore reason (audit trail)")],
) -> None:
    """Mark selected errors as ignored (exclude from phase-failure threshold)."""
    actx: AppContext = ctx.obj
    out = actx.output
    c = actx.client
    rec = _resolve_accurate_company(c, identifier)
    id_list = [int(x) for x in ids.split(",") if x.strip()]
    if not id_list:
        out.error("Empty --ids.")
        raise typer.Exit(2)
    # Verify all belong to this company
    belong = c.search(
        "accurate.migration.error",
        domain=[("id", "in", id_list), ("company_id", "=", rec["id"])],
    )
    if len(belong) != len(id_list):
        out.error(f"Some IDs do not belong to {rec['slug']} (found {len(belong)} of {len(id_list)}).")
        raise typer.Exit(2)
    c.write("accurate.migration.error", id_list, {"retry_status": "ignored", "ignore_reason": reason})
    out.success(f"Marked {len(id_list)} error(s) as ignored for {rec['slug']}.")


# --- parity ------------------------------------------------------------


_PARITY_REPORT_FIELDS = [
    "id",
    "as_of_date",
    "run_date",
    "all_passed",
    "passed_count",
    "total_count",
    "company_id",
]

_PARITY_LINE_FIELDS = [
    "id",
    "check_name",
    "category",
    "passed",
    "accurate_total",
    "odoo_total",
    "delta",
    "deltas_json",
    "details_json",
    "error",
    "remediation_hints",
]


def _fetch_parity_report(client, report_id: int) -> tuple[dict, list[dict]]:
    """Load report header + lines by id."""
    reports = client.read("accurate.parity.report", [report_id], _PARITY_REPORT_FIELDS)
    if not reports:
        raise typer.BadParameter(f"accurate.parity.report id={report_id} not found")
    lines = client.search_read(
        "accurate.parity.line",
        domain=[("report_id", "=", report_id)],
        fields=_PARITY_LINE_FIELDS,
        order="category,check_name",
    )
    return reports[0], lines


def _format_parity_markdown(rec: dict, report: dict, lines: list[dict]) -> str:
    """Pretty markdown rendering of one parity report."""

    def _sym(line: dict) -> str:
        if line.get("error"):
            return "!"
        return "OK" if line.get("passed") else "FAIL"

    def _fmt_money(v) -> str:
        try:
            return f"{float(v or 0):,.2f}"
        except (TypeError, ValueError):
            return str(v or "-")

    header = [
        f"# Parity Report — {rec['slug']} ({rec['name']})",
        "",
        f"- **As of**: {report.get('as_of_date')}",
        f"- **Run at**: {report.get('run_date')}",
        f"- **Passed**: {report.get('passed_count')} / {report.get('total_count')}",
        f"- **Overall**: {'PASSED' if report.get('all_passed') else 'FAILED'}",
        "",
        "| Status | Check | Category | Accurate | Odoo | Delta | Notes |",
        "|---|---|---|---:|---:|---:|---|",
    ]
    acc_sum = 0.0
    odoo_sum = 0.0
    for line in lines:
        note = ""
        if line.get("error"):
            note = f"ERROR: {str(line['error'])[:80]}"
        elif line.get("remediation_hints"):
            note = str(line["remediation_hints"]).splitlines()[0][:80]
        header.append(
            "| {sym} | {check} | {cat} | {acc} | {odoo} | {delta} | {note} |".format(
                sym=_sym(line),
                check=line["check_name"],
                cat=line.get("category") or "-",
                acc=_fmt_money(line.get("accurate_total")),
                odoo=_fmt_money(line.get("odoo_total")),
                delta=_fmt_money(line.get("delta")),
                note=note.replace("|", "\\|"),
            )
        )
        try:
            acc_sum += float(line.get("accurate_total") or 0)
            odoo_sum += float(line.get("odoo_total") or 0)
        except (TypeError, ValueError):
            pass
    header.append(
        "| | **Totals** | | **{acc}** | **{odoo}** | **{delta}** | |".format(
            acc=f"{acc_sum:,.2f}",
            odoo=f"{odoo_sum:,.2f}",
            delta=f"{odoo_sum - acc_sum:,.2f}",
        )
    )
    return "\n".join(header) + "\n"


def _run_parity_for_company(
    client,
    rec: dict,
    as_of: str | None,
    only: list[str] | None,
    tolerance: float,
) -> tuple[dict, list[dict]]:
    """Invoke action_run_parity and load back the persisted report."""
    kwargs: dict = {"tolerance": tolerance}
    kwargs["as_of"] = as_of or date.today().isoformat()
    if only:
        kwargs["only"] = list(only)

    action = client.execute_kw(
        "accurate.company",
        "action_run_parity",
        [[rec["id"]]],
        kwargs,
    )
    report_id = None
    if isinstance(action, dict):
        report_id = action.get("res_id")
    elif isinstance(action, int):
        report_id = action
    if not isinstance(report_id, int):
        # Fallback — grab most recent report for this company
        recent = client.search_read(
            "accurate.parity.report",
            domain=[("company_id", "=", rec["id"])],
            fields=["id"],
            limit=1,
            order="run_date desc",
        )
        if not recent:
            raise typer.Exit(1)
        report_id = recent[0]["id"]
    return _fetch_parity_report(client, report_id)


@app.command("parity")
def parity(
    ctx: typer.Context,
    slug: Annotated[str, typer.Argument(help="accurate.company slug (e.g. sumber-pangan)")],
    check: Annotated[
        Optional[list[str]],
        typer.Option("--check", "-c", help="Only run named check(s). Repeat for multiple."),
    ] = None,
    as_of: Annotated[
        Optional[str],
        typer.Option("--date", "-d", help="As-of date YYYY-MM-DD (default today)"),
    ] = None,
    tolerance: Annotated[
        float,
        typer.Option("--tolerance", help="IDR tolerance for deltas"),
    ] = 1.0,
    output: Annotated[
        str,
        typer.Option("--output", "-o", help="markdown | json | xlsx"),
    ] = "markdown",
    output_file: Annotated[
        Optional[Path],
        typer.Option("--output-file", help="Write output here instead of stdout"),
    ] = None,
) -> None:
    """Run parity suite against Accurate for one tenant + format results."""
    actx: AppContext = ctx.obj
    _bump_client_timeout(actx, seconds=600.0)
    out = actx.output
    c = actx.client
    rec = _resolve_accurate_company(c, slug)
    out.info(f"Running parity suite for {rec['slug']}...")
    try:
        report, lines = _run_parity_for_company(c, rec, as_of, check, tolerance)
    except Exception as exc:  # noqa: BLE001
        out.error(f"Parity run failed: {exc}")
        raise typer.Exit(1)

    fmt = (output or "markdown").lower()
    if fmt == "json":
        payload = json.dumps(
            {
                "company": {"id": rec["id"], "slug": rec["slug"], "name": rec["name"]},
                "report": report,
                "lines": lines,
            },
            indent=2,
            default=str,
        )
        if output_file:
            output_file.write_text(payload)
            out.success(f"JSON written to {output_file}")
        else:
            typer.echo(payload)
    elif fmt == "xlsx":
        out.warn("xlsx output not yet implemented (Task 13). Falling back to markdown.")
        md = _format_parity_markdown(rec, report, lines)
        if output_file:
            output_file.write_text(md)
            out.success(f"Markdown written to {output_file}")
        else:
            typer.echo(md)
    else:
        md = _format_parity_markdown(rec, report, lines)
        if output_file:
            output_file.write_text(md)
            out.success(f"Markdown written to {output_file}")
        else:
            typer.echo(md)

    if not report.get("all_passed"):
        failed = report.get("total_count", 0) - report.get("passed_count", 0)
        out.error(f"{failed} parity check(s) FAILED for {rec['slug']}.")
        raise typer.Exit(1)
    out.success(f"All {report.get('total_count', 0)} parity checks PASSED for {rec['slug']}.")


@app.command("parity-all")
def parity_all(
    ctx: typer.Context,
    as_of: Annotated[
        Optional[str],
        typer.Option("--date", "-d", help="As-of date YYYY-MM-DD (default today)"),
    ] = None,
    tolerance: Annotated[
        float,
        typer.Option("--tolerance", help="IDR tolerance for deltas"),
    ] = 1.0,
    output_file: Annotated[
        Optional[Path],
        typer.Option("--output-file", help="Write consolidated markdown here instead of stdout"),
    ] = None,
) -> None:
    """Run parity suite across every accurate.company; emit consolidated markdown."""
    actx: AppContext = ctx.obj
    _bump_client_timeout(actx, seconds=600.0)
    out = actx.output
    c = actx.client

    companies = c.search_read(
        "accurate.company",
        domain=[],
        fields=["id", "slug", "name", "state"],
        order="name",
    )
    if not companies:
        out.warn("No accurate.company records found.")
        return

    chunks: list[str] = []
    effective_date = as_of or date.today().isoformat()
    chunks.append(f"# Consolidated Parity Scorecard — {effective_date}\n")
    scorecard_rows: list[tuple[str, str, int, int, bool]] = []
    any_failed = False

    for rec in companies:
        out.info(f"→ {rec['slug']}")
        try:
            report, lines = _run_parity_for_company(c, rec, as_of, None, tolerance)
        except Exception as exc:  # noqa: BLE001
            out.error(f"{rec['slug']}: parity run FAILED — {exc}")
            any_failed = True
            chunks.append(f"## {rec['slug']} — ERROR\n\n`{exc}`\n")
            scorecard_rows.append((rec["slug"], rec["name"], 0, 0, False))
            continue
        passed = report.get("passed_count", 0)
        total = report.get("total_count", 0)
        all_passed = bool(report.get("all_passed"))
        if not all_passed:
            any_failed = True
        scorecard_rows.append((rec["slug"], rec["name"], passed, total, all_passed))
        chunks.append(_format_parity_markdown(rec, report, lines))
        chunks.append("")

    summary = [
        "## Summary",
        "",
        "| Slug | Name | Passed | Total | Overall |",
        "|---|---|---:|---:|---|",
    ]
    for slug, name, passed, total, all_passed in scorecard_rows:
        summary.append(f"| {slug} | {name} | {passed} | {total} | {'PASSED' if all_passed else 'FAILED'} |")
    full = "\n".join([*summary, "", *chunks])

    if output_file:
        output_file.write_text(full)
        out.success(f"Consolidated scorecard written to {output_file}")
    else:
        typer.echo(full)

    if any_failed:
        raise typer.Exit(1)
    out.success(f"All {len(companies)} tenant(s) passed every parity check.")


# --------------------------------------------------------------------------
# Phase 1 — snapshot + reconcile
# (spec: 2026-04-21-accurate-migration-cutoff-validation-design.md)
# --------------------------------------------------------------------------

snapshot_app = typer.Typer(help="Snapshot management — pull, list, export snapshots.")
reconcile_app = typer.Typer(help="Daily reconciliation commands.")


@snapshot_app.command("pull")
def snapshot_pull(
    ctx: typer.Context,
    slug: Annotated[str, typer.Argument(help="Accurate company slug")],
    source: Annotated[
        str,
        typer.Option(
            "--source",
            help="Which side(s) to snapshot: accurate, odoo, or both.",
        ),
    ] = "both",
    snapshot_date: Annotated[
        Optional[str],
        typer.Option("--date", "-d", help="Snapshot date YYYY-MM-DD (default today)."),
    ] = None,
) -> None:
    """Pull a snapshot on demand for <slug>."""
    actx: AppContext = ctx.obj
    _bump_client_timeout(actx, seconds=1200.0)

    date_str = snapshot_date or date.today().isoformat()
    sources = ["accurate", "odoo"] if source == "both" else [source]
    for src in sources:
        result = actx.client.execute_kw(
            "accurate.company",
            "cli_snapshot_pull",
            [[("slug", "=", slug)], src, date_str],
        )
        typer.echo(
            f"{src} snapshot created: id={result['snapshot_id']} lines={result['line_count']} state={result['state']}"
        )


@snapshot_app.command("list")
def snapshot_list(
    ctx: typer.Context,
    slug: Annotated[str, typer.Argument(help="Accurate company slug")],
    limit: Annotated[int, typer.Option("--limit", "-l")] = 30,
) -> None:
    """List recent snapshots for <slug>."""
    actx: AppContext = ctx.obj
    rows = actx.client.execute_kw(
        "accurate.snapshot",
        "cli_list_for_slug",
        [slug, limit],
    )
    typer.echo(json.dumps(rows, indent=2, default=str))


@snapshot_app.command("export")
def snapshot_export(
    ctx: typer.Context,
    snapshot_id: Annotated[int, typer.Argument(help="Snapshot record ID")],
    output: Annotated[
        Path,
        typer.Option("--output", "-o", help="Output file path (gzipped JSON)."),
    ],
) -> None:
    """Export snapshot <snapshot_id> to a gzipped JSON file."""
    actx: AppContext = ctx.obj
    import base64

    blob_b64 = actx.client.execute_kw(
        "accurate.snapshot",
        "cli_export_gzip",
        [snapshot_id],
    )
    output.write_bytes(base64.b64decode(blob_b64))
    typer.echo(f"Wrote {output}")


@reconcile_app.command("run")
def reconcile_run(
    ctx: typer.Context,
    slug: Annotated[str, typer.Argument(help="Accurate company slug")],
    validator: Annotated[
        Optional[list[str]],
        typer.Option("--validator", help="Validator name(s) to run. Default: all."),
    ] = None,
    run_date: Annotated[
        Optional[str],
        typer.Option("--date", "-d", help="Run date YYYY-MM-DD (default today)."),
    ] = None,
    tolerance: Annotated[
        float,
        typer.Option("--tolerance", help="Per-delta tolerance in account currency units."),
    ] = 1000.0,
) -> None:
    """Run reconciliation for <slug> with selected validators."""
    actx: AppContext = ctx.obj
    out = actx.output
    _bump_client_timeout(actx, seconds=1800.0)

    date_str = run_date or date.today().isoformat()
    result = actx.client.execute_kw(
        "accurate.company",
        "cli_reconcile_run",
        [
            [("slug", "=", slug)],
            validator or None,
            date_str,
            tolerance,
        ],
    )
    out.info(
        f"Report id={result['report_id']} state={result['state']} "
        f"green={result['green']} amber={result['amber']} red={result['red']}"
    )
    for line in result.get("lines", []):
        typer.echo(f"  {line['validator_name']:20s} {line['status']:6s} Δ={line['delta']:+,.0f}")
    if result.get("state") not in ("passed", "green"):
        raise typer.Exit(1)


# --- Phase 4 — variance + cutover + audit-package ----------------------


variance_app = typer.Typer(help="Variance explanations (Phase 4).")


@variance_app.command("list")
def variance_list(
    ctx: typer.Context,
    slug: Annotated[str, typer.Argument(help="Tenant slug")],
) -> None:
    """List variance explanations for <slug>."""
    actx: AppContext = ctx.obj
    rows = actx.client.execute_kw(
        "accurate.company",
        "cli_variance_list",
        [slug],
    )
    typer.echo(json.dumps(rows, indent=2))


@variance_app.command("add")
def variance_add(
    ctx: typer.Context,
    slug: Annotated[str, typer.Argument(help="Tenant slug")],
    validator: Annotated[str, typer.Option("--validator", "-v", help="Validator name")],
    delta_key: Annotated[str, typer.Option("--delta-key", "-k", help="Delta key (glob or exact)")],
    explanation: Annotated[str, typer.Option("--explanation", "-e", help="Explanation text")],
    category: Annotated[str, typer.Option("--category", "-c", help="Variance category")] = "cutover_timing",
    expires: Annotated[
        Optional[str],
        typer.Option("--expires", help="Expiry date YYYY-MM-DD; default today+30d"),
    ] = None,
) -> None:
    """Create a variance explanation."""
    actx: AppContext = ctx.obj
    result = actx.client.execute_kw(
        "accurate.company",
        "cli_variance_add",
        [slug, validator, delta_key, category, explanation, expires],
    )
    typer.echo(f"Created explanation id={result['id']} delta_key={result['delta_key']}")


@variance_app.command("expire")
def variance_expire(
    ctx: typer.Context,
    explanation_id: Annotated[int, typer.Argument(help="Explanation record ID")],
) -> None:
    """Expire (deactivate) a variance explanation."""
    actx: AppContext = ctx.obj
    result = actx.client.execute_kw(
        "accurate.company",
        "cli_variance_expire",
        [explanation_id],
    )
    typer.echo(f"Expired id={result['id']} active={result['active']}")


cutover_app = typer.Typer(help="Cutover readiness + execution (Phase 4).")


@cutover_app.command("status")
def cutover_status(ctx: typer.Context) -> None:
    """Streak status across all tenants."""
    actx: AppContext = ctx.obj
    rows = actx.client.execute_kw(
        "accurate.company",
        "cli_cutover_status",
        [],
    )
    typer.echo(f"{'slug':12s} {'streak':>8s} {'required':>10s} {'ready?':>8s}")
    typer.echo("-" * 40)
    for r in rows:
        ready = "yes" if r["ready"] else "no"
        typer.echo(f"{r['slug']:12s} {r['streak']:>8d} {r['required']:>10d} {ready:>8s}")


@app.command("audit-package")
def audit_package(
    ctx: typer.Context,
    slug: Annotated[str, typer.Argument(help="Tenant slug")],
    output: Annotated[str, typer.Option("-o", "--output", help="Output zip path")],
) -> None:
    """Generate the audit package zip for <slug>."""
    actx: AppContext = ctx.obj
    import base64

    _bump_client_timeout(actx, seconds=300.0)
    blob_b64 = actx.client.execute_kw(
        "accurate.company",
        "cli_audit_package",
        [[("slug", "=", slug)]],
    )
    with open(output, "wb") as f:
        f.write(base64.b64decode(blob_b64))
    typer.echo(f"Wrote {output}")


@app.command("report")
def migration_report(
    ctx: typer.Context,
    slug: Annotated[str, typer.Argument(help="Tenant slug (or 'all' for all companies)")],
    output: Annotated[
        Optional[Path],
        typer.Option("-o", "--output", help="Output Excel path"),
    ] = None,
) -> None:
    """Generate a migration report Excel for one or all companies."""
    actx: AppContext = ctx.obj
    _bump_client_timeout(actx, seconds=600.0)
    c = actx.client
    out = actx.output

    if slug == "all":
        companies = c.search_read(
            "accurate.company",
            domain=[],
            fields=["id", "slug"],
            order="id",
        )
    else:
        rec = _resolve_accurate_company(c, slug)
        companies = [rec]

    out.info(f"Generating migration report for {len(companies)} company(ies)...")

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        out.error("openpyxl not installed. Run: pip install openpyxl")
        raise typer.Exit(1)

    wb = Workbook()
    hdr_font = Font(bold=True, color="FFFFFF", size=11)
    hdr_fill = PatternFill("solid", fgColor="2F5496")
    pass_fill = PatternFill("solid", fgColor="C6EFCE")
    fail_fill = PatternFill("solid", fgColor="FFC7CE")
    inc_fill = PatternFill("solid", fgColor="FFEB9C")
    border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )

    def write_headers(ws, headers, widths=None):
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = hdr_font
            cell.fill = hdr_fill
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
        if widths:
            for i, w in enumerate(widths, 1):
                ws.column_dimensions[get_column_letter(i)].width = w
        ws.freeze_panes = "A2"

    # ── Sheet 1: Summary ──
    ws1 = wb.active
    ws1.title = "Migration Summary"
    write_headers(
        ws1,
        [
            "Company",
            "Slug",
            "State",
            "Phase",
            "Cutover Date",
            "Total Moves",
            "Opening JE",
            "BS Adjustment",
            "Parity Passed",
            "Parity Total",
            "Parity %",
        ],
        [30, 35, 12, 22, 14, 12, 10, 12, 12, 12, 10],
    )

    row = 2
    all_parity_data = {}
    for comp in companies:
        rec = c.search_read(
            "accurate.company",
            domain=[("id", "=", comp["id"])],
            fields=[
                "name",
                "slug",
                "state",
                "current_phase",
                "cutover_date",
                "odoo_company_id",
            ],
        )[0]
        co_id = rec["odoo_company_id"][0] if rec["odoo_company_id"] else 0

        moves = c.search_count(
            "account.move",
            [
                ("company_id", "=", co_id),
                ("state", "=", "posted"),
            ],
        )
        opening = c.search_count(
            "account.move",
            [
                ("company_id", "=", co_id),
                ("x_accurate_migration_source", "=", "opening_gl_trial_balance"),
            ],
        )
        bs_adj = c.search_count(
            "account.move",
            [
                ("company_id", "=", co_id),
                ("x_accurate_migration_source", "=", "bs_adjustment"),
            ],
        )

        # Latest parity report
        parity_reports = c.search_read(
            "accurate.parity.report",
            domain=[("company_id", "=", comp["id"])],
            fields=["passed_count", "total_count", "run_date", "all_passed"],
            order="id desc",
            limit=1,
        )
        p_passed = parity_reports[0]["passed_count"] if parity_reports else 0
        p_total = parity_reports[0]["total_count"] if parity_reports else 0
        p_pct = f"{p_passed * 100 // p_total}%" if p_total else "N/A"

        ws1.cell(row=row, column=1, value=rec["name"]).font = Font(size=10)
        ws1.cell(row=row, column=2, value=rec["slug"]).font = Font(size=10)
        ws1.cell(row=row, column=3, value=rec["state"]).font = Font(size=10)
        ws1.cell(row=row, column=4, value=rec["current_phase"]).font = Font(size=10)
        ws1.cell(row=row, column=5, value=rec["cutover_date"]).font = Font(size=10)
        ws1.cell(row=row, column=6, value=moves).font = Font(size=10)
        ws1.cell(row=row, column=7, value="Yes" if opening else "No").font = Font(size=10, bold=True)
        ws1.cell(row=row, column=7).fill = pass_fill if opening else fail_fill
        ws1.cell(row=row, column=8, value="Yes" if bs_adj else "No").font = Font(size=10, bold=True)
        ws1.cell(row=row, column=8).fill = pass_fill if bs_adj else fail_fill
        ws1.cell(row=row, column=9, value=p_passed).font = Font(size=10)
        ws1.cell(row=row, column=10, value=p_total).font = Font(size=10)
        ws1.cell(row=row, column=11, value=p_pct).font = Font(size=10, bold=True)

        for col in range(1, 12):
            ws1.cell(row=row, column=col).border = border
        row += 1

        all_parity_data[rec["slug"]] = parity_reports[0] if parity_reports else None
        out.info(f"  {rec['slug']}: {moves} moves, parity {p_passed}/{p_total}")

    # ── Sheet 2: Parity Detail ──
    ws2 = wb.create_sheet("Parity Detail")
    write_headers(
        ws2,
        [
            "Company",
            "Check",
            "Category",
            "Status",
            "Accurate",
            "Odoo",
            "Delta",
            "Source",
            "Notes",
        ],
        [30, 18, 12, 8, 18, 18, 18, 12, 50],
    )

    row = 2
    for comp in companies:
        reports = c.search_read(
            "accurate.parity.report",
            domain=[("company_id", "=", comp["id"])],
            fields=["id"],
            order="id desc",
            limit=1,
        )
        if not reports:
            continue
        lines = c.search_read(
            "accurate.parity.line",
            domain=[("report_id", "=", reports[0]["id"])],
            fields=[
                "check_name",
                "category",
                "status",
                "accurate_total",
                "odoo_total",
                "source",
                "remediation_hints",
            ],
            order="id",
        )
        rec_name = c.search_read(
            "accurate.company",
            [("id", "=", comp["id"])],
            fields=["name"],
        )[0]["name"]

        for ln in lines:
            delta = (ln.get("odoo_total") or 0) - (ln.get("accurate_total") or 0)
            status = ln.get("status") or "?"
            ws2.cell(row=row, column=1, value=rec_name).font = Font(size=10)
            ws2.cell(row=row, column=2, value=ln["check_name"]).font = Font(size=10)
            ws2.cell(row=row, column=3, value=ln.get("category") or "").font = Font(size=10)
            status_cell = ws2.cell(row=row, column=4, value=status.upper())
            status_cell.font = Font(size=10, bold=True)
            if status == "passed":
                status_cell.fill = pass_fill
            elif status == "failed":
                status_cell.fill = fail_fill
            else:
                status_cell.fill = inc_fill
            ws2.cell(row=row, column=5, value=ln.get("accurate_total") or 0).font = Font(size=10)
            ws2.cell(row=row, column=5).number_format = "#,##0.00"
            ws2.cell(row=row, column=6, value=ln.get("odoo_total") or 0).font = Font(size=10)
            ws2.cell(row=row, column=6).number_format = "#,##0.00"
            ws2.cell(row=row, column=7, value=delta).font = Font(size=10)
            ws2.cell(row=row, column=7).number_format = "#,##0.00"
            ws2.cell(row=row, column=8, value=ln.get("source") or "").font = Font(size=10)
            hints = ln.get("remediation_hints") or ""
            ws2.cell(row=row, column=9, value=hints[:200] if hints else "").font = Font(size=9)
            for col in range(1, 10):
                ws2.cell(row=row, column=col).border = border
            row += 1

    ws2.auto_filter.ref = f"A1:I{row - 1}"

    # ── Sheet 3: Move Breakdown ──
    ws3 = wb.create_sheet("Move Breakdown")
    write_headers(
        ws3,
        [
            "Company",
            "Source",
            "Move Count",
        ],
        [30, 30, 12],
    )

    row = 2
    source_labels = [
        ("opening_gl_trial_balance", "Opening GL Trial Balance"),
        ("bs_adjustment", "BS Adjustment"),
        ("current_fy_full", "Current FY Transactions"),
        ("gap_sync", "Gap Sync Backfill"),
        ("ar_ap_reversal", "AR/AP Reversal"),
        ("outstanding_open_items", "Outstanding Open Items"),
        ("pnl_neutralization", "P&L Neutralization"),
    ]
    for comp in companies:
        rec = c.search_read(
            "accurate.company",
            [("id", "=", comp["id"])],
            fields=["name", "odoo_company_id"],
        )[0]
        co_id = rec["odoo_company_id"][0] if rec["odoo_company_id"] else 0
        for src_code, src_label in source_labels:
            cnt = c.search_count(
                "account.move",
                [
                    ("company_id", "=", co_id),
                    ("x_accurate_migration_source", "=", src_code),
                    ("state", "=", "posted"),
                ],
            )
            if cnt:
                ws3.cell(row=row, column=1, value=rec["name"]).font = Font(size=10)
                ws3.cell(row=row, column=2, value=src_label).font = Font(size=10)
                ws3.cell(row=row, column=3, value=cnt).font = Font(size=10)
                for col in range(1, 4):
                    ws3.cell(row=row, column=col).border = border
                row += 1
        # Non-migration moves
        total = c.search_count(
            "account.move",
            [
                ("company_id", "=", co_id),
                ("state", "=", "posted"),
            ],
        )
        migration = c.search_count(
            "account.move",
            [
                ("company_id", "=", co_id),
                ("state", "=", "posted"),
                ("x_accurate_migration_source", "!=", False),
            ],
        )
        other = total - migration
        if other:
            ws3.cell(row=row, column=1, value=rec["name"]).font = Font(size=10)
            ws3.cell(row=row, column=2, value="Other (non-migration)").font = Font(size=10)
            ws3.cell(row=row, column=3, value=other).font = Font(size=10)
            for col in range(1, 4):
                ws3.cell(row=row, column=col).border = border
            row += 1

    # ── Sheet 4: Phase History ──
    ws4 = wb.create_sheet("Phase History")
    write_headers(
        ws4,
        [
            "Company",
            "Phase",
            "State",
            "Started",
            "Duration (s)",
        ],
        [30, 22, 10, 22, 12],
    )

    row = 2
    for comp in companies:
        rec = c.search_read(
            "accurate.company",
            [("id", "=", comp["id"])],
            fields=["name"],
        )[0]
        phases = c.search_read(
            "accurate.company.phase",
            domain=[("company_id", "=", comp["id"]), ("state", "=", "passed")],
            fields=["phase_code", "state", "started_at", "duration_s"],
            order="id desc",
            limit=20,
        )
        for p in phases:
            ws4.cell(row=row, column=1, value=rec["name"]).font = Font(size=10)
            ws4.cell(row=row, column=2, value=p.get("phase_code") or "").font = Font(size=10)
            state_cell = ws4.cell(row=row, column=3, value=p.get("state") or "")
            state_cell.font = Font(size=10)
            state_cell.fill = pass_fill if p.get("state") == "passed" else fail_fill
            ws4.cell(row=row, column=4, value=p.get("started_at") or "").font = Font(size=10)
            ws4.cell(row=row, column=5, value=p.get("duration_s") or 0).font = Font(size=10)
            ws4.cell(row=row, column=5).number_format = "#,##0"
            for col in range(1, 6):
                ws4.cell(row=row, column=col).border = border
            row += 1

    # ── Save ──
    if not output:
        safe_slug = slug.replace("/", "_").replace(" ", "_")
        output = Path(f"migration_report_{safe_slug}_{date.today().isoformat()}.xlsx")
    wb.save(str(output))
    out.info(f"Report saved to {output}")
    out.info(f"  Sheet 1: Migration Summary ({len(companies)} companies)")
    out.info(f"  Sheet 2: Parity Detail (12 checks per company)")
    out.info(f"  Sheet 3: Move Breakdown (by migration source)")
    out.info(f"  Sheet 4: Phase History")


# --- attach sub-apps ---------------------------------------------------


# Direct-Accurate reports group ("reports" plural to avoid colliding with
# the existing flat `accurate report` migration-report command above).
from kctl_odoo.commands._credit_limit.accurate_command import (  # noqa: E402
    credit_limit_report,
)

reports_app = typer.Typer(help="Direct-Accurate pilot reports (no Odoo migration required).")
reports_app.command("credit-limit")(credit_limit_report)


app.add_typer(companies_app, name="companies")
app.add_typer(errors_app, name="errors")
app.add_typer(snapshot_app, name="snapshot")
app.add_typer(reconcile_app, name="reconcile")
app.add_typer(variance_app, name="variance")
app.add_typer(cutover_app, name="cutover")
app.add_typer(reports_app, name="reports")
